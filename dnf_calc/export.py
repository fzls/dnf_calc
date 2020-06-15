#!/usr/bin/env python
# -*- coding: utf-8 -*-
# -------------------------------
# File      : export
# DateTime  : 2020/6/7 0007 8:32
# Author    : Chen Ji
# Email     : fzls.zju@gmail.com
# -------------------------------
import threading
import tkinter.messagebox
from typing import List, Dict, Tuple, Protocol

from openpyxl import Workbook

from .config import *
from .equipment import *
from .export_def import *


def export_result(ele_skill, col_names, job_name, cool, equip_index_to_realname, custom_buf_data, extract_rank_cols_func, rankings):
    """
    # 导出结果到excel
    @type ele_skill: int
    @type col_names: List[str]
    @type job_name: str
    @type cool: str
    @type equip_index_to_realname: Dict[str, str]
    @type custom_buf_data: Dict[str, int]
    @type extract_rank_cols_func: ExtractRankColsFunc
    @type rankings: List[Tuple[str, Any]]
    @return: None
    """
    export_config = config().export_result_as_excel
    if not export_config.enable:
        return

    def _export_reuslt():
        book = Workbook()
        book.remove(book.active)
        for sheet_index, rt in enumerate(rankings):
            ranking_name = rt[0]

            # 为该排行榜创建sheet
            sheet = book.create_sheet(ranking_name, sheet_index)

            row = 1
            # 首行为列名
            for col_index, col_name in enumerate(col_names):
                sheet.cell(row, col_index + 1).value = col_name

            # 其余行为各个排行条目
            for index, ranking in enumerate(rt[1]):
                rank = index + 1
                col_values = extract_rank_cols_func(ele_skill, job_name, cool, equip_index_to_realname, custom_buf_data, ranking_name, rank, ranking)

                row += 1
                for col_index, col_value in enumerate(col_values):
                    sheet.cell(row, col_index + 1).value = col_value

        # 保存文件
        book.save(export_config.export_file_name)
        tkinter.messagebox.showinfo("排行结果已导出", "前{}排行数据已导出到当前目录下的{}，可打开进行查看".format(
            export_config.export_rank_count,
            export_config.export_file_name,
        ))

    threading.Thread(target=_export_reuslt, daemon=True).start()


class ExtractRankColsFunc(Protocol):
    def __call__(self, ele_skill, job_name, cool, equip_index_to_realname, custom_buf_data, ranking_name, rank, ranking_detail):
        """
        @type ele_skill: int
        @type job_name: str
        @type cool: str
        @type equip_index_to_realname: Dict[str, str]
        @type custom_buf_data: Dict[str, int]
        @type ranking_name: str
        @type rank: int
        @type ranking_detail: list
        @rtype: list
        """


def extract_deal_rank_cols(ele_skill, job_name, cool, equip_index_to_realname, custom_buf_data, ranking_name, rank, ranking_detail):
    # (damage, unique_index, [calc_wep, base_array, baibianguai, tuple(not_owned_equips)])
    damage = "{}%".format(int(100 * ranking_detail[0]))
    weapon_index = ranking_detail[2][0][0]
    equip_indexes = ranking_detail[2][0][1:]
    base_array = ranking_detail[2][1]
    baibianguai = ranking_detail[2][2]
    not_owned_equips = ranking_detail[2][3]

    cols = []
    cols.append(rank)  # 排行
    cols.append(damage)  # 伤害倍率
    cols.append(job_name)  # 职业
    cols.append(" | ".join(get_readable_names(equip_index_to_realname, weapon_index, equip_indexes)))  # 搭配概览
    cols.append(equip_index_to_realname[weapon_index])  # 武器
    # 上衣 裤子 头肩 腰带 鞋子 手镯 项链 戒指 辅助装备 魔法石 耳环
    cols.extend(get_slot_names(equip_index_to_realname, equip_indexes))
    bbg = ""  # 百变怪
    if baibianguai is not None:
        bbg = equip_index_to_realname[baibianguai]
    cols.append(bbg)
    cols.append(",".join(equip_index_to_realname[equip_index] for equip_index in not_owned_equips))  # 跨界或升级工作服得来的装备
    for index_deal in range(28):
        cols.append(base_array[index_deal])
    cols.append(cool)  # 冷却补正
    cols.append(ele_skill)  # 技能属强补正
    cols.append("{}%".format(round(100 * (1.05 / (1.05 + int(ele_skill) * 0.0045) - 1), 1)))  # 逆补正

    if len(cols) != len(deal_col_names):
        raise Exception("col number not match")

    return cols


def extract_buf_rank_cols(ele_skill, job_name, cool, equip_index_to_realname, custom_buf_data, ranking_name, rank, ranking_detail):
    # (score, unique_index, [taiyang_calc_wep, [bless_overview, taiyang_overview, first_awaken_passive_overview, all_score_str], baibianguai, tuple(noe), huanzhuang_equip])
    score = ranking_detail[0]
    weapon_index = ranking_detail[2][0][0]
    equip_indexes = ranking_detail[2][0][1:]

    score = ranking_detail[0]  # xx标准
    if ranking_name != "太阳适用面板排行":
        # 除面板得分外，其余的都要除10
        score = score // 10
    bless = ranking_detail[2][1][0]  # 祝福数据
    taiyang = ranking_detail[2][1][1]  # 太阳数据
    taiyang_passive = ranking_detail[2][1][2]  # 太阳被动
    all_score = ranking_detail[2][1][3]  # 三个标准下的得分

    baibianguai = ranking_detail[2][2]
    not_owned_equips = ranking_detail[2][3]
    huanzhuang_equip = ranking_detail[2][4]

    cols = []
    cols.append(rank)  # 排行
    cols.append(score)  # 得分
    cols.append(all_score)  # 祝福得分/一觉得分/综合得分
    cols.append(job_name)  # 职业
    cols.append(" | ".join(get_readable_names(equip_index_to_realname, weapon_index, equip_indexes, huanzhuang_equip)))  # 搭配概览
    cols.append(equip_index_to_realname[weapon_index])  # 武器
    # 上衣 裤子 头肩 腰带 鞋子 手镯 项链 戒指 辅助装备 魔法石 耳环
    cols.extend(get_slot_names(equip_index_to_realname, equip_indexes))
    huanzhuang = ""
    if huanzhuang_equip != "":
        huanzhuang = equip_index_to_realname[huanzhuang_equip]
    cols.append(huanzhuang)  # 祝福切装
    bbg = ""  # 百变怪
    if baibianguai is not None:
        bbg = equip_index_to_realname[baibianguai]
    cols.append(bbg)
    cols.append(",".join(equip_index_to_realname[equip_index] for equip_index in not_owned_equips))  # 跨界或升级工作服得来的装备
    cols.append(bless)  # 祝福
    cols.append(taiyang)  # 一觉
    cols.append(taiyang_passive)  # 一觉被动
    cols.append(custom_buf_data["bless_level"])  # 自定义祝福+X级
    cols.append(custom_buf_data["taiyang_level"])  # 自定义一觉+X级
    cols.append(custom_buf_data["bless_data"])  # 祝福数据+
    cols.append(custom_buf_data["taiyang_data"])  # 太阳数据+

    if len(cols) != len(buf_col_names):
        raise Exception("col number not match")

    return cols
