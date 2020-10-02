#!/usr/bin/env python
# -*- coding: utf-8 -*-

## 코드를 무단으로 복제하여 개조 및 배포하지 말 것##

import queue
import tkinter.messagebox
from zipfile import BadZipFile

import PIL
import PIL.Image
import PIL.ImageTk
from openpyxl import load_workbook

from dnf_calc import *

if __name__ == '__main__':
    multiprocessing.freeze_support()
    configure_bugsnag()
    logger.info("configure_bugsnag done in {}".format(multiprocessing.current_process()))

    # 启动时先读取config和setting
    load_config()
    load_settings()

    # 设置一些常量
    layout_cfg = config().ui.layout
    other_window_x_offset = layout_cfg.window_x_offset + layout_cfg.window_width + 10

    result_window_width = 585
    result_window_readable_result_area_height = 18 * 3
    result_window_height = 402 + result_window_readable_result_area_height
    result_window_x_offset = other_window_x_offset
    result_window_y_offset = layout_cfg.window_y_offset + (layout_cfg.window_height - result_window_height) // 2

    custom_window_width = 620
    custom_window_height = 400
    custom_window_x_offset = other_window_x_offset
    custom_window_y_offset = layout_cfg.window_y_offset + (layout_cfg.window_height - custom_window_height) // 2

    res_txt_readable_result_center_x = result_window_width // 2
    res_txt_readable_result_center_y = result_window_height - result_window_readable_result_area_height // 2


def hide_result_window_if_exists():
    global result_window
    try:
        result_window.destroy()
    except Exception as error:
        pass


def report_bugsnag_with_context(error, extra_context=None, show_error_messagebox=True):
    global exit_calc
    exit_calc.value = 1

    hide_result_window_if_exists()

    traceback_info = traceback.format_exc()

    # 打印错误日志
    logger.error("calc unhandled exception\n{}".format(traceback_info))

    # 弹出错误框
    if show_error_messagebox:
        tkinter.messagebox.showerror("出错啦", "计算过程中出现了未处理的异常\n{}".format(traceback_info))

    # 上报bugsnag
    items, not_select_items, work_uniforms_items = get_equips()
    cpu_name, physical_cpu_cores, manufacturer = get_hardward_info()
    meta_data = {
        "stacktrace_brief": {
            "info": traceback_info,
        },
        "ui_options": {
            "speed": select_speed.get(),
            "weapons": wep_combopicker.get_selected_entrys(),
            "job_name": jobup_select.get(),
            "shuchu_time": time_select.get(),
            "style": style_select.get(),
            "creature": creature_select.get(),
            "cooldown": req_cool.get(),
            "baibianguai": baibianguai_select.get(),
            "work_uniform": can_upgrade_work_unifrom_nums_select.get(),
            "transfer": transfer_equip_combopicker.get_selected_entrys(),
            "max_transfer_count": can_transfer_nums_select.get(),
            "use_pulei": use_pulei_legend_by_default_select.get(),
            "save_name": save_name_list[current_save_name_index],
        },
        "euqips": {
            "items": get_equip_slots_with_name(equip_index_to_realname, items),
            "not_select_items": not_select_items,
            "work_uniforms_items": work_uniforms_items,
        },
        "config": config(),
        "settings": all_settings(),
        "app": {
            "releaseStage": RUN_ENV,
            "version": now_version,
            "release_time": ver_time,
        },
        "device": {
            "uuid": uuid.getnode(),
            "node": platform.node(),
            "osName": platform.system(),
            "osVersion": platform.version(),
            "release": platform.release(),
            "architecture": platform.machine(),
            "processor": platform.processor(),
            "logical_cpu_num": multiprocessing.cpu_count(),
            "physical_cpu_num": physical_cpu_cores,
            "cpu_name": cpu_name,
            "manufacturer": manufacturer,
            "time": time.strftime("%Y-%m-%d %H:%M:%S"),
            "timezone": time.strftime("%z", time.gmtime()),
        },
    }
    if extra_context is not None:
        meta_data["extra_context"] = extra_context
    bugsnag.notify(
        exception=error,
        context="calc",
        meta_data=meta_data,
        user={"id": platform.node(), "uuid": uuid.getnode(), },
    )


def calc_thread():
    threading.Thread(target=calc_with_try_except, daemon=True).start()


def calc_with_try_except():
    if not is_debug_mode():
        try:
            calc()
        except Exception as error:
            report_bugsnag_with_context(error)
    else:
        calc()


minheap_with_queues = []  # type: List[MinHeapWithQueue]


## 计算函数##
def calc():
    hide_result_window_if_exists()

    if select_speed.get() == speed_slow:
        set_perfect = 1
    else:
        set_perfect = 0
    showsta(text="正在准备组合算法驱动...")
    logger.info("开始计算")
    start_time = time.time()

    logger.debug("loading data.xlsx")
    try:
        load_excel = load_workbook("DATA.xlsx", read_only=True, data_only=True)
    except Exception as error:
        notify_error(logger, "data.xlsx文件不见啦，可能是未解压，请解压后再使用,err={}".format(error))
        return

    db_one = load_excel["one"]
    opt_one = {}
    for row in db_one.rows:
        row_value = [cell.value for cell in row]
        if len(row_value) == 0:
            continue

        name = row_value[0]
        row_value_cut = row_value[2:]

        opt_one[name] = row_value_cut[0:20] + row_value_cut[22:23] + row_value_cut[34:35] + row_value_cut[38:46] + row_value_cut[47:52]

    db_set = load_excel["set"]
    opt_set = {}
    for row in db_set.rows:
        row_value = [cell.value for cell in row]
        if len(row_value) == 0:
            continue

        if len(row_value) == 0:
            continue

        set_index = row_value[0]
        row_value_cut = row_value[2:]

        opt_set[set_index] = row_value_cut  ## DB 装入 ##

    db_buf = load_excel["buf"]
    opt_buf = {}
    for row in db_buf.rows:
        row_value = [cell.value for cell in row]
        if len(row_value) == 0:
            continue

        buf_index = row_value[0]
        row_value_cut = row_value[2:]

        opt_buf[buf_index] = row_value_cut  ## DB 装入 ##

    db_buflvl = load_excel["buflvl"]
    opt_buflvl = {}
    for row in db_buflvl.rows:
        row_value = [cell.value for cell in row]
        if len(row_value) == 0:
            continue

        buf_name = row_value[0]
        row_value_cut = row_value[1:]

        opt_buflvl[buf_name] = [0] + row_value_cut  # 首位0用来占位，这样后面就可以直接用等级索引对应等级的数值

    # 读取奶系自定义数据数据
    load_buf_custom_data()

    logger.debug("loading preset.xlsx")
    load_presetc = load_workbook("preset.xlsx", data_only=True)
    db_preset = load_presetc["custom"]
    job_name = jobup_select.get()
    try:
        ele_skill = int(opt_job_ele[job_name][0])
    except Exception as error:
        tkinter.messagebox.showerror('部分参数有误', "未选择职业或职业非法", parent=self)
        logger.warning("job_name=%s invalid", job_name)
        return
    try:
        ele_in = (int(db_preset["B14"].value) + int(db_preset["B15"].value) + int(db_preset["B16"].value) +
                  int(ele_skill) - int(db_preset["B18"].value) + int(db_preset["B19"].value) + 13)
    except Exception as error:
        notify_error(logger, "preset.xlsx的B14、B15、B16、B18、B19均应为整数，请仔细检查，是否填空值了")
        return

    global count_valid, count_invalid, show_number, all_list_num, max_setopt, count_start_time, unique_index, multiprocessingManager
    count_valid = 0
    count_invalid = 0
    show_number = 0
    max_setopt = multiprocessingManager.Value('i', 0)

    if job_name[-4:] == "(奶系)":
        active_eff_one = 15
        active_eff_set = 18 - 3
    else:
        active_eff_one = 21
        active_eff_set = 27 - 3

    if time_select.get() == "60秒(觉醒占比↓)":
        lvl_shift = 6
    else:
        lvl_shift = 0

    job_lv1 = opt_job[job_name][11 + lvl_shift]
    job_lv2 = opt_job[job_name][12 + lvl_shift]
    job_lv3 = opt_job[job_name][13 + lvl_shift]
    job_lv4 = opt_job[job_name][14 + lvl_shift]
    job_lv5 = opt_job[job_name][15 + lvl_shift]
    job_lv6 = opt_job[job_name][16 + lvl_shift]
    job_pas0 = opt_job[job_name][0]
    job_pas1 = opt_job[job_name][1]
    job_pas2 = opt_job[job_name][2]
    job_pas3 = opt_job[job_name][3]

    if req_cool.get() == 'X(纯伤害)':
        cool_on = 0
    else:
        cool_on = 1

    valid_weapon = True
    weapon_indexs = []
    weapon_names = wep_combopicker.get_selected_entrys()
    try:
        for weapon_name in weapon_names:
            weapon_indexs.append(wep_name_to_index[weapon_name])
    except:
        valid_weapon = False
    if len(weapon_indexs) == 0:
        valid_weapon = False

    if not valid_weapon:
        tkinter.messagebox.showerror('部分参数有误', "未选择武器或武器非法", parent=self)
        logger.warning("weapon_names=%s invalid", weapon_names)
        return

    # 获取当前装备、百变怪可选装备、工作服列表
    items, not_select_items, work_uniforms_items = get_equips()

    # 获取选定的账号的各部位所拥有的当前账户未拥有的装备列表
    transfer_slots_equips = get_transfer_slots_equips(items, load_presetc["one"])
    transfer_max_count = get_can_transfer_nums()

    # 根据需求决定是否需要开启将神话放到前面来加快剪枝的方案
    modify_slots_order(items, not_select_items, work_uniforms_items, transfer_slots_equips)

    # 已选装备的搭配数
    original_count = calc_ori_counts(items)
    # 百变怪增加的搭配数
    bbg_count = calc_bbg_add_counts(items, not_select_items)
    # 额外升级的工作服增加的搭配数
    work_uniforms_count = calc_upgrade_work_uniforms_add_counts(items, not_select_items, work_uniforms_items)

    all_list_num = original_count + bbg_count + work_uniforms_count

    showsta(text='开始计算')
    count_start_time = time.time()  # 开始计时
    logger.debug(("\n"
                  "items={}\n"
                  "not_select_items={}\n"
                  "work_uniforms_items={}\n"
                  "transfer_slots_equips={}\n"
                  "transfer_max_count={}").format(
        items, not_select_items, work_uniforms_items, transfer_slots_equips, transfer_max_count
    ))

    global exit_calc
    # 开始计算
    exit_calc.value = 0

    has_baibianguai = baibianguai_select.get() == txt_has_baibianguai
    can_upgrade_work_unifrom_nums = get_can_upgrade_work_unifrom_nums()
    has_uniforms = pre_calc_has_uniforms(items, work_uniforms_items)
    # 超慢速时不进行任何剪枝操作，装备搭配对比的标准是最终计算出的伤害与奶量倍率
    dont_pruning = select_speed.get() == speed_super_slow
    dont_prefer_god = not prefer_god()

    logger.info(("all_list_num={} (original_count={} bbg_count={} work_uniforms_count={})\n"
                 "transfer_max_count={} has_baibianguai={}, can_upgrade_work_unifrom_nums={} dont_pruning={}, dont_prefer_god={}\n"
                 "transfer_slots_equips={}\n"
                 "has_uniforms={}\n"
                 "job_name={} weapon_names={}".format(
        all_list_num, original_count, bbg_count, work_uniforms_count,
        transfer_max_count, has_baibianguai, can_upgrade_work_unifrom_nums, dont_pruning, dont_prefer_god,
        transfer_slots_equips,
        has_uniforms,
        job_name, weapon_names,
    )))

    ui_top_n = 5
    save_top_n = ui_top_n
    cfg = config()
    if cfg.export_result_as_excel.enable:
        save_top_n = max(save_top_n, cfg.export_result_as_excel.export_rank_count)

    m = multiprocessingManager
    global minheap_with_queues
    minheap_with_queues = []  # type: List[MinHeapWithQueue]

    step_data = CalcStepData()

    step_data.items = items
    step_data.has_baibianguai = has_baibianguai
    step_data.not_select_items = not_select_items
    step_data.has_uniforms = has_uniforms
    step_data.can_upgrade_work_unifrom_nums = can_upgrade_work_unifrom_nums
    step_data.work_uniforms_items = work_uniforms_items
    step_data.transfer_max_count = transfer_max_count
    step_data.transfer_slots_equips = transfer_slots_equips

    last_god_slot = get_last_god_slot(items)
    step_data.last_god_slot = last_god_slot

    step_data.current_index = 0
    step_data.has_god = False
    step_data.local_max_setop = 0
    step_data.max_setopt = max_setopt
    step_data.max_possiable_setopt = 3 + 2 + 2 + 1  # 533 以及神话对应的一个词条
    if set_perfect or not prefer_god() or last_god_slot == -1:
        # 如果神话不优先或者没有神话，则不计入最高历史词条
        step_data.max_possiable_setopt -= 1

    calc_data = CalcData()
    calc_data.weapon_indexs = weapon_indexs
    calc_data.exit_calc = exit_calc
    step_data.calc_data = calc_data

    step_data.dont_pruning = dont_pruning
    step_data.set_perfect = set_perfect
    step_data.prefer_god = prefer_god()
    step_data.prune_cfg = cfg.prune
    # 代码中深度从0开始计算，-1则表示不启用
    step_data.start_parallel_computing_at_depth_n = config().multi_threading.start_parallel_computing_at_depth_n - 1

    step_data.config = config()

    step_data.producer = producer
    producer_data.calc_index += 1
    producer_data.produced_count = 0

    finished = False

    def log_result_queue_info(log_func, msg, mq: MinHeapWithQueue):
        log_func("calc#{}: {}: {} remaining_qize={} sync_batch_size={} processed_result={}, speed={:.2f}/s totalWork={}".format(
            producer_data.calc_index,
            mq.name, msg, mq.minheap_queue.qsize(), mq.minheap.batch_size, mq.minheap.processed_result_count, mq.process_results_per_second(), producer_data.produced_count
        ))

    def try_fetch_result(mq: MinHeapWithQueue):
        idx = 1
        while True:
            try:
                minheap_to_merge = mq.minheap_queue.get(block=False)
                mq.minheap.merge(minheap_to_merge)

                if mq.minheap.processed_result_count >= 1000 * idx:
                    log_result_queue_info(logger.info, "try_fetch_result periodly report", mq)
                    idx = mq.minheap.processed_result_count // 1000 + 1
            except queue.Empty as error:
                break

    def try_fetch_result_in_background(mq: MinHeapWithQueue):
        while not finished:
            log_result_queue_info(logger.info, "try_fetch_result_in_background", mq)
            try_fetch_result(mq)
            time.sleep(0.5)

    data_transfer_batch_size = config().data_transfer.batch_size

    is_shuchu_job = job_name not in ["(奶系)神思者", "(奶系)炽天使", "(奶系)冥月女神"]
    if is_shuchu_job:
        unique_index = 0
        show_number = 1

        base_array_with_deal_bonus_attributes = np.array([0.0 for idx in range(len(deal_entry_index_to_name))])
        # 加上基础数据
        # re：韩服作者加的这个默认的额外1%技攻是啥？在没有弄清楚前暂时先注释掉了= =
        # base_array_with_deal_bonus_attributes[index_deal_extra_percent_skill_attack_power] = 1
        base_array_with_deal_bonus_attributes[index_deal_extra_all_element_strength] += ele_in

        # 当前存档的装备补正信息 equip_index => list(attribute_index_str => fixup_value)
        deal_equip_fixup = {}

        # 加入输出职业的特色加成
        add_bonus_attributes_to_base_array("deal", base_array_with_deal_bonus_attributes, style_select.get(), creature_select.get(), save_name_list[current_save_name_index], deal_equip_fixup, equip_index_to_realname, {})

        # 补正装备
        for equip_index, ba in deal_equip_fixup.items():
            if equip_index in opt_one:
                opt_one[equip_index] += ba

        minheap_with_queues = [
            MinHeapWithQueue("输出排行", MinHeap(save_top_n, data_transfer_batch_size), m.Queue()),
        ]

        # 异步排行线程
        fetch_result_threads = []
        for mq in minheap_with_queues:
            thread = threading.Thread(target=try_fetch_result_in_background, args=(mq,), daemon=True)
            thread.start()
            fetch_result_threads.append(thread)

        calc_data = step_data.calc_data
        calc_data.is_shuchu_job = True
        calc_data.base_array_with_deal_bonus_attributes = base_array_with_deal_bonus_attributes
        calc_data.opt_one = opt_one
        calc_data.job_lv1 = job_lv1
        calc_data.job_lv2 = job_lv2
        calc_data.job_lv3 = job_lv3
        calc_data.job_lv4 = job_lv4
        calc_data.job_lv5 = job_lv5
        calc_data.job_lv6 = job_lv6
        calc_data.job_pas0 = job_pas0
        calc_data.job_pas1 = job_pas1
        calc_data.job_pas2 = job_pas2
        calc_data.job_pas3 = job_pas3
        calc_data.cool_on = cool_on
        calc_data.ele_skill = ele_skill
        calc_data.minheaps = [MinHeap(save_top_n, data_transfer_batch_size) for mq in minheap_with_queues]
        calc_data.minheap_queues = [mq.minheap_queue for mq in minheap_with_queues]
        step_data.calc_data = calc_data

        step_data.process_func = process_deal

        parallel_dfs(step_data)

        # 等到所有工作处理完成
        producer_data.work_queue.join()
        finished = True

        # 等待异步排行线程退出
        for thread in fetch_result_threads:
            thread.join()

        # 最终将剩余结果（若有）也加入排序
        for mq in minheap_with_queues:
            log_result_queue_info(logger.info, "after join", mq)
            try_fetch_result(mq)
            log_result_queue_info(logger.info, "after final", mq)

        show_number = 0
        showsta(text='结果统计中')

        all_ranking = minheap_with_queues[0].minheap.getTop()
        ranking = []
        for index, data in enumerate(all_ranking[:ui_top_n]):
            damage = data[0]
            value = data[2]
            ranking.append((damage, value))
            showsta(text='结果统计中 {} / 5'.format(index + 1))

        show_result(ranking, 'deal', ele_skill)

        export_result(ele_skill, deal_col_names, jobup_select.get(), req_cool.get(), equip_index_to_realname, custom_buf_data, extract_deal_rank_cols, [
            ("伤害排行", all_ranking)
        ])

    else:
        # 计算祝福数值时的额外15级转职被动等级（暂时不知道为啥要额外加）
        base_job_passive_lv15_bless = cfg.initital_data.base_job_passive_lv15_bless
        # 计算太阳数值时的额外15级转职被动等级（暂时不知道为啥要额外加）
        base_job_passive_lv15_taiyang = cfg.initital_data.base_job_passive_lv15_taiyang
        # 自定义祝福数据-自定义太阳数据（暂时不清楚这个是干啥的）
        base_stat_custom_bless_data_minus_taiyang_data = custom_buf_data["bless_data"] - custom_buf_data["taiyang_data"]

        show_number = 1

        # 基础体力、精神
        base_stat_physical_and_mental = eval(cfg.initital_data.physical_and_mental) + custom_buf_data["taiyang_data"]
        # 基础智力
        base_stat_intelligence = eval(cfg.initital_data.intelligence) + custom_buf_data["taiyang_data"]
        # 祝福等级
        base_bless_level = 10 + custom_buf_data["bless_level"]
        # 太阳等级
        base_taiyang_level = 12 + custom_buf_data["taiyang_level"]
        # 15级转职被动等级
        base_job_passive_lv15 = 0
        # 基础奶爸25级守护徽章等级
        base_naiba_protect_badge_lv25 = 0

        base_array_with_buf_bonus_attributes = np.array([0.0 for idx in range(len(buf_entry_index_to_name))])
        # 加上基础数据
        base_array_with_buf_bonus_attributes[index_buf_physical_and_mental_strength] += base_stat_physical_and_mental  # 体力、精神 +X
        base_array_with_buf_bonus_attributes[index_buf_intelligence] += base_stat_intelligence  # 智力 +X
        base_array_with_buf_bonus_attributes[index_buf_bless_lv30] += base_bless_level  # 祝福技能祝福等级+X
        base_array_with_buf_bonus_attributes[index_buf_taiyang_lv50] += base_taiyang_level  # 太阳技能太阳等级+X
        base_array_with_buf_bonus_attributes[index_buf_job_passive_lv15] += base_job_passive_lv15  # 15级职业被动Lv+X
        base_array_with_buf_bonus_attributes[index_buf_naiba_protect_badge_lv25] += base_naiba_protect_badge_lv25  # 奶爸25级守护徽章等级+X

        # 当前存档的装备补正信息 equip_index => list(attribute_index_str => fixup_value)
        buf_equip_fixup = {}
        # 当前存档的buff换装槽位补正信息 slot_index => list(attribute_index_str => fixup_value)
        huanzhuang_slot_fixup = {}

        # 增加奶系的国服特色
        add_bonus_attributes_to_base_array("buf", base_array_with_buf_bonus_attributes, style_select.get(), creature_select.get(), save_name_list[current_save_name_index], buf_equip_fixup, equip_index_to_realname, huanzhuang_slot_fixup)

        # 补正装备
        for equip_index, ba in buf_equip_fixup.items():
            if equip_index in opt_buf:
                opt_buf[equip_index] += ba

        minheap_with_queues = [
            MinHeapWithQueue("祝福排行", MinHeap(save_top_n, data_transfer_batch_size), m.Queue()),
            MinHeapWithQueue("太阳排行", MinHeap(save_top_n, data_transfer_batch_size), m.Queue()),
            MinHeapWithQueue("综合排行", MinHeap(save_top_n, data_transfer_batch_size), m.Queue()),
            MinHeapWithQueue("面板排行", MinHeap(save_top_n, data_transfer_batch_size), m.Queue()),
        ]

        for mq in minheap_with_queues:
            threading.Thread(target=try_fetch_result_in_background, args=(mq,), daemon=True).start()

        calc_data = step_data.calc_data
        calc_data.is_shuchu_job = False
        calc_data.base_array_with_buf_bonus_attributes = base_array_with_buf_bonus_attributes
        calc_data.job_name = job_name
        calc_data.const = cfg.const
        calc_data.huan_zhuang = cfg.huan_zhuang
        calc_data.huanzhuang_slot_fixup = huanzhuang_slot_fixup
        calc_data.opt_buf = opt_buf
        calc_data.opt_buflvl = opt_buflvl
        calc_data.base_job_passive_lv15_bless = base_job_passive_lv15_bless
        calc_data.base_job_passive_lv15_taiyang = base_job_passive_lv15_taiyang
        calc_data.base_stat_custom_bless_data_minus_taiyang_data = base_stat_custom_bless_data_minus_taiyang_data
        calc_data.base_stat_physical_and_mental = base_stat_physical_and_mental
        calc_data.base_stat_intelligence = base_stat_intelligence
        calc_data.base_bless_level = base_bless_level
        calc_data.base_taiyang_level = base_taiyang_level
        calc_data.base_job_passive_lv15 = base_job_passive_lv15
        calc_data.base_naiba_protect_badge_lv25 = base_naiba_protect_badge_lv25
        calc_data.minheaps = [MinHeap(save_top_n, data_transfer_batch_size) for mq in minheap_with_queues]
        calc_data.minheap_queues = [mq.minheap_queue for mq in minheap_with_queues]
        step_data.calc_data = calc_data

        # 预计算套装信息
        owned_set_2_equips_map = {}
        for slot_equip_indexes in step_data.items:
            for equip_index in slot_equip_indexes:
                set_index = get_set_name(equip_index)
                if set_index not in owned_set_2_equips_map:
                    owned_set_2_equips_map[set_index] = set()
                owned_set_2_equips_map[set_index].add(equip_index)
        step_data.owned_set_2_equips_map = owned_set_2_equips_map

        step_data.process_func = process_buf

        parallel_dfs(step_data)

        # 等到所有工作处理完成
        producer_data.work_queue.join()
        finished = True

        # 最终将剩余结果也加入排序
        for mq in minheap_with_queues:
            log_result_queue_info(logger.info, "after join", mq)
            try_fetch_result(mq)
            log_result_queue_info(logger.info, "after final", mq)

        show_number = 0
        showsta(text='结果统计中')

        all_rankings = [mq.minheap.getTop() for mq in minheap_with_queues]
        rankings = [[] for x in range(len(all_rankings))]
        for rank_type_index in range(len(all_rankings)):
            for index, data in enumerate(all_rankings[rank_type_index][:ui_top_n]):
                score = data[0]
                value = data[2]
                rankings[rank_type_index].append((score, value))

        show_result(rankings, 'buf', ele_skill)

        export_result(ele_skill, buf_col_names, jobup_select.get(), req_cool.get(), equip_index_to_realname, custom_buf_data, extract_buf_rank_cols, [
            ("祝福排行", all_rankings[0]),
            ("太阳排行", all_rankings[1]),
            ("综合排行", all_rankings[2]),
            ("太阳适用面板排行", all_rankings[3]),
        ])

    load_presetc.close()
    load_excel.close()
    showsta(text='输出完成' + "时间 = " + format_time(time.time() - start_time))
    # 结束计算
    exit_calc.value = 1
    logger.info("工作进程数：{}个 共处理workItem：{}个".format(max_thread, producer_data.produced_count))
    logger.info("计算耗时时间 = " + str(time.time() - start_time) + "秒")


def stop_calc():
    # global exit_calc
    # exit_calc.value = 1
    logger.info("手动停止计算")
    tkinter.messagebox.showinfo("提示", (
        "多进程中同步状态性能开销比较大，因此处于性能考虑，目前废弃了停止功能\n"
        "请直接点击右上角关闭程序再开启来实现中途停止计算"
    ))


def get_set_code(equip_index_except_first_digit):
    return str(equip_index_except_first_digit)[1:3]


def get_equips():
    # 选中的装备
    list11, list12, list13, list14, list15 = [[] for i in range(5)]
    list21, list22, list23 = [[] for i in range(3)]
    list31, list32, list33 = [[] for i in range(3)]
    list41, list42, list43 = [[] for i in range(3)]
    # 套装
    list_setcode = []
    set_has_god = {}
    # 未选中的装备
    listns11, listns12, listns13, listns14, listns15 = [[] for i in range(5)]
    listns21, listns22, listns23 = [[] for i in range(3)]
    listns31, listns32, listns33 = [[] for i in range(3)]
    listns41, listns42, listns43 = [[] for i in range(3)]

    local_vals = locals()

    def _get_equips(layout_cfg):
        """
        :type layout_cfg: LayoutConfig
        """
        for block_info in layout_cfg.equip_block_infos:
            if block_info.type == EquipBlockType_Set:
                for set_code in range(block_info.set_code_start, block_info.set_code_end + 1):
                    for idx, val in enumerate(block_info.slot_orders):
                        slot, god = val
                        equip_index = "{0:02}{1:02}{2:1}".format(slot, set_code, god)
                        equip_btn_index = 'tg{0}'.format(equip_index)

                        if select_item[equip_btn_index] == 1:
                            ls = eval("list{}".format(slot), local_vals)  # type: list
                            ls.append(equip_index)
                            list_setcode.append(str(set_code))
                            if god == 1:
                                set_has_god[str(set_code)] = True
                        elif can_convert_from_baibianguai(equip_index):
                            ls = eval("listns{}".format(slot), local_vals)  # type: list
                            ls.append(equip_index)
            elif block_info.type == EquipBlockType_Single:
                for idx, equip_index in enumerate(block_info.equips):
                    equip_index = str(equip_index)
                    slot = equip_index[0:2]
                    equip_btn_index = 'tg{0}'.format(equip_index)

                    if select_item[equip_btn_index] == 1:
                        ls = eval("list{}".format(slot), local_vals)  # type: list
                        ls.append(equip_index)
            elif block_info.type == EquipBlockType_Nested:
                _get_equips(block_info.nested_block)
            else:
                notify_error(logger, "ui布局配置有误，不支持类型为{}的block".format(block_info.type))
                sys.exit(0)

    _get_equips(config().ui.layout)

    equip_slot_indexes = ['11', '12', '13', '14', '15', '21', '22', '23', '31', '32', '33', '41', '42', '43']
    if select_speed.get() == speed_quick:
        for set_code, count in collections.Counter(list_setcode).items():
            if count > 1:
                # 非散件，略过
                continue
            if set_code in set_has_god:
                # 只有一件神话散件，略过
                continue
            for equip_slot_index in equip_slot_indexes:
                try:
                    eval("list{}.remove('{}{}0')".format(equip_slot_index, equip_slot_index, set_code))
                except ValueError as error:
                    c = 1

    def append_if_not_in(ls, equip_index):
        if equip_index not in ls:
            ls.append(equip_index)

    if use_pulei_legend_by_default():
        append_if_not_in(list11, '11360')
        append_if_not_in(list12, '12360')
        append_if_not_in(list13, '13360')
        append_if_not_in(list14, '14360')
        append_if_not_in(list15, '15360')

        append_if_not_in(list21, '21370')
        append_if_not_in(list22, '22370')
        append_if_not_in(list23, '23370')

        append_if_not_in(list31, '31380')
        append_if_not_in(list32, '32380')
        append_if_not_in(list33, '33380')
    else:
        # 如果不默认使用普雷传说，则仅在对应槽位没有其他任何可选装备的时候才加入
        # 防具任意部位不存在可选装备时将所有传说防具加入备选池
        if len(list11) == 0 or len(list12) == 0 or len(list13) == 0 or len(list14) == 0 or len(list15) == 0:
            append_if_not_in(list11, '11360')
            append_if_not_in(list12, '12360')
            append_if_not_in(list13, '13360')
            append_if_not_in(list14, '14360')
            append_if_not_in(list15, '15360')

        # 如果首饰至少两个部位不存在可选装备，则将所有普雷首饰加入备选池
        if len(list21) == 0 and len(list22) == 0 or len(list22) == 0 and len(list23) == 0 or len(list23) == 0 and len(list21) == 0:
            append_if_not_in(list21, '21370')
            append_if_not_in(list22, '22370')
            append_if_not_in(list23, '23370')

        # 如果特殊装备至少两个部位不存在可选装备，则将所有普雷特殊装备加入备选池
        if len(list31) == 0 and len(list32) == 0 or len(list32) == 0 and len(list33) == 0 or len(list33) == 0 and len(list31) == 0:
            append_if_not_in(list31, '31380')
            append_if_not_in(list32, '32380')
            append_if_not_in(list33, '33380')

        # 若首饰特殊某个部位不存在可选装备，则将对应槽位的普雷装备加入备选池
        if len(list21) == 0:
            append_if_not_in(list21, '21370')
        if len(list22) == 0:
            append_if_not_in(list22, '22370')
        if len(list23) == 0:
            append_if_not_in(list23, '23370')
        if len(list31) == 0:
            append_if_not_in(list31, '31380')
        if len(list32) == 0:
            append_if_not_in(list32, '32380')
        if len(list33) == 0:
            append_if_not_in(list33, '33380')

    # 所有已选装备
    items = [list11, list12, list13, list14, list15, list21, list22, list23, list31, list32, list33]
    # 百变怪的各部位可选装备需要与上面的部位顺序一致
    not_select_items = [listns11, listns12, listns13, listns14, listns15, listns21, listns22, listns23, listns31, listns32, listns33]
    # 可升级得到的各部位工作服
    work_uniforms_items = ["11150", "12150", "13150", "14150", "15150", "21190", "22190", "23190", "31230", "32230", "33230"]

    return items, not_select_items, work_uniforms_items


def calc_ori_counts(all_slots_equips):
    cnt = 1
    for one_slot_equips in all_slots_equips:
        cnt *= len(one_slot_equips)
    return cnt


def calc_bbg_add_counts(slots_equips, slots_not_select_equips):
    if baibianguai_select.get() != txt_has_baibianguai:
        return 0

    ori_counts = calc_ori_counts(slots_equips)

    # 百变怪增加的搭配数
    bbg_add_num = 0
    for i in range(0, len(slots_not_select_equips)):
        bbg_add_num += ori_counts / len(slots_equips[i]) * len(slots_not_select_equips[i])

    return bbg_add_num


# 玩家各个部位是否已经升级了工作服
def pre_calc_has_uniforms(items, work_uniforms_items):
    return [work_uniforms_items[idx] in items[idx] for idx in range(len(items))]


def get_can_upgrade_work_unifrom_nums():
    # 用户配置的当前可升级的工作服数目
    can_upgrade_work_unifrom_nums = 0
    if can_upgrade_work_unifrom_nums_select.get() in can_upgrade_work_unifrom_nums_str_2_int:
        can_upgrade_work_unifrom_nums = can_upgrade_work_unifrom_nums_str_2_int[
            can_upgrade_work_unifrom_nums_select.get()]
    return can_upgrade_work_unifrom_nums


def get_can_transfer_nums():
    # 用户配置的当前可升级的工作服数目
    can_transfer_nums = 0
    if can_transfer_nums_select.get() in can_transfer_nums_str_2_int:
        can_transfer_nums = can_transfer_nums_str_2_int[
            can_transfer_nums_select.get()]
    return can_transfer_nums


# 是否默认将普雷传说装备加入备选池
def use_pulei_legend_by_default():
    return use_pulei_legend_by_default_select.get() == txt_use_pulei_legend_by_default


# 是否偏好神话
def prefer_god():
    return select_speed.get() != speed_middle_not_prefer_god


# 获取选定的账号的各部位所拥有的当前账户未拥有的装备列表
def get_transfer_slots_equips(items, sheet):
    # 获取各存档的装备信息
    slot_name_list = ['11', '12', '13', '14', '15', '21', '22', '23', '31', '32', '33']
    slot_name_to_index = {}
    for idx, name in enumerate(slot_name_list):
        slot_name_to_index[name] = idx

    transfer_slots_equips = [set() for i in range(0, 11)]
    for account_name in transfer_equip_combopicker.get_selected_entrys():
        if account_name == "":
            continue
        # 获取当前存档的index
        account_index = 0
        try:
            account_index = save_name_list.index(account_name)
        except Exception as err:
            logger.warning("get_transfer_slots_equips 无法找到存档{}, err={}".format(account_name, err))
            continue

        # 读取各个装备的点亮情况
        for i in range(1, max_save_equips + 1):
            has_equip = sheet.cell(i, 2 + account_index).value == 1
            equip_index = sheet.cell(i, 1).value
            if equip_index is None:
                continue
            if len(equip_index) == 6:
                # 六位为武器，过滤掉武器
                continue

            if has_equip:
                try:
                    slot_index = slot_name_to_index[equip_index[:2]]
                    # 如果该装备当前账号未拥有，且之前的账户中未添加过，则加入备选集
                    if equip_index not in items[slot_index] and equip_index not in transfer_slots_equips[slot_index] and not is_god(equip_index):
                        transfer_slots_equips[slot_index].add(equip_index)
                except KeyError as error:
                    pass

    return transfer_slots_equips


def calc_upgrade_work_uniforms_add_counts(slots_equips, slots_not_select_equips, slots_work_uniforms):
    # 找出所有尚未升级工作服的部位
    not_has_uniform_slots = []
    for slot, work_uniform in enumerate(slots_work_uniforms):
        if work_uniform not in slots_equips[slot]:
            not_has_uniform_slots.append(slot)

    total_add_counts = 0

    # 穷举尚未升级部位的大小小于等于最大可升级数目的所有组合
    max_upgrade_count = min(get_can_upgrade_work_unifrom_nums(), len(not_has_uniform_slots))
    # 遍历所有可能升级的件数
    for upgrade_count in range(1, max_upgrade_count + 1):
        # 遍历升级该件数的所有部位的组合
        for upgrade_slots in itertools.combinations(not_has_uniform_slots, upgrade_count):
            # 获取非升级部位的已有装备
            other_slots_equips = []
            for slot, slot_equips in enumerate(slots_equips):
                if slot not in upgrade_slots:
                    other_slots_equips.append(slot_equips)
            # 获取非升级部位的未选择装备
            other_slots_not_select_equips = []
            for slot, slot_not_select_equips in enumerate(slots_not_select_equips):
                if slot not in upgrade_slots:
                    other_slots_not_select_equips.append(slot_not_select_equips)

            # 计算该升级方案下其余部位的可能搭配数目
            total_add_counts += calc_bbg_add_counts(other_slots_equips, other_slots_not_select_equips)

    return total_add_counts


if __name__ == '__main__':
    res_txt_readable_result = None


# 用文字方式写成当前搭配，避免每次都得一个个肉眼对比图标来确认是啥装备
def change_readable_result_area(weapon, equips, is_create, huanzhuang_equips=[]):
    global res_txt_readable_result, canvas_res

    readable_names = get_readable_names(equip_index_to_realname, weapon, equips, huanzhuang_equips)
    content = pretty_words(readable_names, 40, ' | ')
    if is_create:
        res_txt_readable_result = canvas_res.create_text(res_txt_readable_result_center_x, res_txt_readable_result_center_y,
                                                         text=content,
                                                         font=guide_font, fill='white')
    else:
        canvas_res.itemconfig(res_txt_readable_result, text=content)


# 展示当前搭配的各装备名称
def show_name():
    global g_rank_equips, g_current_rank, g_current_job, g_current_buff_type, rank_type_buf, rank_huanzhuang_equips

    equips = None
    if g_current_job == "deal":
        equips = g_rank_equips[g_current_rank]
    else:
        equips = g_rank_equips[g_current_buff_type][g_current_rank]

    # 确保按照正常排列的顺序展示
    ordered_equip_indexes = list(equips[1:])
    reverse_modify_slots_order_(ordered_equip_indexes)

    ordered_equip_indexes.insert(0, equips[0])

    readable_names = []
    for equip in ordered_equip_indexes:
        readable_names.append(equip_index_to_realname[equip])

    if g_current_job != "deal":
        rank_type_index = rank_type_buf - 1
        hz_equips = rank_huanzhuang_equips[rank_type_index][g_current_rank]
        if len(hz_equips) != 0:
            readable_names.append("(之后为祝福切装)")
            for hz in hz_equips:
                readable_names.append(equip_index_to_realname[hz])

    tkinter.messagebox.showinfo("装备详细信息", pretty_words(readable_names, 30, ' | '), parent=result_window)


# 保证一行不会有太多词
def pretty_words(words, max_line_word_count, delimiter):
    pretty_result = ""
    line_word_count = 0
    for word in words:
        if line_word_count + len(word) >= max_line_word_count:
            line_word_count = 0
            pretty_result += "\n"
        elif line_word_count != 0:
            pretty_result += delimiter

        pretty_result += word
        line_word_count += len(word)

    return pretty_result


if __name__ == '__main__':
    # 奶系自定义数据
    custom_buf_data = {}


def load_buf_custom_data():
    global custom_buf_data

    load_presetr = load_workbook("preset.xlsx", data_only=True)
    r_preset = load_presetr["custom"]

    try:
        custom_buf_data = {
            "bless_level": int(r_preset['H2'].value) + int(r_preset['H4'].value) + int(r_preset['H5'].value),
            "taiyang_level": int(r_preset['H3'].value),
            "bless_data": int(r_preset['H6'].value),
            "taiyang_data": int(r_preset['H1'].value),
        }
    except Exception as error:
        notify_error(logger, "preset.xlsx中custom表单中奶妈相关参数需要为整数，指H1->H6，请仔细检查，是否填空值了")
        sys.exit(-1)

    load_presetr.close()


def get_score_to_damage_rate():
    # 获取当前存档名
    current_save_name = save_name_list[current_save_name_index]

    damage_cfg = config().twenty_seconds_damage

    # 尝试使用默认的打桩系数配置
    cfg = damage_cfg.score_to_damage_rate

    for rate_cfg in damage_cfg.save_name_configs:
        # 若配置了当前存档的打桩系数，则用这个
        if rate_cfg.save_name == current_save_name:
            cfg = rate_cfg.score_to_damage_rate
            break

    return eval(cfg)


def format_damage(score):
    if config().twenty_seconds_damage.enable:
        return "{}% {}亿".format(int(100 * score), int(score * get_score_to_damage_rate()))
    else:
        return "{}%".format(int(100 * score))


def extract_score_from_score_damage(score_damage):
    if config().twenty_seconds_damage.enable:
        return score_damage.split(" ")[0]
    else:
        return score_damage


def show_result(rank_list, job_type, ele_skill):
    global g_rank_equips, g_current_rank, g_current_job, g_current_buff_type
    g_current_rank = 0
    g_current_job = job_type

    logger.debug("show_result: job_type={}, ele_skill={}\nrank_list={}".format(job_type, ele_skill, rank_list))

    # 可能还没有任何结果的时候直接点了结束计算，这时候就不要走后续流程了
    if (job_type == 'deal' and len(rank_list) == 0) or (job_type == 'buf' and len(rank_list[0]) == 0):
        return

    global result_window
    result_window = tkinter.Toplevel(self)
    result_window.title("计算结果 - ver" + now_version)
    result_window.attributes("-topmost", True)
    result_window.focus_force()
    result_window.geometry("{}x{}+{}+{}".format(result_window_width, result_window_height, result_window_x_offset, result_window_y_offset))
    result_window.resizable(config().main_window_resizable, config().main_window_resizable)
    global canvas_res
    canvas_width = result_window_width + 2
    canvas_height = result_window_height + 2
    canvas_res = Canvas(result_window, width=canvas_width, height=canvas_height, bd=0)
    canvas_res.place(x=-2, y=-2)
    if job_type == 'deal':
        result_bg = tkinter.PhotoImage(file='ext_img/bg_result.png')
    else:
        result_bg = tkinter.PhotoImage(file='ext_img/bg_result2.png')
    canvas_res.create_image(canvas_width // 2, canvas_height // 2, image=result_bg)

    gif_image_ids.clear()

    global image_list, image_list2
    global res_img_weapon, res_img11, res_img12, res_img13, res_img14, res_img15, res_img21, res_img22, res_img23, res_img31, res_img32, res_img33, res_txtbbgs, res_imgbbgs, wep_combopicker, jobup_select, res_txt_weapon

    wep_index = ""
    wep_image = None
    length = 0

    if job_type == 'deal':  ###########################

        global result_image_on, rank_dam, rank_stat, rank_stat2, req_cool, res_dam, res_stat, res_stat2, rank_baibiaoguai, rank_setting
        total_count = len(rank_list)
        rank_baibiaoguai = [0 for x in range(total_count)]
        rank_not_owned_equips = [0 for x in range(total_count)]
        rank_dam = [0 for x in range(total_count)]
        rank_setting = [0 for x in range(total_count)]
        rss = [0 for x in range(total_count)]
        result_image_on = [{} for x in range(total_count)]
        # rank => [score, [calc_wep, base_array, baibianguai, not_owned_equips]]
        for rank in range(total_count):
            rank_baibiaoguai[rank] = rank_list[rank][1][2]
            rank_not_owned_equips[rank] = rank_list[rank][1][3]
            rank_dam[rank] = format_damage(rank_list[rank][0])
            rank_setting[rank] = rank_list[rank][1][0]  ##0号是排名
            rss[rank] = rank_list[rank][1][1]
            for equip_slot_index in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                for equip_index in rank_setting[rank]:
                    if len(equip_index) != 6:
                        if equip_index[0:2] == str(equip_slot_index):
                            result_image_on[rank][str(equip_slot_index)] = image_list[equip_index]
                            # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                            if equip_index in rank_not_owned_equips[rank]:
                                result_image_on[rank][str(equip_slot_index)] = image_list2[equip_index]
            if rank_baibiaoguai[rank] is not None:
                result_image_on[rank]["bbg"] = image_list[rank_baibiaoguai[rank]]
            weapon_index = rank_setting[rank][0]  # 装备列表的第一个是武器的编号
            result_image_on[rank]["weapon"] = image_list[weapon_index]

        # 0추스탯 1추공 2증 3크 4추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 쿨감보정/21 二觉캐특수액티브 /22~27 액티브레벨링
        rank_stat = [0, 0, 0, 0, 0]
        rank_stat2 = [0, 0, 0, 0, 0]
        for rank in range(total_count):
            dark_resistance = int(rss[rank][index_deal_extra_dark_resistance])
            if dark_resistance < 81:
                dark_resistance = "{}(低于81)".format(dark_resistance)
            moving_speed = int(rss[rank][index_deal_extra_percent_moving_speed])
            if moving_speed < 100:
                moving_speed = "{}(低于100)".format(moving_speed)
            rank_stat[rank] = (
                "\n"
                "增伤={zengsu}%\n"
                "爆伤={baoshang}%\n"
                "白字={baizi}%\n"
                "所攻={suogong}%\n"
                "三攻={sangong}%\n"
                "力智={lizhi}%\n"
                "属强={shuqiang}\n"
                "持续={chixu}%\n"
                "技攻={jigong}%\n"
                "特殊={teshu}%\n"
                "攻速={gongsu}%\n"
                "暴击率={baojilv}%\n"
                "暗抗={dark_resistance}\n"
                "移速={moving_speed}%"
            ).format(
                zengsu=int(rss[rank][index_deal_extra_percent_attack_damage] + rss[rank][index_deal_increase_percent_attack_damage]),
                baoshang=int(rss[rank][index_deal_extra_percent_crit_damage] + rss[rank][index_deal_increase_percent_crit_damage]),
                baizi=int(rss[rank][index_deal_extra_percent_addtional_damage]),
                suogong=int(rss[rank][index_deal_extra_percent_final_damage]),
                sangong=int(rss[rank][index_deal_extra_percent_physical_magical_independent_attack_power]),
                lizhi=int(rss[rank][index_deal_extra_percent_strength_and_intelligence]),
                shuqiang=int(rss[rank][index_deal_extra_all_element_strength]),
                chixu=int(rss[rank][index_deal_extra_percent_continued_damage]),
                jigong=int(rss[rank][index_deal_extra_percent_skill_attack_power]),
                teshu=int(rss[rank][index_deal_extra_percent_special_effect]),
                gongsu=int(rss[rank][index_deal_extra_percent_attack_speed]),
                baojilv=int(rss[rank][index_deal_extra_percent_magic_physical_crit_rate]),
                dark_resistance=dark_resistance,
                moving_speed=moving_speed,
            )
            rank_stat2[rank] = (
                "   <主动>\n"
                "  1~45技能= {lv_1_45}级\n"
                "    50技能= {lv_50}级\n"
                " 60~80技能= {lv_60_80}级\n"
                "    85技能= {lv_85}级\n"
                "    95技能= {lv_95}级\n"
                "   100技能= {lv_100}级\n"
                "\n"
                "   <被动>\n"
                "  转职被动= {passive_lv_15}级\n"
                "  一觉被动= {passive_lv_48}级\n"
                "  二觉被动= {passive_lv_85}级\n"
                "  三觉被动= {passive_lv_95}级"
            ).format(
                lv_1_45=round(rss[rank][index_deal_extra_active_skill_lv_1_45], 1),
                lv_50=int(rss[rank][index_deal_extra_active_skill_lv_50]),
                lv_60_80=round(rss[rank][index_deal_extra_active_skill_lv_60_80], 1),
                lv_85=int(rss[rank][index_deal_extra_active_skill_lv_85]),
                lv_95=int(rss[rank][index_deal_extra_active_skill_lv_95]),
                lv_100=int(rss[rank][index_deal_extra_active_skill_lv_100]),
                passive_lv_15=round(rss[rank][index_deal_extra_passive_transfer_skill], 1),
                passive_lv_48=int(rss[rank][index_deal_extra_passive_first_awaken_skill]),
                passive_lv_85=int(rss[rank][index_deal_extra_passive_second_awaken_skill]),
                passive_lv_95=int(rss[rank][index_deal_extra_passive_third_awaken_skill]),
            )

        cool_check = req_cool.get()
        cool_txt = ""
        if cool_check == "O(打开)":
            cool_txt = "冷却补正"
        else:
            cool_txt = "纯伤害"
        canvas_res.create_text(122, 145, text=cool_txt, font=guide_font, fill='white')
        if int(ele_skill) != 0:
            canvas_res.create_text(122, 170, font=guide_font, fill='white',
                                   text="技能属强补正={} / 逆校正%={}%".format(int(ele_skill), round(100 * (1.05 / (1.05 + int(ele_skill) * 0.0045) - 1), 1)))
        res_dam = canvas_res.create_text(122, 125, text=extract_score_from_score_damage(rank_dam[0]), font=mid_font, fill='white')
        res_stat = canvas_res.create_text(65, 293, text=rank_stat[0], fill='white')
        res_stat2 = canvas_res.create_text(185, 293, text=rank_stat2[0], fill='white')

        res_img11 = canvas_res.create_image(57, 57, image=result_image_on[0]['11'])  # 上衣
        res_img12 = canvas_res.create_image(27, 87, image=result_image_on[0]['12'])  # 裤子
        res_img13 = canvas_res.create_image(27, 57, image=result_image_on[0]['13'])  # 头肩
        res_img14 = canvas_res.create_image(57, 87, image=result_image_on[0]['14'])  # 腰带
        res_img15 = canvas_res.create_image(27, 117, image=result_image_on[0]['15'])  # 鞋子
        res_img21 = canvas_res.create_image(189, 57, image=result_image_on[0]['21'])  # 手镯
        res_img22 = canvas_res.create_image(219, 57, image=result_image_on[0]['22'])  # 项链
        res_img23 = canvas_res.create_image(219, 87, image=result_image_on[0]['23'])  # 戒指
        res_img31 = canvas_res.create_image(189, 87, image=result_image_on[0]['31'])  # 辅助装备
        res_img32 = canvas_res.create_image(219, 117, image=result_image_on[0]['32'])  # 魔法石
        res_img33 = canvas_res.create_image(189, 117, image=result_image_on[0]['33'])  # 耳环
        gif_image_ids.extend([res_img11, res_img12, res_img13, res_img14, res_img15])
        gif_image_ids.extend([res_img21, res_img22, res_img23])
        gif_image_ids.extend([res_img31, res_img32, res_img33])

        slot_equip_name_dict = get_slot_names_dict(equip_index_to_realname, rank_setting[0][1:])
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img11, slot_equip_name_dict['11'], 57, 57))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img12, slot_equip_name_dict['12'], 27, 87))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img13, slot_equip_name_dict['13'], 27, 57))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img14, slot_equip_name_dict['14'], 57, 87))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img15, slot_equip_name_dict['15'], 27, 117))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img21, slot_equip_name_dict['21'], 189, 57))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img22, slot_equip_name_dict['22'], 219, 57))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img23, slot_equip_name_dict['23'], 219, 87))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img31, slot_equip_name_dict['31'], 189, 87))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img32, slot_equip_name_dict['32'], 219, 117))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img33, slot_equip_name_dict['33'], 189, 117))

        res_txtbbgs = [None, None, None, None, None, None]  # 0-4 => 右边的展示区间， 5 => 左边的那个百变怪
        res_imgbbgs = [None, None, None, None, None, None]  # 0-4 => 右边的展示区间， 5 => 左边的那个百变怪
        if 'bbg' in result_image_on[0]:
            res_txtbbgs[5] = canvas_res.create_text(178, 147, text="百变怪=>", font=guide_font, fill='white')
            res_imgbbgs[5] = canvas_res.create_image(219, 147, image=result_image_on[0]['bbg'])  # 百变怪
            bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_imgbbgs[5], equip_index_to_realname[rank_baibiaoguai[0]], 219, 147))
        cn1 = 0
        for rank in range(total_count):
            slot_equip_name_dict = get_slot_names_dict(equip_index_to_realname, rank_setting[rank][1:])
            for equip_slot_index in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                x, y = 268 + cn1 * 29, 67 + 78 * rank
                res_img = canvas_res.create_image(x, y, image=result_image_on[rank][str(equip_slot_index)])
                gif_image_ids.append(res_img)
                bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img, slot_equip_name_dict[str(equip_slot_index)], x, y))
                cn1 = cn1 + 1
            x, y = 268 + 1 * 29 + 4, 37 + 78 * rank
            res_img = canvas_res.create_image(x, y, image=result_image_on[rank]['weapon'])
            gif_image_ids.append(res_img)
            bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img, equip_index_to_realname[rank_setting[rank][0]], x, y))
            if 'bbg' in result_image_on[rank]:
                # res_txtbbgs[j] = canvas_res.create_text(268 + 5 * 29 + 14, 38 + 78 * j, text="百变怪=>", font=guide_font, fill='white')
                x, y = 268 + 7 * 29 + 40, 37 + 78 * rank
                res_imgbbgs[rank] = canvas_res.create_image(x, y, image=result_image_on[rank]['bbg'])
                bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_imgbbgs[rank], equip_index_to_realname[rank_baibiaoguai[rank]], x, y))
            cn1 = 0
            canvas_res.create_text(366 + 44, 34 + 78 * rank, text=rank_dam[rank], font=mid_font, fill='white')

        weapon = rank_setting[0][0]
        equips = rank_setting[0][1:]
        change_readable_result_area(weapon, equips, True)

        wep_index = weapon
        wep_image = result_image_on[0]['weapon']

        g_rank_equips = {}
        for rank in range(total_count):
            g_rank_equips[rank] = rank_setting[rank]

        length = total_count

    elif job_type == 'buf':  ##########################
        global result_image_ons, rank_bufs, rank_type_buf, res_buf, res_img_list, res_buf_list, res_buf_exs, rank_buf_exs, res_buf_type_what, rank_huanzhuang_equips, rank_not_owned_equipss, rank_baibiaoguais, rank_settings
        rank_type_buf = 3
        ## rank_setting[rank]=rank_list[a][rank][b][c]
        ## a: 0=祝福,1=크오,2=합계
        ## b: 0=계수,1=스펙or증가량
        ## c: b에서 1 선택시, 0=스펙, 1=증가량
        # ranking = [ranking1, ranking2, ranking3]
        # ranking1 = rank => [score, [taiyang_calc_wep, [bless_overview, taiyang_overview, first_awaken_passive_overview, all_score_str], baibianguai, tuple(noe), huanzhuang_equip]]
        total_rank_type_count = len(rank_list)
        rank_baibiaoguais = [0 for x in range(total_rank_type_count)]
        rank_not_owned_equipss = [0 for x in range(total_rank_type_count)]
        rank_huanzhuang_equips = [0 for x in range(total_rank_type_count)]
        rank_settings = [0 for x in range(total_rank_type_count)]
        result_image_ons = [0 for x in range(total_rank_type_count)]
        rank_bufs = [0 for x in range(total_rank_type_count)]
        rank_buf_exs = [0 for x in range(total_rank_type_count)]
        for rank_type_index in range(total_rank_type_count):
            total_count = len(rank_list[rank_type_index])
            rank_baibiaoguais[rank_type_index] = [0 for x in range(total_count)]
            rank_not_owned_equipss[rank_type_index] = [0 for x in range(total_count)]
            rank_huanzhuang_equips[rank_type_index] = [0 for x in range(total_count)]
            rank_settings[rank_type_index] = [0 for x in range(total_count)]
            result_image_ons[rank_type_index] = [{} for x in range(total_count)]
            rank_bufs[rank_type_index] = [0 for x in range(total_count)]
            rank_buf_exs[rank_type_index] = [0 for x in range(total_count)]
            for rank in range(total_count):
                rank_baibiaoguais[rank_type_index][rank] = rank_list[rank_type_index][rank][1][2]
                rank_not_owned_equipss[rank_type_index][rank] = rank_list[rank_type_index][rank][1][3]
                rank_huanzhuang_equips[rank_type_index][rank] = rank_list[rank_type_index][rank][1][4]
                rank_settings[rank_type_index][rank] = rank_list[rank_type_index][rank][1][0]
                score = rank_list[rank_type_index][rank][0]
                # 0-祝福，1-一觉，2-综合，3-祝福适用面板
                # 除面板得分外，其余的都要除10
                if rank_type_index != 3:
                    score = score / 10
                rank_bufs[rank_type_index][rank] = int(score)
                rank_buf_exs[rank_type_index][rank] = rank_list[rank_type_index][rank][1][1]
                for equip_slot_index in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                    for equip_index in rank_settings[rank_type_index][rank]:
                        if len(equip_index) != 6:
                            if equip_index[0:2] == str(equip_slot_index):
                                result_image_ons[rank_type_index][rank][str(equip_slot_index)] = image_list[equip_index]
                                # 如果该装备在额外升级的工作服或跨界装备列表中，则将其图片设为未点亮的图片，这样可以很快分辨出来
                                if equip_index in rank_not_owned_equipss[rank_type_index][rank]:
                                    result_image_ons[rank_type_index][rank][str(equip_slot_index)] = image_list2[equip_index]
                if rank_baibiaoguais[rank_type_index][rank] is not None:
                    result_image_ons[rank_type_index][rank]["bbg"] = image_list[rank_baibiaoguais[rank_type_index][rank]]
                weapon_index = rank_settings[rank_type_index][rank][0]  # 装备列表的第一个是武器的编号
                result_image_ons[rank_type_index][rank]["weapon"] = image_list[weapon_index]

        canvas_res.create_text(122, 193, font=guide_font, fill='white', text=(
            "自定义 祝福+{}级 / 自定义 一觉+{}级\n"
            "祝福数据+{} / 一觉数据+{}"
        ).format(
            custom_buf_data["bless_level"], custom_buf_data["taiyang_level"],
            custom_buf_data["bless_data"], custom_buf_data["taiyang_data"],
        ))

        res_buf = canvas_res.create_text(122, 125, text=rank_bufs[2][0], font=mid_font, fill='white')
        res_buf_type_what = canvas_res.create_text(122, 145, text="综合标准", font=guide_font, fill='white')
        res_buf_exs = [0, 0, 0]
        res_buf_exs[0] = canvas_res.create_text(123, 247, text="祝福={}".format(rank_buf_exs[2][0][0]), font=guide_font, fill='white')
        res_buf_exs[1] = canvas_res.create_text(20, 275, text="一觉={}".format(rank_buf_exs[2][0][1]), font=guide_font, anchor="w",
                                                fill='white')
        taiyang_pass_and_buf_huanzhuang = "一觉被动={}\n总计: 力智={} 三攻={}".format(rank_buf_exs[2][0][2], rank_buf_exs[2][0][4], rank_buf_exs[2][0][5])
        hz_equips = rank_huanzhuang_equips[2][0]
        if len(hz_equips) != 0:
            taiyang_pass_and_buf_huanzhuang += "\n祝福切装={}".format('/'.join([equip_index_to_realname[hz] for hz in hz_equips]))

            from_baibianguai = False
            from_transfer_or_upgrade = False
            for hz in hz_equips:
                if hz == rank_baibiaoguais[2][0]:
                    from_baibianguai = True
                elif hz in rank_not_owned_equipss[2][0]:
                    from_transfer_or_upgrade = True

            if from_baibianguai:
                taiyang_pass_and_buf_huanzhuang += "(百变怪)"
            if from_transfer_or_upgrade:
                taiyang_pass_and_buf_huanzhuang += "(跨界/升级)"
        res_buf_exs[2] = canvas_res.create_text(20, 285, text=taiyang_pass_and_buf_huanzhuang, font=guide_font, anchor="nw",
                                                fill='white')

        res_img11 = canvas_res.create_image(57, 57, image=result_image_ons[2][0]['11'])
        res_img12 = canvas_res.create_image(27, 87, image=result_image_ons[2][0]['12'])
        res_img13 = canvas_res.create_image(27, 57, image=result_image_ons[2][0]['13'])
        res_img14 = canvas_res.create_image(57, 87, image=result_image_ons[2][0]['14'])
        res_img15 = canvas_res.create_image(27, 117, image=result_image_ons[2][0]['15'])
        res_img21 = canvas_res.create_image(189, 57, image=result_image_ons[2][0]['21'])
        res_img22 = canvas_res.create_image(219, 57, image=result_image_ons[2][0]['22'])
        res_img23 = canvas_res.create_image(219, 87, image=result_image_ons[2][0]['23'])
        res_img31 = canvas_res.create_image(189, 87, image=result_image_ons[2][0]['31'])
        res_img32 = canvas_res.create_image(219, 117, image=result_image_ons[2][0]['32'])
        res_img33 = canvas_res.create_image(189, 117, image=result_image_ons[2][0]['33'])
        gif_image_ids.extend([res_img11, res_img12, res_img13, res_img14, res_img15])
        gif_image_ids.extend([res_img21, res_img22, res_img23])
        gif_image_ids.extend([res_img31, res_img32, res_img33])

        slot_equip_name_dict = get_slot_names_dict(equip_index_to_realname, rank_settings[2][0][1:])
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img11, slot_equip_name_dict['11'], 57, 57))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img12, slot_equip_name_dict['12'], 27, 87))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img13, slot_equip_name_dict['13'], 27, 57))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img14, slot_equip_name_dict['14'], 57, 87))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img15, slot_equip_name_dict['15'], 27, 117))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img21, slot_equip_name_dict['21'], 189, 57))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img22, slot_equip_name_dict['22'], 219, 57))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img23, slot_equip_name_dict['23'], 219, 87))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img31, slot_equip_name_dict['31'], 189, 87))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img32, slot_equip_name_dict['32'], 219, 117))
        bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img33, slot_equip_name_dict['33'], 189, 117))

        res_txtbbgs = [None, None, None, None, None, None]  # 0-4 => 右边的展示区间， 5 => 左边的那个百变怪
        res_imgbbgs = [None, None, None, None, None, None]  # 0-4 => 右边的展示区间， 5 => 左边的那个百变怪
        if 'bbg' in result_image_ons[2][0]:
            res_txtbbgs[5] = canvas_res.create_text(178, 147, text="百变怪=>", font=guide_font, fill='white')
            res_imgbbgs[5] = canvas_res.create_image(219, 147, image=result_image_ons[2][0]['bbg'])  # 百变怪
            bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_imgbbgs[5], equip_index_to_realname[rank_baibiaoguais[2][0]], 219, 147))
        cn1 = 0
        res_img_list = {}
        res_buf_list = {}
        for rank in range(len(rank_list[2])):
            slot_equip_name_dict = get_slot_names_dict(equip_index_to_realname, rank_settings[2][rank][1:])
            for equip_slot_index in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
                x, y = 268 + cn1 * 29, 67 + 78 * rank
                temp_res = canvas_res.create_image(x, y, image=result_image_ons[2][rank][str(equip_slot_index)])
                res_img_list[str(rank) + str(equip_slot_index)] = temp_res
                gif_image_ids.append(temp_res)
                bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, temp_res, slot_equip_name_dict[str(equip_slot_index)], x, y))
                cn1 = cn1 + 1
            x, y = 268 + 1 * 29 + 4, 37 + 78 * rank
            temp_res = canvas_res.create_image(x, y, image=result_image_ons[2][rank]['weapon'])
            res_img_list[str(rank) + 'weapon'] = temp_res
            gif_image_ids.append(temp_res)
            bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, temp_res, equip_index_to_realname[rank_settings[2][rank][0]], x, y))
            if 'bbg' in result_image_ons[2][rank]:
                x, y = 268 + 7 * 29 + 40, 37 + 78 * rank
                res_txtbbgs[rank] = canvas_res.create_text(268 + 5 * 29 + 14 + 40, 38 + 78 * rank, text="百变怪=>", font=guide_font, fill='white')
                res_imgbbgs[rank] = canvas_res.create_image(x, y, image=result_image_ons[2][rank]['bbg'])
                bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_imgbbgs[rank], equip_index_to_realname[rank_baibiaoguais[2][rank]], x, y))
            cn1 = 0
            temp_buf = canvas_res.create_text(346 + 22, 34 + 78 * rank, text=rank_bufs[2][rank], font=mid_font, fill='white')
            res_buf_list[rank] = temp_buf
        length = len(rank_list[2])
        type1_img = tkinter.PhotoImage(file='ext_img/type_bless.png')
        type2_img = tkinter.PhotoImage(file='ext_img/type_crux.png')
        type3_img = tkinter.PhotoImage(file='ext_img/type_all.png')
        type4_img = tkinter.PhotoImage(file='ext_img/type_mianban.png')
        rank_type_but1 = tkinter.Button(result_window, command=lambda: change_rank_type(1), image=type1_img, bg=dark_main, borderwidth=0, activebackground=dark_main)
        rank_type_but1.place(x=8, y=337)
        rank_type_but2 = tkinter.Button(result_window, command=lambda: change_rank_type(2), image=type2_img, bg=dark_main, borderwidth=0, activebackground=dark_main)
        rank_type_but2.place(x=8 + 1 * 57, y=337)
        rank_type_but3 = tkinter.Button(result_window, command=lambda: change_rank_type(3), image=type3_img, bg=dark_main, borderwidth=0, activebackground=dark_main)
        rank_type_but3.place(x=8 + 2 * 57, y=337)
        rank_type_but4 = tkinter.Button(result_window, command=lambda: change_rank_type(4), image=type4_img, bg=dark_main, borderwidth=0, activebackground=dark_main)
        rank_type_but4.place(x=8 + 3 * 57, y=337)
        rank_type_but1.image = type1_img
        rank_type_but2.image = type2_img
        rank_type_but3.image = type3_img
        rank_type_but4.image = type4_img

        weapon = rank_settings[2][0][0]
        equips = rank_settings[2][0][1:]
        change_readable_result_area(weapon, equips, True, hz_equips)

        wep_index = weapon
        wep_image = result_image_ons[2][0]['weapon']

        g_rank_equips = {}
        g_current_buff_type = "综合"
        for buf_type, rank_setting in [("祝福", rank_settings[0]), ("一觉", rank_settings[1]), ("综合", rank_settings[2]), ("适用面板", rank_settings[3])]:
            ranks = {}
            for rank in range(len(rank_setting)):
                ranks[rank] = rank_setting[rank]
            g_rank_equips[buf_type] = ranks

    wep_name = equip_index_to_realname[wep_index]
    job_name = jobup_select.get()
    res_txt_weapon = canvas_res.create_text(122, 20, text="(●—●)", font=guide_font, fill='white')
    res_img_weapon = canvas_res.create_image(189, 27, image=wep_image)  # 武器
    gif_image_ids.append(res_img_weapon)
    bind_tip_on_canvas_area(CanvasTipInfo(canvas_res, res_img_weapon, wep_name, 189, 27))
    canvas_res.create_text(122, 50, text="<职业>", font=guide_font, fill='white')
    canvas_res.create_text(122, 87, text=job_name, font=guide_font, fill='white')

    show_name_img = tkinter.PhotoImage(file='ext_img/show_name.png')
    res_bt_show_name = tkinter.Button(result_window, command=lambda: show_name(), image=show_name_img,
                                      bg=dark_blue, borderwidth=0, activebackground=dark_blue);
    res_bt_show_name.place(x=8, y=135)

    show_detail_img = tkinter.PhotoImage(file='ext_img/show_detail.png')

    res_bt1 = tkinter.Button(result_window, command=lambda: change_rank(0, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue);
    res_bt1.place(x=486 + 42, y=20 + 78 * 0)
    res_bt2 = tkinter.Button(result_window, command=lambda: change_rank(1, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue)
    res_bt3 = tkinter.Button(result_window, command=lambda: change_rank(2, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue)
    res_bt4 = tkinter.Button(result_window, command=lambda: change_rank(3, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue)
    res_bt5 = tkinter.Button(result_window, command=lambda: change_rank(4, job_type), image=show_detail_img,
                             bg=dark_blue, borderwidth=0, activebackground=dark_blue)
    if length > 1:
        res_bt2.place(x=486 + 42, y=20 + 78 * 1)
    if length > 2:
        res_bt3.place(x=486 + 42, y=20 + 78 * 2)
    if length > 3:
        res_bt4.place(x=486 + 42, y=20 + 78 * 3)
    if length > 4:
        res_bt5.place(x=486 + 42, y=20 + 78 * 4)

    canvas_res.image = result_bg
    res_bt1.image = show_detail_img
    res_bt_show_name.image = show_name_img


def change_rank(now, job_type):
    global g_current_rank
    g_current_rank = now

    global image_list, canvas_res, res_img_weapon, res_img11, res_img12, res_img13, res_img14, res_img15, res_img21, res_img22, res_img23, res_img31, res_img32, res_img33, res_txtbbgs, res_imgbbgs
    if job_type == 'deal':
        global res_dam, res_stat, res_stat2, rank_stat, rank_stat2, result_image_on, rank_baibiaoguai, rank_setting
        image_changed = result_image_on[now]
        canvas_res.itemconfig(res_img11, image=image_changed['11'])
        canvas_res.itemconfig(res_img12, image=image_changed['12'])
        canvas_res.itemconfig(res_img13, image=image_changed['13'])
        canvas_res.itemconfig(res_img14, image=image_changed['14'])
        canvas_res.itemconfig(res_img15, image=image_changed['15'])
        canvas_res.itemconfig(res_img21, image=image_changed['21'])
        canvas_res.itemconfig(res_img22, image=image_changed['22'])
        canvas_res.itemconfig(res_img23, image=image_changed['23'])
        canvas_res.itemconfig(res_img31, image=image_changed['31'])
        canvas_res.itemconfig(res_img32, image=image_changed['32'])
        canvas_res.itemconfig(res_img33, image=image_changed['33'])

        slot_equip_name_dict = get_slot_names_dict(equip_index_to_realname, rank_setting[now][1:])
        replace_tip(canvas_res, res_img11, slot_equip_name_dict['11'])
        replace_tip(canvas_res, res_img12, slot_equip_name_dict['12'])
        replace_tip(canvas_res, res_img13, slot_equip_name_dict['13'])
        replace_tip(canvas_res, res_img14, slot_equip_name_dict['14'])
        replace_tip(canvas_res, res_img15, slot_equip_name_dict['15'])
        replace_tip(canvas_res, res_img21, slot_equip_name_dict['21'])
        replace_tip(canvas_res, res_img22, slot_equip_name_dict['22'])
        replace_tip(canvas_res, res_img23, slot_equip_name_dict['23'])
        replace_tip(canvas_res, res_img31, slot_equip_name_dict['31'])
        replace_tip(canvas_res, res_img32, slot_equip_name_dict['32'])
        replace_tip(canvas_res, res_img33, slot_equip_name_dict['33'])

        if res_txtbbgs[5] is not None:
            canvas_res.delete(res_txtbbgs[5])
        if res_imgbbgs[5] is not None:
            canvas_res.delete(res_imgbbgs[5])
            replace_tip(canvas_res, res_imgbbgs[5], "")
        if 'bbg' in image_changed:
            res_txtbbgs[5] = canvas_res.create_text(178, 147, text="百变怪=>", fill='white')
            res_imgbbgs[5] = canvas_res.create_image(219, 147, image=image_changed['bbg'])  # 百变怪
            replace_tip(canvas_res, res_imgbbgs[5], equip_index_to_realname[rank_baibiaoguai[now]])
        else:
            res_txtbbgs[5] = None
            res_imgbbgs[5] = None
        canvas_res.itemconfig(res_dam, text=extract_score_from_score_damage(rank_dam[now]))
        canvas_res.itemconfig(res_stat, text=rank_stat[now])
        canvas_res.itemconfig(res_stat2, text=rank_stat2[now])

        current_weapon = g_rank_equips[now][0]
        current_equips = g_rank_equips[now][1:]
        # canvas_res.itemconfig(res_txt_weapon, text=equip_index_to_realname[current_weapon])
        canvas_res.itemconfig(res_img_weapon, image=image_changed['weapon'])
        replace_tip(canvas_res, res_img_weapon, equip_index_to_realname[current_weapon])
        change_readable_result_area(current_weapon, current_equips, False)

    elif job_type == 'buf':
        global result_image_ons, rank_bufs, rank_type_buf, res_buf, res_buf_exs, rank_buf_exs, rank_huanzhuang_equips, rank_not_owned_equipss, rank_baibiaoguais, rank_settings

        rank_type_index = rank_type_buf - 1

        image_changed = result_image_ons[rank_type_index][now]
        rank_changed = rank_bufs[rank_type_index][now]
        rank_buf_ex_changed = rank_buf_exs[rank_type_index]

        taiyang_pass_and_buf_huanzhuang = "一觉被动={}\n总计: 力智={} 三攻={}".format(rank_buf_ex_changed[now][2], rank_buf_ex_changed[now][4], rank_buf_ex_changed[now][5])

        hz_equips = rank_huanzhuang_equips[rank_type_index][now]
        if len(hz_equips) != 0:
            taiyang_pass_and_buf_huanzhuang += "\n祝福切装={}".format('/'.join([equip_index_to_realname[hz] for hz in hz_equips]))

            from_baibianguai = False
            from_transfer_or_upgrade = False
            for hz in hz_equips:
                if hz == rank_baibiaoguais[rank_type_index][now]:
                    from_baibianguai = True
                elif hz in rank_not_owned_equipss[rank_type_index][now]:
                    from_transfer_or_upgrade = True

            if from_baibianguai:
                taiyang_pass_and_buf_huanzhuang += "(百变怪)"
            if from_transfer_or_upgrade:
                taiyang_pass_and_buf_huanzhuang += "(跨界/升级)"

        canvas_res.itemconfig(res_buf, text=rank_changed)
        canvas_res.itemconfig(res_buf_exs[0], text="祝福=" + rank_buf_ex_changed[now][0])
        canvas_res.itemconfig(res_buf_exs[1], text="一觉=" + rank_buf_ex_changed[now][1])
        canvas_res.itemconfig(res_buf_exs[2], text=taiyang_pass_and_buf_huanzhuang)
        canvas_res.itemconfig(res_img11, image=image_changed['11'])
        canvas_res.itemconfig(res_img12, image=image_changed['12'])
        canvas_res.itemconfig(res_img13, image=image_changed['13'])
        canvas_res.itemconfig(res_img14, image=image_changed['14'])
        canvas_res.itemconfig(res_img15, image=image_changed['15'])
        canvas_res.itemconfig(res_img21, image=image_changed['21'])
        canvas_res.itemconfig(res_img22, image=image_changed['22'])
        canvas_res.itemconfig(res_img23, image=image_changed['23'])
        canvas_res.itemconfig(res_img31, image=image_changed['31'])
        canvas_res.itemconfig(res_img32, image=image_changed['32'])
        canvas_res.itemconfig(res_img33, image=image_changed['33'])

        slot_equip_name_dict = get_slot_names_dict(equip_index_to_realname, rank_settings[rank_type_index][now][1:])
        replace_tip(canvas_res, res_img11, slot_equip_name_dict['11'])
        replace_tip(canvas_res, res_img12, slot_equip_name_dict['12'])
        replace_tip(canvas_res, res_img13, slot_equip_name_dict['13'])
        replace_tip(canvas_res, res_img14, slot_equip_name_dict['14'])
        replace_tip(canvas_res, res_img15, slot_equip_name_dict['15'])
        replace_tip(canvas_res, res_img21, slot_equip_name_dict['21'])
        replace_tip(canvas_res, res_img22, slot_equip_name_dict['22'])
        replace_tip(canvas_res, res_img23, slot_equip_name_dict['23'])
        replace_tip(canvas_res, res_img31, slot_equip_name_dict['31'])
        replace_tip(canvas_res, res_img32, slot_equip_name_dict['32'])
        replace_tip(canvas_res, res_img33, slot_equip_name_dict['33'])

        if res_txtbbgs[5] is not None:
            canvas_res.delete(res_txtbbgs[5])
        if res_imgbbgs[5] is not None:
            canvas_res.delete(res_imgbbgs[5])
            replace_tip(canvas_res, res_imgbbgs[5], "")
        if 'bbg' in image_changed:
            res_txtbbgs[5] = canvas_res.create_text(178, 147, text="百变怪=>", fill='white')
            res_imgbbgs[5] = canvas_res.create_image(219, 147, image=image_changed['bbg'])  # 百变怪
            replace_tip(canvas_res, res_imgbbgs[5], equip_index_to_realname[rank_baibiaoguais[rank_type_index][0]])
        else:
            res_txtbbgs[5] = None
            res_imgbbgs[5] = None

        current_weapon = g_rank_equips[g_current_buff_type][now][0]
        current_equips = g_rank_equips[g_current_buff_type][now][1:]
        # canvas_res.itemconfig(res_txt_weapon, text=equip_index_to_realname[current_weapon])
        canvas_res.itemconfig(res_img_weapon, image=image_changed['weapon'])
        replace_tip(canvas_res, res_img_weapon, equip_index_to_realname[current_weapon])
        change_readable_result_area(current_weapon, current_equips, False, hz_equips)


def change_rank_type(in_type):
    global g_current_rank
    g_current_rank = 0
    global g_current_buff_type
    global image_list, canvas_res, res_img_weapon, res_img11, res_img12, res_img13, res_img14, res_img15, res_img21, res_img22, res_img23, res_img31, res_img32, res_img33, res_txtbbgs, res_imgbbgs
    global result_image_ons, rank_bufs, rank_type_buf, res_img_list, res_buf_list, res_buf_exs, rank_buf_exs, res_buf_type_what, rank_huanzhuang_equips, rank_not_owned_equipss, rank_baibiaoguais, rank_settings

    rank_type_index = in_type - 1

    rank_type_buf = in_type
    image_changed = result_image_ons[rank_type_index][0]
    image_changed_all = result_image_ons[rank_type_index]
    rank_changed = rank_bufs[rank_type_index]
    rank_buf_ex_changed = rank_buf_exs[rank_type_index]

    taiyang_pass_and_buf_huanzhuang = "一觉被动={}\n总计: 力智={} 三攻={}".format(rank_buf_ex_changed[0][2], rank_buf_ex_changed[0][4], rank_buf_ex_changed[0][5])

    hz_equips = rank_huanzhuang_equips[rank_type_index][0]
    if len(hz_equips) != 0:
        taiyang_pass_and_buf_huanzhuang += "\n祝福切装={}".format('/'.join([equip_index_to_realname[hz] for hz in hz_equips]))

        from_baibianguai = False
        from_transfer_or_upgrade = False
        for hz in hz_equips:
            if hz == rank_baibiaoguais[rank_type_index][0]:
                from_baibianguai = True
            elif hz in rank_not_owned_equipss[rank_type_index][0]:
                from_transfer_or_upgrade = True

        if from_baibianguai:
            taiyang_pass_and_buf_huanzhuang += "(百变怪)"
        if from_transfer_or_upgrade:
            taiyang_pass_and_buf_huanzhuang += "(跨界/升级)"

    if in_type == 1:
        type_changed = "祝福标准"
        g_current_buff_type = "祝福"
    elif in_type == 2:
        type_changed = "一觉标准"
        g_current_buff_type = "一觉"
    elif in_type == 3:
        type_changed = "综合标准"
        g_current_buff_type = "综合"
    elif in_type == 4:
        type_changed = "面板标准"
        g_current_buff_type = "适用面板"
    canvas_res.itemconfig(res_buf_type_what, text=type_changed)
    canvas_res.itemconfig(res_buf_exs[0], text="祝福=" + rank_buf_ex_changed[0][0])
    canvas_res.itemconfig(res_buf_exs[1], text="一觉=" + rank_buf_ex_changed[0][1])
    canvas_res.itemconfig(res_buf_exs[2], text=taiyang_pass_and_buf_huanzhuang)
    canvas_res.itemconfig(res_buf, text=rank_changed[0])
    canvas_res.itemconfig(res_img11, image=image_changed['11'])
    canvas_res.itemconfig(res_img12, image=image_changed['12'])
    canvas_res.itemconfig(res_img13, image=image_changed['13'])
    canvas_res.itemconfig(res_img14, image=image_changed['14'])
    canvas_res.itemconfig(res_img15, image=image_changed['15'])
    canvas_res.itemconfig(res_img21, image=image_changed['21'])
    canvas_res.itemconfig(res_img22, image=image_changed['22'])
    canvas_res.itemconfig(res_img23, image=image_changed['23'])
    canvas_res.itemconfig(res_img31, image=image_changed['31'])
    canvas_res.itemconfig(res_img32, image=image_changed['32'])
    canvas_res.itemconfig(res_img33, image=image_changed['33'])

    slot_equip_name_dict = get_slot_names_dict(equip_index_to_realname, rank_settings[rank_type_index][0][1:])
    replace_tip(canvas_res, res_img11, slot_equip_name_dict['11'])
    replace_tip(canvas_res, res_img12, slot_equip_name_dict['12'])
    replace_tip(canvas_res, res_img13, slot_equip_name_dict['13'])
    replace_tip(canvas_res, res_img14, slot_equip_name_dict['14'])
    replace_tip(canvas_res, res_img15, slot_equip_name_dict['15'])
    replace_tip(canvas_res, res_img21, slot_equip_name_dict['21'])
    replace_tip(canvas_res, res_img22, slot_equip_name_dict['22'])
    replace_tip(canvas_res, res_img23, slot_equip_name_dict['23'])
    replace_tip(canvas_res, res_img31, slot_equip_name_dict['31'])
    replace_tip(canvas_res, res_img32, slot_equip_name_dict['32'])
    replace_tip(canvas_res, res_img33, slot_equip_name_dict['33'])

    if res_txtbbgs[5] is not None:
        canvas_res.delete(res_txtbbgs[5])
    if res_imgbbgs[5] is not None:
        canvas_res.delete(res_imgbbgs[5])
        replace_tip(canvas_res, res_imgbbgs[5], "")
    if 'bbg' in image_changed:
        res_txtbbgs[5] = canvas_res.create_text(178, 147, text="百变怪=>", fill='white')
        res_imgbbgs[5] = canvas_res.create_image(219, 147, image=image_changed['bbg'])  # 百变怪
        replace_tip(canvas_res, res_imgbbgs[5], equip_index_to_realname[rank_baibiaoguais[rank_type_index][0]])
    else:
        res_txtbbgs[5] = None
        res_imgbbgs[5] = None

    cn2 = 0
    for rank in range(len(rank_changed)):
        slot_equip_name_dict = get_slot_names_dict(equip_index_to_realname, rank_settings[rank_type_index][rank][1:])
        for equip_slot_index in [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 33]:
            canvas_res.itemconfig(res_img_list[str(rank) + str(equip_slot_index)], image=image_changed_all[rank][str(equip_slot_index)])
            replace_tip(canvas_res, res_img_list[str(rank) + str(equip_slot_index)], slot_equip_name_dict[str(equip_slot_index)])
            cn2 = cn2 + 2
        canvas_res.itemconfig(res_img_list[str(rank) + 'weapon'], image=image_changed_all[rank]['weapon'])
        replace_tip(canvas_res, res_img_list[str(rank) + 'weapon'], equip_index_to_realname[rank_settings[rank_type_index][rank][0]])

        if res_txtbbgs[rank] is not None:
            canvas_res.delete(res_txtbbgs[rank])
        if res_imgbbgs[rank] is not None:
            canvas_res.delete(res_imgbbgs[rank])
            replace_tip(canvas_res, res_imgbbgs[rank], "")
        if 'bbg' in image_changed_all[rank]:
            res_txtbbgs[rank] = canvas_res.create_text(268 + 5 * 29 + 14 + 40, 38 + 78 * rank, text="百变怪=>", font=guide_font, fill='white')
            res_imgbbgs[rank] = canvas_res.create_image(268 + 7 * 29 + 40, 37 + 78 * rank, image=image_changed_all[rank]['bbg'])
            replace_tip(canvas_res, res_imgbbgs[rank], equip_index_to_realname[rank_baibiaoguais[rank_type_index][rank]])
        else:
            res_txtbbgs[rank] = None
            res_imgbbgs[rank] = None

        cn2 = 0
        canvas_res.itemconfig(res_buf_list[rank], text=rank_changed[rank], font=mid_font, fill='white')

    current_weapon = g_rank_equips[g_current_buff_type][0][0]
    current_equips = g_rank_equips[g_current_buff_type][0][1:]
    # canvas_res.itemconfig(res_txt_weapon, text=equip_index_to_realname[current_weapon])
    canvas_res.itemconfig(res_img_weapon, image=image_changed['weapon'])
    replace_tip(canvas_res, res_img_weapon, equip_index_to_realname[current_weapon])
    change_readable_result_area(current_weapon, current_equips, False, hz_equips)


def costum():
    global custom_window
    try:
        custom_window.destroy()
    except Exception as error:
        pass
    custom_window = tkinter.Toplevel(self)
    custom_window.title("统一自定义")
    custom_window.attributes("-topmost", True)
    custom_window.focus_force()
    custom_window.geometry("{}x{}+{}+{}".format(custom_window_width, custom_window_height, custom_window_x_offset, custom_window_y_offset))
    custom_window.resizable(config().main_window_resizable, config().main_window_resizable)

    load_preset = load_workbook("preset.xlsx", data_only=True)
    db_preset = load_preset["custom"]

    tkinter.Label(custom_window, text="<输出环境>", font=mid_font).place(x=100, y=10)
    tkinter.Label(custom_window, text="属性攻击=", font=guide_font).place(x=10, y=50)
    ele_list = ['火', '冰', '光', '暗']
    ele_type = tkinter.ttk.Combobox(custom_window, width=5, values=ele_list)
    ele_type.place(x=80, y=52)  ##
    ele_type.set(db_preset['B1'].value)
    tkinter.Label(custom_window, text="冷却补正比例=          %", font=guide_font).place(x=160, y=50)  ##Y11/Z11
    cool_con = tkinter.Entry(custom_window, width=5)
    cool_con.place(x=255, y=52)
    cool_con.insert(END, db_preset['B2'].value)

    tkinter.Label(custom_window, text="<特殊装备补正>", font=mid_font).place(x=100, y=85)
    tkinter.Label(custom_window, text=" 输入窗口的数值会以对应百分比加成最终数值", fg="Red").place(x=30, y=120)
    tkinter.Label(custom_window, text="歧路腰带=          %", font=guide_font).place(x=160, y=155)  ##O164
    cus1 = tkinter.Entry(custom_window, width=5)
    cus1.place(x=230, y=157)
    cus1.insert(END, db_preset['B3'].value)
    tkinter.Label(custom_window, text="歧路鞋子=          %", font=guide_font).place(x=160, y=185)  ##O180
    cus2 = tkinter.Entry(custom_window, width=5)
    cus2.place(x=230, y=187)
    cus2.insert(END, db_preset['B4'].value)
    tkinter.Label(custom_window, text="经验等级=          ", font=guide_font).place(x=160, y=215)  ##G276
    lvl_list = ['传说↓', '英雄↑']
    cus3 = tkinter.ttk.Combobox(custom_window, width=5, values=lvl_list)
    cus3.place(x=230, y=217)
    cus3.set(db_preset['B12'].value)
    tkinter.Label(custom_window, text="恍惚增幅=", font=guide_font).place(x=160, y=245)
    lvl_list = [lv for lv in range(0, 13 + 1)]
    cus4 = tkinter.ttk.Combobox(custom_window, width=2, values=lvl_list)
    cus4.place(x=230, y=247)
    cus4.set(db_preset['B13'].value)

    tkinter.Label(custom_window, text="不息上衣=          %", font=guide_font).place(x=10, y=155)  ##O100
    cus6 = tkinter.Entry(custom_window, width=5)
    cus6.place(x=80, y=157)
    cus6.insert(END, db_preset['B5'].value)
    tkinter.Label(custom_window, text="不息裤子=          %", font=guide_font).place(x=10, y=185)  ##O127
    cus7 = tkinter.Entry(custom_window, width=5)
    cus7.place(x=80, y=187)
    cus7.insert(END, db_preset['B6'].value)
    tkinter.Label(custom_window, text="不息护肩=          %", font=guide_font).place(x=10, y=215)  ##O147
    cus8 = tkinter.Entry(custom_window, width=5)
    cus8.place(x=80, y=217)
    cus8.insert(END, db_preset['B7'].value)
    tkinter.Label(custom_window, text="不息腰带=          %", font=guide_font).place(x=10, y=245)  ##O163
    cus9 = tkinter.Entry(custom_window, width=5)
    cus9.place(x=80, y=247)
    cus9.insert(END, db_preset['B8'].value)
    tkinter.Label(custom_window, text="不息鞋子=          %", font=guide_font).place(x=10, y=275)  ##O179
    cus10 = tkinter.Entry(custom_window, width=5)
    cus10.place(x=80, y=277)
    cus10.insert(END, db_preset['B9'].value)
    tkinter.Label(custom_window, text="不息2件套=         %", font=guide_font).place(x=10, y=305)  ##O295
    cus11 = tkinter.Entry(custom_window, width=5)
    cus11.place(x=80, y=307)
    cus11.insert(END, db_preset['B10'].value)
    tkinter.Label(custom_window, text="不息3件套=         %", font=guide_font).place(x=10, y=335)  ##O296,O297
    cus12 = tkinter.Entry(custom_window, width=5)
    cus12.place(x=80, y=337)
    cus12.insert(END, db_preset['B11'].value)

    tkinter.Label(custom_window, text="<奶量增幅相关>", font=mid_font, fg='blue').place(x=410, y=5)
    tkinter.Label(custom_window, text="补正辅助角色的表现", fg="Red").place(x=350, y=33)
    tkinter.Label(custom_window, text="面板体精智+          ", font=guide_font).place(x=320, y=80)  ##
    c_stat = tkinter.Entry(custom_window, width=7)
    c_stat.place(x=398, y=80)
    c_stat.insert(END, db_preset['H1'].value)
    tkinter.Label(custom_window, text="面板体精智+          ", font=guide_font).place(x=470, y=80)  ##
    b_stat = tkinter.Entry(custom_window, width=7)
    b_stat.place(x=548, y=80)
    b_stat.insert(END, db_preset['H6'].value)
    three = [0, 1, 2, 3]
    two = [0, 1, 2]
    tkinter.Label(custom_window, text="祝福称号=", font=guide_font).place(x=320, y=110)
    b_style_lvl = tkinter.ttk.Combobox(custom_window, width=5, values=three)
    b_style_lvl.place(x=390, y=112)  ##
    b_style_lvl.set(db_preset['H2'].value)
    tkinter.Label(custom_window, text="一觉称号=", font=guide_font).place(x=470, y=110)
    c_style_lvl = tkinter.ttk.Combobox(custom_window, width=5, values=two)
    c_style_lvl.place(x=540, y=112)  ##
    c_style_lvl.set(db_preset['H3'].value)
    tkinter.Label(custom_window, text="祝福等级=", font=guide_font).place(x=320, y=140)
    b_plt = tkinter.ttk.Combobox(custom_window, width=5, values=two)
    b_plt.place(x=390, y=142)  ##
    b_plt.set(db_preset['H4'].value)
    tkinter.Label(custom_window, text="祝福等级=", font=guide_font).place(x=470, y=140)
    b_cri = tkinter.ttk.Combobox(custom_window, width=5, values=[0, 1])
    b_cri.place(x=540, y=142)  ##
    b_cri.set(db_preset['H5'].value)

    tkinter.Label(custom_window, text="<属强提升>", font=mid_font).place(x=410, y=175)
    tkinter.Label(custom_window, text="基础属强=", font=guide_font).place(x=470, y=210)
    ele1 = tkinter.Entry(custom_window, width=7)
    ele1.place(x=540, y=212)  ##
    ele1.insert(END, db_preset['B14'].value)
    tkinter.Label(custom_window, text="其他属强=", font=guide_font).place(x=470, y=240)
    ele2 = tkinter.Entry(custom_window, width=7)
    ele2.place(x=540, y=242)  ##
    ele2.insert(END, db_preset['B15'].value)
    tkinter.Label(custom_window, text="勋章属强=", font=guide_font).place(x=470, y=270)
    ele3 = tkinter.Entry(custom_window, width=7)
    ele3.place(x=540, y=272)  ##
    ele3.insert(END, db_preset['B16'].value)
    tkinter.Label(custom_window, text="技能属强= ", font=guide_font).place(x=320, y=210)
    ele4 = tkinter.Entry(custom_window, width=7)
    ele4.place(x=390, y=212)  ##
    ele4.insert(END, db_preset['B17'].value)
    tkinter.Label(custom_window, text="怪物属抗=", font=guide_font).place(x=320, y=240)
    ele5 = tkinter.Entry(custom_window, width=7)
    ele5.place(x=390, y=242)  ##
    ele5.insert(END, db_preset['B18'].value)
    tkinter.Label(custom_window, text="辅助减抗=", font=guide_font).place(x=320, y=270)
    ele6 = tkinter.Entry(custom_window, width=7)
    ele6.place(x=390, y=272)  ##
    ele6.insert(END, db_preset['B19'].value)

    load_preset.close()
    save_command = lambda: save_custom(ele_type.get(), cool_con.get(), cus1.get(), cus2.get(), cus3.get(), cus4.get(),
                                       cus6.get(), cus7.get(), cus8.get(), cus9.get(), cus10.get(), cus11.get(),
                                       cus12.get(),
                                       c_stat.get(), b_stat.get(), b_style_lvl.get(), c_style_lvl.get(), b_plt.get(),
                                       b_cri.get(),
                                       ele1.get(), ele2.get(), ele3.get(), ele4.get(), ele5.get(), ele6.get())
    tkinter.Button(custom_window, text="保存", font=mid_font, command=save_command, bg="lightyellow").place(x=190, y=295)


# 根据装备编码，获取装备对应行号的字符串
def get_row(equip_index):
    return str(equip_index_to_row_index[equip_index])


def save_custom(ele_type, cool_con, cus1, cus2, cus3, cus4, cus6, cus7, cus8, cus9, cus10, cus11, cus12, c_stat, b_stat,
                b_style_lvl, c_style_lvl, b_plt, b_cri, ele1, ele2, ele3, ele4, ele5, ele6):
    global custom_window
    try:
        load_excel3 = load_workbook("DATA.xlsx")
        load_preset1 = load_workbook("preset.xlsx")
        db_custom1 = load_preset1["custom"]
        db_save_one = load_excel3["one"]
        db_save_set = load_excel3["set"]

        #########################################################
        #                     输出环境                           #
        #########################################################

        # 属性攻击
        db_custom1['B1'] = ele_type
        # 大自然防具、天劫套装、天御之灾套装会根据属性不同部位有不同的属强加成
        db_save_one["L" + get_row("12150")] = 0  # 工作服裤子
        db_save_one["L" + get_row("13150")] = 0  # 工作服头肩
        db_save_one["L" + get_row("14150")] = 0  # 工作服腰带
        db_save_one["L" + get_row("15150")] = 0  # 工作服鞋子
        db_save_one["L" + get_row("11570")] = 0  # 天劫上衣
        db_save_one["L" + get_row("12570")] = 0  # 天劫裤子
        db_save_one["L" + get_row("13570")] = 0  # 天劫头肩
        db_save_one["L" + get_row("14570")] = 0  # 天劫腰带
        db_save_one["L" + get_row("11530")] = 0  # 天御之灾上衣
        db_save_one["L" + get_row("12530")] = 0  # 天御之灾裤子
        db_save_one["L" + get_row("13530")] = 0  # 天御之灾头肩
        db_save_one["L" + get_row("14530")] = 0  # 天御之灾腰带

        if ele_type == '火':
            # 工作服头肩在火属性攻击时会增加火属性属强24点
            db_save_one["L" + get_row("13150")] = 24  # 工作服头肩
            # 天劫裤子增加火属性强化12点
            db_save_one["L" + get_row("12570")] = 12  # 天劫裤子
            # 天御之灾裤子增加火属性强化24点
            db_save_one["L" + get_row("12530")] = 24  # 天御之灾裤子
        elif ele_type == '冰':
            # 工作服腰带在冰属性攻击时会增加冰属性属强24点
            db_save_one["L" + get_row("14150")] = 24  # 工作服腰带
            # 天劫上衣增加冰属性强化12点
            db_save_one["L" + get_row("11570")] = 12  # 天劫上衣
            # 天御之灾腰带在冰属性攻击时会增加冰属性属强24点
            db_save_one["L" + get_row("14530")] = 24  # 天御之灾腰带
        elif ele_type == '光':
            # 工作服鞋子在光属性攻击时会增加光属性属强24点
            db_save_one["L" + get_row("15150")] = 24  # 工作服鞋子
            # 天劫头肩增加光属性强化12点
            db_save_one["L" + get_row("13570")] = 12  # 天劫头肩
            # 天御之灾上衣增加光属性强化24点
            db_save_one["L" + get_row("11530")] = 24  # 天御之灾上衣
        elif ele_type == '暗':
            # 工作服裤子在暗属性攻击时会增加暗属性属强24点
            db_save_one["L" + get_row("12150")] = 24  # 工作服裤子
            # 天劫腰带增加暗属性强化12点
            db_save_one["L" + get_row("14570")] = 12  # 天劫腰带
            # 天御之灾头肩增加暗属性强化24点
            db_save_one["L" + get_row("13530")] = 24  # 天御之灾头肩

        # 冷却补正比例
        db_custom1['B2'] = float(cool_con)
        # 行1：中文列名（我新加的）
        # 行2-257：装备列表
        # 行258-259：原作者加的英文缩写列名与序号
        # 行260-351：套装列表
        # 行352-353：原作者加的英文缩写列名与序号
        # 行354-361：智慧产物列表
        # 行362：套装编号与套装名称列
        # 行363-400：套装编号与套装名称
        for i in range(1, 400 + 50 + 1):  # 保险起见，多写几行，避免日后添加新装备忘记更新，反正多写不会错
            try:
                db_save_one.cell(i, 25).value = db_save_one.cell(i, 26).value * float(cool_con) / 100
            except:
                pass

        for i in range(1, 93 + 50 + 1):  # 保险起见，多写几行，避免日后添加新装备忘记更新，反正多写不会错
            try:
                db_save_set.cell(i, 25).value = db_save_one.cell(i, 26).value * float(cool_con) / 100
            except:
                pass

        #########################################################
        #                     特殊装备补正                       #
        #########################################################

        # 歧路腰带=X%
        db_custom1['B3'] = float(cus1)
        db_save_one['O' + get_row("14140")] = float(cus1)
        # 歧路鞋子=X%
        db_custom1['B4'] = float(cus2)
        db_save_one['O' + get_row("15140")] = float(cus2)
        # 不息上衣=X%
        db_custom1['B5'] = float(cus6)
        db_save_one['O' + get_row("11130")] = float(cus6)
        db_save_one['O' + get_row("11131")] = float(cus6)
        # 不息裤子=X%
        db_custom1['B6'] = float(cus7)
        db_save_one['O' + get_row("12130")] = float(cus7)
        # 不息护肩=X%
        db_custom1['B7'] = float(cus8)
        db_save_one['O' + get_row("13130")] = float(cus8)
        # 不息腰带=X%
        db_custom1['B8'] = float(cus9)
        db_save_one['O' + get_row("14130")] = float(cus9)
        # 不息鞋子=X%
        db_custom1['B9'] = float(cus10)
        db_save_one['O' + get_row("15130")] = float(cus10)
        # 不息2件套=X%
        db_custom1['B10'] = float(cus11)
        db_save_one['O' + get_row("1131")] = float(cus11)
        # 不息3件套=X%
        db_custom1['B11'] = float(cus12)
        db_save_one['O' + get_row("1132")] = float(cus12)
        db_save_one['O' + get_row("1133")] = float(cus12)
        # 经验等级=英雄↑ 或 传说↓
        db_custom1['B12'] = cus3
        if cus3 == '传说↓':
            # 传说↓
            db_save_one['J' + get_row("11060")] = 34  # 龙血玄黄-上衣
            db_save_one['F' + get_row("12060")] = 34  # 龙血玄黄-裤子
            db_save_one['N' + get_row("13060")] = 34  # 龙血玄黄-头肩
            db_save_one['L' + get_row("14060")] = 68  # 龙血玄黄-腰带
            db_save_one['K' + get_row("15060")] = 34  # 龙血玄黄-鞋子
            db_save_one['G' + get_row("1063")] = 40  # 龙血玄黄5
        else:
            # 英雄↑
            db_save_one['J' + get_row("11060")] = 35  # 龙血玄黄-上衣
            db_save_one['F' + get_row("12060")] = 35  # 龙血玄黄-裤子
            db_save_one['N' + get_row("13060")] = 35  # 龙血玄黄-头肩
            db_save_one['L' + get_row("14060")] = 72  # 龙血玄黄-腰带
            db_save_one['K' + get_row("15060")] = 35  # 龙血玄黄-鞋子
            db_save_one['G' + get_row("1063")] = 41  # 龙血玄黄5
        # 恍惚增幅等级
        db_custom1['B13'] = cus4
        db_save_one['N' + get_row("21170")] = int(cus4) + 4  # 破晓-手镯
        db_save_one['N' + get_row("21171")] = int(cus4) + 4  # 破晓-神话手镯
        db_save_one['K' + get_row("22170")] = int(cus4) + 4  # 破晓-项链
        db_save_one['E' + get_row("23170")] = int(cus4) + 4  # 破晓-戒指

        #########################################################
        #                     奶量增幅相关                       #
        #########################################################

        # 面板体精智（左边的）
        db_custom1['H1'] = c_stat
        # 面板体精智（右边的）
        db_custom1['H6'] = b_stat
        # 祝福称号等级
        db_custom1['H2'] = b_style_lvl
        # 一觉称号等级
        db_custom1['H3'] = c_style_lvl
        # 祝福等级（左）
        db_custom1['H4'] = b_plt
        # 祝福等级（右）
        db_custom1['H5'] = b_cri

        #########################################################
        #                        属强相关                        #
        #########################################################

        # 基础属强
        db_custom1['B14'] = ele1
        # 其他属强
        db_custom1['B15'] = ele2
        # 勋章属强
        db_custom1['B16'] = ele3
        # 技能属强
        db_custom1['B17'] = ele4
        # 怪物属抗
        db_custom1['B18'] = ele5
        # 辅助减抗（如奶妈-60抗性）
        db_custom1['B19'] = ele6

        load_preset1.save("preset.xlsx")
        load_preset1.close()
        load_excel3.save("DATA.xlsx")
        load_excel3.close()
        custom_window.destroy()
        tkinter.messagebox.showinfo("通知", "保存完成", parent=self)
        logger.info("save_custom({}) success".format(", ".join(str(arg) for arg in [
            ele_type, cool_con, cus1, cus2, cus3, cus4, cus6, cus7, cus8, cus9, cus10, cus11, cus12, c_stat, b_stat,
            b_style_lvl, c_style_lvl, b_plt, b_cri, ele1, ele2, ele3, ele4, ele5, ele6
        ])))
    except Exception as error:
        msg = "保存统一自定义出现错误，err={}".format(error)
        logger.error(msg)
        tkinter.messagebox.showerror("错误", msg, parent=self)


if __name__ == '__main__':
    # 上次读档/存档时的存档名
    g_save_name_index_on_last_load_or_save = 0


def load_checklist():
    if config().destroy_result_windows_when_click_load_checklist_button:
        hide_result_window_if_exists()

    ask_msg1 = tkinter.messagebox.askquestion('确认', "确认读取存档吗？", parent=self)
    if ask_msg1 == 'yes':
        load_checklist_noconfirm(current_save_name_index)


def load_checklist_noconfirm(account_index):
    global g_save_name_index_on_last_load_or_save

    try:
        load_preset3 = load_workbook("preset.xlsx")
        db_load_check = load_preset3["one"]
        load_cell = db_load_check.cell

        # 存档所对应的列，从第二列开始
        account_column = 2 + account_index

        # 读取各个装备的点亮情况
        # 1-263行为各个装备在各个存档下的点亮情况
        for row in range(1, max_save_equips + 1):
            equip_index = load_cell(row, 1).value
            if equip_index is None:
                continue
            if load_cell(row, account_column).value == 1:
                try:
                    select_item['tg{}'.format(equip_index)] = 1
                except KeyError as error:
                    pass
            else:
                try:
                    select_item['tg{}'.format(equip_index)] = 0
                except KeyError as error:
                    pass

        # 301行开始为自定义存档内容
        # 如果存在老版本自定义内容(通过判断N2各自内容是否为 武器 )，先转为新版存档格式
        if load_cell(g_old_row_custom_save_start + g_row_custom_save_weapon, g_old_col_custom_save_key).value == "武器":
            sheet_one = db_load_check
            # 转换老存档格式为新存档格式
            transfer_old_custom_save(sheet_one)
            # 保存
            load_preset3.save("preset.xlsx")

        col_custom_save_value = g_col_custom_save_value_begin + account_index

        # 职业
        job_name = load_cell(g_row_custom_save_start + g_row_custom_save_job, col_custom_save_value).value
        # 调整名称后，为保证兼容之前的存档，需要替换存档中的名称为新的名字
        if job_name is not None:
            job_name = job_name.replace("(奶系)奶妈", "(奶系)炽天使").replace("(奶系)奶萝", "(奶系)冥月女神").replace("(奶系)奶爸", "(奶系)神思者") \
                .replace("剑神", "极诣·剑魂").replace("黑暗君主", "极诣·鬼泣").replace("帝血弑天", "极诣·狂战士").replace("天帝", "极诣·阿修罗").replace("夜见罗刹", "极诣·剑影") \
                .replace("剑皇", "极诣·驭剑士").replace("裁决女神", "极诣·暗殿骑士").replace("弑神者", "极诣·契魔者").replace("剑帝", "极诣·流浪武士") \
                .replace("铁血教父", "铁血统帅") \
                .replace("极诣·气功师（男）", "归元·气功师").replace("极诣·散打（男）", "归元·散打").replace("极诣·街霸（男）", "归元·街霸").replace("极诣·柔道家（男）", "归元·柔道家")
        if job_name == "职业选择":
            job_name = jobs[0]
        jobup_select.set(job_name or jobs[0])
        current_weapons = wep_combopicker.get_selected_entrys()
        set_job_weapons()

        # 武器
        weapons = (load_cell(g_row_custom_save_start + g_row_custom_save_weapon, col_custom_save_value).value or "").split(',')
        if weapons != [""] and weapons != ["选择武器"]:
            wep_combopicker.set(weapons)

        # 输出时间
        time_select.set(load_cell(g_row_custom_save_start + g_row_custom_save_fight_time, col_custom_save_value).value or shuchu_times[0])

        # 称号
        style = load_cell(g_row_custom_save_start + g_row_custom_save_title, col_custom_save_value).value
        # 由于调整了国服特色的实现，若找不到之前版本存档的称号，则换为第一个称号
        if style not in styles():
            style = styles()[0]
        style_select.set(style)

        # 宠物
        creature = load_cell(g_row_custom_save_start + g_row_custom_save_pet, col_custom_save_value).value
        # 由于调整了国服特色的实现，若找不到之前版本存档的称号，则换为第一个称号
        if creature not in creatures():
            creature = creatures()[0]
        creature_select.set(creature)

        # 冷却补正
        req_cool.set(load_cell(g_row_custom_save_start + g_row_custom_save_cd, col_custom_save_value).value or cool_list[0])

        # 速度设置
        select_speed.set(load_cell(g_row_custom_save_start + g_row_custom_save_speed, col_custom_save_value).value or speed_middle)

        # 是否拥有百变怪
        baibianguai_select.set(
            load_cell(g_row_custom_save_start + g_row_custom_save_has_baibianguai, col_custom_save_value).value or txt_no_baibianguai)

        # 可升级的工作服数目
        can_upgrade_work_unifrom_nums_select.set(
            load_cell(g_row_custom_save_start + g_row_custom_save_can_upgrade_work_uniforms_nums, col_custom_save_value).value or
            txt_can_upgrade_work_unifrom_nums[0])

        # 跨界的来源账号（存档）列表
        transfer_equip_combopicker.set_values(get_other_account_names())
        transfer_equip_combopicker.set((load_cell(g_row_custom_save_start + g_row_custom_save_transfer_from, col_custom_save_value).value or "").split(','))

        # 最大可跨界的数目
        can_transfer_nums_select.set(load_cell(g_row_custom_save_start + g_row_custom_save_max_transfer_count, col_custom_save_value).value or txt_can_transfer_nums[0])

        # 是否默认将普雷传说加入备选池
        use_pulei_legend_by_default_select.set(load_cell(g_row_custom_save_start + g_row_custom_save_use_pulei_legend_by_default, col_custom_save_value).value or txt_not_use_pulei_legend_by_default)

        load_preset3.close()
        check_equipment()
        for set_code in set_index_2_equip_indexes:
            check_set(set_code)

        # 读档成功时更新上次存读档的存档名
        g_save_name_index_on_last_load_or_save = account_index
        logger.info("load_checklist({}) success".format(save_name_list[account_index]))
    except Exception as error:
        msg = "读档出现错误，err={}".format(error)
        logger.error(msg)
        tkinter.messagebox.showerror("错误", msg, parent=self)


def transfer_old_custom_save(sheet_one):
    # 旧版本共有10个存档，其自定义存档数据保存在N1到X13区域，
    # 也就是cell(g_old_row_custom_save_start+g_row_custom_save_save_name, g_old_col_custom_save_key)到cell(g_old_row_custom_save_start+g_row_custom_save_save_name+12, g_old_col_custom_save_key+10)
    for account_index in range(0, 10):
        ####################################### 读取历史存档的数据 #######################################
        col_old_custom_save_value = g_old_col_custom_save_value_begin + account_index
        # 武器
        weapons = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_weapon, col_old_custom_save_value).value or ""

        # 职业
        job_name = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_job, col_old_custom_save_value).value
        if job_name is not None:
            # 调整名称后，为保证兼容之前的存档，需要替换存档中的名称为新的名字
            job_name = job_name.replace("(奶系)奶妈", "(奶系)炽天使").replace("(奶系)奶萝", "(奶系)冥月女神").replace("(奶系)奶爸", "(奶系)神思者") \
                .replace("剑神", "极诣·剑魂").replace("黑暗君主", "极诣·鬼泣").replace("帝血弑天", "极诣·狂战士").replace("天帝", "极诣·阿修罗").replace("夜见罗刹", "极诣·剑影") \
                .replace("剑皇", "极诣·驭剑士").replace("裁决女神", "极诣·暗殿骑士").replace("弑神者", "极诣·契魔者").replace("剑帝", "极诣·流浪武士") \
                .replace("铁血教父", "铁血统帅")
        job_name = job_name or "职业选择"

        # 输出时间
        shuchu_time = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_fight_time, col_old_custom_save_value).value or shuchu_times[0]

        # 称号
        style = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_title, col_old_custom_save_value).value
        # 由于调整了国服特色的实现，若找不到之前版本存档的称号，则换为第一个称号
        if style not in styles():
            style = styles()[0]

        # 宠物
        creature = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_pet, col_old_custom_save_value).value
        # 由于调整了国服特色的实现，若找不到之前版本存档的称号，则换为第一个称号
        if creature not in creatures():
            creature = creatures()[0]

        # 冷却补正
        cool = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_cd, col_old_custom_save_value).value or cool_list[0]

        # 速度设置
        speed = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_speed, col_old_custom_save_value).value or speed_middle

        # 是否拥有百变怪
        baibianguai = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_has_baibianguai, col_old_custom_save_value).value or txt_no_baibianguai

        # 可升级的工作服数目
        can_upgrade_work_unifrom_nums = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_can_upgrade_work_uniforms_nums, col_old_custom_save_value).value or txt_can_upgrade_work_unifrom_nums[0]

        # 跨界的来源账号（存档）列表
        transfer_equip = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_transfer_from, col_old_custom_save_value).value or ""

        # 最大可跨界的数目
        can_transfer_nums = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_max_transfer_count, col_old_custom_save_value).value or txt_can_transfer_nums[0]

        # 是否默认将普雷传说加入备选池
        use_pulei_legend_by_default = sheet_one.cell(g_old_row_custom_save_start + g_row_custom_save_use_pulei_legend_by_default, col_old_custom_save_value).value or txt_not_use_pulei_legend_by_default

        ####################################### 保存到新的存档区域 #######################################
        col_custom_save_value = g_col_custom_save_value_begin + account_index
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_save_name, col_custom_save_value, "存档名", save_name_list[account_index])
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_weapon, col_custom_save_value, "武器", weapons)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_job, col_custom_save_value, "职业选择", job_name)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_fight_time, col_custom_save_value, "输出时间", shuchu_time)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_title, col_custom_save_value, "称号选择", style)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_pet, col_custom_save_value, "宠物选择", creature)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_cd, col_custom_save_value, "冷却补正", cool)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_speed, col_custom_save_value, "选择速度", speed)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_has_baibianguai, col_custom_save_value, "是否拥有百变怪", baibianguai)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_can_upgrade_work_uniforms_nums, col_custom_save_value, "材料够升级的工作服数目", can_upgrade_work_unifrom_nums)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_transfer_from, col_custom_save_value, "跨界来源账户列表", transfer_equip)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_max_transfer_count, col_custom_save_value, "最大跨界数目", can_transfer_nums)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_use_pulei_legend_by_default, col_custom_save_value, "是否默认将普雷传说装备加入备选池", use_pulei_legend_by_default)

    ####################################### 清空历史存档区域 #######################################
    # 旧版本共有10个存档，其自定义存档数据保存在N1到X13区域，
    # 也就是cell(g_old_row_custom_save_start+g_row_custom_save_save_name, g_old_col_custom_save_key)到cell(g_old_row_custom_save_start+g_row_custom_save_save_name+12, g_old_col_custom_save_key+10)
    for row in range(g_old_row_custom_save_start + g_row_custom_save_save_name, g_old_row_custom_save_start + g_row_custom_save_save_name + 12 + 1):
        for col in range(g_old_col_custom_save_key, g_old_col_custom_save_key + 10 + 1):
            sheet_one.cell(row, col).value = ""


def transfer_old_custom_save_ver2(sheet_one):
    # 第二代旧版本共有config().max_save_count（默认为61）个存档，其自定义存档数据保存在A301到BJ313区域，
    max_save_count = config().max_save_count
    for account_index in range(0, max_save_count):
        ####################################### 读取历史存档的数据 #######################################
        col_old_custom_save_value = g_col_custom_save_value_begin + account_index
        # 武器
        weapons = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_weapon, col_old_custom_save_value).value or ""

        # 职业
        job_name = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_job, col_old_custom_save_value).value
        if job_name is not None:
            # 调整名称后，为保证兼容之前的存档，需要替换存档中的名称为新的名字
            job_name = job_name.replace("(奶系)奶妈", "(奶系)炽天使").replace("(奶系)奶萝", "(奶系)冥月女神").replace("(奶系)奶爸", "(奶系)神思者") \
                .replace("剑神", "极诣·剑魂").replace("黑暗君主", "极诣·鬼泣").replace("帝血弑天", "极诣·狂战士").replace("天帝", "极诣·阿修罗").replace("夜见罗刹", "极诣·剑影") \
                .replace("剑皇", "极诣·驭剑士").replace("裁决女神", "极诣·暗殿骑士").replace("弑神者", "极诣·契魔者").replace("剑帝", "极诣·流浪武士") \
                .replace("铁血教父", "铁血统帅")
        job_name = job_name or "职业选择"

        # 输出时间
        shuchu_time = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_fight_time, col_old_custom_save_value).value or '20秒(觉醒占比↑)'

        # 称号
        style = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_title, col_old_custom_save_value).value
        # 由于调整了国服特色的实现，若找不到之前版本存档的称号，则换为第一个称号
        if style not in styles():
            style = styles()[0]

        # 宠物
        creature = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_pet, col_old_custom_save_value).value
        # 由于调整了国服特色的实现，若找不到之前版本存档的称号，则换为第一个称号
        if creature not in creatures():
            creature = creatures()[0]

        # 冷却补正
        cool = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_cd, col_old_custom_save_value).value or 'X(纯伤害)'

        # 速度设置
        speed = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_speed, col_old_custom_save_value).value or '中速'

        # 是否拥有百变怪
        baibianguai = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_has_baibianguai, col_old_custom_save_value).value or 'No(没有百变怪)'

        # 可升级的工作服数目
        can_upgrade_work_unifrom_nums = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_can_upgrade_work_uniforms_nums, col_old_custom_save_value).value or '材料够升级零件'

        # 跨界的来源账号（存档）列表
        transfer_equip = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_transfer_from, col_old_custom_save_value).value or ""

        # 最大可跨界的数目
        can_transfer_nums = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_max_transfer_count, col_old_custom_save_value).value or '0'

        # 是否默认将普雷传说加入备选池
        use_pulei_legend_by_default = sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_use_pulei_legend_by_default, col_old_custom_save_value).value or "不加入备选池"

        ####################################### 保存到新的存档区域 #######################################
        col_custom_save_value = g_col_custom_save_value_begin + account_index
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_save_name, col_custom_save_value, "存档名", save_name_list[account_index])
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_weapon, col_custom_save_value, "武器", weapons)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_job, col_custom_save_value, "职业选择", job_name)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_fight_time, col_custom_save_value, "输出时间", shuchu_time)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_title, col_custom_save_value, "称号选择", style)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_pet, col_custom_save_value, "宠物选择", creature)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_cd, col_custom_save_value, "冷却补正", cool)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_speed, col_custom_save_value, "选择速度", speed)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_has_baibianguai, col_custom_save_value, "是否拥有百变怪", baibianguai)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_can_upgrade_work_uniforms_nums, col_custom_save_value, "材料够升级的工作服数目", can_upgrade_work_unifrom_nums)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_transfer_from, col_custom_save_value, "跨界来源账户列表", transfer_equip)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_max_transfer_count, col_custom_save_value, "最大跨界数目", can_transfer_nums)
        save_my_custom(sheet_one.cell, g_row_custom_save_start + g_row_custom_save_use_pulei_legend_by_default, col_custom_save_value, "是否默认将普雷传说装备加入备选池", use_pulei_legend_by_default)

    ####################################### 清空历史存档区域 #######################################
    # 也就是cell(g_old_row_custom_save_start_ver2+g_row_custom_save_save_name, g_col_custom_save_key)到cell(g_old_row_custom_save_start_ver2+g_row_custom_save_use_pulei_legend_by_default, g_col_custom_save_key+61)
    for row in range(g_old_row_custom_save_start_ver2 + g_row_custom_save_save_name, g_old_row_custom_save_start_ver2 + g_row_custom_save_use_pulei_legend_by_default + 1):
        for col in range(g_col_custom_save_key, g_col_custom_save_key + max_save_count + 1):
            sheet_one.cell(row, col).value = ""


# save_idx为存档的下标，从0到9
def save_my_custom(sc, row, col_custom_save_value, name, value):
    sc(row, g_col_custom_save_key).value = name
    sc(row, col_custom_save_value).value = value


def save_checklist():
    global g_save_name_index_on_last_load_or_save

    ask_msg2 = tkinter.messagebox.askquestion('确认', "确认保存吗？", parent=self)
    if ask_msg2 == "yes" and g_save_name_index_on_last_load_or_save != current_save_name_index:
        # 如果上次读档时的存档名与当前要存档的存档名不一致时，很可能是误操作
        # 比如我选了角色A，读档，看了看，后面过了会我改为角色B，但是没有点读档，直接开始点亮操作，最后点存档，这时候会导致B的很多数据被A覆盖
        # 这货在哪个情况下额外谈弹一个确认框
        if not tkinter.messagebox.askokcancel("误操作提醒", (
                "你上次执行读档的存档名为{0}\n"
                "本次执行存档的读档名为{1}\n"
                "两者不一致，执行存档后，前者({0})的内容与上次读档、本次存档之间的操作改动将覆盖到存档{1}中\n"
                "你确定要这样做吗？").format(save_name_list[g_save_name_index_on_last_load_or_save], save_name_list[current_save_name_index]), parent=self):
            return
    try:
        if ask_msg2 == 'yes':
            load_preset4 = load_workbook("preset.xlsx")
            db_save_check = load_preset4["one"]
            save_cell = db_save_check.cell

            account_index = current_save_name_index

            # 保存装备按钮的点亮情况
            # 1-263行为各个装备在各个存档下的点亮情况
            opt_save = {}  # 装备按钮的index => 对应的行号（1-263）
            for i in range(1, max_save_equips + 1):
                equip_index = save_cell(i, 1).value
                if equip_index is None or equip_index == "":
                    continue
                opt_save[equip_index] = i

            for code in opt_save.keys():
                try:
                    if eval("select_item['tg{}']".format(code)) == 1:
                        save_cell(opt_save[code], 2 + account_index).value = 1
                except KeyError as error:
                    passss1 = 1

                try:
                    if eval("select_item['tg{}']".format(code)) == 0:
                        save_cell(opt_save[code], 2 + account_index).value = 0
                except KeyError as error:
                    passss1 = 1

                passss = 1

            # 301行开始为自定义存档内容
            col_custom_save_value = g_col_custom_save_value_begin + account_index
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_save_name, col_custom_save_value, "存档名", save_name_list[current_save_name_index])
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_weapon, col_custom_save_value, "武器", wep_combopicker.current_value)
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_job, col_custom_save_value, "职业选择", jobup_select.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_fight_time, col_custom_save_value, "输出时间", time_select.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_title, col_custom_save_value, "称号选择", style_select.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_pet, col_custom_save_value, "宠物选择", creature_select.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_cd, col_custom_save_value, "冷却补正", req_cool.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_speed, col_custom_save_value, "选择速度", select_speed.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_has_baibianguai, col_custom_save_value, "是否拥有百变怪", baibianguai_select.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_can_upgrade_work_uniforms_nums, col_custom_save_value, "材料够升级的工作服数目", can_upgrade_work_unifrom_nums_select.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_transfer_from, col_custom_save_value, "跨界来源账户列表", transfer_equip_combopicker.current_value)
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_max_transfer_count, col_custom_save_value, "最大跨界数目", can_transfer_nums_select.get())
            save_my_custom(save_cell, g_row_custom_save_start + g_row_custom_save_use_pulei_legend_by_default, col_custom_save_value, "是否默认将普雷传说装备加入备选池", use_pulei_legend_by_default_select.get())

            load_preset4.save("preset.xlsx")
            load_preset4.close()
            tkinter.messagebox.showinfo("通知", "保存完成", parent=self)

            # 存档成功时更新上次存读档的存档名
            g_save_name_index_on_last_load_or_save = current_save_name_index
            logger.info("save_checklist({}) success".format(save_name_list[current_save_name_index]))
    except PermissionError as error:
        msg = "存档出现错误，err={}".format(error)
        logger.error(msg)
        tkinter.messagebox.showerror("错误", msg, parent=self)


# 修改当前存档的存档名为新输入的名称
def change_save_name():
    global save_name_list
    try:
        current_index = current_save_name_index
        new_save_name = save_select.get()
        old_save_name = save_name_list[current_index]

        if old_save_name == new_save_name:
            tkinter.messagebox.showinfo("提示", "存档名并未改变，无需操作")
            return

        if ',' in new_save_name:
            tkinter.messagebox.showerror("错误", "存档名不允许使用 ,（英文小写逗号）")
            return

        # 保存到preset表中
        load_preset5 = load_workbook("preset.xlsx", data_only=True)

        db_custom2 = load_preset5["custom"]
        db_custom2.cell(current_index + 1, 5).value = new_save_name

        db_one = load_preset5["one"]
        db_one.cell(g_row_custom_save_start + g_row_custom_save_save_name, g_col_custom_save_value_begin + current_index, new_save_name)

        load_preset5.save("preset.xlsx")
        load_preset5.close()

        # 更新内存信息和界面信息
        save_name_list[current_index] = new_save_name
        save_select.set(new_save_name)
        save_select['values'] = save_name_list
        tkinter.messagebox.showinfo("通知", "保存完成", parent=self)
        logger.info("change save name {} => {}".format(old_save_name, new_save_name))
    except PermissionError as error:
        msg = "更改存档名出现错误，err={}".format(error)
        logger.error(msg)
        tkinter.messagebox.showerror("错误", msg, parent=self)


def update_count():
    global count_valid, count_invalid, show_number
    global showcon, all_list_num, count_start_time
    global exit_calc
    global minheap_with_queues  # type: List[MinHeapWithQueue]
    hours = 0
    minutes = 0
    seconds = 0
    using_time_str = "0s"
    processed_count = "0"
    while True:
        try:
            show_str = "{}有效搭配/{}无效".format(count_valid, count_invalid)
            if exit_calc.value == 0:
                using_time = time.time() - count_start_time
                using_time_str = format_time(using_time)

                processed_count = sum(mq.minheap.processed_result_count for mq in minheap_with_queues)

            showcon(text=(
                "用时={}\n"
                "已计算有效搭配={}"
            ).format(
                using_time_str,
                processed_count,
            ))
            time.sleep(0.1)
        except Exception as e:
            logger.warning("update_count except: {}".format(e))


def display_realtime_counting_info():
    while True:
        try:
            items, not_select_items, work_uniforms_items = get_equips()

            # 已选装备的搭配数
            all_list_num = calc_ori_counts(items)
            # 百变怪增加的搭配数
            all_list_num += calc_bbg_add_counts(items, not_select_items)
            # 额外升级的工作服增加的搭配数
            all_list_num += calc_upgrade_work_uniforms_add_counts(items, not_select_items, work_uniforms_items)

            current_equips = 0
            for slot_equips in items:
                for equip_index in slot_equips:
                    if get_set_name(equip_index) in ["36", "37", "38"]:
                        # 过滤掉100传说、普雷首饰、普雷特殊装备
                        continue
                    current_equips += 1
            total_equips = len(equip_index_2_set_index)
            percent = current_equips / total_equips * 100

            show_txt = "{}/{}({:.2f}%) N={}".format(current_equips, total_equips, percent, int(all_list_num))
            showcon2(text=show_txt)
            time.sleep(1)
        except Exception as e:
            logger.warning("display_realtime_counting_info except: {}".format(e))


# 启动时自动读取第一个配置
def load_checklist_on_start():
    load_checklist_noconfirm(0)
    logger.info("启动时自动读取首个配置完成")


# # tk image实例的名称与文件名的映射
# tkimage_name_2_filename = {}
# # gif文件的帧列表
# gif_frames = {}

def gif_ticker():
    global canvas_res
    frame_index = 0
    while True:
        try:
            cfg = config().gif
            if cfg.enable:
                invalid_btns = []
                for btn in gif_buttons:
                    try:
                        image_filename = tkimage_name_2_filename[btn["image"]]
                        if not image_filename.endswith(".gif"):
                            continue
                        # 处理一个因时序产生的问题，若在重置装备时，恰巧这里也切换了gif帧，可能使前者失效，导致图片继续为gif
                        equip_index = image_filename[:-5]
                        if select_item["tg{}".format(equip_index)] == 0:
                            btn["image"] = image_list2[equip_index]
                            continue

                        frames = gif_frames[image_filename]
                        btn["image"] = frames[frame_index % len(frames)]
                    except Exception as err:
                        invalid_btns.append(btn)
                        logger.debug("gif_ticker will remove btn={}, 可能由于关闭了额外可选的窗口，如超界普雷, err={}".format(btn, err))

                for btn in invalid_btns:
                    gif_buttons.remove(btn)

                invalid_image_ids = []
                for image_id in gif_image_ids:
                    try:
                        image_filename = tkimage_name_2_filename[canvas_res.itemcget(image_id, "image")]
                        if not image_filename.endswith(".gif"):
                            continue
                        frames = gif_frames[image_filename]
                        canvas_res.itemconfig(image_id, image=frames[frame_index % len(frames)])
                    except Exception as err:
                        invalid_image_ids.append(image_id)
                        logger.debug("gif_ticker will remove image_id={}, 可能由于关闭了结果界面, err={}".format(image_id, err))

                for image_id in invalid_image_ids:
                    gif_image_ids.remove(image_id)

            frame_index += 1
            time.sleep(1.0 / cfg.frame_rate)
        except Exception as err:
            logger.error("gif_ticker err={}".format(err))
            time.sleep(1.0)
            continue
    pass


def update_thread():
    threading.Thread(target=update_count, daemon=True).start()
    threading.Thread(target=display_realtime_counting_info, daemon=True).start()
    threading.Thread(target=load_checklist_on_start, daemon=True).start()
    threading.Thread(target=check_update_on_start, daemon=True).start()
    threading.Thread(target=gif_ticker, daemon=True).start()
    threading.Thread(target=process_async_bind_tips, daemon=True).start()


def reset_equips(layout_cfg):
    """
    :type layout_cfg: LayoutConfig
    """
    including_god = config().ui.batch_op.reset_including_god

    for block_info in layout_cfg.equip_block_infos:
        if block_info.type == EquipBlockType_Set:
            for set_code in range(block_info.set_code_start, block_info.set_code_end + 1):
                for idx, val in enumerate(block_info.slot_orders):
                    slot, god = val
                    if god == 1 and not including_god:
                        continue
                    equip_index = "{0:02}{1:02}{2:1}".format(slot, set_code, god)
                    equip_btn_index = 'tg{0}'.format(equip_index)
                    select_item[equip_btn_index] = 0
        elif block_info.type == EquipBlockType_Single:
            for idx, equip_index in enumerate(block_info.equips):
                equip_index = str(equip_index)
                equip_btn_index = 'tg{0}'.format(equip_index)
                select_item[equip_btn_index] = 0
        elif block_info.type == EquipBlockType_Nested:
            reset_equips(block_info.nested_block)
        else:
            notify_error(logger, "ui布局配置有误，不支持类型为{}的block".format(block_info.type))
            sys.exit(0)


def reset_all_equips():
    reset_equips(layout_cfg)

    # 状态检查
    check_equipment()
    for set_code in set_index_2_equip_indexes:
        check_set(set_code)


def reset():
    reset_all_equips()

    # 处理百变怪与工作服升级数目
    baibianguai_select.set(txt_no_baibianguai)
    can_upgrade_work_unifrom_nums_select.set(txt_can_upgrade_work_unifrom_nums[0])

    transfer_equip_combopicker.set(None)
    can_transfer_nums_select.set(txt_can_transfer_nums[0])

    if config().ui.batch_op.reset_including_weapon:
        wep_combopicker.set(get_job_allowed_weapons(jobup_select.get())[:1])


def check_all():
    reset_all_equips()

    # 点亮各个套装
    for set_code in set_index_2_equip_indexes:
        click_set(set_code)

    if config().ui.batch_op.check_all_including_weapon:
        wep_combopicker.set(get_job_allowed_weapons(jobup_select.get()))


###########################################################
#                         逻辑初始化                       #
###########################################################

if __name__ == '__main__':
    multiprocessingManager = multiprocessing.Manager()
    exit_calc = multiprocessingManager.Value('i', 1)
    count_valid = 0
    unique_index = 0
    count_invalid = 0
    show_number = 0
    all_list_num = 0
    g_current_rank = 0
    g_current_job = ""
    g_current_buff_type = "祝福"  # 祝福 一觉 综合
    g_rank_equips = {}
    count_start_time = time.time()  # 开始计算的时间点

    # 由于这里不需要对data.xlsx写入，设置read_only为True可以大幅度加快读取速度，在我的电脑上改动前读取耗时0.67s，占启动时间32%，改动之后用时0.1s，占启动时间4%
    try:
        load_excel1 = load_workbook("DATA.xlsx", read_only=True, data_only=True)
    except (FileNotFoundError, BadZipFile) as error:
        notify_error(logger, "data.xlsx文件不见了或格式不对，可能是未解压，请解压后再使用,err={}".format(error))
        sys.exit(-1)
    db_one = load_excel1["one"]
    name_one = {}
    equip_index_to_realname = {}
    equip_index_to_row_index = {}
    need_save_equip_indexes = set([])
    for row in db_one.rows:
        row_value = [cell.value for cell in row]
        if len(row_value) == 0 or row_value[0] is None:
            continue

        index = row_value[0]
        realname = row_value[1]
        need_save = row_value[48] == 1

        name_one[index] = row_value
        equip_index_to_realname[index] = realname
        if need_save:
            need_save_equip_indexes.add(index)
        if len(row) != 0:
            try:
                equip_index_to_row_index[index] = row[0].row
            except Exception as err:
                logger.warning("load row index failed, err={}".format(err))

    db_job = load_excel1["lvl"]
    # 角色可以使用的武器类型列表
    opt_job_allowed_weapon_types = {}
    all_job_can_use_weapon_types = ["夜雨黑瞳武器", "原初之梦武器（输出改5/奶系改3）"]
    # 角色的属强信息：0-属强，1-树强成长
    opt_job_ele = {}
    # 角色的数据
    # 0 	    1 	    2 	    3 	    4 	    5 	    6 	7 	    8 	    9 	        10 	        11 	    12 	13 	    14 	15 	16 	    17 	    18 	19 	    20 	21 	22
    # 职业被动	1觉被动	2觉被动	3觉被动	真觉醒	二觉	1觉	20秒	60秒	20秒比重	    60秒比重	    1~45	50 	60~80	85 	95 	100 	1~45	50 	60~80	85 	95 	100
    opt_job = {}
    jobs = []

    for row in db_job.rows:
        row_value = [cell.value for cell in row]
        if len(row_value) == 0:
            continue

        # 第一列为职业，第二列为可使用的武器列表，第三列为属强，第四列为属强成长，之后为该职业各个与伤害计算相关的系数
        job = row_value[0]
        if job in ["20/60s", "下标", "职业系数下标（除属强外）", "职业"]:
            continue

        opt_job_allowed_weapon_types[job] = str(row_value[1]).split("|")
        opt_job_ele[job] = row_value[2:4]
        opt_job[job] = row_value[4:]
        jobs.append(job)

    load_excel1.close()

    load_preset0 = load_workbook("preset.xlsx", data_only=True)
    db_custom = load_preset0["custom"]

    save_name_list = []
    for save_index in range(0, config().max_save_count):
        save_name = db_custom.cell(save_index + 1, 5).value
        save_name_list.append(save_name or "存档{}".format(save_index + 1))

    changed = False

    sheet_one = load_preset0["one"]

    # 检查存档是否为旧版，若是则平移到新的位置

    # 301行开始为自定义存档内容
    # 如果存在老版本自定义内容(通过判断N2各自内容是否为 武器 )，先转为新版存档格式
    if sheet_one.cell(g_old_row_custom_save_start_ver2 + g_row_custom_save_weapon, g_col_custom_save_key).value == "武器":
        # 转换老存档格式为新存档格式
        transfer_old_custom_save_ver2(sheet_one)
        changed = True

    # 检查是否有新增的存盘项
    current_save_equips = set([])
    for i in range(1, max_save_equips + 1):
        equip_index = sheet_one.cell(i, 1).value
        if equip_index is None:
            continue
        if len(equip_index) == 6:
            # 六位为武器，武器通过其他方式存储，空出位置给其他装备
            sheet_one.cell(i, 1).value = None
            changed = True
            continue
        current_save_equips.add(equip_index)

    add_save_set = need_save_equip_indexes.difference(current_save_equips)
    if len(add_save_set) != 0:
        for i in range(1, max_save_equips + 1):
            equip_index = sheet_one.cell(i, 1).value
            if equip_index is not None and equip_index != "":
                continue
            add_save = add_save_set.pop()
            sheet_one.cell(i, 1).value = add_save
            changed = True
            if len(add_save_set) == 0:
                break

    if changed:
        load_preset0.save("preset.xlsx")
    load_preset0.close()

###########################################################
#                        ui相关变量                        #
###########################################################
if __name__ == '__main__':
    select_item = {}


###########################################################
#                        ui相关函数                        #
###########################################################


def guide_speed():
    tkinter.messagebox.showinfo("准确度选择", (
        "快速=不太精确-删除单一散件\n"
        "中速=稍精确-包括散件, 神话优先\n"
        "中速(不偏好神话)=稍精确-包括散件, 神话同等优先级\n"
        "慢速=比较精确-所有限制解除(非常慢)(保留价值预估函数过滤)\n"
        "超慢速=非常精确-所有限制解除(天荒地老海枯石烂的慢)"), parent=self)


def click_equipment(code):
    code = str(code)
    equip_btn_index = 'tg{}'.format(code)
    btn = eval('select_{}'.format(code))

    if select_item[equip_btn_index] == 0:
        btn['image'] = image_list[code]
        select_item[equip_btn_index] = 1
    elif select_item[equip_btn_index] == 1:
        btn['image'] = image_list2[code]
        select_item[equip_btn_index] = 0

    if code in equip_index_2_set_index:
        check_set(equip_index_2_set_index[code])


def check_equipment():
    for equip_index, set_index in equip_index_2_set_index.items():
        equip_btn_index = 'tg{}'.format(equip_index)
        btn = eval('select_{}'.format(equip_index))
        try:
            if select_item[equip_btn_index] == 1:
                btn['image'] = image_list[equip_index]
            else:
                btn['image'] = image_list2[equip_index]
        except Exception as error:
            logger.error(str(error))


def click_set(code):
    code = str(code)
    if code not in set_index_2_equip_indexes:
        logger.error("无法找到套装{}相关的信息".format(code))
        return

    equips = set_index_2_equip_indexes[code]
    including_god = config().ui.batch_op.check_all_including_god

    # 统计套装内当前被选择的数目
    set_checked = 0
    total = 0
    for equip_index in equips:
        if is_god(equip_index) and not including_god:
            continue
        equip_btn_index = 'tg{}'.format(equip_index)
        total += 1
        if select_item[equip_btn_index] == 1:
            set_checked += 1

    # 若未满，则全选，否则反之
    if set_checked < total:
        _image_list = image_list
        _image_list_set = image_list_set
        select_result = 1
    else:
        _image_list = image_list2
        _image_list_set = image_list_set2
        select_result = 0

    for equip_index in equips:
        if is_god(equip_index) and not including_god:
            continue
        equip_btn_index = 'tg{}'.format(equip_index)
        btn = eval('select_{}'.format(equip_index))

        try:
            btn["image"] = _image_list[equip_index]
        except Exception as error:
            logger.error(str(error))
        select_item[equip_btn_index] = select_result
    set_btn = eval('set{}'.format(code))
    try:
        set_btn["image"] = _image_list_set[code]
    except Exception as error:
        logger.error(str(error))


def check_set(code):
    code = str(code)
    if code not in set_index_2_equip_indexes:
        logger.error("无法找到套装{}相关的信息".format(code))
        return

    equips = set_index_2_equip_indexes[code]

    # 统计套装内当前被选择的数目
    slot_set = set([])
    selected_equips = []
    has_god = False
    for equip_index in equips:
        equip_btn_index = 'tg{}'.format(equip_index)
        if select_item[equip_btn_index] == 1:
            slot_set.add(get_slot_index(equip_index))
            selected_equips.append(equip_index)
        if is_god(equip_index):
            has_god = True

    if has_god:
        all_checked = len(slot_set) == len(equips) - 1
    else:
        all_checked = len(selected_equips) == len(equips)

    set_btn = eval('set{}'.format(code))
    try:
        if all_checked:
            set_btn["image"] = image_list_set[code]
        else:
            set_btn["image"] = image_list_set2[code]
    except Exception as error:
        logger.error(str(error))


def donate():
    os.popen("支持一下.png")


def dunfaoff():
    webbrowser.open('https://space.bilibili.com/4952736')


def blog():
    webbrowser.open('https://blog.naver.com/dawnclass16/221837654941')


def hamjung():
    tkinter.messagebox.showinfo("제작자 크레딧",
                                "총제작자=Dawnclass(새벽반)\n이미지/그래픽=경철부동산\n직업/버퍼DB=대략볼록할철\n서버제공=던파오프\n기타조언=히든 도비 4,5,6호\n\n오류 제보는 블로그 덧글이나 던조 쪽지로", parent=self)


def get_other_account_names():
    return [name for name in save_name_list if name != save_name_list[current_save_name_index]]


def show_usage():
    webbrowser.open(os.path.realpath("./使用说明"))


# 从武器名中提取出武器类型，如夜雨黑瞳武器、光剑-星之海：巴德纳尔、短剑-信念徽章：自由分别对应夜雨黑瞳武器、光剑、短剑
def get_weapon_type(weapon_name: str) -> str:
    return weapon_name.split("-")[0]


def get_job_allowed_weapons(job_name: str):
    allowed_weapon_types = opt_job_allowed_weapon_types[job_name]

    allowed_weapons = []  # type: list[str]

    for weapon_name in wep_list:
        weapon_type = get_weapon_type(weapon_name)
        if weapon_type in allowed_weapon_types or weapon_type in all_job_can_use_weapon_types:
            allowed_weapons.append(weapon_name)

    return allowed_weapons


if __name__ == '__main__':
    # 某职业已选择的武器列表
    job_selected_weapons = {}


def on_weapon_change():
    global job_selected_weapons, wep_select_img, empty_weapon_image
    weapons = wep_combopicker.get_selected_entrys()
    job_selected_weapons[jobup_select.get()] = weapons

    # 清空原有武器
    for i, sel_wep in enumerate(wep_select_img):
        if sel_wep in gif_buttons:
            gif_buttons.remove(sel_wep)
        sel_wep.configure(image=empty_weapon_image)
        unbind_tip(wep_select_img[i])

    # 设置新的武器
    for i, weapon_name in enumerate(weapons):
        try:
            wep_select_img[i].configure(image=image_list[wep_name_to_index[weapon_name]])
            if wep_select_img[i] not in gif_buttons:
                gif_buttons.append(wep_select_img[i])

            bind_tip(wep_select_img[i], weapon_name, 32 + 31 * i, 70, 28, -28, sync=True)
        except:
            pass


# 职业选择
def set_job_weapons():
    job_name = jobup_select.get()
    current_job_weapons = get_job_allowed_weapons(job_name)
    wep_combopicker.set_values(current_job_weapons)
    selected_weapons = current_job_weapons[:1]
    if job_name in job_selected_weapons and job_selected_weapons[job_name] != [""]:
        selected_weapons = job_selected_weapons[job_name]
    wep_combopicker.set(selected_weapons)


def on_job_selected(event):
    set_job_weapons()


def reload_config_and_setting():
    load_config()
    load_settings()
    logger.info("reload_config_and_setting")
    tkinter.messagebox.showinfo("提示", "配置已重载，可继续使用")


current_save_name_index = 0


def on_save_select_change(event):
    global current_save_name_index
    current_save_name_index = event.widget.current()


def get_gif_frames(gif_path):
    im = PIL.Image.open(gif_path)
    frames = []
    for f in range(im.n_frames):
        im.seek(f)
        photoframe = PIL.ImageTk.PhotoImage(im.copy().convert('RGBA'), name="{}_frame_{}".format(gif_path, f))
        frames.append(photoframe)

    return frames


def open_setting_tool():
    threading.Thread(target=open_setting_tool_sync, daemon=True).start()


def open_setting_tool_sync():
    import subprocess
    subprocess.call("dnf_calc_setting_tool_py/setting_tool.exe", cwd="./dnf_calc_setting_tool_py")


def format_title(title: str):
    return title.replace("$version$", now_version)


if __name__ == '__main__':
    ###########################################################
    #                        tkinter初始化                    #
    ###########################################################
    cfg = config()
    bg_cfg = cfg.ui.background
    layout_cfg = cfg.ui.layout
    dark_main = from_rgb(bg_cfg.main)
    dark_sub = from_rgb(bg_cfg.sub)
    dark_blue = from_rgb(bg_cfg.blue)

    self = tkinter.Tk()
    self.title(format_title(layout_cfg.title))
    self.geometry("{}x{}+{}+{}".format(layout_cfg.window_width, layout_cfg.window_height, layout_cfg.window_x_offset, layout_cfg.window_y_offset))
    self.resizable(layout_cfg.window_width_resizable, layout_cfg.window_height_resizable)
    self.configure(bg=dark_main)
    self.iconbitmap(r'ext_img/icon.ico')

    ###########################################################
    #                      拼接ui的琐碎代码                    #
    ###########################################################
    # 鼠标在装备图标上时弹tip
    tips_image = tkinter.PhotoImage(file='ext_img/tips.png')
    async_bind_tips = []
    toplevel_2_canvas_tips_dict = {}


    def bind_tip(btn, tip, x=0, y=0, x_offset=0, y_offset=0, auto_position=False, sync=False):
        if sync:
            bind_tip_sync(btn, tip, x, y, x_offset, y_offset, auto_position)
        else:
            async_bind_tips.append((btn, tip, x, y, x_offset, y_offset, auto_position))


    def process_async_bind_tips():
        while True:
            try:
                if len(async_bind_tips) != 0:
                    for async_bind_tip in async_bind_tips:
                        bind_tip_sync(*async_bind_tip)

                    async_bind_tips.clear()
                time.sleep(1)
            except Exception as err:
                logger.error("process_async_bind_tips err={}".format(err))


    def bind_tip_sync(btn, tip, x=0, y=0, x_offset=0, y_offset=0, auto_position=False):
        if auto_position:
            btn.master.update()
            if x == 0:
                x = btn.winfo_rootx()
            if y == 0:
                y = btn.winfo_rooty()
            if x_offset == 0:
                x_offset = btn.winfo_width()
            if y_offset == 0:
                y_offset = -btn.winfo_height()
        btn.bind("<Enter>", show_equip_name_tip(btn, tip, x, y, x_offset, y_offset))
        btn.bind("<Leave>", hide_equip_name_tip(btn))


    def unbind_tip(btn):
        btn.unbind("<Enter>")
        btn.unbind("<Leave>")


    def get_canvas_tips(root_window):
        if root_window not in toplevel_2_canvas_tips_dict:
            canvas_tips = Canvas(root_window, width=78, height=20, bd=0, bg=dark_main)
            canvas_tips.create_image(0, 0, image=tips_image, anchor='nw')
            canvas_tips.place(x=-200, y=-200)
            tkinter.Misc.lift(canvas_tips)
            toplevel_2_canvas_tips_dict[root_window] = canvas_tips

        return toplevel_2_canvas_tips_dict[root_window]


    def show_equip_name_tip(btn, tip, x, y, x_offset, y_offset):
        def _show_equip_name_tip(event):
            _show_tip(btn.winfo_toplevel(), tip, x, y, x_offset, y_offset)

        return _show_equip_name_tip


    def hide_equip_name_tip(btn):
        def _hide_equip_name_tip(event):
            canvas_tips = get_canvas_tips(btn.winfo_toplevel())
            canvas_tips.delete("mouse_overlap")
            canvas_tips.place(x=-200, y=-200)

        return _hide_equip_name_tip


    # 对某个window上的某个canvas的特定区域添加tips
    class CanvasTipInfo:
        def __init__(self, canvas, item_id, tip, center_x, center_y, width=28, height=28, tip_x_offset=28, tip_y_offset=-28):
            self.window = canvas.winfo_toplevel()
            self.canvas = canvas
            self.item_id = item_id
            self.center_x = center_x
            self.center_y = center_y
            self.width = width
            self.height = height

            self.tip = tip
            self.tip_x_offset = tip_x_offset
            self.tip_y_offset = tip_y_offset


    allCanvasTipInfoList = []  # type: list[CanvasTipInfo]
    regsiteredCanvas = set([])
    regsiteredWindows = set([])


    def bind_tip_on_canvas_area(canvasTipInfo: CanvasTipInfo):
        allCanvasTipInfoList.append(canvasTipInfo)

        if canvasTipInfo.canvas not in regsiteredCanvas:
            canvasTipInfo.canvas.bind("<Motion>", _on_move_mouse_on_canvas(canvasTipInfo.window, canvasTipInfo.canvas))
            regsiteredCanvas.add(canvasTipInfo.canvas)
        if canvasTipInfo.window not in regsiteredWindows:
            canvasTipInfo.window.bind("<Destroy>", _unbind_canvas_tip(canvasTipInfo.window))
            regsiteredWindows.add(canvasTipInfo.window)


    def replace_tip(canvas, item_id, tip):
        for info in allCanvasTipInfoList:
            if info.window == canvas.winfo_toplevel() and info.canvas == canvas and info.item_id == item_id:
                info.tip = tip
                break


    def _unbind_canvas_tip(window):
        def _destroy(event):
            global allCanvasTipInfoList
            allCanvasTipInfoList = [tipInfo for tipInfo in allCanvasTipInfoList if tipInfo.window != window]

        return _destroy


    def _on_move_mouse_on_canvas(window, canvas):
        def _move(event):
            if not cfg.ui.show_equip_tips.enable:
                return

            x, y = event.x, event.y

            hide = True
            for info in allCanvasTipInfoList:
                if info.window != window or info.canvas != canvas or info.tip == "":
                    continue
                # 仅处理在对应窗口对应canvas内的事件
                if info.center_x - info.width / 2 <= x <= info.center_x + info.width / 2 and info.center_y - info.height / 2 <= y <= info.center_y + info.height / 2:
                    # 在对应区域内，则按照提示弹出tip
                    _show_tip(info.window, info.tip, info.center_x - info.width / 2, info.center_y - info.height / 2, info.tip_x_offset, info.tip_y_offset)
                    hide = False

            if hide:
                # 否则，移除tip
                _hide_tip(window)
            pass

        return _move


    def _show_tip(root_window, tip, x, y, x_offset, y_offset):
        if not cfg.ui.show_equip_tips.enable:
            return

        _hide_tip(root_window)

        lines = tip.split("\n")
        height = 20 * len(lines)
        width = 15
        for line in lines:
            width = max(width, 15 * len(line))

        canvas_tips = get_canvas_tips(root_window)
        canvas_tips.config(width=width, height=height)
        canvas_tips.create_text(width / 2, height / 2, text=tip, tags=('mouse_overlap'), font=guide_font, fill='white')
        _x = x + x_offset + 5
        _y = y + y_offset
        if _x + width >= root_window.winfo_width():
            _x = x - width - 5
        if _y <= 0:
            _y = y
        canvas_tips.place(x=_x, y=_y)


    def _hide_tip(root_window):
        canvas_tips = get_canvas_tips(root_window)
        canvas_tips.delete("mouse_overlap")
        canvas_tips.place(x=-200, y=-200)


    font_cfg = cfg.ui.fonts

    guide_font_cfg = font_cfg.guide_font
    guide_font = tkinter.font.Font(family=guide_font_cfg.family, size=guide_font_cfg.size, weight=guide_font_cfg.weight)

    mid_font_cfg = font_cfg.mid_font
    mid_font = tkinter.font.Font(family=mid_font_cfg.family, size=mid_font_cfg.size, weight=mid_font_cfg.weight)

    big_font_cfg = font_cfg.big_font
    big_font = tkinter.font.Font(family=big_font_cfg.family, size=big_font_cfg.size, weight=big_font_cfg.weight)

    ## 내부 구조 ##
    image_list = {}
    image_list2 = {}
    image_list_set = {}
    image_list_set2 = {}
    # tk image实例的名称与文件名的映射
    tkimage_name_2_filename = {}
    # gif文件的帧列表
    gif_frames = {}
    # gif按钮列表
    gif_buttons = []
    gif_image_ids = []

    # 读取装备图片
    # 通过遍历文件夹来实现加载所需的图片，而不是穷举所有可能，最后导致启动时要卡顿两秒，根据测试，目前读取图片共使用0:00:01.780298秒, 总共尝试加载6749个， 有效的加载为351个
    for filename in os.listdir("image"):
        # 目前只处理图片目录中的gif和png文件
        is_image = filename.endswith(".gif") or filename.endswith(".png")
        if not is_image:
            continue

        # 示例文件：22390240f.png
        index = filename[:-5]  # 装备的key(除去后五位后剩余的字符串)：22390240
        file_path = "image/{}".format(filename)
        newImage = PhotoImage(file=file_path, name=filename)  #
        if filename[-5] == "n":  # 根据倒数第五位决定使用哪个list
            # 神话装备会有三个文件，以11011为例，分别为11011f.png/11011n.gif/11011n.png，其中后面两个为点亮时的样式，
            # 为了跟原版一致，当是神话装备时，加载点亮样式时，优先使用gif版本的
            if is_god(index) and index in image_list and filename.endswith(".png"):
                continue
            image_list[index] = newImage
        else:
            image_list2[index] = newImage

        tkimage_name_2_filename[newImage.name] = filename
        if filename.endswith(".gif"):
            frames = get_gif_frames(file_path)
            gif_frames[filename] = frames
            for frame in frames:
                tkimage_name_2_filename[str(frame)] = filename

    # 读取套装图片
    for filename in os.listdir("set_name"):
        # 目前只处理图片目录中的gif和png文件
        is_image = filename.endswith(".png")
        if not is_image:
            continue

        # 示例文件：22390240f.png
        if filename.endswith("f.png"):
            is_light = False
            index = filename[:-5]
        else:
            is_light = True
            index = filename[:-4]
        file_path = "set_name/{}".format(filename)
        newImage = PhotoImage(file=file_path, name=filename)  #
        if is_light:
            image_list_set[index] = newImage
        else:
            image_list_set2[index] = newImage

    bg_img = PhotoImage(file="ext_img/bg_img.png")
    bg_wall = tkinter.Label(self, image=bg_img)
    bg_wall.place(x=0, y=0)

    select_speed = tkinter.ttk.Combobox(self, values=speeds, width=15)
    select_speed.place(x=145, y=11 - 4)
    select_speed.set(speed_middle)

    show_usage_img = PhotoImage(file="ext_img/show_usage.png")
    tkinter.Button(self, command=show_usage, image=show_usage_img, borderwidth=0, activebackground=dark_main,
                   bg=dark_main).place(x=29, y=7 - 4)

    reset_img = PhotoImage(file="ext_img/reset.png")
    tkinter.Button(self, command=reset, image=reset_img, borderwidth=0, activebackground=dark_main, bg=dark_main).place(
        x=300, y=476)
    check_all_img = PhotoImage(file="ext_img/check_all.png")
    tkinter.Button(self, command=check_all, image=check_all_img, borderwidth=0, activebackground=dark_main, bg=dark_main).place(
        x=360, y=476)

    wep_list = []
    wep_name_to_index = {}
    for equip_index, name in equip_index_to_realname.items():
        if not is_weapon(str(equip_index)):
            continue
        wep_list.append(name)
        wep_name_to_index[name] = equip_index

    # 输出时间
    shuchu_times = ['20秒(觉醒占比↑)', '60秒(觉醒占比↓)']
    time_select = tkinter.ttk.Combobox(self, width=13, values=shuchu_times)
    time_select.set(shuchu_times[0])
    time_select.place(x=390 - 17, y=220 + 52)

    # 武器选择
    wep_image = PhotoImage(file="ext_img/wep.png")
    wep_g = tkinter.Label(self, image=wep_image, borderwidth=0, activebackground=dark_main, bg=dark_main)
    wep_g.place(x=29, y=55 - 12 - 4)
    wep_combopicker = Combopicker(self, entrywidth=30)
    wep_combopicker.on_change = on_weapon_change
    wep_combopicker.place(x=110, y=60 - 12 - 4)

    wep_select_img = [0 for i in range(15)]
    empty_weapon_image = PhotoImage(file="ext_img/empty_wep.png")
    for i in range(len(wep_select_img)):
        wep_select_img[i] = tkinter.Label(self, bg=dark_main, bd=0, image=empty_weapon_image)
        wep_select_img[i].place(x=32 + 31 * i, y=70)

    jobup_select = tkinter.ttk.Combobox(self, width=13, values=jobs)
    jobup_select.set(jobs[0])
    set_job_weapons()
    jobup_select.place(x=390 - 17, y=190 + 52)
    jobup_select.bind("<<ComboboxSelected>>", on_job_selected)

    # 称号选择
    style_list = styles()
    style_select = tkinter.ttk.Combobox(self, width=13, values=style_list)
    style_select.set(styles()[0])
    style_select.place(x=390 - 17, y=250 + 52)

    # 宠物选择
    creature_list = creatures()
    creature_select = tkinter.ttk.Combobox(self, width=13, values=creature_list)
    creature_select.set(creatures()[0])
    creature_select.place(x=390 - 17, y=280 + 52)

    # 冷却补正
    cool_list = ['X(纯伤害)', 'O(打开)']
    req_cool = tkinter.ttk.Combobox(self, width=13, values=cool_list)
    req_cool.set(cool_list[0])
    req_cool.place(x=390 - 17, y=310 + 52)

    calc_img = PhotoImage(file="ext_img/calc.png")
    calc_btn = tkinter.Button(self, image=calc_img, borderwidth=0, activebackground=dark_main, command=calc_thread,
                              bg=dark_main)
    calc_btn.place(x=390 - 35, y=7)
    stop_img = PhotoImage(file="ext_img/stop.png")
    tkinter.Button(self, image=stop_img, borderwidth=0, activebackground=dark_main, command=stop_calc, bg=dark_main).place(
        x=390 - 35, y=62)

    # 更多国服特色
    reload_config_and_setting_img = PhotoImage(file="ext_img/reload_config_and_setting.png")
    reload_config_btn = tkinter.Button(self, image=reload_config_and_setting_img, borderwidth=0, activebackground=dark_main, command=reload_config_and_setting, bg=dark_main)
    reload_config_btn.place(x=275, y=10 - 4)
    bind_tip(reload_config_btn, "修改配置表或配置文件后\n需要点击该按钮方可生效", auto_position=True)

    custom_img = PhotoImage(file="ext_img/custom.png")
    select_custom2 = tkinter.Button(self, image=custom_img, borderwidth=0, activebackground=dark_main, command=costum,
                                    bg=dark_sub)
    select_custom2.place(x=435 + 165, y=340 - 100)

    save_select = tkinter.ttk.Combobox(self, width=8, values=save_name_list)
    save_select.place(x=345 + 165, y=410 - 100);
    save_select.set(save_name_list[0])
    save_select.bind('<<ComboboxSelected>>', on_save_select_change)
    save_img = PhotoImage(file="ext_img/SAVE.png")
    save = tkinter.Button(self, image=save_img, borderwidth=0, activebackground=dark_main, command=save_checklist,
                          bg=dark_sub)
    save.place(x=345 + 165, y=440 - 100)
    load_img = PhotoImage(file="ext_img/LOAD.png")
    load = tkinter.Button(self, image=load_img, borderwidth=0, activebackground=dark_main, command=load_checklist,
                          bg=dark_sub)
    load.place(x=435 + 165, y=440 - 100)
    change_name_img = PhotoImage(file="ext_img/name_change.png")
    change_list_but = tkinter.Button(self, image=change_name_img, borderwidth=0, activebackground=dark_main,
                                     command=change_save_name, bg=dark_sub)
    change_list_but.place(x=435 + 165, y=405 - 100)

    # 百变怪选项
    txt_no_baibianguai = 'No(没有百变怪)'
    txt_has_baibianguai = 'Yes(拥有百变怪)'
    baibianguai_txt = tkinter.Label(self, text="  百变怪  ", font=guide_font, fg="white", bg=dark_sub)
    baibianguai_txt.place(x=300, y=395)
    baibianguai_select = tkinter.ttk.Combobox(self, width=13, values=[txt_no_baibianguai, txt_has_baibianguai])
    baibianguai_select.set(txt_no_baibianguai)
    baibianguai_select.place(x=390 - 17, y=395)

    can_upgrade_work_unifrom_nums_txt = tkinter.Label(self, text="  工作服  ", font=guide_font, fg="white", bg=dark_sub)
    can_upgrade_work_unifrom_nums_txt.place(x=300, y=421)
    can_upgrade_work_unifrom_nums_select = tkinter.ttk.Combobox(self, width=13,
                                                                values=txt_can_upgrade_work_unifrom_nums)
    can_upgrade_work_unifrom_nums_select.set(txt_can_upgrade_work_unifrom_nums[0])
    can_upgrade_work_unifrom_nums_select.place(x=390 - 17, y=421)

    transfer_equip_txt = tkinter.Label(self, text="  跨界  ", font=guide_font, fg="white", bg=dark_sub)
    transfer_equip_txt.place(x=300, y=447)
    transfer_equip_combopicker = Combopicker(self, values=get_other_account_names(), entrywidth=11)
    transfer_equip_combopicker.place(x=390 - 17, y=447)

    can_transfer_nums_select = tkinter.ttk.Combobox(self, width=2, values=txt_can_transfer_nums)
    can_transfer_nums_select.set(txt_can_transfer_nums[0])
    can_transfer_nums_select.place(x=457, y=447)

    use_pulei_legend_by_default_txt = tkinter.Label(self, text="传说普雷默认", font=guide_font, fg="white", bg=dark_sub)
    use_pulei_legend_by_default_txt.place(x=510, y=240)
    use_pulei_legend_by_default_select = tkinter.ttk.Combobox(self, width=8,
                                                              values=[txt_not_use_pulei_legend_by_default, txt_use_pulei_legend_by_default])
    use_pulei_legend_by_default_select.set(txt_not_use_pulei_legend_by_default)
    use_pulei_legend_by_default_select.place(x=510, y=270)

    show_count = tkinter.Label(self, font=guide_font, fg="white", bg=dark_sub)
    show_count.place(x=490, y=40)
    showcon = show_count.configure
    show_state = tkinter.Label(self, text="计算栏", font=guide_font, fg="white", bg=dark_sub)
    show_state.place(x=490, y=20)
    showsta = show_state.configure

    display_realtime_counting_info_label = tkinter.Label(self, font=guide_font, fg="white", bg=dark_sub)
    display_realtime_counting_info_label.place(x=430, y=480)
    showcon2 = display_realtime_counting_info_label.configure

    equip_index_2_set_index = {}
    set_index_2_equip_indexes = {}


    def create_ui_layout(master, layout_cfg, is_startup=False):
        """
        根据布局配置创建ui配置
        :type master: Tk
        :type layout_cfg: LayoutConfig
        """
        for block_info in layout_cfg.equip_block_infos:
            if block_info.type == EquipBlockType_Set:
                for set_code in range(block_info.set_code_start, block_info.set_code_end + 1):
                    # 设置套装图标
                    set_index = '1{0:02}'.format(set_code)
                    x = block_info.topleft_x
                    y = block_info.topleft_y + block_info.set_equip_icon_height * (set_code - block_info.set_code_start)

                    set_btn = tkinter.Button(master, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2[set_index],
                                             command=lambda set_index=set_index: click_set(set_index))
                    set_btn.place(x=x, y=y)

                    exec("""global set{0}; set{0} = set_btn;""".format(set_index))

                    # 设置各个装备图标
                    for idx, val in enumerate(block_info.slot_orders):
                        slot, god = val
                        equip_index = "{0:02}{1:02}{2:1}".format(slot, set_code, god)

                        x = block_info.topleft_x + block_info.set_icon_width + block_info.equip_icon_width * idx

                        equip_btn_index = 'tg{0}'.format(equip_index)
                        if equip_btn_index not in select_item:
                            select_item[equip_btn_index] = 0
                        if select_item[equip_btn_index] == 1:
                            _image_list = image_list
                        else:
                            _image_list = image_list2
                        select_btn = tkinter.Button(master, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main, image=_image_list[equip_index],
                                                    command=lambda equip_index=equip_index: click_equipment(equip_index))
                        select_btn.place(x=x, y=y)

                        image_filename = tkimage_name_2_filename[image_list[equip_index].name]
                        if image_filename.endswith(".gif") and select_btn not in gif_buttons:
                            gif_buttons.append(select_btn)

                        # 设置tip
                        bind_tip(select_btn, equip_index_to_realname[equip_index], x, y, 28, -28)

                        exec("""global select_{0}; select_{0} = select_btn""".format(equip_index))

                        init_equip(equip_index, set_index)

                    check_set(set_index)

            elif block_info.type == EquipBlockType_Single:
                exec("""global single_equip_list_{0}; single_equip_list_{0} = copy.deepcopy(block_info.equips);""".format(block_info.set_index))

                if block_info.set_icon_width != 0:
                    set_btn = tkinter.Button(master, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2[block_info.set_index],
                                             command=lambda set_index=block_info.set_index: click_set(set_index))
                    set_btn.place(x=block_info.topleft_x, y=block_info.topleft_y)

                    exec("""global set{0}; set{0} = set_btn;""".format(block_info.set_index))

                for idx, equip_index in enumerate(block_info.equips):
                    equip_index = str(equip_index)
                    x = block_info.topleft_x + block_info.set_icon_width + block_info.equip_icon_width * idx
                    y = block_info.topleft_y

                    # 创建按钮
                    equip_btn_index = 'tg{0}'.format(equip_index)
                    if equip_btn_index not in select_item:
                        select_item[equip_btn_index] = 0
                    if select_item[equip_btn_index] == 1:
                        _image_list = image_list
                    else:
                        _image_list = image_list2
                    select_btn = tkinter.Button(master, relief='flat', borderwidth=0, activebackground=dark_main, bg=dark_main, image=_image_list[equip_index],
                                                command=lambda equip_index=equip_index: click_equipment(equip_index))
                    select_btn.place(x=x, y=y)

                    # 设置tip
                    bind_tip(select_btn, equip_index_to_realname[equip_index], x, y, 28, -28)

                    exec("""global select_{0}; select_{0} = select_btn""".format(equip_index))

                    init_equip(equip_index, block_info.set_index)

                check_set(block_info.set_index)

            elif block_info.type == EquipBlockType_Nested:
                nested_btn = tkinter.Button(master, bg=dark_main, borderwidth=0, activebackground=dark_main, image=image_list_set2[block_info.set_index],
                                            command=lambda block_info=block_info: open_nested_block(block_info))
                nested_btn.place(x=block_info.topleft_x, y=block_info.topleft_y)
                exec("""global nested_btn_{0}; nested_btn_{0} = nested_btn""".format(block_info.set_index))
                open_nested_block(block_info, is_startup=is_startup)
            else:
                notify_error(logger, "ui布局配置有误，不支持类型为{}的block".format(block_info.type))
                sys.exit(0)


    def init_equip(equip_index, set_index):
        equip_index = str(equip_index)

        equip_index_2_set_index[equip_index] = set_index
        if set_index not in set_index_2_equip_indexes:
            set_index_2_equip_indexes[set_index] = []
        if equip_index not in set_index_2_equip_indexes[set_index]:
            set_index_2_equip_indexes[set_index].append(equip_index)


    def open_nested_block(block_info, is_startup=False):
        """
        创建一个新的窗口，并绘制对应装备
        :type block_info: EquipBlockInfoConfig
        """
        try:
            global old_nested_window
            exec("""global old_nested_window;old_nested_window = nested_window_{0};""".format(block_info.name))
            old_nested_window.destroy()
        except Exception as error:
            pass

        dont_show = is_startup and not block_info.show_on_startup

        nested_block = block_info.nested_block
        nested_window = tkinter.Toplevel(self)
        nested_window.title(nested_block.title)
        nested_window.geometry("{}x{}+{}+{}".format(nested_block.window_width, nested_block.window_height, nested_block.window_x_offset, nested_block.window_y_offset))
        nested_window.resizable(nested_block.window_width_resizable, nested_block.window_height_resizable)
        nested_window.attributes("-topmost", True)
        nested_window.focus_force()
        nested_window.configure(bg=dark_main)
        if dont_show:
            nested_window.withdraw()

        exec("""global nested_window_{0};nested_window_{0} = nested_window;""".format(block_info.name))

        create_ui_layout(nested_window, nested_block)


    # 根据布局信息创建ui并初始化装备相关信息
    create_ui_layout(self, layout_cfg, is_startup=True)


    # # 希洛克武器词条
    # def update_inv(event):
    #     global inv_tg
    #     if inv_mod.get() == "无" or inv_mod.get() == "自动择优":
    #         if inv_mod.get() == "无":
    #             inv_tg = 0
    #         elif inv_mod.get() == "自动择优":
    #             inv_tg = 2
    #         inv_select1_1['state'] = 'disabled'
    #         inv_select1_2['state'] = 'disabled'
    #         inv_select2_1['state'] = 'disabled'
    #         inv_select2_2['state'] = 'disabled'
    #         inv_select3_1['state'] = 'disabled'
    #         inv_select3_2['state'] = 'disabled'
    #         inv_select4_1['state'] = 'disabled'
    #         inv_select4_2['state'] = 'disabled'
    #     elif inv_mod.get() == "手动选择":
    #         inv_tg = 1
    #         inv_select1_1['state'] = 'normal'
    #         inv_select1_2['state'] = 'normal'
    #         inv_select2_1['state'] = 'normal'
    #         inv_select2_2['state'] = 'normal'
    #         inv_select3_1['state'] = 'normal'
    #         inv_select3_2['state'] = 'normal'
    #         inv_select4_1['state'] = 'normal'
    #         inv_select4_2['state'] = 'normal'
    #
    #
    # inv_select3_2_values = [
    #     ['3%/60', '3%/40', '3%/20'],
    #     ['4%/3%', '3%/3%', '2%/3%'],
    #     ['4%/25', '3%/25', '2%/25'],
    #     ['3%/3%', '3%/2%', '3%/1%'],
    #     ['+185', '+155', '+125'],
    #     ['3%/+1', '2%/+1', '1%/+1'],
    # ]
    #
    #
    # def update_inv_buf(event):
    #     new_index = inv_select3_1.current()
    #     values = inv_select3_2_values[new_index]
    #     inv_select3_2['values'] = values
    #     inv_select3_2.set(values[0])
    #
    #
    # inv_select4_2_values = [
    #     ['3%/40', '3%/30', '3%/20'],
    #     ['4%/2%', '3%/2%', '2%/2%'],
    #     ['3%/25', '2%/25', '1%/25'],
    #     ['2%/3%', '2%/2%', '2%/1%'],
    #     ['+145', '+115', '+85'],
    #     ['+1/30', '+1/20', '+1/10'],
    # ]
    #
    #
    # def update_inv_buf2(event):
    #     new_index = inv_select4_1.current()
    #     values = inv_select4_2_values[new_index]
    #     inv_select4_2['values'] = values
    #     inv_select4_2.set(values[0])
    #
    #
    # tkinter.Label(self, text="希洛克武器词条", font=big_font, fg="white", bg=dark_sub).place(x=740, y=220)
    #
    # tkinter.Label(self, text="输出职业", font=guide_font, fg="white", bg=dark_sub).place(x=780, y=260)
    #
    # inv_left_x = 715
    # inv_top_y = 285
    # inv_mod_list = ["无", "手动选择", "自动择优"]
    # inv_mod = tkinter.ttk.Combobox(self, width=10, values=inv_mod_list)
    # inv_mod.place(x=inv_left_x + 60, y=inv_top_y)
    # inv_mod.set("无")
    # inv_mod.bind("<<ComboboxSelected>>", update_inv)
    #
    # inv_type_list = ["攻击时，增加X%的伤害（黄字）", "暴击时，增加X%的伤害（爆伤）", "攻击时，附加X%的伤害（白字）", "最终伤害+X%", "力智+X%", "三攻+X%"]
    # inv_value_list1 = [6, 8, 10]
    # inv_value_list2 = [3, 4, 5]
    # inv_select1_1 = tkinter.ttk.Combobox(self, width=23, values=inv_type_list)
    # inv_select1_1.place(x=inv_left_x, y=inv_top_y + 30)
    # inv_select1_1.set(inv_type_list[0])
    # inv_select1_2 = tkinter.ttk.Combobox(self, width=2, values=inv_value_list1)
    # inv_select1_2.place(x=inv_left_x + 185, y=inv_top_y + 30)
    # inv_select1_2.set(10)
    # inv_select2_1 = tkinter.ttk.Combobox(self, width=23, values=inv_type_list)
    # inv_select2_1.place(x=inv_left_x, y=inv_top_y + 30 * 2)
    # inv_select2_1.set(inv_type_list[0])
    # inv_select2_2 = tkinter.ttk.Combobox(self, width=2, values=inv_value_list2)
    # inv_select2_2.place(x=inv_left_x + 185, y=inv_top_y + 30 * 2)
    # inv_select2_2.set(5)
    #
    # tkinter.Label(self, text="奶系职业", font=guide_font, fg="white", bg=dark_sub).place(x=780, y=375)
    #
    # inv_top_y = 400
    # inv_type_list2 = ["祝福力智+X%/太阳力智+X", "祝福力智+X%/太阳力智+X%", "祝福三攻+X%/太阳力智+X", "祝福三攻+X%/太阳力智+X%", "Lv15转职被动四维+X", "祝福力智+X%/太阳Lv+X"]
    # inv_type_list2_1 = ["祝福力智+X%/太阳力智+X", "祝福力智+X%/太阳力智+X%", "祝福三攻+X%/太阳力智+X", "祝福三攻+X%/太阳力智+X%", "Lv15转职被动四维+X", "祝福Lv+X/太阳力智+X"]
    # inv_value_list3 = inv_select3_2_values[0]
    # inv_value_list3_1 = inv_select4_2_values[0]
    # inv_select3_1 = tkinter.ttk.Combobox(self, width=28, values=inv_type_list2)
    # inv_select3_1.place(x=inv_left_x, y=inv_top_y)
    # inv_select3_1.set(inv_type_list2[0])
    # inv_select3_2 = tkinter.ttk.Combobox(self, width=28, values=inv_value_list3)
    # inv_select3_2.place(x=inv_left_x, y=inv_top_y + 30)
    # inv_select3_2.set(inv_value_list3[0])
    # inv_select4_1 = tkinter.ttk.Combobox(self, width=28, values=inv_type_list2_1)
    # inv_select4_1.place(x=inv_left_x, y=inv_top_y + 30 * 2)
    # inv_select4_1.set(inv_type_list2_1[0])
    # inv_select4_2 = tkinter.ttk.Combobox(self, width=28, values=inv_value_list3_1)
    # inv_select4_2.place(x=inv_left_x, y=inv_top_y + 30 * 3)
    # inv_select4_2.set(inv_value_list3_1[0])
    # inv_select3_1.bind("<<ComboboxSelected>>", update_inv_buf)
    # inv_select4_1.bind("<<ComboboxSelected>>", update_inv_buf2)
    # update_inv(0)

    donate_image = PhotoImage(file='ext_img/donate.png')
    donate_bt = tkinter.Button(self, image=donate_image, command=donate, borderwidth=0, bg=dark_main,
                               activebackground=dark_main)
    donate_bt.place(x=625, y=550 - 28)

    open_setting_tool_image = PhotoImage(file='ext_img/open_setting_tool.png')
    open_setting_tool_btn = tkinter.Button(self, image=open_setting_tool_image, command=open_setting_tool, borderwidth=0, bg=dark_main,
                                           activebackground=dark_main)
    open_setting_tool_btn.place(x=500, y=400)
    bind_tip(open_setting_tool_btn, "点击即可打开配置工具\n设置相关内容", 500, 400, -40, -64)

    dunfaoff_image = PhotoImage(file='ext_img/dunfaoff.png')
    dunfaoff_url = tkinter.Button(self, image=dunfaoff_image, command=dunfaoff, borderwidth=0, bg=dark_main,
                                  activebackground=dark_main)
    dunfaoff_url.place(x=500 + 65, y=406)

    blog_image = PhotoImage(file='ext_img/blog.png')
    blog_url = tkinter.Button(self, image=blog_image, command=blog, borderwidth=0, bg=dark_main,
                              activebackground=dark_main)
    blog_url.place(x=500 + 135, y=408)

    maker_image = PhotoImage(file='ext_img/maker.png')
    maker = tkinter.Button(self, image=maker_image, command=hamjung, borderwidth=0, bg=dark_main,
                           activebackground=dark_main)
    version = tkinter.Label(self, text='V ' + str(now_version) + '\n' + ver_time, font=guide_font, fg="white",
                            bg=dark_main)

    maker.place(x=625, y=590)
    version.place(x=630, y=650)

###########################################################
#                 启动工作线程并进入ui主循环                #
###########################################################

if __name__ == "__main__":
    # 工作队列
    work_queue = multiprocessing.JoinableQueue()
    work_queue.cancel_join_thread()  # or else thread that puts data will not term
    producer_data.work_queue = work_queue
    # 工作进程
    workers = []
    max_thread = config().multi_threading.max_thread
    for i in range(max_thread):
        p = multiprocessing.Process(target=consumer, args=(work_queue, exit_calc, parallel_dfs), daemon=True, name="worker#{}".format(i + 1))
        p.start()
        workers.append(p)

    logger.info("已启动{}个工作进程".format(max_thread))

    # 启动主进程的一些后台线程
    update_thread()

    # 程序启动完毕
    logger.info("计算器已成功启动，欢迎使用")

    self.mainloop()
    self.quit()
