#!/usr/bin/env python
# -*- coding: utf-8 -*-
# -------------------------------
# File   : export_excel_to_txt
# Date   : 2020/4/26 0026
# Author : Chen Ji
# Email  : fzls.zju@gmail.com
# Usage  : export given excel files to txt file, such that version compare is much easier
# -------------------------------

from openpyxl import load_workbook


def dump_excels(workbook_names, save_filename):
    lines = []
    for workbook_name in sorted(workbook_names):
        lines.append('{}workbook: {}{}'.format('-' * 30, workbook_name, '-' * 30))

        load_excel = load_workbook(workbook_name, read_only=True, data_only=True)
        for sheet_name in sorted(load_excel.sheetnames):
            lines.append('{}sheet: {}{}'.format('-' * 30, sheet_name, '-' * 30))

            for row_index, row in enumerate(load_excel[sheet_name].rows):
                lines.append(','.join([str(cell.value).replace('\n', '$CRLF$') for cell in row]))

        load_excel.close()

    with open(save_filename, 'w', encoding='utf-8') as save_file:
        save_file.write('\n'.join(lines))


if __name__ == '__main__':
    dump_excels(["DATA.xlsx", "preset.XLSX"], "txt_from_excel.txt")
