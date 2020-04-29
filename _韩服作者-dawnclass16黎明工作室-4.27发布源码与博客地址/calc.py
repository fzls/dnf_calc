now_version="2.3.2"
ver_time='200416'

## 코드를 무단으로 복제하여 개조 및 배포하지 말 것##
## Copyright ⓒ 2020 Dawnclass(새벽반) dawnclass16@naver.com

import requests
from bs4 import BeautifulSoup
import urllib.request
from urllib import parse
from json import loads
import tkinter.ttk
import tkinter.font
import tkinter.messagebox
from tkinter import *
import openpyxl
from openpyxl import load_workbook
import itertools
import threading
import time
import numpy as np
from collections import Counter
from math import floor
import webbrowser
import calc_update


#https://dunfaoff.com/DawnClass.df

def _from_rgb(rgb):
    """translates an rgb tuple of int to a tkinter friendly color code
    """
    return "#%02x%02x%02x" % rgb

dark_main=_from_rgb((32, 34, 37))
dark_sub=_from_rgb((46, 49, 52))
dark_blue=_from_rgb((29, 30, 36))

    
apikey = "TQS79U4MT11jswCLHq7G260XzXU0JhGC" ##보안코드

auto_saved=0
exit_calc=0
save_name_list=[]
save_select=0
count_num=0
count_all=0
show_number=0
all_list_num=0
all_list_list_num=0


self = tkinter.Tk()
self.title("에픽 조합 자동 계산기")
self.geometry("710x720+0+0")
self.resizable(False, False)
self.configure(bg=dark_main)
self.iconbitmap(r'ext_img/icon.ico')
guide_font=tkinter.font.Font(family="맑은 고딕", size=10, weight='bold')
small_font=tkinter.font.Font(family="맑은 고딕", size=8, weight='bold')
mid_font=tkinter.font.Font(family="맑은 고딕", size=14, weight='bold')
big_font=tkinter.font.Font(family="맑은 고딕", size=18, weight='bold')

bg_img=PhotoImage(file = "ext_img/bg_img.png")
bg_wall=tkinter.Label(self,image=bg_img)
bg_wall.place(x=0,y=0)

load_excel1=load_workbook("DATA.xlsx", data_only=True)


db_one=load_excel1["one"]
opt_one={}
name_one={}
a=1
for row in db_one.rows:
    row_value=[]
    row_value_cut=[]
    for cell in row:
        row_value.append(cell.value)
        row_value_cut = row_value[2:]
    opt_one[db_one.cell(a,1).value]=row_value_cut
    name_one[db_one.cell(a,1).value]=row_value
    a=a+1

db_job=load_excel1["lvl"]
opt_job={}
opt_job_ele={}
u=1
for row in db_job.rows:
    row_value=[]
    for cell in row:
        row_value.append(cell.value)
    opt_job[db_job.cell(u,1).value]=row_value[3:]
    opt_job_ele[db_job.cell(u,1).value]=row_value[:3]
    u=u+1
del opt_job["empty"]
del opt_job["직업명"]
jobs=list(opt_job.keys())
    
    
load_preset0=load_workbook("preset.xlsx", data_only=True)
db_custom=load_preset0["custom"]
db_save=load_preset0["one"]
save_name_list=[]
for i in range(1,11):
    save_name_list.append(db_custom.cell(i,5).value)

auto_custom=0
########## 버전 최초 구동 프리셋 업데이트 ###########
try:
    print("Preset 엑셀 버전= "+str(db_custom['K1'].value))
    print("클라이언트 버전= "+now_version)
    if str(db_custom['K1'].value) != now_version:
        print("DB 업데이트")
        db_custom['K1']=now_version
        auto_custom=1
        load_preset0.save("preset.xlsx")
        load_preset0.close()
        calc_update.update_preset ## 외부모듈
    
except PermissionError as error:
    tkinter.messagebox.showerror("에러","업데이트 실패. 엑셀을 닫고 다시 실행해주세요.")
    self.destroy()
load_excel1.close()


    
## 계산 함수 ##
def calc():
    global result_window
    try:
        result_window.destroy()
    except NameError as error:
        pass
    if select_perfect.get() == '세트필터↓(느림)':
        set_perfect=1
    else:
        set_perfect=0
    showsta(text="조합 알고리즘 구동 준비중...")
    start_time = time.time()
    load_excel=load_workbook("DATA.xlsx",data_only=True)

    db_one=load_excel["one"]
    opt_one={}
    name_one={}
    a=1
    for row in db_one.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_one[db_one.cell(a,1).value]=row_value_cut
        name_one[db_one.cell(a,1).value]=row_value
        a=a+1

    b=1        
    db_set=load_excel["set"]
    opt_set={}
    for row in db_set.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_set[db_set.cell(b,1).value]=row_value_cut ## DB 불러오기 ##
        b=b+1

    c=1        
    db_buf=load_excel["buf"]
    opt_buf={}
    name_buf={}
    for row in db_buf.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_buf[db_buf.cell(c,1).value]=row_value_cut ## DB 불러오기 ##
        name_buf[db_buf.cell(c,1).value]=row_value
        c=c+1

    d=1        
    db_buflvl=load_excel["buflvl"]
    opt_buflvl={}
    for row in db_buflvl.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = [0] + row_value[1:]
        opt_buflvl[db_buflvl.cell(d,1).value]=row_value_cut
        d=d+1

    load_presetc=load_workbook("preset.xlsx", data_only=True)
    db_preset=load_presetc["custom"]
    ele_skill=int(opt_job_ele[jobup_select.get()][1])
    ele_in=(int(db_preset["B14"].value)+int(db_preset["B15"].value)+int(db_preset["B16"].value)+
            int(ele_skill)-int(db_preset["B18"].value)+int(db_preset["B19"].value)+13)
    cool_eff=float(db_preset["B2"].value)/100

    betterang=int(db_one["J86"].value)
    
    global count_num, count_all, show_number,all_list_list_num, max_setopt
    count_num=0;count_all=0;show_number=0;metamong=0
    
    
    if jobup_select.get()[-4:] == "(진각)":
        active_eff_one=15
        active_eff_set=18-3
    else:
        active_eff_one=21
        active_eff_set=27-3

    if time_select.get() == "60초(각성비중↓)":
        lvl_shift=6
    else:
        lvl_shift=0

    job_lv1=opt_job[jobup_select.get()][11+lvl_shift]
    job_lv2=opt_job[jobup_select.get()][12+lvl_shift]
    job_lv3=opt_job[jobup_select.get()][13+lvl_shift]
    job_lv4=opt_job[jobup_select.get()][14+lvl_shift]
    job_lv5=opt_job[jobup_select.get()][15+lvl_shift]
    job_lv6=opt_job[jobup_select.get()][16+lvl_shift]
    job_pas0=opt_job[jobup_select.get()][0]
    job_pas1=opt_job[jobup_select.get()][1]
    job_pas2=opt_job[jobup_select.get()][2]
    job_pas3=opt_job[jobup_select.get()][3]
    
    
    extra_dam=0
    extra_cri=0
    extra_bon=0
    extra_all=0
    extra_pas2=0
    if style_select.get() == '증뎀칭호':
        extra_dam=extra_dam+10
    if style_select.get() == '추뎀칭호':
        extra_bon=extra_bon+10
    if creature_select.get() == '모공크리쳐':
        extra_all=extra_all+15
    if creature_select.get() == '크증크리쳐':
        extra_cri=extra_cri+18
        extra_pas2=extra_pas2+1

    if req_cool.get()=='X(순데미지)':
        cool_on=0
    else:
        cool_on=1
    list11=[];list12=[];list13=[];list14=[];list15=[]
    list21=[];list22=[];list23=[];list31=[];list32=[];list33=[]
    list11_0=[];list11_1=[];list21_0=[];list21_1=[];list33_0=[];list33_1=[]
    list_setnum=[];list_num=[]
    for i in range(101,199):
        try:
            if eval('select_item["tg1{}0"]'.format(i)) == 1:
                list11.append('1'+str(i)+'0')
                list11_0.append('1'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    
    for i in range(201,299):
        try:
            if eval('select_item["tg1{}0"]'.format(i)) == 1:
                list12.append('1'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(301,399):
        try:
            if eval('select_item["tg1{}0"]'.format(i)) == 1:
                list13.append('1'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(401,499):
        try:
            if eval('select_item["tg1{}0"]'.format(i)) == 1:
                list14.append('1'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(501,599):
        try:
            if eval('select_item["tg1{}0"]'.format(i)) == 1:
                list15.append('1'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(101,199):
        try:
            if eval('select_item["tg2{}0"]'.format(i)) == 1:
                list21.append('2'+str(i)+'0')
                list21_0.append('2'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(201,299):
        try:
            if eval('select_item["tg2{}0"]'.format(i)) == 1:
                list22.append('2'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(301,399):
        try:
            if eval('select_item["tg2{}0"]'.format(i)) == 1:
                list23.append('2'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(101,199):
        try:
            if eval('select_item["tg3{}0"]'.format(i)) == 1:
                list31.append('3'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(201,299):
        try:
            if eval('select_item["tg3{}0"]'.format(i)) == 1:
                list32.append('3'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(301,399):
        try:
            if eval('select_item["tg3{}0"]'.format(i)) == 1:
                list33.append('3'+str(i)+'0')
                list33_0.append('3'+str(i)+'0')
                list_num.append(str(i)[1:]+'0')
                list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
        
    algo_list=['11','12','13','14','15','21','22','23','31','32','33']
    if select_perfect.get() == '단품제외(빠름)':
        for i in list_num:
            if list_num.count(i)==1:
                if i[-1]!='1':
                    for ca in algo_list:
                        try:
                            eval("list{}.remove('{}{}')".format(ca,ca,i))
                            eval("list{}_0.remove('{}{}')".format(ca,ca,i))
                        except:
                            c=1

    list_setnum2=list_setnum
    set_num_dict2=Counter(list_setnum2)
                            
    for i in range(101,199):
        try:
            if eval('select_item["tg1{}1"]'.format(i)) == 1:
                list11.append('1'+str(i)+'1')
                list11_1.append('1'+str(i)+'1')
                if list11.count('1'+str(i)+'0')==0:
                    list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(101,199):
        try:
            if eval('select_item["tg2{}1"]'.format(i)) == 1:
                list21.append('2'+str(i)+'1')
                list21_1.append('2'+str(i)+'1')
                if list21.count('2'+str(i)+'0')==0:
                    list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1
    for i in range(301,399):
        try:
            if eval('select_item["tg3{}1"]'.format(i)) == 1:
                list33.append('3'+str(i)+'1')
                list33_1.append('3'+str(i)+'1')
                if list33.count('3'+str(i)+'0')==0:
                    list_setnum.append('1'+str(i)[1:3])
        except KeyError as error:
            c=1

    set_num_dict=Counter(list_setnum)
    
    for know_one in know_list:
        if eval('select_item["tg{}"]'.format(know_one)) == 1:
            eval('list{}.append(str({}))'.format(know_one[0:2],know_one))
            if know_one[0:2] =='11':
                eval('list{}_0.append(str({}))'.format(know_one[0:2],know_one))
            if know_one[0:2] =='21':
                eval('list{}_0.append(str({}))'.format(know_one[0:2],know_one))
            if know_one[0:2] =='33':
                eval('list{}_0.append(str({}))'.format(know_one[0:2],know_one))

    all_list_num=0
    all_list_list_num=0
    all_list_list=[]

##레전기본값##
##########################################################################################################################
                
    global default_legend,default_chawon,default_old
    if default_legend==1:
        df11='11360';df12='12360';df13='13360';df14='14360';df15='15360'
        df21='21370';df22='22370';df23='23370';df31='31380';df32='32380';df33='33380'
    elif default_chawon==1:
        df11='11440';df12='12440';df13='13440';df14='14440';df15='15440'
        df21='21450';df22='22450';df23='23450';df31='31460';df32='32460';df33='33460'
    elif default_old==1:
        df11='11470';df12='12470';df13='13470';df14='14470';df15='15470'
        df21='21480';df22='22480';df23='23480';df31='31490';df32='32490';df33='33490'

    if len(list11_0)==0:
        list11.append(df11);list12.append(df12);list13.append(df13);list14.append(df14);list15.append(df15);list11_0.append(df11)
    elif len(list12)==0:
        list11.append(df11);list12.append(df12);list13.append(df13);list14.append(df14);list15.append(df15);list11_0.append(df11)
    elif len(list13)==0:
        list11.append(df11);list12.append(df12);list13.append(df13);list14.append(df14);list15.append(df15);list11_0.append(df11)
    elif len(list14)==0:
        list11.append(df11);list12.append(df12);list13.append(df13);list14.append(df14);list15.append(df15);list11_0.append(df11)
    elif len(list15)==0:
        list11.append(df11);list12.append(df12);list13.append(df13);list14.append(df14);list15.append(df15);list11_0.append(df11)
        
    if len(list21_0)==0 and len(list22)==0:
        list21.append(df21);list22.append(df22);list23.append(df23);list21_0.append(df21)
    elif len(list22)==0 and len(list23)==0:
        list21.append(df21);list22.append(df22);list23.append(df23);list21_0.append(df21)
    elif len(list23)==0 and len(list21_0)==0:
        list21.append(df21);list22.append(df22);list23.append(df23);list21_0.append(df21)
        
    if len(list31)==0 and len(list32)==0:
        list31.append(df31);list32.append(df32);list33.append(df33);list33_0.append(df33)
    elif len(list32)==0 and len(list33_0)==0:
        list31.append(df31);list32.append(df32);list33.append(df33);list33_0.append(df33)
    elif len(list33_0)==0 and len(list31)==0:
        list31.append(df31);list32.append(df32);list33.append(df33);list33_0.append(df33)

    if len(list21_0)==0:
        list21.append(df21);list21_0.append(df21)
    if len(list22)==0:
        list22.append(df22)
    if len(list23)==0:
        list23.append(df23)
    if len(list31)==0:
        list31.append(df31)
    if len(list32)==0:
        list32.append(df32)
    if len(list33_0)==0:
        list33.append(df33);list33_0.append(df33)

    set_max_list1=[]
    set_max_list2=[]
    set_max_list3=[]
    print(set_num_dict)
    for i in range(0,24):
        if set_num_dict.get(str(i+100)) != None:
            if i < 16:
                set_max_list1.append(set_num_dict.get(str(i+100)))
            elif i < 20:
                set_max_list2.append(set_num_dict.get(str(i+100)))
            else:
                set_max_list3.append(set_num_dict.get(str(i+100)))
        else:
            if i < 16:
                set_max_list1.append(0)
            elif i < 20:
                set_max_list2.append(0)
            else:
                set_max_list3.append(0)
                
    if max(set_max_list1) < 2:
        items0=[[df11],[df12],[df13],[df14],[df15],list21_0,list22,list23,list31,list32,list33_0]
        items1=[]
        items2=[[df11],[df12],[df13],[df14],[df15],list21_1,list22,list23,list31,list32,list33_0]
        items3=[[df11],[df12],[df13],[df14],[df15],list21_0,list22,list23,list31,list32,list33_1]
        all_list=list(itertools.product(*items0))
        all_list1=[]
        if len(list21_1) != 0:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        if len(list33_1) != 0:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
    if max(set_max_list2) < 2:
        items0=[list11_0,list12,list13,list14,list15,[df21],[df22],[df23],list31,list32,list33_0]
        items1=[list11_1,list12,list13,list14,list15,[df21],[df22],[df23],list31,list32,list33_0]
        items2=[]
        items3=[list11_0,list12,list13,list14,list15,[df21],[df22],[df23],list31,list32,list33_1]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        all_list2=[]
        if len(list33_1) != 0:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
    if max(set_max_list3) < 2:
        items0=[list11_0,list12,list13,list14,list15,list21_0,list22,list23,[df31],[df32],[df33]]
        items1=[list11_1,list12,list13,list14,list15,list21_0,list22,list23,[df31],[df32],[df33]]
        items2=[list11_0,list12,list13,list14,list15,list21_1,list22,list23,[df31],[df32],[df33]]
        items3=[]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        if len(list21_1) != 0:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
    if max(set_max_list2) < 2 and max(set_max_list3) < 2:
        items0=[list11_0,list12,list13,list14,list15,[df21],[df22],[df23],[df31],[df32],[df33]]
        items1=[list11_1,list12,list13,list14,list15,[df21],[df22],[df23],[df31],[df32],[df33]]
        items2=[]
        items3=[]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        all_list2=[]
        all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])

##메타몽모드##
##########################################################################################################################
    

    if select_perfect.get() == '메타몽모드(느림)' or select_perfect.get() == '메타몽모드(중간)':
        showsta(text="메타몽 계산중...")
        metamong=1
        ask_meta=tkinter.messagebox.askquestion('시작하기 앞서...',"주의: 본 기능은 참고용입니다.\n\n어느정도 세트가 완성될 각이 있어야 정상작동합니다.\n또한 현재 체크된 에픽 중 사용중인or사용했던 메타몽 에픽이 전부 체크 해제되있는지 확인하십시오. 안되있으면 취소를 누르고 다시 진행해주세요.")
        if ask_meta == 'no':
            showsta(text='중지됨')
            return
        full_list=[]
        semifull_list=[]
        semigod_list=[]
        bang_danpum_list=[];god_danpum_list=[]
        bang_eset_list=[];god_eset_list=[]
        bang_samset_list=[];god_samset_list=[]
        notem_list=[]
        
        
        for i in range(1,16):
            if set_num_dict[str(i+100)]==5:
                full_list.append(str(i+100)) #
        for i in range(16,36):
            if set_num_dict[str(i+100)]==3:
                full_list.append(str(i+100)) #
        for i in range(1,16):
            if set_num_dict2[str(i+100)]==4:
                semifull_list.append(str(i+100)) ##
            elif set_num_dict[str(i+100)]==4:
                semigod_list.append(str(i+100)) ##
        for i in range(16,36):
            if set_num_dict2[str(i+100)]==2:
                semifull_list.append(str(i+100)) ##
            elif set_num_dict[str(i+100)]==2:
                semigod_list.append(str(i+100)) ##
        for i in range(1,16):
            if set_num_dict[str(i+100)]==1:
                bang_danpum_list.append(str(i+100)) ###
        for i in range(1,16):
            if set_num_dict[str(i+100)]==2:
                bang_eset_list.append(str(i+100)) ###
        for i in range(1,16):
            if set_num_dict[str(i+100)]==3 or set_num_dict[str(i+100)]==4:
                bang_samset_list.append(str(i+100)) ###
        for i in range(1,36):
            if set_num_dict[str(i+100)]==0:
                notem_list.append(str(i+100)) ####
        all_bang_dict={}
        for i2 in range(1,16):
            i=str(i2+100)
            set_array=np.array([0,0,0,0,0])
            for j in [list11,list12,list13,list14,list15]:
                if j.count(j[0][0:2]+i[1:3]+'0')+j.count(j[0][0:2]+i[1:3]+'1')!=0:
                    set_array[int(j[0][1:2])-1]=1
            all_bang_dict[i]=set_array

        if len(semifull_list) != 0:  ## 경우1: 영고해제
            for metamong_full in semifull_list:
                tlist11=list11;tlist12=list12;tlist13=list13;tlist14=list14;tlist15=list15;tlist11_0=list11_0;tlist11_1=list11_1
                tlist21=list21;tlist22=list22;tlist23=list23;tlist21_0=list21_0;tlist21_1=list21_1
                tlist31=list31;tlist32=list32;tlist33=list33;tlist33_0=list33_0;tlist33_1=list33_1
                meta_set_num=metamong_full[1:3]
                if int(meta_set_num)<15:
                    tlist11=[];tlist11_0=[]
                    tlist11.append('11'+meta_set_num+'0')
                    tlist11_0.append('11'+meta_set_num+'0')
                    tlist12=[];tlist12.append('12'+meta_set_num+'0')
                    tlist13=[];tlist13.append('13'+meta_set_num+'0')
                    tlist14=[];tlist14.append('14'+meta_set_num+'0')
                    tlist15=[];tlist15.append('15'+meta_set_num+'0')
                elif 15<int(meta_set_num)<19:
                    tlist21=[];tlist21_0=[]
                    tlist21.append('21'+meta_set_num+'0')
                    tlist21_0.append('21'+meta_set_num+'0')
                    tlist22=[];tlist22.append('22'+meta_set_num+'0')
                    tlist23=[];tlist23.append('23'+meta_set_num+'0')
                elif 19<int(meta_set_num)<23:
                    tlist31=[];tlist31.append('31'+meta_set_num+'0')
                    tlist32=[];tlist32.append('32'+meta_set_num+'0')
                    tlist33=[];tlist33_0=[]
                    tlist33.append('33'+meta_set_num+'0')
                    tlist33_0.append('33'+meta_set_num+'0')
                elif 23<int(meta_set_num)<28:
                    tlist12=[];tlist12.append('12'+meta_set_num+'0')
                    tlist21=[];tlist21_0=[]
                    tlist21.append('21'+meta_set_num+'0')
                    tlist21_0.append('21'+meta_set_num+'0')
                    tlist32=[];tlist32.append('32'+meta_set_num+'0')
                elif 27<int(meta_set_num)<32:
                    tlist11=[];tlist11_0=[]
                    tlist11.append('11'+meta_set_num+'0')
                    tlist11_0.append('11'+meta_set_num+'0')
                    tlist22=[];tlist22.append('22'+meta_set_num+'0')
                    tlist31=[];tlist31.append('31'+meta_set_num+'0')
                elif 31<int(meta_set_num)<36:
                    tlist15=[];tlist15.append('15'+meta_set_num+'0')
                    tlist23=[];tlist23.append('23'+meta_set_num+'0')
                    tlist33=[];tlist33_0=[]
                    tlist33.append('33'+meta_set_num+'0')
                    tlist33_0.append('33'+meta_set_num+'0')

                items0=[tlist11_0,tlist12,tlist13,tlist14,tlist15,tlist21_0,tlist22,tlist23,tlist31,tlist32,tlist33_0]
                items1=[tlist11_1,tlist12,tlist13,tlist14,tlist15,tlist21_0,tlist22,tlist23,tlist31,tlist32,tlist33_0]
                items2=[tlist11_0,tlist12,tlist13,tlist14,tlist15,tlist21_1,tlist22,tlist23,tlist31,tlist32,tlist33_0]
                items3=[tlist11_0,tlist12,tlist13,tlist14,tlist15,tlist21_0,tlist22,tlist23,tlist31,tlist32,tlist33_1]
                all_list=list(itertools.product(*items0))
                if len(list11_1) != 0:
                    all_list1=list(itertools.product(*items1))
                else:
                    all_list1=[]
                if len(list21_1) != 0:
                    all_list2=list(itertools.product(*items2))
                else:
                    all_list2=[]
                if len(list33_1) != 0:
                    all_list3=list(itertools.product(*items3))
                else:
                    all_list3=[]
                all_list_god=all_list1+all_list2+all_list3
                del all_list1,all_list2,all_list3
                all_list_num=len(all_list_god)+len(all_list)
                all_list_list.append([all_list,all_list_god,all_list_num])

        if len(semigod_list) != 0:  ## 경우1-1: 신화영고해제
            for metamong_full in semigod_list:
                tlist11=list11;tlist12=list12;tlist13=list13;tlist14=list14;tlist15=list15;tlist11_0=list11_0;tlist11_1=list11_1
                tlist21=list21;tlist22=list22;tlist23=list23;tlist21_0=list21_0;tlist21_1=list21_1
                tlist31=list31;tlist32=list32;tlist33=list33;tlist33_0=list33_0;tlist33_1=list33_1
                meta_set_num=metamong_full[1:3]
                if int(meta_set_num)<15:
                    tlist11=[];tlist11_1=[]
                    tlist11.append('11'+meta_set_num+'1')
                    tlist11_1.append('11'+meta_set_num+'1')
                    tlist12=[];tlist12.append('12'+meta_set_num+'0')
                    tlist13=[];tlist13.append('13'+meta_set_num+'0')
                    tlist14=[];tlist14.append('14'+meta_set_num+'0')
                    tlist15=[];tlist15.append('15'+meta_set_num+'0')
                elif 15<int(meta_set_num)<19:
                    tlist21=[];tlist21_1=[]
                    tlist21.append('21'+meta_set_num+'1')
                    tlist21_1.append('21'+meta_set_num+'1')
                    tlist22=[];tlist22.append('22'+meta_set_num+'0')
                    tlist23=[];tlist23.append('23'+meta_set_num+'0')
                elif 19<int(meta_set_num)<23:
                    tlist31=[];tlist31.append('31'+meta_set_num+'0')
                    tlist32=[];tlist32.append('32'+meta_set_num+'0')
                    tlist33=[];tlist33_1=[]
                    tlist33.append('33'+meta_set_num+'1')
                    tlist33_1.append('33'+meta_set_num+'1')
                elif 23<int(meta_set_num)<28:
                    tlist12=[];tlist12.append('12'+meta_set_num+'0')
                    tlist21=[];tlist21_1=[]
                    tlist21.append('21'+meta_set_num+'1')
                    tlist21_1.append('21'+meta_set_num+'1')
                    tlist32=[];tlist32.append('32'+meta_set_num+'0')
                elif 27<int(meta_set_num)<32:
                    tlist11=[];tlist11_1=[]
                    tlist11.append('11'+meta_set_num+'1')
                    tlist11_1.append('11'+meta_set_num+'1')
                    tlist22=[];tlist22.append('22'+meta_set_num+'0')
                    tlist31=[];tlist31.append('31'+meta_set_num+'0')
                elif 31<int(meta_set_num)<36:
                    tlist15=[];tlist15.append('15'+meta_set_num+'0')
                    tlist23=[];tlist23.append('23'+meta_set_num+'0')
                    tlist33=[];tlist33_1=[]
                    tlist33.append('33'+meta_set_num+'1')
                    tlist33_1.append('33'+meta_set_num+'1')

                items0=[tlist11_0,tlist12,tlist13,tlist14,tlist15,tlist21_0,tlist22,tlist23,tlist31,tlist32,tlist33_0]
                items1=[tlist11_1,tlist12,tlist13,tlist14,tlist15,tlist21_0,tlist22,tlist23,tlist31,tlist32,tlist33_0]
                items2=[tlist11_0,tlist12,tlist13,tlist14,tlist15,tlist21_1,tlist22,tlist23,tlist31,tlist32,tlist33_0]
                items3=[tlist11_0,tlist12,tlist13,tlist14,tlist15,tlist21_0,tlist22,tlist23,tlist31,tlist32,tlist33_1]
                all_list=list(itertools.product(*items0))
                if len(list11_1) != 0:
                    all_list1=list(itertools.product(*items1))
                else:
                    all_list1=[]
                if len(list21_1) != 0:
                    all_list2=list(itertools.product(*items2))
                else:
                    all_list2=[]
                if len(list33_1) != 0:
                    all_list3=list(itertools.product(*items3))
                else:
                    all_list3=[]
                all_list_god=all_list1+all_list2+all_list3
                del all_list1,all_list2,all_list3
                all_list_num=len(all_list_god)+len(all_list)
                all_list_list.append([all_list,all_list_god,all_list_num])
        
        if len(bang_danpum_list)!=0  :  ## 경우2: 방어구 단일+3셋
            for i in bang_danpum_list:
                tlist11=list11;tlist12=list12;tlist13=list13;tlist14=list14;tlist15=list15;tlist11_0=list11_0;tlist11_1=list11_1
                array_now=all_bang_dict.get(i)
                for j in range(101,116):   
                    if str(j) != i:
                        array_j=all_bang_dict.get(str(j))
                        sum_array=array_now+array_j
                        keep_go=0
                        if array_j.tolist().count(1)==3 and sum_array.tolist().count(2)==0:
                            keep_go=1
                        elif array_j.tolist().count(1)==4 and sum_array.tolist().count(0)==0:
                            keep_go=2
                        elif array_j.tolist().count(1)==4 and sum_array.tolist().count(0)!=0:
                            keep_go=1
                        elif array_j.tolist().count(1)==5:
                            keep_go=2
                            
                        if keep_go==1:
                            tlist12=[];tlist13=[];tlist14=[];tlist15=[];tlist11_0=[];tlist11_1=list11_1
                            if array_now[0]==1:
                                if list11_0.count('11'+i[1:3]+'0')==1:
                                    tlist11_0.append('11'+i[1:3]+'0')
                            if array_now[1]==1:
                                tlist12.append('12'+i[1:3]+'0')
                            if array_now[2]==1:
                                tlist13.append('13'+i[1:3]+'0')
                            if array_now[3]==1:
                                tlist14.append('14'+i[1:3]+'0')
                            if array_now[4]==1:
                                tlist15.append('15'+i[1:3]+'0') ##
                            if array_j[0]==1 and array_now[0]==0:
                                if list11_0.count('11'+str(j)[1:3]+'0')==1:
                                    tlist11_0.append('11'+str(j)[1:3]+'0')
                            if array_j[1]==1 and array_now[1]==0:
                                tlist12.append('12'+str(j)[1:3]+'0')
                            if array_j[2]==1 and array_now[2]==0:
                                tlist13.append('13'+str(j)[1:3]+'0')
                            if array_j[3]==1 and array_now[3]==0:
                                tlist14.append('14'+str(j)[1:3]+'0')
                            if array_j[4]==1 and array_now[4]==0:
                                tlist15.append('15'+str(j)[1:3]+'0') ##
                            if i!='115':
                                if sum_array[0]==0:
                                    tlist11_0.append('11'+i[1:3]+'0')
                                elif sum_array[1]==0:
                                    tlist12.append('12'+i[1:3]+'0')
                                elif sum_array[2]==0:
                                    tlist13.append('13'+i[1:3]+'0')
                                elif sum_array[3]==0:
                                    tlist14.append('14'+i[1:3]+'0')
                                elif sum_array[4]==0:
                                    tlist15.append('15'+i[1:3]+'0') ##

                            items0=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                            items1=[list11_1,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                            items2=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_1,list22,list23,list31,list32,list33_0]
                            items3=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_1]
                            all_list=list(itertools.product(*items0))
                            if len(list11_1) != 0:
                                all_list1=list(itertools.product(*items1))
                            else:
                                all_list1=[]
                            if len(list21_1) != 0:
                                all_list2=list(itertools.product(*items2))
                            else:
                                all_list2=[]
                            if len(list33_1) != 0:
                                all_list3=list(itertools.product(*items3))
                            else:
                                all_list3=[]
                            all_list_god=all_list1+all_list2+all_list3
                            del all_list1,all_list2,all_list3
                            all_list_num=len(all_list_god)+len(all_list)
                            all_list_list.append([all_list,all_list_god,all_list_num])

                        if keep_go==2:
                            for k in range(0,5):
                                tlist12=[];tlist13=[];tlist14=[];tlist15=[];tlist11_0=[];tlist11_1=list11_1
                                if array_now[k]==0:
                                    if array_now[0]==1:
                                        if list11_0.count('11'+i[1:3]+'0')==1:
                                            tlist11_0.append('11'+i[1:3]+'0')
                                    if array_now[1]==1:
                                        tlist12.append('12'+i[1:3]+'0')
                                    if array_now[2]==1:
                                        tlist13.append('13'+i[1:3]+'0')
                                    if array_now[3]==1:
                                        tlist14.append('14'+i[1:3]+'0')
                                    if array_now[4]==1:
                                        tlist15.append('15'+i[1:3]+'0')
                                    if i!='115':
                                        if k==0:
                                            tlist11_0.append('11'+i[1:3]+'0')
                                        elif k==1:
                                            tlist12.append('12'+i[1:3]+'0')
                                        elif k==2:
                                            tlist13.append('13'+i[1:3]+'0')
                                        elif k==3:
                                            tlist14.append('14'+i[1:3]+'0')
                                        elif k==4:
                                            tlist15.append('15'+i[1:3]+'0')
                                        if len(tlist11_0)==0:
                                            if list11_0.count('11'+str(j)[1:3]+'0')==1:
                                                tlist11_0.append('11'+str(j)[1:3]+'0')
                                        if len(tlist12)==0:
                                            tlist12.append('12'+str(j)[1:3]+'0')
                                        if len(tlist13)==0:
                                            tlist13.append('13'+str(j)[1:3]+'0')
                                        if len(tlist14)==0:
                                            tlist14.append('14'+str(j)[1:3]+'0')
                                        if len(tlist15)==0:
                                            tlist15.append('15'+str(j)[1:3]+'0')
                                        
                                    items0=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                                    items1=[list11_1,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                                    items2=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_1,list22,list23,list31,list32,list33_0]
                                    items3=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_1]
                                    all_list=list(itertools.product(*items0))
                                    if len(list11_1) != 0:
                                        all_list1=list(itertools.product(*items1))
                                    else:
                                        all_list1=[]
                                    if len(list21_1) != 0:
                                        all_list2=list(itertools.product(*items2))
                                    else:
                                        all_list2=[]
                                    if len(list33_1) != 0:
                                        all_list3=list(itertools.product(*items3))
                                    else:
                                        all_list3=[]
                                    all_list_god=all_list1+all_list2+all_list3
                                    del all_list1,all_list2,all_list3
                                    all_list_num=len(all_list_god)+len(all_list)
                                    all_list_list.append([all_list,all_list_god,all_list_num])

        if len(bang_eset_list)!=0 : #경우3: 방어구2셋+2~5개
            for i in bang_eset_list:
                array_now=all_bang_dict.get(i)
                for j in range(101,116):
                    if str(j) != i:
                        array_j=all_bang_dict.get(str(j))
                        array_sum=array_now+array_j
                        keep_go=0
                        if array_j.tolist().count(1) >=2:
                            if array_sum.tolist().count(0)==1:
                                keep_go=1
                            if array_sum.tolist().count(0)==0:
                                keep_go=2
                        if keep_go==1:
                            tlist12=[];tlist13=[];tlist14=[];tlist15=[];tlist11_0=[];tlist11_1=list11_1
                            if array_now[0]==1:
                                if list11_0.count('11'+i[1:3]+'0')==1:
                                    tlist11_0.append('11'+i[1:3]+'0')
                            elif array_j[0]==1:
                                if list11_0.count('11'+str(j)[1:3]+'0')==1:
                                    tlist11_0.append('11'+str(j)[1:3]+'0')
                            elif array_j[0]==0:
                                if i!='115':
                                    tlist11_0.append('11'+i[1:3]+'0')
                                if array_j.tolist().count(1) >=3 and  str(j)!='115':
                                    tlist11_0.append('11'+str(j)[1:3]+'0')
                            if array_now[1]==1:
                                tlist12.append('12'+i[1:3]+'0')
                            elif array_j[1]==1:
                                tlist12.append('12'+str(j)[1:3]+'0')
                            elif array_j[1]==0:
                                if i!='115':
                                    tlist12.append('12'+i[1:3]+'0')
                                if array_j.tolist().count(1) >=3 and  str(j)!='115':
                                    tlist12.append('12'+str(j)[1:3]+'0')
                            if array_now[2]==1:
                                tlist13.append('13'+i[1:3]+'0')
                            elif array_j[2]==1:
                                tlist13.append('13'+str(j)[1:3]+'0')
                            elif array_j[2]==0:
                                if i!='115':
                                    tlist13.append('13'+i[1:3]+'0')
                                if array_j.tolist().count(1) >=3 and  str(j)!='115':
                                    tlist13.append('13'+str(j)[1:3]+'0')
                            if array_now[3]==1:
                                tlist14.append('14'+i[1:3]+'0')
                            elif array_j[3]==1:
                                tlist14.append('14'+str(j)[1:3]+'0')
                            elif array_j[3]==0:
                                if i!='115':
                                    tlist14.append('14'+i[1:3]+'0')
                                if array_j.tolist().count(1) >=3 and  str(j)!='115':
                                    tlist14.append('14'+str(j)[1:3]+'0')
                            if array_now[4]==1:
                                tlist15.append('15'+i[1:3]+'0') ##
                            elif array_j[4]==1:
                                tlist15.append('15'+str(j)[1:3]+'0')
                            elif array_j[4]==0:
                                if i!='115':
                                    tlist15.append('15'+i[1:3]+'0')
                                if array_j.tolist().count(1) >=3 and  str(j)!='115':
                                    tlist15.append('15'+str(j)[1:3]+'0')
                                
                            items0=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                            items1=[list11_1,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                            items2=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_1,list22,list23,list31,list32,list33_0]
                            items3=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_1]
                            all_list=list(itertools.product(*items0))
                            if len(list11_1) != 0:
                                all_list1=list(itertools.product(*items1))
                            else:
                                all_list1=[]
                            if len(list21_1) != 0:
                                all_list2=list(itertools.product(*items2))
                            else:
                                all_list2=[]
                            if len(list33_1) != 0:
                                all_list3=list(itertools.product(*items3))
                            else:
                                all_list3=[]
                            all_list_god=all_list1+all_list2+all_list3
                            del all_list1,all_list2,all_list3
                            all_list_num=len(all_list_god)+len(all_list)
                            all_list_list.append([all_list,all_list_god,all_list_num])
                            
                        if keep_go==2:
                            for k in range(0,5):
                                tlist12=[];tlist13=[];tlist14=[];tlist15=[];tlist11_0=[];tlist11_1=list11_1
                                if array_now[k]==0:
                                    if array_now[0]==1:
                                        if list11_0.count('11'+i[1:3]+'0')==1:
                                            tlist11_0.append('11'+i[1:3]+'0')
                                    if array_now[1]==1:
                                        tlist12.append('12'+i[1:3]+'0')
                                    if array_now[2]==1:
                                        tlist13.append('13'+i[1:3]+'0')
                                    if array_now[3]==1:
                                        tlist14.append('14'+i[1:3]+'0')
                                    if array_now[4]==1:
                                        tlist15.append('15'+i[1:3]+'0')
                                    if i!='115':
                                        if k==0:
                                            tlist11_0.append('11'+i[1:3]+'0')
                                        elif k==1:
                                            tlist12.append('12'+i[1:3]+'0')
                                        elif k==2:
                                            tlist13.append('13'+i[1:3]+'0')
                                        elif k==3:
                                            tlist14.append('14'+i[1:3]+'0')
                                        elif k==4:
                                            tlist15.append('15'+i[1:3]+'0')
                                        if len(tlist11_0)==0:
                                            if list11_0.count('11'+str(j)[1:3]+'0')==1:
                                                tlist11_0.append('11'+str(j)[1:3]+'0')
                                        if len(tlist12)==0:
                                            tlist12.append('12'+str(j)[1:3]+'0')
                                        if len(tlist13)==0:
                                            tlist13.append('13'+str(j)[1:3]+'0')
                                        if len(tlist14)==0:
                                            tlist14.append('14'+str(j)[1:3]+'0')
                                        if len(tlist15)==0:
                                            tlist15.append('15'+str(j)[1:3]+'0')

                                    items0=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                                    items1=[list11_1,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                                    items2=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_1,list22,list23,list31,list32,list33_0]
                                    items3=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_1]
                                    all_list=list(itertools.product(*items0))
                                    if len(list11_1) != 0:
                                        all_list1=list(itertools.product(*items1))
                                    else:
                                        all_list1=[]
                                    if len(list21_1) != 0:
                                        all_list2=list(itertools.product(*items2))
                                    else:
                                        all_list2=[]
                                    if len(list33_1) != 0:
                                        all_list3=list(itertools.product(*items3))
                                    else:
                                        all_list3=[]
                                    all_list_god=all_list1+all_list2+all_list3
                                    del all_list1,all_list2,all_list3
                                    all_list_num=len(all_list_god)+len(all_list)
                                    all_list_list.append([all_list,all_list_god,all_list_num])


        if len(bang_samset_list)!=0 : #경우4: 방어구3or4셋+3or4~5개
            for i in bang_samset_list:
                temp_number=set_num_dict[i]
                array_now=all_bang_dict.get(i)
                for j in range(101,116):
                    if str(j) != i:
                        array_j=all_bang_dict.get(str(j))
                        array_sum=array_now+array_j
                        keep_go=0
                        if array_j.tolist().count(1) >=temp_number:
                            if array_sum.tolist().count(0)==1:
                                keep_go=1
                            if array_sum.tolist().count(0)==0:
                                keep_go=2
                        if keep_go==1:
                            tlist12=[];tlist13=[];tlist14=[];tlist15=[];tlist11_0=[];tlist11_1=list11_1
                            if array_now[0]==1:
                                if list11_0.count('11'+i[1:3]+'0')==1:
                                    tlist11_0.append('11'+i[1:3]+'0')
                            if array_now[1]==1:
                                tlist12.append('12'+i[1:3]+'0')
                            if array_now[2]==1:
                                tlist13.append('13'+i[1:3]+'0')
                            if array_now[3]==1:
                                tlist14.append('14'+i[1:3]+'0')
                            if array_now[4]==1:
                                tlist15.append('15'+i[1:3]+'0')
                            if array_j[0]==1:
                                if list11_0.count('11'+str(j)[1:3]+'0')==1:
                                    tlist11_0.append('11'+str(j)[1:3]+'0')
                            if array_j[1]==1:
                                tlist12.append('12'+str(j)[1:3]+'0')
                            if array_j[2]==1:
                                tlist13.append('13'+str(j)[1:3]+'0')
                            if array_j[3]==1:
                                tlist14.append('14'+str(j)[1:3]+'0')
                            if array_j[4]==1:
                                tlist15.append('15'+str(j)[1:3]+'0')
                            
                            if i !='115':
                                if array_sum[0]==0:
                                    tlist11_0.append('11'+i[1:3]+'0')
                                elif array_sum[1]==0:
                                    tlist12.append('12'+i[1:3]+'0')
                                elif array_sum[2]==0:
                                    tlist13.append('13'+i[1:3]+'0')
                                elif array_sum[3]==0:
                                    tlist14.append('14'+i[1:3]+'0')
                                elif array_sum[4]==0:
                                    tlist15.append('15'+i[1:3]+'0') ##

                            items0=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                            items1=[list11_1,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                            items2=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_1,list22,list23,list31,list32,list33_0]
                            items3=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_1]
                            all_list=list(itertools.product(*items0))
                            if len(list11_1) != 0:
                                all_list1=list(itertools.product(*items1))
                            else:
                                all_list1=[]
                            if len(list21_1) != 0:
                                all_list2=list(itertools.product(*items2))
                            else:
                                all_list2=[]
                            if len(list33_1) != 0:
                                all_list3=list(itertools.product(*items3))
                            else:
                                all_list3=[]
                            all_list_god=all_list1+all_list2+all_list3
                            del all_list1,all_list2,all_list3
                            all_list_num=len(all_list_god)+len(all_list)
                            all_list_list.append([all_list,all_list_god,all_list_num])
                            
                        if keep_go==2:
                            for k in range(0,5):
                                tlist12=[];tlist13=[];tlist14=[];tlist15=[];tlist11_0=[];tlist11_1=list11_1
                                if array_now[k]==0:
                                    if array_now[0]==1:
                                        if list11_0.count('11'+i[1:3]+'0')==1:
                                            tlist11_0.append('11'+i[1:3]+'0')
                                    if array_now[1]==1:
                                        tlist12.append('12'+i[1:3]+'0')
                                    if array_now[2]==1:
                                        tlist13.append('13'+i[1:3]+'0')
                                    if array_now[3]==1:
                                        tlist14.append('14'+i[1:3]+'0')
                                    if array_now[4]==1:
                                        tlist15.append('15'+i[1:3]+'0')
                                    if array_j[0]==1:
                                        if list11_0.count('11'+str(j)[1:3]+'0')==1:
                                            tlist11_0.append('11'+str(j)[1:3]+'0')
                                    if array_j[1]==1:
                                        tlist12.append('12'+str(j)[1:3]+'0')
                                    if array_j[2]==1:
                                        tlist13.append('13'+str(j)[1:3]+'0')
                                    if array_j[3]==1:
                                        tlist14.append('14'+str(j)[1:3]+'0')
                                    if array_j[4]==1:
                                        tlist15.append('15'+str(j)[1:3]+'0')
                                    if i!='115':
                                        if k==0:
                                            tlist11_0.append('11'+i[1:3]+'0')
                                        elif k==1:
                                            tlist12.append('12'+i[1:3]+'0')
                                        elif k==2:
                                            tlist13.append('13'+i[1:3]+'0')
                                        elif k==3:
                                            tlist14.append('14'+i[1:3]+'0')
                                        elif k==4:
                                            tlist15.append('15'+i[1:3]+'0')

                                    items0=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                                    items1=[list11_1,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_0]
                                    items2=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_1,list22,list23,list31,list32,list33_0]
                                    items3=[tlist11_0,tlist12,tlist13,tlist14,tlist15,list21_0,list22,list23,list31,list32,list33_1]
                                    all_list=list(itertools.product(*items0))
                                    if len(list11_1) != 0:
                                        all_list1=list(itertools.product(*items1))
                                    else:
                                        all_list1=[]
                                    if len(list21_1) != 0:
                                        all_list2=list(itertools.product(*items2))
                                    else:
                                        all_list2=[]
                                    if len(list33_1) != 0:
                                        all_list3=list(itertools.product(*items3))
                                    else:
                                        all_list3=[]
                                    all_list_god=all_list1+all_list2+all_list3
                                    del all_list1,all_list2,all_list3
                                    all_list_num=len(all_list_god)+len(all_list)
                                    all_list_list.append([all_list,all_list_god,all_list_num])
        setnum_ha=[0]
        setnum_sang=[0]
        setnum_sin=[0]
        keep_go_dict={}

        for i in range(24,28):
            setnum_ha.append(set_num_dict[str(i+100)])
        for i in range(28,32):
            setnum_sang.append(set_num_dict[str(i+100)])
        for i in range(32,36):
            setnum_sin.append(set_num_dict[str(i+100)])
        for i in range(101,115):
            temp_list=all_bang_dict[str(i)].tolist()
            if temp_list[2]==1 and temp_list[3]==0:
                keep_go_dict[str(i)]=1  ## 벨트x
            if temp_list[2]==0 and temp_list[3]==1:
                keep_go_dict[str(i)]=2  ## 어깨x
            if temp_list[2]==1 and temp_list[3]==1:
                keep_go_dict[str(i)]=3  ## 변형3332 (변형부위x)

        tlist11_0=[];tlist12=[];tlist13=[];tlist14=[];tlist15=[];
        tlist21_0=[];tlist22=[];tlist23=[];
        tlist31=[];tlist32=[];tlist33_0=[];
        list_all=[]
        hapal=0;sanmop=0;sinban=0
        for i in range(124,136):
            if set_num_dict[str(i)]==3:
                if i < 128:
                    hapal=1
                elif i < 132:
                    sanmop=1
                elif i < 136:
                    sinban=1
                
        if hapal==1:
            list_all=list_all+[[list12,tlist12],[list21_0,tlist21_0],[list32,tlist32]]
        if sanmop==1:
            list_all=list_all+[[list11_0,tlist11_0],[list22,tlist22],[list31,tlist31]]
        if sinban==1:
            list_all=list_all+[[list15,tlist15],[list23,tlist23],[list33_0,tlist33_0]]
        
        for i in list_all:
            for j in i[0]:
                if int(j[2:4])>23:
                    i[1].append(j)
        
                    
        tlist_all=[[list12,tlist12],[list21_0,tlist21_0],[list32,tlist32],[list11_0,tlist11_0],[list22,tlist22],[list31,tlist31],[list15,tlist15],[list23,tlist23],[list33_0,tlist33_0]]


        if len(tlist11_0)==0:
            tlist11_0=list11_0
        if len(tlist12)==0:
            tlist12=list12
        if len(tlist13)==0:
            tlist13=list13
        if len(tlist14)==0:
            tlist14=list14
        if len(tlist15)==0:
            tlist15=list15
        if len(tlist21_0)==0:
            tlist21_0=list21_0
        if len(tlist22)==0:
            tlist22=list22
        if len(tlist23)==0:
            tlist23=list23
        if len(tlist31)==0:
            tlist31=list31
        if len(tlist32)==0:
            tlist32=list32
        if len(tlist33_0)==0:
            tlist33_0=list33_0

        
        for i in list(keep_go_dict.keys()):
            plist11_0=tlist11_0;plist12=tlist12;plist15=tlist15
            plist21_0=tlist21_0;plist22=tlist22;plist23=tlist23
            plist33_0=tlist33_0;plist31=tlist31;plist32=tlist32
            
            if keep_go_dict.get(i)==1 or keep_go_dict.get(i)==2:
                if list11_0.count('11'+i[1:3]+'0') ==1 and plist11_0.count('11'+i[1:3]+'0') ==0:
                    plist11_0.append('11'+i[1:3]+'0')
                if list12.count('12'+i[1:3]+'0') ==1 and plist12.count('12'+i[1:3]+'0') ==0:
                    plist12.append('12'+i[1:3]+'0')
                plist13=['13'+i[1:3]+'0']
                plist14=['14'+i[1:3]+'0']
                if list15.count('15'+i[1:3]+'0') ==1 and plist15.count('15'+i[1:3]+'0') ==0:
                    plist15.append('15'+i[1:3]+'0')
                
                items0=[plist11_0,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,plist33_0]
                items1=[list11_1,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,plist33_0]
                items2=[plist11_0,plist12,plist13,plist14,plist15,list21_1,plist22,plist23,plist31,plist32,plist33_0]
                items3=[plist11_0,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,list33_1]
                all_list=list(itertools.product(*items0))
                if len(list11_1) != 0:
                    all_list1=list(itertools.product(*items1))
                else:
                    all_list1=[]
                if len(list21_1) != 0:
                    all_list2=list(itertools.product(*items2))
                else:
                    all_list2=[]
                if len(list33_1) != 0:
                    all_list3=list(itertools.product(*items3))
                else:
                    all_list3=[]
                all_list_god=all_list1+all_list2+all_list3
                del all_list1,all_list2,all_list3
                all_list_num=len(all_list_god)+len(all_list)
                all_list_list.append([all_list,all_list_god,all_list_num])
                
            if keep_go_dict.get(i)==3:
                if plist11_0.count('11'+i[1:3]+'0')==0 and list11_1.count('11'+i[1:3]+'1')==0:
                    plist11_0=tlist11_0;plist12=tlist12;plist13=tlist13;plist14=tlist14;plist15=tlist15
                    plist11_0=['11'+i[1:3]+'0']
                    plist13=['13'+i[1:3]+'0']
                    plist14=['14'+i[1:3]+'0']
                    items0=[plist11_0,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,plist33_0]
                    items1=[list11_1,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,plist33_0]
                    items2=[plist11_0,plist12,plist13,plist14,plist15,list21_1,plist22,plist23,plist31,plist32,plist33_0]
                    items3=[plist11_0,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,list33_1]
                    all_list=list(itertools.product(*items0))
                    if len(list11_1) != 0:
                        all_list1=list(itertools.product(*items1))
                    else:
                        all_list1=[]
                    if len(list21_1) != 0:
                        all_list2=list(itertools.product(*items2))
                    else:
                        all_list2=[]
                    if len(list33_1) != 0:
                        all_list3=list(itertools.product(*items3))
                    else:
                        all_list3=[]
                    all_list_god=all_list1+all_list2+all_list3
                    del all_list1,all_list2,all_list3
                    all_list_num=len(all_list_god)+len(all_list)
                    all_list_list.append([all_list,all_list_god,all_list_num])
                if plist12.count('12'+i[1:3]+'0')==0:
                    plist11_0=tlist11_0;plist12=tlist12;plist13=tlist13;plist14=tlist14;plist15=tlist15
                    plist12=['12'+i[1:3]+'0']
                    plist13=['13'+i[1:3]+'0']
                    plist14=['14'+i[1:3]+'0']
                    items0=[plist11_0,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,plist33_0]
                    items1=[list11_1,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,plist33_0]
                    items2=[plist11_0,plist12,plist13,plist14,plist15,list21_1,plist22,plist23,plist31,plist32,plist33_0]
                    items3=[plist11_0,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,list33_1]
                    all_list=list(itertools.product(*items0))
                    if len(list11_1) != 0:
                        all_list1=list(itertools.product(*items1))
                    else:
                        all_list1=[]
                    if len(list21_1) != 0:
                        all_list2=list(itertools.product(*items2))
                    else:
                        all_list2=[]
                    if len(list33_1) != 0:
                        all_list3=list(itertools.product(*items3))
                    else:
                        all_list3=[]
                    all_list_god=all_list1+all_list2+all_list3
                    del all_list1,all_list2,all_list3
                    all_list_num=len(all_list_god)+len(all_list)
                    all_list_list.append([all_list,all_list_god,all_list_num])
                if plist15.count('15'+i[1:3]+'0')==0:
                    plist11_0=tlist11_0;plist12=tlist12;plist13=tlist13;plist14=tlist14;plist15=tlist15
                    plist15=['15'+i[1:3]+'0']
                    plist13=['13'+i[1:3]+'0']
                    plist14=['14'+i[1:3]+'0']
                    items0=[plist11_0,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,plist33_0]
                    items1=[list11_1,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,plist33_0]
                    items2=[plist11_0,plist12,plist13,plist14,plist15,list21_1,plist22,plist23,plist31,plist32,plist33_0]
                    items3=[plist11_0,plist12,plist13,plist14,plist15,plist21_0,plist22,plist23,plist31,plist32,list33_1]
                    all_list=list(itertools.product(*items0))
                    if len(list11_1) != 0:
                        all_list1=list(itertools.product(*items1))
                    else:
                        all_list1=[]
                    if len(list21_1) != 0:
                        all_list2=list(itertools.product(*items2))
                    else:
                        all_list2=[]
                    if len(list33_1) != 0:
                        all_list3=list(itertools.product(*items3))
                    else:
                        all_list3=[]
                    all_list_god=all_list1+all_list2+all_list3
                    del all_list1,all_list2,all_list3
                    all_list_num=len(all_list_god)+len(all_list)
                    all_list_list.append([all_list,all_list_god,all_list_num])

##세트산물 계산##                
#########################################################################################################################
    know_set_list=['22400150','22400250','22400350','22400450','22400550','21400640','31400750',
                   '31400850','31400950','31401050','31401150','32401240','32401340','32401440']
    know_bang1_list=['22400150','22400250','22400350','22400450','22400550']
    know_bang2_list=['31400750','31400850','31400950','31401050','31401150']
    know_acc_list=['32401240','32401340','32401440']
    know_jin_list=['11410100','11410110','11410120','11410130','11410140','11410150',
                   '21420100','21420110','21420120','21420130','21420140','21420150',
                   '33430100','33430110','33430120','33430130','33430140','33430150']

    for i in know_set_list: ##경우1:산물 하나
        if select_item['tg'+i]==1:
            if int(i[4:6]) <6:
                items0=[['99990'],['99990'],['99990'],['99990'],['99990'],list21_0,[i],list23,list31,list32,list33_0]
                items1=[]
                items2=[['99990'],['99990'],['99990'],['99990'],['99990'],list21_1,[i],list23,list31,list32,list33_0]
                items3=[['99990'],['99990'],['99990'],['99990'],['99990'],list21_0,[i],list23,list31,list32,list33_1]
            elif int(i[4:6])==6:
                items0=[list11_0,list12,list13,list14,list15,[i],list22,list23,['99990'],['99990'],['99990']]
                items1=[list11_1,list12,list13,list14,list15,[i],list22,list23,['99990'],['99990'],['99990']]
                items2=[]
                items3=[]
            elif int(i[4:6]) <12:
                items0=[['99990'],['99990'],['99990'],['99990'],['99990'],list21_0,list22,list23,[i],list32,list33_0]
                items1=[]
                items2=[['99990'],['99990'],['99990'],['99990'],['99990'],list21_1,list22,list23,[i],list32,list33_0]
                items3=[['99990'],['99990'],['99990'],['99990'],['99990'],list21_0,list22,list23,[i],list32,list33_1]
            elif int(i[4:6]) <15:
                items0=[list11_0,list12,list13,list14,list15,list31,[i],list33_0,['99990'],['99990'],['99990']]
                items1=[list11_1,list12,list13,list14,list15,list31,[i],list33_0,['99990'],['99990'],['99990']]
                items2=[]
                items3=[list11_0,list12,list13,list14,list15,list31,[i],list33_1,['99990'],['99990'],['99990']]
            all_list=list(itertools.product(*items0))
            if len(list11_1) != 0 and items1 !=[]:
                all_list1=list(itertools.product(*items1))
            else:
                all_list1=[]
            if len(list21_1) != 0 and items2 !=[]:
                all_list2=list(itertools.product(*items2))
            else:
                all_list2=[]
            if len(list33_1) != 0 and items3 !=[]:
                all_list3=list(itertools.product(*items3))
            else:
                all_list3=[]
            all_list_god=all_list1+all_list2+all_list3
            del all_list1,all_list2,all_list3
            all_list_num=len(all_list_god)+len(all_list)
            all_list_list.append([all_list,all_list_god,all_list_num])

    know_bang1_on=[]
    for i in know_bang1_list:
        if select_item['tg'+i]==1:
            know_bang1_on.append(i)
    if select_item['tg21400640']==1:  ##경우2:만유(팔찌)+방어구(목걸이)
        items0=[['99990'],['99990'],['99990'],['99990'],['99990'],['21400640'],know_bang1_on,list23,['99990'],['99990'],['99990']]
        items1=[]
        items2=[]
        items3=[]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0 and items1 !=[]:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        if len(list21_1) != 0 and items2 !=[]:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        if len(list33_1) != 0 and items3 !=[]:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
                
    know_acc_on=[]
    know_bang2_on=[]
    for i in know_acc_list:
        if select_item['tg'+i]==1:
            know_acc_on.append(i)
    for i in know_bang2_list:
        if select_item['tg'+i]==1:
            know_bang2_on.append(i)
    if len(know_acc_list)!=0 and len(know_bang2_on)!=0:  ##경우3: 악세(법석)+방어구(보장)
        items0=[['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],know_bang2_on,know_acc_on,list33_0]
        items1=[]
        items2=[]
        items3=[['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],know_bang2_on,know_acc_on,list33_1]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0 and items1 !=[]:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        if len(list21_1) != 0 and items2 !=[]:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        if len(list33_1) != 0 and items3 !=[]:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
        
    jin_sang=[]
    jin_pal=[]
    jin_gui=[]
    for i in know_jin_list:
        if select_item['tg'+i]==1:
            if i[0:2]=='11':
                jin_sang.append(i)
            elif i[0:2]=='21':
                jin_pal.append(i)
            elif i[0:2]=='33':
                jin_gui.append(i)

    for i in know_jin_list:
        if select_item['tg'+i]==1: ##경우4: 진레전산물
            if i[0:2]=='11': ##상의만
                items0=[[i],['12410'],['13410'],['14410'],['15410'],list21_0,list22,list23,list31,list32,list33_0]
                items1=[]
                items2=[[i],['12410'],['13410'],['14410'],['15410'],list21_1,list22,list23,list31,list32,list33_0]
                items3=[[i],['12410'],['13410'],['14410'],['15410'],list21_0,list22,list23,list31,list32,list33_1]
                all_list=list(itertools.product(*items0))
                if len(list11_1) != 0 and items1 !=[]:
                    all_list1=list(itertools.product(*items1))
                else:
                    all_list1=[]
                if len(list21_1) != 0 and items2 !=[]:
                    all_list2=list(itertools.product(*items2))
                else:
                    all_list2=[]
                if len(list33_1) != 0 and items3 !=[]:
                    all_list3=list(itertools.product(*items3))
                else:
                    all_list3=[]
                all_list_god=all_list1+all_list2+all_list3
                del all_list1,all_list2,all_list3
                all_list_num=len(all_list_god)+len(all_list)
                all_list_list.append([all_list,all_list_god,all_list_num])
            if i[0:2]=='21': ##팔찌만
                items0=[list11_0,list12,list13,list14,list15,[i],['22420'],['23420'],list31,list32,list33_0]
                items1=[list11_1,list12,list13,list14,list15,[i],['22420'],['23420'],list31,list32,list33_0]
                items2=[]
                items3=[list11_0,list12,list13,list14,list15,[i],['22420'],['23420'],list31,list32,list33_1]
                all_list=list(itertools.product(*items0))
                if len(list11_1) != 0 and items1 !=[]:
                    all_list1=list(itertools.product(*items1))
                else:
                    all_list1=[]
                if len(list21_1) != 0 and items2 !=[]:
                    all_list2=list(itertools.product(*items2))
                else:
                    all_list2=[]
                if len(list33_1) != 0 and items3 !=[]:
                    all_list3=list(itertools.product(*items3))
                else:
                    all_list3=[]
                all_list_god=all_list1+all_list2+all_list3
                del all_list1,all_list2,all_list3
                all_list_num=len(all_list_god)+len(all_list)
                all_list_list.append([all_list,all_list_god,all_list_num])
            if i[0:2]=='33': ##귀걸만
                items0=[list11_0,list12,list13,list14,list15,list21_0,list22,list23,['31430'],['32430'],[i]]
                items1=[list11_1,list12,list13,list14,list15,list21_0,list22,list23,['31430'],['32430'],[i]]
                items2=[list11_0,list12,list13,list14,list15,list21_1,list22,list23,['31430'],['32430'],[i]]
                items3=[]
                all_list=list(itertools.product(*items0))
                if len(list11_1) != 0 and items1 !=[]:
                    all_list1=list(itertools.product(*items1))
                else:
                    all_list1=[]
                if len(list21_1) != 0 and items2 !=[]:
                    all_list2=list(itertools.product(*items2))
                else:
                    all_list2=[]
                if len(list33_1) != 0 and items3 !=[]:
                    all_list3=list(itertools.product(*items3))
                else:
                    all_list3=[]
                all_list_god=all_list1+all_list2+all_list3
                del all_list1,all_list2,all_list3
                all_list_num=len(all_list_god)+len(all_list)
                all_list_list.append([all_list,all_list_god,all_list_num])

    if len(jin_sang)!=0 and len(jin_pal)!=0: ##상의+팔찌
        items0=[jin_sang,['12410'],['13410'],['14410'],['15410'],jin_pal,['22420'],['23420'],list31,list32,list33_0]
        items1=[]
        items2=[]
        items3=[jin_sang,['12410'],['13410'],['14410'],['15410'],jin_pal,['22420'],['23420'],list31,list32,list33_1]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0 and items1 !=[]:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        if len(list21_1) != 0 and items2 !=[]:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        if len(list33_1) != 0 and items3 !=[]:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
    if len(jin_sang)!=0 and len(jin_gui)!=0: ##상의+귀걸
        items0=[jin_sang,['12410'],['13410'],['14410'],['15410'],list21_0,list22,list23,['31430'],['32430'],jin_gui]
        items1=[]
        items2=[jin_sang,['12410'],['13410'],['14410'],['15410'],list21_1,list22,list23,['31430'],['32430'],jin_gui]
        items3=[]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0 and items1 !=[]:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        if len(list21_1) != 0 and items2 !=[]:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        if len(list33_1) != 0 and items3 !=[]:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
    if len(jin_pal)!=0 and len(jin_gui)!=0: ##팔찌+귀걸
        items0=[list11_0,list12,list13,list14,list15,jin_pal,['22420'],['23420'],['31430'],['32430'],jin_gui]
        items1=[list11_1,list12,list13,list14,list15,jin_pal,['22420'],['23420'],['31430'],['32430'],jin_gui]
        items2=[]
        items3=[]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0 and items1 !=[]:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        if len(list21_1) != 0 and items2 !=[]:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        if len(list33_1) != 0 and items3 !=[]:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
    if len(jin_sang)!=0 and len(jin_pal)!=0 and len(jin_gui)!=0: ##3개 전부
        items0=[jin_sang,['12410'],['13410'],['14410'],['15410'],jin_pal,['22420'],['23420'],['31430'],['32430'],jin_gui]
        items1=[]
        items2=[]
        items3=[]
        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0 and items1 !=[]:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        if len(list21_1) != 0 and items2 !=[]:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        if len(list33_1) != 0 and items3 !=[]:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3
        all_list_num=len(all_list_god)+len(all_list)
        all_list_list.append([all_list,all_list_god,all_list_num])
        
        

##일반 경우의 수##
#########################################################################################################################
    timp_list_num=0
    for i in all_list_list:
        timp_list_num=timp_list_num+int(i[2])

    if select_perfect.get() != '메타몽모드(중간)' or timp_list_num < 1000000:
        items=[list11,list12,list13,list14,list15,list21,list22,list23,list31,list32,list33]
        items0=[list11_0,list12,list13,list14,list15,list21_0,list22,list23,list31,list32,list33_0]
        items1=[list11_1,list12,list13,list14,list15,list21_0,list22,list23,list31,list32,list33_0]
        items2=[list11_0,list12,list13,list14,list15,list21_1,list22,list23,list31,list32,list33_0]
        items3=[list11_0,list12,list13,list14,list15,list21_0,list22,list23,list31,list32,list33_1]

        all_list=list(itertools.product(*items0))
        if len(list11_1) != 0:
            all_list1=list(itertools.product(*items1))
        else:
            all_list1=[]
        if len(list21_1) != 0:
            all_list2=list(itertools.product(*items2))
        else:
            all_list2=[]
        if len(list33_1) != 0:
            all_list3=list(itertools.product(*items3))
        else:
            all_list3=[]
        all_list_god=all_list1+all_list2+all_list3
        del all_list1,all_list2,all_list3

        all_list_num=len(all_list_god)+len(all_list)
        if all_list_num==0:
            all_list=[('11360','12360','13360','14360','15360','11360','21370','22370','23370','31380','32380','33380')]

        all_list_list.append([all_list,all_list_god,all_list_num])

#########################################################################################################################

    for i in all_list_list:
        all_list_list_num=all_list_list_num+int(i[2])

    if all_list_list_num > 500000000:
        tkinter.messagebox.showerror('에러',"경우의 수가 5억가지가 넘습니다.\n진행이 불가능합니다.\n안 쓸 에픽 체크를 풀어주세요")
        showsta(text='중지됨')
        return
    elif all_list_list_num > 100000000:
        ask_msg2=tkinter.messagebox.askquestion('확인',"경우의 수가 1억가지가 넘습니다.\n메모리 에러가 날 수 있고 30분이상 걸릴 수 있습니다.\n진행하시겠습니까?")
        if ask_msg2 == 'no':
            showsta(text='중지됨')
            return
    elif all_list_list_num > 30000000:
        ask_msg2=tkinter.messagebox.askquestion('확인',"경우의 수가 3천만가지가 넘습니다.\n다소 오래 걸릴 수 있습니다.\n진행하시겠습니까?")
        if ask_msg2 == 'no':
            showsta(text='중지됨')
            return
    if set_perfect ==1 and all_list_list_num > 30000000:
        tkinter.messagebox.showerror('에러',"정확도 높음 기능은 많은 경우의 수를 지원하지 않습니다.")
        showsta(text='중지됨')
        return
    if set_perfect !=1 and all_list_list_num < 10000:
        set_perfect=1


#########################################################################################################################계산 시작

    for i in range(0,75):  ## 무기 
        if wep_select.get() == wep_list[i]:
            wep_num=(str(i+111001),)
            
    global exit_calc
    showsta(text='계산 시작')
    save_list={}
    save_list1={}
    save_list2={}
    save_list3={}
    max_setopt=0
    loop_counter=0

    
    for now_all_list in all_list_list:
        loop_counter=loop_counter+1
        print(str(loop_counter)+'회차')
        all_list_god=now_all_list[1]
        all_list=now_all_list[0]
        all_list_num=now_all_list[2]
        print(all_list_num)
    
        if jobup_select.get()[4:7] != "세인트" and jobup_select.get()[4:7] != "세라핌" and jobup_select.get()[4:7] != "헤카테":
            getone=opt_one.get
            
            if len(all_list_god)!=0:
                for calc_now in all_list_god:
                    if exit_calc==1:
                        showsta(text='중지됨')
                        return
                    set_list=["1"+str(calc_now[x][2:4]) for x in range(0,11)] 
                    set_val=Counter(set_list)
                    del set_val['136'],set_val['137'],set_val['138'],set_val['144'],set_val['145'],set_val['146'],set_val['147'],set_val['148'],set_val['149']
                    setopt_num=sum([floor(x*0.7) for x in set_val.values()])+1
                    if setopt_num >= max_setopt-set_perfect :
                        set_on=[];setapp=set_on.append
                        setcount=set_list.count
                        set_oncount=set_on.count
                        onecount=calc_now.count
                        for i in range(101,136):
                            if setcount(str(i))==2:
                                setapp(str(i)+"1")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"2")
                            if setcount(str(i))==5:
                                setapp(str(i)+"3")
                        for i in range(141,144):
                            if setcount(str(i))==2:
                                setapp(str(i)+"1")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"2")
                            if setcount(str(i))==5:
                                setapp(str(i)+"3")
                        for i in range(136,139):
                            if setcount(str(i))==2:
                                setapp(str(i)+"0")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"1")
                            if setcount(str(i))==5:
                                setapp(str(i)+"2")
                        for i in range(144,150):
                            if setcount(str(i))==2:
                                setapp(str(i)+"0")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"1")
                            if setcount(str(i))==5:
                                setapp(str(i)+"2")
                        if onecount('32390650')==1:
                            if onecount('21390340')==1:
                                setapp('1401')
                            elif onecount('31390540')==1:
                                setapp('1401')
                        if setopt_num >= max_setopt:
                            max_setopt = setopt_num
                        calc_wep=wep_num+calc_now
                        damage=0
                        base_array=np.array([0,0,0,0,extra_bon,0,extra_all,0,0,ele_in,0,1,0,0,0,0,0,0,extra_pas2,0,0,0,0,0,0,0,0,0])
                        max_damper=extra_dam
                        max_criper=extra_cri
                        skiper=0
                        coolper=0
                        for_calc=tuple(set_on)+calc_wep
                        oneone=len(for_calc)
                        oneonelist=[]
                        for i in range(oneone):
                            no_cut=getone(for_calc[i])               ## 11번 스증 ## 20번 쿨감
                            cut=np.array(no_cut[0:20]+no_cut[22:23]+no_cut[34:35]+no_cut[38:44])
                            skiper=(skiper/100+1)*(cut[11]/100+1)*100-100
                            coolper=(1-(100-coolper)/100*(100-cut[20])/100)*100
                            max_damper=max([no_cut[44],max_damper])
                            max_criper=max([no_cut[45],max_criper])
                            oneonelist.append(cut)
                        for i in range(oneone):
                            base_array=base_array+oneonelist[i]

                        if set_oncount('1201')==1 and onecount('32200')==1:
                            base_array[3]=base_array[3]-5
                        if onecount('33200')==1 and onecount('31200')==0:
                            base_array[8]=base_array[8]-10
                        if onecount('33230')==1 or onecount('33231')==1:
                            if onecount('31230')==0:
                                base_array[4]=base_array[4]-10
                            if onecount('32230')==0:
                                base_array[9]=base_array[9]-40
                        if onecount('15340')==1 or onecount('23340')==1 or onecount('33340')==1 or onecount('33341')==1:
                            if set_oncount('1341')==0 and set_oncount('1342') ==0:
                                if onecount('15340')==1:
                                    base_array[9]=base_array[9]-20
                                elif onecount('23340')==1:
                                    base_array[2]=base_array[2]-10
                                elif onecount('33340')==1:
                                    base_array[6]=base_array[6]-5
                                else:
                                    base_array[9]=base_array[9]-4
                                    base_array[2]=base_array[2]-2
                                    base_array[6]=base_array[6]-1
                                    base_array[8]=base_array[8]-1.93
                        if onecount('11111')==1:
                            if set_oncount('1112')==1 or set_oncount('1113')==1:
                                coolper=(1-(100-coolper)/100*(100-11)/100)*100
                        if onecount('11301')==1:
                            if onecount('22300')!=1:
                                base_array[4]=base_array[4]-10
                                base_array[7]=base_array[7]+10
                            if onecount('31300')!=1:
                                base_array[4]=base_array[4]-10
                                base_array[7]=base_array[7]+10
                        if onecount('11061')==1:
                            if betterang ==34:
                                if onecount('12060')==1:
                                    base_array[3]=base_array[3]+1
                                if onecount('13060')==1:
                                    skiper=skiper/1.34*1.35
                                if onecount('14060')==1:
                                    base_array[9]=base_array[9]+4
                                if onecount('15060')==1:
                                    base_array[8]=base_array[8]+1
                                if set_oncount('1063')==1:
                                    base_array[4]=base_array[4]+1
                        if set_oncount('1441') ==1:
                            if onecount('11440')!=1: ##3셋 공3% 모공5% 감소
                                base_array[7]=base_array[7]-3
                                base_array[6]=base_array[6]-5
                                
                        base_array[11]=skiper
                        base_array[2]=max_damper+base_array[2]
                        base_array[3]=max_criper+base_array[3]
                        real_bon=base_array[4]+base_array[5]*(base_array[9]*0.0045+1.05)
                        actlvl=((base_array[active_eff_one]+base_array[22]*job_lv1+base_array[23]*job_lv2+base_array[24]*job_lv3+
                                base_array[25]*job_lv4+base_array[26]*job_lv5+base_array[27]*job_lv6)/100+1)
                        paslvl=((100+base_array[16]*job_pas0)/100)*((100+base_array[17]*job_pas1)/100)*((100+base_array[18]*job_pas2)/100)*((100+base_array[19]*job_pas3)/100)
                        damage=((base_array[2]/100+1)*(base_array[3]/100+1)*(real_bon/100+1)*(base_array[6]/100+1)*(base_array[7]/100+1)*
                                (base_array[8]/100+1)*(base_array[9]*0.0045+1.05)*(base_array[10]/100+1)*(skiper/100+1)*
                                paslvl*((54500+3.31*base_array[0])/54500)*((4800+base_array[1])/4800)/(1.05+0.0045*int(ele_skill)))
                        final_damage=actlvl*damage*((100/(100-coolper)-1)*cool_eff*cool_on+1)*(base_array[12]/100+1)
                        
                        
                        base_array[4]=real_bon
                        save_list[final_damage]=[calc_wep,base_array,damage]
                        count_num=count_num+1
                    else:
                        count_all=count_all+1
            # 코드 이름
            # 0추스탯 1추공 2증 3크 4추 5속추
            # 6모 7공 8스탯 9속강 10지속 11스증 12특수
            # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 쿨감보정/21 2각캐특수액티브 /22~27 액티브레벨링/
            if max_setopt != 8 or set_perfect==1:
                for calc_now in all_list:
                    if exit_calc==1:
                        showsta(text='중지됨')
                        return
                    set_list=["1"+str(calc_now[x][2:4]) for x in range(0,11)] 
                    set_val=Counter(set_list)
                    del set_val['136'],set_val['137'],set_val['138'],set_val['144'],set_val['145'],set_val['146'],set_val['147'],set_val['148'],set_val['149']
                    setopt_num=sum([floor(x*0.7) for x in set_val.values()])
                    if setopt_num >= max_setopt-set_perfect :
                        set_on=[];setapp=set_on.append
                        setcount=set_list.count
                        set_oncount=set_on.count
                        onecount=calc_now.count
                        for i in range(101,136):
                            if setcount(str(i))==2:
                                setapp(str(i)+"1")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"2")
                            if setcount(str(i))==5:
                                setapp(str(i)+"3")
                        for i in range(141,144):
                            if setcount(str(i))==2:
                                setapp(str(i)+"1")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"2")
                            if setcount(str(i))==5:
                                setapp(str(i)+"3")
                        for i in range(136,139):
                            if setcount(str(i))==2:
                                setapp(str(i)+"0")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"1")
                            if setcount(str(i))==5:
                                setapp(str(i)+"2")
                        for i in range(144,150):
                            if setcount(str(i))==2:
                                setapp(str(i)+"0")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"1")
                            if setcount(str(i))==5:
                                setapp(str(i)+"2")
                        if onecount('32390650')==1:
                            if onecount('21390340')==1:
                                setapp('1401')
                            elif onecount('31390540')==1:
                                setapp('1401')
                        if setopt_num >= max_setopt:
                            max_setopt = setopt_num
                        calc_wep=wep_num+calc_now
                        damage=0
                        base_array=np.array([0,0,0,0,extra_bon,0,extra_all,0,0,ele_in,0,1,0,0,0,0,0,0,extra_pas2,0,0,0,0,0,0,0,0,0])
                        
                        skiper=0
                        coolper=0
                        for_calc=tuple(set_on)+calc_wep
                        oneone=len(for_calc)
                        oneonelist=[]
                        max_damper=extra_dam
                        max_criper=extra_cri
                        for i in range(oneone):
                            no_cut=getone(for_calc[i])               ## 11번 스증
                            cut=np.array(no_cut[0:20]+no_cut[22:23]+no_cut[34:35]+no_cut[38:44])
                            skiper=(skiper/100+1)*(cut[11]/100+1)*100-100
                            coolper=(1-(100-coolper)/100*(100-cut[20])/100)*100
                            max_damper=max([no_cut[44],max_damper])
                            max_criper=max([no_cut[45],max_criper])
                            oneonelist.append(cut)
                        for i in range(oneone):
                            base_array=base_array+oneonelist[i]
                        

                        if set_oncount('1201')==1 and onecount('32200')==1:
                            base_array[3]=base_array[3]-5
                        if onecount('33200')==1 and onecount('31200')==0:
                            base_array[8]=base_array[8]-10
                        if onecount('33230')==1 or onecount('33231')==1:
                            if onecount('31230')==0:
                                base_array[4]=base_array[4]-10
                            if onecount('32230')==0:
                                base_array[9]=base_array[9]-40
                        if onecount('15340')==1 or onecount('23340')==1 or onecount('33340')==1 or onecount('33341')==1:
                            if set_oncount('1341')==0 and set_oncount('1342') ==0:
                                if onecount('15340')==1:
                                    base_array[9]=base_array[9]-20
                                elif onecount('23340')==1:
                                    base_array[2]=base_array[2]-10
                                elif onecount('33340')==1:
                                    base_array[6]=base_array[6]-5
                                else:
                                    base_array[9]=base_array[9]-4
                                    base_array[2]=base_array[2]-2
                                    base_array[6]=base_array[6]-1
                                    base_array[8]=base_array[8]-1.93
                        if onecount('11111')==1:
                            if set_oncount('1112')==1 or set_oncount('1113')==1:
                                coolper=(1-(100-coolper)/100*(100-11)/100)*100
                        if onecount('11301')==1:
                            if onecount('22300')!=1:
                                base_array[4]=base_array[4]-10
                                base_array[7]=base_array[7]+10
                            if onecount('31300')!=1:
                                base_array[4]=base_array[4]-10
                                base_array[7]=base_array[7]+10
                        if onecount('11061')==1:
                            if betterang ==34:
                                if onecount('12060')==1:
                                    base_array[3]=base_array[3]+1
                                if onecount('13060')==1:
                                    skiper=skiper/1.34*1.35
                                if onecount('14060')==1:
                                    base_array[9]=base_array[9]+4
                                if onecount('15060')==1:
                                    base_array[8]=base_array[8]+1
                                if set_oncount('1063')==1:
                                    base_array[4]=base_array[4]+1
                        if set_oncount('1441') ==1:
                            if onecount('11440')!=1: ##3셋 공3% 모공5% 감소
                                base_array[7]=base_array[7]-3
                                base_array[6]=base_array[6]-5
                        
                        base_array[11]=skiper
                        base_array[2]=max_damper+base_array[2]
                        base_array[3]=max_criper+base_array[3]
                        real_bon=base_array[4]+base_array[5]*(base_array[9]*0.0045+1.05)
                        actlvl=((base_array[active_eff_one]+base_array[22]*job_lv1+base_array[23]*job_lv2+base_array[24]*job_lv3+
                                base_array[25]*job_lv4+base_array[26]*job_lv5+base_array[27]*job_lv6)/100+1)
                        paslvl=((100+base_array[16]*job_pas0)/100)*((100+base_array[17]*job_pas1)/100)*((100+base_array[18]*job_pas2)/100)*((100+base_array[19]*job_pas3)/100)
                        damage=((base_array[2]/100+1)*(base_array[3]/100+1)*(real_bon/100+1)*(base_array[6]/100+1)*(base_array[7]/100+1)*
                                (base_array[8]/100+1)*(base_array[9]*0.0045+1.05)*(base_array[10]/100+1)*(skiper/100+1)*
                                paslvl*((54500+3.31*base_array[0])/54500)*((4800+base_array[1])/4800)/(1.05+0.0045*int(ele_skill)))
                        final_damage=actlvl*damage*((100/(100-coolper)-1)*cool_eff*cool_on+1)*(base_array[12]/100+1)
                        
                        base_array[4]=real_bon
                        save_list[final_damage]=[calc_wep,base_array,damage]
                        count_num=count_num+1
                    else:
                        count_all=count_all+1
            else:
                print('스킵됨')
                count_all=count_all+len(all_list)
                
        else: ##버퍼
            base_b=10+int(db_preset['H2'].value)+int(db_preset['H4'].value)+int(db_preset['H5'].value)+1
            base_c=12+int(db_preset['H3'].value)+1
            base_pas0=0
            base_pas0_c=3
            base_pas0_b=0
            base_stat_s=4166+74-126+int(db_preset['H1'].value)
            base_stat_d=int(db_preset['H6'].value)-int(db_preset['H1'].value)
            base_stat_h=4308-45-83+int(db_preset['H1'].value)  ##2각 꺼지면 -528
            base_pas0_1=0
            load_presetc.close()
            lvlget=opt_buflvl.get
            #코드 이름
            #0 체정 1 지능
            #축복 2 스탯% 3 물공% 4 마공% 5 독공%
            #아포 6 고정 7 스탯%
            #8 축렙 9 포렙
            #10 아리아/보징증폭
            #11 전직패 12 보징 13 각패1 14 각패2 15 2각 16 각패3
            #17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨
            
            setget=opt_buf.get
            if len(all_list_god)!=0:
                for calc_now in all_list_god:
                    if exit_calc==1:
                        showsta(text='중지됨')
                        return
                    set_list=["1"+str(calc_now[x][2:4]) for x in range(0,11)]
                    set_val=Counter(set_list)
                    del set_val['136'],set_val['137'],set_val['138'],set_val['144'],set_val['145'],set_val['146'],set_val['147'],set_val['148'],set_val['149']
                    setopt_num=sum([floor(x*0.7) for x in set_val.values()])+1
                    if setopt_num >= max_setopt-set_perfect :
                        calc_wep=wep_num+calc_now
                        base_array=np.array([base_stat_h,base_stat_s,0,0,0,0,0,0,base_b,base_c,0,base_pas0,base_pas0_1,0,0,0,0,0,0,0,0])
                        set_on=[];setapp=set_on.append
                        setcount=set_list.count
                        set_oncount=set_on.count
                        onecount=calc_now.count
                        for i in range(101,136):
                            if setcount(str(i))==2:
                                setapp(str(i)+"1")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"2")
                            if setcount(str(i))==5:
                                setapp(str(i)+"3")
                        for i in range(141,144):
                            if setcount(str(i))==2:
                                setapp(str(i)+"1")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"2")
                            if setcount(str(i))==5:
                                setapp(str(i)+"3")
                        for i in range(136,139):
                            if setcount(str(i))==2:
                                setapp(str(i)+"0")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"1")
                            if setcount(str(i))==5:
                                setapp(str(i)+"2")
                        for i in range(144,150):
                            if setcount(str(i))==2:
                                setapp(str(i)+"0")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"1")
                            if setcount(str(i))==5:
                                setapp(str(i)+"2")
                        if setopt_num >= max_setopt:
                            max_setopt = setopt_num
                        b_stat=10.24 ##탈리스만
                        b_phy=0
                        b_mag=0
                        b_ind=0
                        c_per=0
                        for_calc=tuple(set_on)+calc_wep
                        oneone=len(for_calc)
                        oneonelist=[]
                        for i in range(oneone):
                            no_cut=np.array(setget(for_calc[i]))             ## 2 3 4 5 7
                            base_array=base_array+no_cut
                            b_stat=(b_stat/100+1)*(no_cut[2]/100+1)*100-100
                            b_phy=(b_phy/100+1)*(no_cut[3]/100+1)*100-100
                            b_mag=(b_mag/100+1)*(no_cut[4]/100+1)*100-100
                            b_ind=(b_ind/100+1)*(no_cut[5]/100+1)*100-100
                            c_per=(c_per/100+1)*(no_cut[7]/100+1)*100-100
                            oneonelist.append(no_cut)

                        if set_oncount('1441') ==1:
                            if onecount('11440')!=1: ##3셋 스탯160, 영축힘지8%, 물마독3% 감소
                                base_array[0]=base_array[0]-160
                                base_array[1]=base_array[1]-160
                                b_stat=(b_stat/100+1)/1.08*100-100
                                b_phy=(b_phy/100+1)/1.03*100-100
                                b_mag=(b_mag/100+1)/1.03*100-100
                                b_ind=(b_ind/100+1)/1.03*100-100
                        
                        if jobup_select.get()[4:7] == "세인트":
                            b_base_att=lvlget('hol_b_atta')[int(base_array[8])]
                            stat_pas0lvl_b=lvlget('pas0')[int(base_array[11])+base_pas0_b]+lvlget('hol_pas0_1')[int(base_array[12])]
                            stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+base_pas0_c]+lvlget('hol_pas0_1')[int(base_array[12])]
                            stat_pas1lvl=lvlget('hol_pas1')[int(base_array[13])]
                            stat_pas2lvl=lvlget('hol_act2')[int(base_array[15])]
                            stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
                            stat_b=base_array[0]+stat_pas0lvl_b+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+19*base_array[10]+base_stat_d
                            stat_c=base_array[0]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+19*base_array[10]
                            b_stat_calc=int(int(lvlget('hol_b_stat')[int(base_array[8])]*(b_stat/100+1))*(stat_b/630+1))
                            b_phy_calc=int(int(b_base_att*(b_phy/100+1))*(stat_b/630+1))
                            b_mag_calc=int(int(b_base_att*(b_mag/100+1))*(stat_b/630+1))
                            b_ind_calc=int(int(b_base_att*(b_ind/100+1))*(stat_b/630+1))
                            b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
                            c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
                            pas1_calc=int(lvlget('hol_pas1_out')[int(base_array[13])]+273+base_array[17])
                            pas1_out=str(int(lvlget('hol_pas1_out')[int(base_array[13])]+273+base_array[17]))+"("+str(int(20+base_array[13]))+"렙)"
                            save1='스탯='+str(b_stat_calc)+"\n앞뎀="+str(b_average)+"\n\n적용스탯= "+str(int(stat_b))+"\n적용레벨= "+str(int(base_array[8]))+"렙"

                        else:
                            if jobup_select.get()[4:7] == "세라핌":
                                b_value=675
                                aria=1.25+0.05*base_array[10]
                            if jobup_select.get()[4:7] == "헤카테":
                                b_value=665
                                aria=(1.20+0.05*base_array[10])*1.15
                                
                            b_base_att=lvlget('se_b_atta')[int(base_array[8])]
                            stat_pas0lvl_b=lvlget('pas0')[int(base_array[11])+int(base_pas0_b)]
                            stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+int(base_pas0_c)]
                            stat_pas1lvl=lvlget('se_pas1')[int(base_array[13])]+base_array[18]
                            stat_pas2lvl=lvlget('se_pas2')[int(base_array[14])]
                            stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
                            stat_b=base_array[1]+stat_pas0lvl_b+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+base_stat_d
                            stat_c=base_array[1]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl
                            b_stat_calc=int(int(lvlget('se_b_stat')[int(base_array[8])]*(b_stat/100+1))*(stat_b/b_value+1)*aria)
                            b_phy_calc=int(int(b_base_att*(b_phy/100+1)*(stat_b/b_value+1))*aria)
                            b_mag_calc=int(int(b_base_att*(b_mag/100+1)*(stat_b/b_value+1))*aria)
                            b_ind_calc=int(int(b_base_att*(b_ind/100+1)*(stat_b/b_value+1))*aria)
                            b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
                            c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
                            pas1_calc=int(stat_pas1lvl+442)
                            pas1_out=str(int(stat_pas1lvl+442))+"("+str(int(20+base_array[13]))+"렙)"
                            save1='스탯='+str(b_stat_calc)+"("+str(int(b_stat_calc/aria))+")\n앞뎀="+str(b_average)+"("+str(int(b_average/aria))+")\n\n적용스탯= "+str(int(stat_b))+"\n적용레벨= "+str(int(base_array[8]))+"렙"

                        save2='스탯= '+str(c_calc)+"\n\n적용스탯= "+str(int(stat_c))+"\n적용레벨= "+str(int(base_array[9]))+"렙"
                        ##1축 2포 3합
                        save_list1[((15000+b_stat_calc)/250+1)*(2650+b_average)]=[calc_wep,[save1,save2,pas1_out]]
                        save_list2[((15000+c_calc)/250+1)*2650]=[calc_wep,[save1,save2,pas1_out]]
                        save_list3[((15000+pas1_calc+c_calc+b_stat_calc)/250+1)*(2650+b_average)]=[calc_wep,[save1,save2,pas1_out]]
                                                
                        count_num=count_num+1
                    else:
                        count_all=count_all+1
                        
            if max_setopt != 8 or set_perfect==1:
                for calc_now in all_list:
                    if exit_calc==1:
                        showsta(text='중지됨')
                        return
                    set_list=["1"+str(calc_now[x][2:4]) for x in range(0,11)]
                    set_val=Counter(set_list)
                    del set_val['136'],set_val['137'],set_val['138'],set_val['144'],set_val['145'],set_val['146'],set_val['147'],set_val['148'],set_val['149']
                    setopt_num=sum([floor(x*0.7) for x in set_val.values()])
                    print(setopt_num)
                    if setopt_num >= max_setopt-set_perfect :
                        base_array=np.array([base_stat_h,base_stat_s,0,0,0,0,0,0,base_b,base_c,0,base_pas0,base_pas0_1,0,0,0,0,0,0,0,0])
                        calc_wep=wep_num+calc_now
                        set_on=[];setapp=set_on.append
                        setcount=set_list.count
                        set_oncount=set_on.count
                        onecount=calc_now.count
                        for i in range(101,136):
                            if setcount(str(i))==2:
                                setapp(str(i)+"1")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"2")
                            if setcount(str(i))==5:
                                setapp(str(i)+"3")
                        for i in range(141,144):
                            if setcount(str(i))==2:
                                setapp(str(i)+"1")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"2")
                            if setcount(str(i))==5:
                                setapp(str(i)+"3")
                        for i in range(136,139):
                            if setcount(str(i))==2:
                                setapp(str(i)+"0")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"1")
                            if setcount(str(i))==5:
                                setapp(str(i)+"2")
                        for i in range(144,150):
                            if setcount(str(i))==2:
                                setapp(str(i)+"0")
                            if 4>=setcount(str(i))>=3:
                                setapp(str(i)+"1")
                            if setcount(str(i))==5:
                                setapp(str(i)+"2")
                        if setopt_num >= max_setopt:
                            max_setopt = setopt_num
                        b_stat=10.24 ##탈리스만
                        b_phy=0
                        b_mag=0
                        b_ind=0
                        c_per=0
                        for_calc=tuple(set_on)+calc_wep
                        oneone=len(for_calc)
                        oneonelist=[]
                        for i in range(oneone):
                            no_cut=np.array(setget(for_calc[i]))             ## 2 3 4 5 7
                            base_array=base_array+no_cut
                            b_stat=(b_stat/100+1)*(no_cut[2]/100+1)*100-100
                            b_phy=(b_phy/100+1)*(no_cut[3]/100+1)*100-100
                            b_mag=(b_mag/100+1)*(no_cut[4]/100+1)*100-100
                            b_ind=(b_ind/100+1)*(no_cut[5]/100+1)*100-100
                            c_per=(c_per/100+1)*(no_cut[7]/100+1)*100-100
                            oneonelist.append(no_cut)

                        if set_oncount('1441') ==1:
                            if onecount('11440')!=1: ##3셋 스탯160, 영축힘지8%, 물마독3% 감소
                                base_array[0]=base_array[0]-160
                                base_array[1]=base_array[1]-160
                                b_stat=(b_stat/100+1)/1.08*100-100
                                b_phy=(b_phy/100+1)/1.03*100-100
                                b_mag=(b_mag/100+1)/1.03*100-100
                                b_ind=(b_ind/100+1)/1.03*100-100
                            
                        if jobup_select.get()[4:7] == "세인트":
                            b_base_att=lvlget('hol_b_atta')[int(base_array[8])]
                            stat_pas0lvl_b=lvlget('pas0')[int(base_array[11])+base_pas0_b]+lvlget('hol_pas0_1')[int(base_array[12])]
                            stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+base_pas0_c]+lvlget('hol_pas0_1')[int(base_array[12])]
                            stat_pas1lvl=lvlget('hol_pas1')[int(base_array[13])]
                            stat_pas2lvl=lvlget('hol_act2')[int(base_array[15])]
                            stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
                            stat_b=base_array[0]+stat_pas0lvl_b+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+19*base_array[10]+base_stat_d
                            stat_c=base_array[0]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+19*base_array[10]
                            b_stat_calc=int(int(lvlget('hol_b_stat')[int(base_array[8])]*(b_stat/100+1))*(stat_b/630+1))
                            b_phy_calc=int(int(b_base_att*(b_phy/100+1))*(stat_b/630+1))
                            b_mag_calc=int(int(b_base_att*(b_mag/100+1))*(stat_b/630+1))
                            b_ind_calc=int(int(b_base_att*(b_ind/100+1))*(stat_b/630+1))
                            b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
                            c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
                            pas1_calc=int(lvlget('hol_pas1_out')[int(base_array[13])]+273+base_array[17])
                            pas1_out=str(int(lvlget('hol_pas1_out')[int(base_array[13])]+273+base_array[17]))+"("+str(int(20+base_array[13]))+"렙)"
                            save1='스탯='+str(b_stat_calc)+"\n앞뎀="+str(b_average)+"\n\n적용스탯= "+str(int(stat_b))+"\n적용레벨= "+str(int(base_array[8]))+"렙"

                        else:
                            if jobup_select.get()[4:7] == "세라핌":
                                b_value=675
                                aria=1.25+0.05*base_array[10]
                            if jobup_select.get()[4:7] == "헤카테":
                                b_value=665
                                aria=(1.20+0.05*base_array[10])*1.15
                                
                            b_base_att=lvlget('se_b_atta')[int(base_array[8])]
                            stat_pas0lvl_b=lvlget('pas0')[int(base_array[11])+int(base_pas0_b)]
                            stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+int(base_pas0_c)]
                            stat_pas1lvl=lvlget('se_pas1')[int(base_array[13])]+base_array[18]
                            stat_pas2lvl=lvlget('se_pas2')[int(base_array[14])]
                            stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
                            stat_b=base_array[1]+stat_pas0lvl_b+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+base_stat_d
                            stat_c=base_array[1]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl
                            b_stat_calc=int(int(lvlget('se_b_stat')[int(base_array[8])]*(b_stat/100+1))*(stat_b/b_value+1)*aria)
                            b_phy_calc=int(int(b_base_att*(b_phy/100+1)*(stat_b/b_value+1))*aria)
                            b_mag_calc=int(int(b_base_att*(b_mag/100+1)*(stat_b/b_value+1))*aria)
                            b_ind_calc=int(int(b_base_att*(b_ind/100+1)*(stat_b/b_value+1))*aria)
                            b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
                            c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
                            pas1_calc=int(stat_pas1lvl+442)
                            pas1_out=str(int(stat_pas1lvl+442))+"("+str(int(20+base_array[13]))+"렙)"
                            save1='스탯='+str(b_stat_calc)+"("+str(int(b_stat_calc/aria))+")\n앞뎀="+str(b_average)+"("+str(int(b_average/aria))+")\n\n적용스탯= "+str(int(stat_b))+"\n적용레벨= "+str(int(base_array[8]))+"렙"

                        save2='스탯= '+str(c_calc)+"\n\n적용스탯= "+str(int(stat_c))+"\n적용레벨= "+str(int(base_array[9]))+"렙"
                        ##1축 2포 3합
                        save_list1[((15000+b_stat_calc)/250+1)*(2650+b_average)]=[calc_wep,[save1,save2,pas1_out]]
                        save_list2[((15000+c_calc)/250+1)*2650]=[calc_wep,[save1,save2,pas1_out]]
                        save_list3[((15000+pas1_calc+c_calc+b_stat_calc)/250+1)*(2650+b_average)]=[calc_wep,[save1,save2,pas1_out]]
                                                
                        count_num=count_num+1
                    else:
                        count_all=count_all+1
            else:
                print('스킵됨')
                count_all=count_all+len(all_list_god)



    if jobup_select.get()[4:7] != "세인트" and jobup_select.get()[4:7] != "세라핌" and jobup_select.get()[4:7] != "헤카테":
                    
        show_number=0
        showsta(text='결과 집계중')

        ranking=[]
        for j in range(0,5):
            try:
                now_max=max(list(save_list.keys()))
                ranking.append((now_max,save_list.get(now_max)))
                del save_list[now_max]
                showsta(text='결과 집계중'+str(j)+" / 5")
            except ValueError as error:
                pass
                
        show_result(ranking,'deal',ele_skill,cool_eff)


    else: ##버퍼
        
        show_number=0
        showsta(text='결과 집계중')

        ranking1=[];ranking2=[];ranking3=[]
        for j in range(0,5):
            try:
                now_max1=max(list(save_list1.keys()))
                ranking1.append((now_max1,save_list1.get(now_max1)))
                del save_list1[now_max1]
            except ValueError as error:
                pass
        for j in range(0,5):
            try:
                now_max2=max(list(save_list2.keys()))
                ranking2.append((now_max2,save_list2.get(now_max2)))
                del save_list2[now_max2]
            except ValueError as error:
                pass

        for j in range(0,5):
            try:
                now_max3=max(list(save_list3.keys()))
                ranking3.append((now_max3,save_list3.get(now_max3)))
                del save_list3[now_max3]
            except ValueError as error:
                pass
        ranking=[ranking1,ranking2,ranking3]
        show_result(ranking,'buf',ele_skill,cool_eff)
    load_excel.close()
    showsta(text='출력 완료')
    print("걸린 시간 = "+str(time.time() - start_time)+"초")
    
def calc_thread():
    threading.Thread(target=calc,daemon=True).start()

def show_result(rank_list,job_type,ele_skill,cool_eff):
    global result_window
    result_window=tkinter.Toplevel(self)
    result_window.attributes("-topmost", True)
    result_window.geometry("585x402+800+400")
    result_window.title("결과값")
    result_window.resizable(False,False)
    global canvas_res
    canvas_res = Canvas(result_window, width=587, height=404, bd=0)
    canvas_res.place(x=-2,y=-2)
    if job_type=='deal':
        result_bg=tkinter.PhotoImage(file='ext_img/bg_result.png')
    else:
        result_bg=tkinter.PhotoImage(file='ext_img/bg_result2.png')
    canvas_res.create_image(293,202,image=result_bg)

    global image_list, set_name_toggle, image_list_tag, now_version
    global res_img11,res_img12,res_img13,res_img14,res_img15,res_img21,res_img22,res_img23,res_img31,res_img32,res_img33,wep_select,jobup_select, now_rank_num, res_wep
    now_rank_num=0
    set_name_toggle=0
    
    wep_name=wep_select.get()
    job_name=jobup_select.get()[:-4]
    job_up_name=jobup_select.get()[-4:]
    res_wep=canvas_res.create_text(122,20,text=wep_name,font=guide_font,fill='white')
    canvas_res.create_text(122,50,text="<직업>",font=guide_font,fill='white')
    canvas_res.create_text(122,70,text=job_name,font=guide_font,fill='white')
    canvas_res.create_text(122,87,text=job_up_name,font=guide_font,fill='white')
    know_set_list=['22400150','22400250','22400350','22400450','22400550','21400640','31400750',
                   '31400850','31400950','31401050','31401150','32401240','32401340','32401440']

    if job_type=='deal': ###########################
        global tagkgum_exist,tagk_tg
        tagkgum_exist=0
        tagk_tg=0
        if wep_name=="(도)태극천제검":
            global  tagkgum
            tagkgum_exist=1
            tagk_tg=0
            tagkgum_img=tkinter.PhotoImage(file='ext_img/tagk_um.png')
            tagkgum=tkinter.Button(result_window,command=change_tagk,image=tagkgum_img,bg=dark_main,borderwidth=0,activebackground=dark_main)
            tagkgum.place(x=182,y=7)
            tagkgum.image=tagkgum_img
            tagkgum.command=change_tagk
            
        global result_image_on,result_image_tag,rank_dam,rank_stat,rank_stat2,rank_stat3,req_cool,res_dam,res_stat,res_stat2, res_stat3, rank_dam_tagk
        cool_check=req_cool.get()[2:6]
        canvas_res.create_text(122,145,text=cool_check,font=guide_font,fill='white')
        if cool_check=='쿨감보정':
            cool_eff_check=1
        else:
            cool_eff_check=0

        rank_dam=[0,0,0,0,0]
        rank_dam_tagk=[0,0,0,0,0]
        rank_dam_nolv=[0,0,0,0,0]
        rank_dam_tagk_nolv=[0,0,0,0,0]
        rank_setting=[0,0,0,0,0]
        rss=[0,0,0,0,0]
        result_image_on=[{},{},{},{},{}]
        result_image_tag=[{},{},{},{},{}]
        try:
            rank_dam[0]=str(int(100*rank_list[0][0]))+"%"
            rank_dam_nolv[0]=int(100*rank_list[0][1][2])
            rank_setting[0]=rank_list[0][1][0] ##0번이 랭킹이다
            rss[0]=rank_list[0][1][1]
            rank_dam_tagk[0]=str(int((100*rank_list[0][0])*(100+rss[0][6])/(121+rss[0][6])*1.23))+"%"
            rank_dam_tagk_nolv[0]=int((100+rss[0][6])/(121+rss[0][6])*1.23*(rank_dam_nolv[0]))
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[0][str(i)]=image_list[j]
                            result_image_tag[0][str(i)]=j
                            
            rank_dam[1]=str(int(100*rank_list[1][0]))+"%"
            rank_dam_nolv[1]=int(100*rank_list[1][1][2])
            rank_setting[1]=rank_list[1][1][0] 
            rss[1]=rank_list[1][1][1]
            rank_dam_tagk[1]=str(int((100*rank_list[1][0])*(100+rss[1][6])/(121+rss[1][6])*1.23))+"%"
            rank_dam_tagk_nolv[1]=int((100+rss[1][6])/(121+rss[1][6])*1.23*(rank_dam_nolv[1]))
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[1][str(i)]=image_list[j]
                            result_image_tag[1][str(i)]=j
                            
            rank_dam[2]=str(int(100*rank_list[2][0]))+"%"
            rank_dam_nolv[2]=int(100*rank_list[2][1][2])
            rank_setting[2]=rank_list[2][1][0] 
            rss[2]=rank_list[2][1][1]
            rank_dam_tagk[2]=str(int((100*rank_list[2][0])*(100+rss[2][6])/(121+rss[2][6])*1.23))+"%"
            rank_dam_tagk_nolv[2]=int((100+rss[2][6])/(121+rss[2][6])*1.23*(rank_dam_nolv[2]))
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[2][str(i)]=image_list[j]
                            result_image_tag[2][str(i)]=j

            rank_dam[3]=str(int(100*rank_list[3][0]))+"%"
            rank_dam_nolv[3]=int(100*rank_list[3][1][2])
            rank_setting[3]=rank_list[3][1][0] 
            rss[3]=rank_list[3][1][1]
            rank_dam_tagk[3]=str(int((100*rank_list[3][0])*(100+rss[3][6])/(121+rss[3][6])*1.23))+"%"
            rank_dam_tagk_nolv[3]=int((100+rss[3][6])/(121+rss[3][6])*1.23*(rank_dam_nolv[3]))
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[3][str(i)]=image_list[j]
                            result_image_tag[3][str(i)]=j

            rank_dam[4]=str(int(100*rank_list[4][0]))+"%"
            rank_dam_nolv[4]=int(100*rank_list[4][1][2])
            rank_setting[4]=rank_list[4][1][0] 
            rss[4]=rank_list[4][1][1]
            rank_dam_tagk[4]=str(int((100*rank_list[4][0])*(100+rss[4][6])/(121+rss[4][6])*1.23))+"%"
            rank_dam_tagk_nolv[4]=int((100+rss[4][6])/(121+rss[4][6])*1.23*(rank_dam_nolv[4]))
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on[4][str(i)]=image_list[j]
                            result_image_tag[4][str(i)]=j

        except IndexError as error:
            c=1

        for i in range(0,5):
            try:
                for j in rank_setting[i]:
                    for k in know_set_list:
                        if j==k:
                            if int(j[4:6])<12 and int(j[4:6])!=6:
                                result_image_on[i]['11']=image_list['n11'+j[4:6]];result_image_tag[i]['11']='n11'+j[4:6]
                                result_image_on[i]['12']=image_list['n12'+j[4:6]];result_image_tag[i]['12']='n12'+j[4:6]
                                result_image_on[i]['13']=image_list['n13'+j[4:6]];result_image_tag[i]['13']='n13'+j[4:6]
                                result_image_on[i]['14']=image_list['n14'+j[4:6]];result_image_tag[i]['14']='n14'+j[4:6]
                                result_image_on[i]['15']=image_list['n15'+j[4:6]];result_image_tag[i]['15']='n15'+j[4:6]
                            elif int(j[4:6])==6:
                                result_image_on[i]['31']=image_list['n31'+j[4:6]];result_image_tag[i]['31']='n31'+j[4:6]
                                result_image_on[i]['32']=image_list['n32'+j[4:6]];result_image_tag[i]['32']='n32'+j[4:6]
                                result_image_on[i]['33']=image_list['n33'+j[4:6]];result_image_tag[i]['33']='n33'+j[4:6]
                            elif int(j[4:6])<15:
                                result_image_on[i]['21']=image_list['n21'+j[4:6]];result_image_tag[i]['21']='n21'+j[4:6]
                                result_image_on[i]['22']=image_list['n22'+j[4:6]];result_image_tag[i]['22']='n22'+j[4:6]
                                result_image_on[i]['23']=image_list['n23'+j[4:6]];result_image_tag[i]['23']='n23'+j[4:6]
            except:
                pass
        # 0추스탯 1추공 2증 3크 4추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 쿨감보정/21 2각캐특수액티브 /22~27 액티브레벨링

        if job_up_name =='(진각)':
            simari=0
            jingakgi=1
        else:
            simari=1
            jingakgi=0
        rank_stat=[0,0,0,0,0]
        rank_stat2=[0,0,0,0,0]
        rank_stat3=[0,0,0,0,0]
        global rank_stat_tagk, rank_stat_tagk2
        rank_stat_tagk=[0,0,0,0,0]
        rank_stat_tagk2=[0,0,0,0,0]
        for i in range(0,5):
            try:
                rank_stat[i]=("증뎀= "+str(int(round(rss[i][2],0)))+
                              "%\n크증= "+str(int(round(rss[i][3],0)))+
                              "%\n추뎀= "+str(int(round(rss[i][4],0)))+
                              "%\n모공= "+str(int(round(rss[i][6],0)))+
                              "%\n공%= "+str(int(round(rss[i][7],0)))+
                              "%\n스탯= "+str(int(round(rss[i][8],0)))+
                              "%\n속강= "+str(int(round(rss[i][9],0)))+
                              "\n지속= "+str(int(round(rss[i][10],0)))+
                              "%\n스증= "+str(int(round(rss[i][11],0)))+
                              "%\n특수= "+str(int(round(rss[i][12],0)))+
                              "%\n\n공속= "+str(int(round(rss[i][13],0)))+
                              "%\n크확= "+str(int(round(rss[i][14],0)))+"%")
                if wep_name=="(도)태극천제검":
                    rank_stat_tagk[i]=("증뎀= "+str(int(round(rss[i][2],0)))+
                                      "%\n크증= "+str(int(round(rss[i][3],0)))+
                                      "%\n추뎀= "+str(int(round(rss[i][4],0)))+
                                      "%\n모공= "+str(-21+int(round(rss[i][6],0)))+
                                      "%\n공%= "+str(int(round(rss[i][7],0)))+
                                      "%\n스탯= "+str(int(round(rss[i][8],0)))+
                                      "%\n속강= "+str(int(round(rss[i][9],0)))+
                                      "\n지속= "+str(int(round(rss[i][10],0)))+
                                      "%\n스증= "+str(int(1.23*(100+round(rss[i][11],0))-100))+
                                      "%\n특수= "+str(int(round(rss[i][12],0)))+
                                      "%\n\n공속= "+str(-70+int(round(rss[i][13],0)))+
                                      "%\n크확= "+str(int(round(rss[i][14],0)))+"%")
                rank_stat2[i]=(str(int(round(rss[i][22],0)))+' : '+str(int(rank_dam_nolv[i]*(1+0.1014953549605047*(rss[i][22]+30))/(1+0.1014953549605047*30)))+'%'+
                               "\n"+str(int(round(rss[i][23],0)))+' : '+str(int(rank_dam_nolv[i]*(1+0.2318898690189789*(rss[i][23]+11))/(1+0.2318898690189789*11)))+'%'+
                               "\n"+str(int(round(rss[i][24],0)))+' : '+str(int(rank_dam_nolv[i]*(1+0.1014953549605047*(rss[i][24]+17))/(1+0.1014953549605047*17)))+'%'+
                               "\n"+str(int(round(rss[i][25],0)))+' : '+str(int(rank_dam_nolv[i]*(1+simari*(0.2+0.05*rss[i][27]))/(1+simari*0.2)*(1+0.2318898690189789*(rss[i][25]+4))/(1+0.2318898690189789*4)))+'%'+
                               "\n"+str(int(round(rss[i][26],0)))+' : '+str(int(rank_dam_nolv[i]*jingakgi*(1+0.1014953549605047*(rss[i][26]+5))/(1+0.1014953549605047*5)))+'%'+
                               "\n"+str(int(round(rss[i][27],0)))+' : '+str(int(rank_dam_nolv[i]*jingakgi*(1+0.2318898690189789*(rss[i][27]+1))/(1+0.2318898690189789*1)))+'%')
                if wep_name=="(도)태극천제검":
                    rank_stat_tagk2[i]=(str(int(round(rss[i][22],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*(1+0.1014953549605047*(rss[i][22]+30))/(1+0.1014953549605047*30)))+'%'+
                                       "\n"+str(int(round(rss[i][23],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*(1+0.2318898690189789*(rss[i][23]+11))/(1+0.2318898690189789*11)))+'%'+
                                       "\n"+str(int(round(rss[i][24],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*(1+0.1014953549605047*(rss[i][24]+17))/(1+0.1014953549605047*17)))+'%'+
                                       "\n"+str(int(round(rss[i][25],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*(1+simari*(0.2+0.05*rss[i][27]))/(1+simari*0.2)*(1+0.2318898690189789*(rss[i][25]+4))/(1+0.2318898690189789*4)))+'%'+
                                       "\n"+str(int(round(rss[i][26],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*jingakgi*(1+0.1014953549605047*(rss[i][26]+5))/(1+0.1014953549605047*5)))+'%'+
                                       "\n"+str(int(round(rss[i][27],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*jingakgi*(1+0.2318898690189789*(rss[i][27]+1))/(1+0.2318898690189789*1)))+'%')
                               
                rank_stat3[i]=(str(round(rss[i][16],1))+
                               "\n"+str(int(round(rss[i][17],0)))+
                               "\n"+str(int(round(rss[i][18],0)))+
                               "\n"+str(int(round(rss[i][19],0))))
            except TypeError as error:
                c=1

        
        if int(ele_skill) != 0:
            canvas_res.create_text(122,170,text="스킬 자속강="+str(int(ele_skill))+" / 역보정%="+str(round(100*(1.05/(1.05+int(ele_skill)*0.0045)-1),1))+"%",font=guide_font,fill='white')
        res_dam=canvas_res.create_text(122,125,text=rank_dam[0],font=mid_font,fill='white')
        res_stat=canvas_res.create_text(50,293,text=rank_stat[0],fill='white')
        res_stat2=canvas_res.create_text(193,263,text=rank_stat2[0],fill='white')
        res_stat3=canvas_res.create_text(145+24,361,text=rank_stat3[0],fill='white')

        res_img11=canvas_res.create_image(57,57,image=result_image_on[0]['11'])
        res_img12=canvas_res.create_image(27,87,image=result_image_on[0]['12'])
        res_img13=canvas_res.create_image(27,57,image=result_image_on[0]['13'])
        res_img14=canvas_res.create_image(57,87,image=result_image_on[0]['14'])
        res_img15=canvas_res.create_image(27,117,image=result_image_on[0]['15'])
        res_img21=canvas_res.create_image(189,57,image=result_image_on[0]['21'])
        res_img22=canvas_res.create_image(219,57,image=result_image_on[0]['22'])
        res_img23=canvas_res.create_image(219,87,image=result_image_on[0]['23'])
        res_img31=canvas_res.create_image(189,87,image=result_image_on[0]['31'])
        res_img32=canvas_res.create_image(219,117,image=result_image_on[0]['32'])
        res_img33=canvas_res.create_image(189,117,image=result_image_on[0]['33'])
        cn1=0
        global res_dam_list
        res_dam_list=[0,0,0,0,0]
        for j in range(0,5):
            try:
                for i in [11,12,13,14,15,21,22,23,31,32,33]:
                    canvas_res.create_image(268+cn1*29,67+78*j,image=result_image_on[j][str(i)])
                    cn1=cn1+1
                cn1=0
                res_dam_list[j]=canvas_res.create_text(346,34+78*j,text=rank_dam[j],font=mid_font,fill='white')
            except KeyError as error:
                c=1
        length=len(rank_list)
        
        canvas_res.create_text(217,361,text="계산기\n버전=\n "+str(now_version),fill='white', anchor='c')

    elif job_type=='buf': ##########################
        load_presetr=load_workbook("preset.xlsx", data_only=True)
        r_preset=load_presetr["custom"]
        global result_image_on1,result_image_on2,result_image_on3,rank_buf1,rank_buf2,rank_buf3, rank_type_buf, res_buf, res_img_list, res_buf_list, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3, res_buf_type_what
        global result_image_on1_tag,result_image_on2_tag,result_image_on3_tag
        rank_type_buf=3
        rank_setting1=[0,0,0,0,0]
        rank_setting2=[0,0,0,0,0]
        rank_setting3=[0,0,0,0,0]
        result_image_on1=[{},{},{},{},{}]
        result_image_on2=[{},{},{},{},{}]
        result_image_on3=[{},{},{},{},{}]
        result_image_on1_tag=[{},{},{},{},{}]
        result_image_on2_tag=[{},{},{},{},{}]
        result_image_on3_tag=[{},{},{},{},{}]
        rank_buf1=[0,0,0,0,0]
        rank_buf2=[0,0,0,0,0]
        rank_buf3=[0,0,0,0,0]
        rank_buf_ex1=[0,0,0,0,0]
        rank_buf_ex2=[0,0,0,0,0]
        rank_buf_ex3=[0,0,0,0,0]
        ## rank_setting[rank]=rank_list[a][rank][b][c]
        ## a: 0=축복,1=크오,2=합계
        ## b: 0=계수,1=스펙or증가량
        ## c: b에서 1 선택시, 0=스펙, 1=증가량
        try:
            rank_setting3[0]=rank_list[2][0][1][0]  ##2번째 숫자가 랭킹임
            rank_setting2[0]=rank_list[1][0][1][0]
            rank_setting1[0]=rank_list[0][0][1][0]
            rank_buf3[0]=int(rank_list[2][0][0]/10)
            rank_buf2[0]=int(rank_list[1][0][0]/10)
            rank_buf1[0]=int(rank_list[0][0][0]/10)
            rank_buf_ex3[0]=rank_list[2][0][1][1]
            rank_buf_ex2[0]=rank_list[1][0][1][1]
            rank_buf_ex1[0]=rank_list[0][0][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[0][str(i)]=image_list[j]
                            result_image_on3_tag[0][str(i)]=j
                for j in rank_setting2[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[0][str(i)]=image_list[j]
                            result_image_on2_tag[0][str(i)]=j
                for j in rank_setting1[0]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[0][str(i)]=image_list[j] ##
                            result_image_on1_tag[0][str(i)]=j
            rank_setting3[1]=rank_list[2][1][1][0]
            rank_setting2[1]=rank_list[1][1][1][0]
            rank_setting1[1]=rank_list[0][1][1][0]
            rank_buf3[1]=int(rank_list[2][1][0]/10)
            rank_buf2[1]=int(rank_list[1][1][0]/10)
            rank_buf1[1]=int(rank_list[0][1][0]/10)
            rank_buf_ex3[1]=rank_list[2][1][1][1]
            rank_buf_ex2[1]=rank_list[1][1][1][1]
            rank_buf_ex1[1]=rank_list[0][1][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[1][str(i)]=image_list[j]
                            result_image_on3_tag[1][str(i)]=j
                for j in rank_setting2[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[1][str(i)]=image_list[j]
                            result_image_on2_tag[1][str(i)]=j
                for j in rank_setting1[1]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[1][str(i)]=image_list[j] ##
                            result_image_on1_tag[1][str(i)]=j
            rank_setting3[2]=rank_list[2][2][1][0]
            rank_setting2[2]=rank_list[1][2][1][0]
            rank_setting1[2]=rank_list[0][2][1][0]
            rank_buf3[2]=int(rank_list[2][2][0]/10)
            rank_buf2[2]=int(rank_list[1][2][0]/10)
            rank_buf1[2]=int(rank_list[0][2][0]/10)
            rank_buf_ex3[2]=rank_list[2][2][1][1]
            rank_buf_ex2[2]=rank_list[1][2][1][1]
            rank_buf_ex1[2]=rank_list[0][2][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[2][str(i)]=image_list[j]
                            result_image_on3_tag[2][str(i)]=j
                for j in rank_setting2[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[2][str(i)]=image_list[j]
                            result_image_on2_tag[2][str(i)]=j
                for j in rank_setting1[2]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[2][str(i)]=image_list[j] ##
                            result_image_on1_tag[2][str(i)]=j
            rank_setting3[3]=rank_list[2][3][1][0]
            rank_setting2[3]=rank_list[1][3][1][0]
            rank_setting1[3]=rank_list[0][3][1][0]
            rank_buf3[3]=int(rank_list[2][3][0]/10)
            rank_buf2[3]=int(rank_list[1][3][0]/10)
            rank_buf1[3]=int(rank_list[0][3][0]/10)
            rank_buf_ex3[3]=rank_list[2][3][1][1]
            rank_buf_ex2[3]=rank_list[1][3][1][1]
            rank_buf_ex1[3]=rank_list[0][3][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[3][str(i)]=image_list[j]
                            result_image_on3_tag[3][str(i)]=j
                for j in rank_setting2[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[3][str(i)]=image_list[j]
                            result_image_on2_tag[3][str(i)]=j
                for j in rank_setting1[3]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[3][str(i)]=image_list[j] ##
                            result_image_on1_tag[3][str(i)]=j
            rank_setting3[4]=rank_list[2][4][1][0]
            rank_setting2[4]=rank_list[1][4][1][0]
            rank_setting1[4]=rank_list[0][4][1][0]
            rank_buf3[4]=int(rank_list[2][4][0]/10)
            rank_buf2[4]=int(rank_list[1][4][0]/10)
            rank_buf1[4]=int(rank_list[0][4][0]/10)
            rank_buf_ex3[4]=rank_list[2][4][1][1]
            rank_buf_ex2[4]=rank_list[1][4][1][1]
            rank_buf_ex1[4]=rank_list[0][4][1][1]
            for i in [11,12,13,14,15,21,22,23,31,32,33]:
                for j in rank_setting3[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on3[4][str(i)]=image_list[j]
                            result_image_on3_tag[4][str(i)]=j
                for j in rank_setting2[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on2[4][str(i)]=image_list[j]
                            result_image_on2_tag[4][str(i)]=j
                for j in rank_setting1[4]:
                    if len(j) != 6:
                        if j[0:2] == str(i):
                            result_image_on1[4][str(i)]=image_list[j] ##
                            result_image_on1_tag[4][str(i)]=j
        except IndexError as error:
            c=1
            
        canvas_res.create_text(122,193,text="커스텀 축복+"+str(int(r_preset['H2'].value)+int(r_preset['H4'].value)+int(r_preset['H5'].value))+"렙 / "+"커스텀 1각+"+str(int(r_preset['H3'].value))+"렙\n축복스탯+"+str(int(r_preset['H6'].value))+" / 1각 스탯+"+str(int(r_preset['H1'].value)),font=guide_font,fill='white')

        res_buf=canvas_res.create_text(122,125,text=rank_buf3[0],font=mid_font,fill='white')
        res_buf_type_what=canvas_res.create_text(122,145,text="총합 기준",font=guide_font,fill='white')
        res_buf_ex1=canvas_res.create_text(64,283,text=rank_buf_ex3[0][0],font=small_font,fill='white')
        res_buf_ex2=canvas_res.create_text(183,261,text=rank_buf_ex3[0][1],font=small_font,fill='white')
        res_buf_ex3=canvas_res.create_text(183,318,text=rank_buf_ex3[0][2],font=small_font,fill='white')

        res_img11=canvas_res.create_image(57,57,image=result_image_on3[0]['11'])
        res_img12=canvas_res.create_image(27,87,image=result_image_on3[0]['12'])
        res_img13=canvas_res.create_image(27,57,image=result_image_on3[0]['13'])
        res_img14=canvas_res.create_image(57,87,image=result_image_on3[0]['14'])
        res_img15=canvas_res.create_image(27,117,image=result_image_on3[0]['15'])
        res_img21=canvas_res.create_image(189,57,image=result_image_on3[0]['21'])
        res_img22=canvas_res.create_image(219,57,image=result_image_on3[0]['22'])
        res_img23=canvas_res.create_image(219,87,image=result_image_on3[0]['23'])
        res_img31=canvas_res.create_image(189,87,image=result_image_on3[0]['31'])
        res_img32=canvas_res.create_image(219,117,image=result_image_on3[0]['32'])
        res_img33=canvas_res.create_image(189,117,image=result_image_on3[0]['33'])
        cn1=0
        res_img_list={}
        res_buf_list={}
        for j in range(0,5):
            try:
                for i in [11,12,13,14,15,21,22,23,31,32,33]:
                    temp_res=canvas_res.create_image(268+cn1*29,67+78*j,image=result_image_on3[j][str(i)])
                    res_img_list[str(j)+str(i)]=temp_res
                    cn1=cn1+1
                cn1=0
                temp_buf=canvas_res.create_text(346,34+78*j,text=rank_buf3[j],font=mid_font,fill='white')
                res_buf_list[j]=temp_buf
            except KeyError as error:
                c=1
        length=len(rank_list[0])
        type1_img=tkinter.PhotoImage(file='ext_img/type_bless.png')
        type2_img=tkinter.PhotoImage(file='ext_img/type_crux.png')
        type3_img=tkinter.PhotoImage(file='ext_img/type_all.png')
        rank_type_but1=tkinter.Button(result_window,command=lambda:change_rank_type(1),image=type1_img,bg=dark_main,borderwidth=0,activebackground=dark_main);rank_type_but1.place(x=8,y=337)
        rank_type_but2=tkinter.Button(result_window,command=lambda:change_rank_type(2),image=type2_img,bg=dark_main,borderwidth=0,activebackground=dark_main);rank_type_but2.place(x=84,y=337)
        rank_type_but3=tkinter.Button(result_window,command=lambda:change_rank_type(3),image=type3_img,bg=dark_main,borderwidth=0,activebackground=dark_main);rank_type_but3.place(x=160,y=337)
        rank_type_but1.image=type1_img
        rank_type_but2.image=type2_img
        rank_type_but3.image=type3_img
        
        load_presetr.close()

    
    
    show_detail_img=tkinter.PhotoImage(file='ext_img/show_detail.png')
    
    res_bt1=tkinter.Button(result_window,command=lambda:change_rank(0,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue);res_bt1.place(x=486,y=20+78*0)
    res_bt2=tkinter.Button(result_window,command=lambda:change_rank(1,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue)
    res_bt3=tkinter.Button(result_window,command=lambda:change_rank(2,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue)
    res_bt4=tkinter.Button(result_window,command=lambda:change_rank(3,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue)
    res_bt5=tkinter.Button(result_window,command=lambda:change_rank(4,job_type),image=show_detail_img,bg=dark_blue,borderwidth=0,activebackground=dark_blue)
    if length>1:
        res_bt2.place(x=486,y=20+78*1)
    if length>2:
        res_bt3.place(x=486,y=20+78*2)
    if length>3:
        res_bt4.place(x=486,y=20+78*3)
    if length>4:
        res_bt5.place(x=486,y=20+78*4)

    show_tag_img=tkinter.PhotoImage(file='ext_img/show_set_tag.png')
    show_tag_but=tkinter.Button(result_window,command=lambda:show_set_name(job_type),image=show_tag_img,bg=dark_sub,borderwidth=0,activebackground=dark_sub)
    show_tag_but.place(x=173,y=133)
    show_tag_but.image=show_tag_img

    canvas_res.image=result_bg
    res_bt1.image=show_detail_img

def change_tagk():
    global tagk_tg, tagkgum
    global res_stat,res_stat2,rank_stat_tagk, rank_stat, rank_stat_tagk2, rank_stat2
    global res_dam, rank_dam_tagk, rank_dam
    global now_rank_num
    global res_dam_list
    now=now_rank_num
    tagkgum_img=tkinter.PhotoImage(file='ext_img/tagk_um.png')
    tagkgum_img2=tkinter.PhotoImage(file='ext_img/tagk_ang.png')
    if tagk_tg==0:
        canvas_res.itemconfig(res_dam,text=rank_dam_tagk[now])
        canvas_res.itemconfig(res_stat,text=rank_stat_tagk[now])
        canvas_res.itemconfig(res_stat2,text=rank_stat_tagk2[now])
        tagkgum['image']=tagkgum_img2
        tagkgum.image=tagkgum_img2
        tagk_tg=1
        canvas_res.itemconfig(res_wep,fill='red')
        for i in range(0,5):
            try:
                canvas_res.itemconfig(res_dam_list[i],text=rank_dam_tagk[i],fill='red')
            except:
                pass

    elif tagk_tg==1:
        canvas_res.itemconfig(res_dam,text=rank_dam[now])
        canvas_res.itemconfig(res_stat,text=rank_stat[now])
        canvas_res.itemconfig(res_stat2,text=rank_stat2[now])
        tagkgum['image']=tagkgum_img
        tagkgum.image=tagkgum_img
        tagk_tg=0
        canvas_res.itemconfig(res_wep,fill='white')
        for i in range(0,5):
            try:
                canvas_res.itemconfig(res_dam_list[i],text=rank_dam[i],fill='white')
            except:
                pass

        

def change_rank(now,job_type):
    global image_list,canvas_res, res_img11,res_img12,res_img13,res_img14,res_img15,res_img21,res_img22,res_img23,res_img31,res_img32,res_img33, now_rank_num, res_wep, res_dam_list
    now_rank_num=now
    if job_type =='deal':
        global tagk_tg, tagkgum, tagkgum_exist, rank_dam_tagk, rank_stat_tagk, rank_stat_tagk2
        global res_dam,res_stat,res_stat2,res_stat3,rank_stat,rank_stat2,rank_stat3,result_image_on
        try:      
            image_changed=result_image_on[now]
            canvas_res.itemconfig(res_img11,image=image_changed['11'])
            canvas_res.itemconfig(res_img12,image=image_changed['12'])
            canvas_res.itemconfig(res_img13,image=image_changed['13'])
            canvas_res.itemconfig(res_img14,image=image_changed['14'])
            canvas_res.itemconfig(res_img15,image=image_changed['15'])
            canvas_res.itemconfig(res_img21,image=image_changed['21'])
            canvas_res.itemconfig(res_img22,image=image_changed['22'])
            canvas_res.itemconfig(res_img23,image=image_changed['23'])
            canvas_res.itemconfig(res_img31,image=image_changed['31'])
            canvas_res.itemconfig(res_img32,image=image_changed['32'])
            canvas_res.itemconfig(res_img33,image=image_changed['33'])
            if tagkgum_exist==1 and tagk_tg==1:
                canvas_res.itemconfig(res_dam,text=rank_dam_tagk[now])
                canvas_res.itemconfig(res_stat,text=rank_stat_tagk[now])
                canvas_res.itemconfig(res_stat2,text=rank_stat_tagk2[now])
            else:
                canvas_res.itemconfig(res_dam,text=rank_dam[now])
                canvas_res.itemconfig(res_stat,text=rank_stat[now])
                canvas_res.itemconfig(res_stat2,text=rank_stat2[now])
            canvas_res.itemconfig(res_stat3,text=rank_stat3[now])
        except KeyError as error:
            c=1

    elif job_type =='buf':
        global result_image_on1,result_image_on2,result_image_on3,rank_buf1,rank_buf2,rank_buf3, rank_type_buf, res_buf, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3
        try:
            if rank_type_buf==1:
                image_changed=result_image_on1[now]
                rank_changed=rank_buf1[now]
                rank_buf_ex_changed=rank_buf_ex1
            elif rank_type_buf==2:
                image_changed=result_image_on2[now]
                rank_changed=rank_buf2[now]
                rank_buf_ex_changed=rank_buf_ex2
            elif rank_type_buf==3:
                image_changed=result_image_on3[now]
                rank_changed=rank_buf3[now]
                rank_buf_ex_changed=rank_buf_ex3
            canvas_res.itemconfig(res_buf,text=rank_changed)
            canvas_res.itemconfig(res_buf_ex1,text=rank_buf_ex_changed[now][0])
            canvas_res.itemconfig(res_buf_ex2,text=rank_buf_ex_changed[now][1])
            canvas_res.itemconfig(res_buf_ex3,text=rank_buf_ex_changed[now][2])                
            canvas_res.itemconfig(res_img11,image=image_changed['11'])
            canvas_res.itemconfig(res_img12,image=image_changed['12'])
            canvas_res.itemconfig(res_img13,image=image_changed['13'])
            canvas_res.itemconfig(res_img14,image=image_changed['14'])
            canvas_res.itemconfig(res_img15,image=image_changed['15'])
            canvas_res.itemconfig(res_img21,image=image_changed['21'])
            canvas_res.itemconfig(res_img22,image=image_changed['22'])
            canvas_res.itemconfig(res_img23,image=image_changed['23'])
            canvas_res.itemconfig(res_img31,image=image_changed['31'])
            canvas_res.itemconfig(res_img32,image=image_changed['32'])
            canvas_res.itemconfig(res_img33,image=image_changed['33'])
        except KeyError as error:
            c=1

def show_set_name(job_type):
    global image_list,canvas_res,res_img11,res_img12,res_img13,res_img14,res_img15,res_img21,res_img22,res_img23,res_img31,res_img32,res_img33, now_rank_num
    global set_name_toggle, image_list_tag, result_image_on, result_image_tag
    if job_type == "deal":
        global result_image_tag
        if set_name_toggle ==0:
            set_name_toggle=1
            canvas_res.itemconfig(res_img11,image=image_list_tag[result_image_tag[now_rank_num]['11']])
            canvas_res.itemconfig(res_img12,image=image_list_tag[result_image_tag[now_rank_num]['12']])
            canvas_res.itemconfig(res_img13,image=image_list_tag[result_image_tag[now_rank_num]['13']])
            canvas_res.itemconfig(res_img14,image=image_list_tag[result_image_tag[now_rank_num]['14']])
            canvas_res.itemconfig(res_img15,image=image_list_tag[result_image_tag[now_rank_num]['15']])
            canvas_res.itemconfig(res_img21,image=image_list_tag[result_image_tag[now_rank_num]['21']])
            canvas_res.itemconfig(res_img22,image=image_list_tag[result_image_tag[now_rank_num]['22']])
            canvas_res.itemconfig(res_img23,image=image_list_tag[result_image_tag[now_rank_num]['23']])
            canvas_res.itemconfig(res_img31,image=image_list_tag[result_image_tag[now_rank_num]['31']])
            canvas_res.itemconfig(res_img32,image=image_list_tag[result_image_tag[now_rank_num]['32']])
            canvas_res.itemconfig(res_img33,image=image_list_tag[result_image_tag[now_rank_num]['33']])
        elif set_name_toggle ==1:
            set_name_toggle=0
            canvas_res.itemconfig(res_img11,image=image_list[result_image_tag[now_rank_num]['11']])
            canvas_res.itemconfig(res_img12,image=image_list[result_image_tag[now_rank_num]['12']])
            canvas_res.itemconfig(res_img13,image=image_list[result_image_tag[now_rank_num]['13']])
            canvas_res.itemconfig(res_img14,image=image_list[result_image_tag[now_rank_num]['14']])
            canvas_res.itemconfig(res_img15,image=image_list[result_image_tag[now_rank_num]['15']])
            canvas_res.itemconfig(res_img21,image=image_list[result_image_tag[now_rank_num]['21']])
            canvas_res.itemconfig(res_img22,image=image_list[result_image_tag[now_rank_num]['22']])
            canvas_res.itemconfig(res_img23,image=image_list[result_image_tag[now_rank_num]['23']])
            canvas_res.itemconfig(res_img31,image=image_list[result_image_tag[now_rank_num]['31']])
            canvas_res.itemconfig(res_img32,image=image_list[result_image_tag[now_rank_num]['32']])
            canvas_res.itemconfig(res_img33,image=image_list[result_image_tag[now_rank_num]['33']])
    elif job_type == "buf":
        global result_image_on1_tag,result_image_on2_tag,result_image_on3_tag, rank_type_buf
        if rank_type_buf==1:
            temp_image_tag=result_image_on1_tag
        elif rank_type_buf==2:
            temp_image_tag=result_image_on2_tag
        elif rank_type_buf==3:
            temp_image_tag=result_image_on3_tag
        if set_name_toggle ==0:
            set_name_toggle=1
            canvas_res.itemconfig(res_img11,image=image_list_tag[temp_image_tag[now_rank_num]['11']])
            canvas_res.itemconfig(res_img12,image=image_list_tag[temp_image_tag[now_rank_num]['12']])
            canvas_res.itemconfig(res_img13,image=image_list_tag[temp_image_tag[now_rank_num]['13']])
            canvas_res.itemconfig(res_img14,image=image_list_tag[temp_image_tag[now_rank_num]['14']])
            canvas_res.itemconfig(res_img15,image=image_list_tag[temp_image_tag[now_rank_num]['15']])
            canvas_res.itemconfig(res_img21,image=image_list_tag[temp_image_tag[now_rank_num]['21']])
            canvas_res.itemconfig(res_img22,image=image_list_tag[temp_image_tag[now_rank_num]['22']])
            canvas_res.itemconfig(res_img23,image=image_list_tag[temp_image_tag[now_rank_num]['23']])
            canvas_res.itemconfig(res_img31,image=image_list_tag[temp_image_tag[now_rank_num]['31']])
            canvas_res.itemconfig(res_img32,image=image_list_tag[temp_image_tag[now_rank_num]['32']])
            canvas_res.itemconfig(res_img33,image=image_list_tag[temp_image_tag[now_rank_num]['33']])
        elif set_name_toggle ==1:
            set_name_toggle=0
            canvas_res.itemconfig(res_img11,image=image_list[temp_image_tag[now_rank_num]['11']])
            canvas_res.itemconfig(res_img12,image=image_list[temp_image_tag[now_rank_num]['12']])
            canvas_res.itemconfig(res_img13,image=image_list[temp_image_tag[now_rank_num]['13']])
            canvas_res.itemconfig(res_img14,image=image_list[temp_image_tag[now_rank_num]['14']])
            canvas_res.itemconfig(res_img15,image=image_list[temp_image_tag[now_rank_num]['15']])
            canvas_res.itemconfig(res_img21,image=image_list[temp_image_tag[now_rank_num]['21']])
            canvas_res.itemconfig(res_img22,image=image_list[temp_image_tag[now_rank_num]['22']])
            canvas_res.itemconfig(res_img23,image=image_list[temp_image_tag[now_rank_num]['23']])
            canvas_res.itemconfig(res_img31,image=image_list[temp_image_tag[now_rank_num]['31']])
            canvas_res.itemconfig(res_img32,image=image_list[temp_image_tag[now_rank_num]['32']])
            canvas_res.itemconfig(res_img33,image=image_list[temp_image_tag[now_rank_num]['33']])

        
def change_rank_type(in_type):
    global image_list,canvas_res, res_img11,res_img12,res_img13,res_img14,res_img15,res_img21,res_img22,res_img23,res_img31,res_img32,res_img33
    global result_image_on1,result_image_on2,result_image_on3,rank_buf1,rank_buf2,rank_buf3, rank_type_buf, res_img_list, res_buf_list, res_buf_ex1, res_buf_ex2, res_buf_ex3, rank_buf_ex1, rank_buf_ex2, rank_buf_ex3, res_buf_type_what
    if in_type==1:
        rank_type_buf=1
        image_changed=result_image_on1[0]
        image_changed_all=result_image_on1
        rank_changed=rank_buf1
        rank_buf_ex_changed=rank_buf_ex1
        type_changed="축복 기준"
    elif in_type==2:
        rank_type_buf=2
        image_changed=result_image_on2[0]
        image_changed_all=result_image_on2
        rank_changed=rank_buf2
        rank_buf_ex_changed=rank_buf_ex2
        type_changed="1각 기준"
    elif in_type==3:
        rank_type_buf=3
        image_changed=result_image_on3[0]
        image_changed_all=result_image_on3
        rank_changed=rank_buf3
        rank_buf_ex_changed=rank_buf_ex3
        type_changed="총합 기준"
    canvas_res.itemconfig(res_buf_type_what,text=type_changed)
    canvas_res.itemconfig(res_buf_ex1,text=rank_buf_ex_changed[0][0])
    canvas_res.itemconfig(res_buf_ex2,text=rank_buf_ex_changed[0][1])
    canvas_res.itemconfig(res_buf_ex3,text=rank_buf_ex_changed[0][2])            
    canvas_res.itemconfig(res_buf,text=rank_changed[0])
    canvas_res.itemconfig(res_img11,image=image_changed['11'])
    canvas_res.itemconfig(res_img12,image=image_changed['12'])
    canvas_res.itemconfig(res_img13,image=image_changed['13'])
    canvas_res.itemconfig(res_img14,image=image_changed['14'])
    canvas_res.itemconfig(res_img15,image=image_changed['15'])
    canvas_res.itemconfig(res_img21,image=image_changed['21'])
    canvas_res.itemconfig(res_img22,image=image_changed['22'])
    canvas_res.itemconfig(res_img23,image=image_changed['23'])
    canvas_res.itemconfig(res_img31,image=image_changed['31'])
    canvas_res.itemconfig(res_img32,image=image_changed['32'])
    canvas_res.itemconfig(res_img33,image=image_changed['33'])
    cn2=0
    for j in range(0,5):
            try:
                for i in [11,12,13,14,15,21,22,23,31,32,33]:
                    canvas_res.itemconfig(res_img_list[str(j)+str(i)],image=image_changed_all[j][str(i)])
                    cn2=cn2+2
                cn2=0
                canvas_res.itemconfig(res_buf_list[j],text=rank_changed[j],font=mid_font,fill='white')
            except KeyError as error:
                c=1
   
def costum(auto):
    global custom_window
    try:
        custom_window.destroy()
    except:
        pass
    custom_window=tkinter.Toplevel(self)
    custom_window.attributes("-topmost", True) 
    custom_window.geometry("620x400+750+20")

    load_preset=load_workbook("preset.xlsx",data_only=True)
    db_preset=load_preset["custom"]
    
    tkinter.Label(custom_window,text="<딜러환경>",font=mid_font).place(x=100,y=10)
    tkinter.Label(custom_window,text="주속성=",font=guide_font).place(x=10,y=50)
    ele_list=['화','수','명','암']
    ele_type=tkinter.ttk.Combobox(custom_window,width=5,values=ele_list); ele_type.place(x=80,y=52) ##     
    ele_type.set(db_preset['B1'].value)
    tkinter.Label(custom_window,text="쿨감보정=          %",font=guide_font).place(x=160,y=50) ##Y11/Z11
    cool_con=tkinter.Entry(custom_window,width=5);cool_con.place(x=230,y=52)
    cool_con.insert(END,db_preset['B2'].value)
    
    tkinter.Label(custom_window,text="<딜러장비>",font=mid_font).place(x=100,y=85)
    tkinter.Label(custom_window,text="% 입력창은 그만큼 스증으로 환산한다는 뜻",fg="Red").place(x=30,y=120)
    tkinter.Label(custom_window,text="선택벨트=          %",font=guide_font).place(x=160,y=155) ##O164
    cus1=tkinter.Entry(custom_window,width=5);cus1.place(x=230,y=157)
    cus1.insert(END,db_preset['B3'].value)
    tkinter.Label(custom_window,text="선택신발=          %",font=guide_font).place(x=160,y=185) ##O180
    cus2=tkinter.Entry(custom_window,width=5);cus2.place(x=230,y=187)
    cus2.insert(END,db_preset['B4'].value)
    tkinter.Label(custom_window,text="베테랑=",font=guide_font).place(x=160,y=215) ##G276
    lvl_list=['전설↓','영웅↑']
    cus3=tkinter.ttk.Combobox(custom_window,width=5,values=lvl_list); cus3.place(x=230,y=217)
    cus3.set(db_preset['B12'].value)
    tkinter.Label(custom_window,text="먼동강화=          강",font=guide_font).place(x=160,y=245)
    lvl_list=[10,11,12,13]
    cus4=tkinter.ttk.Combobox(custom_window,width=2,values=lvl_list); cus4.place(x=230,y=247)
    cus4.set(db_preset['B13'].value)

    
    tkinter.Label(custom_window,text="흐름상의=          %",font=guide_font).place(x=10,y=155) ##O100
    cus6=tkinter.Entry(custom_window,width=5);cus6.place(x=80,y=157)
    cus6.insert(END,db_preset['B5'].value)
    tkinter.Label(custom_window,text="흐름하의=          %",font=guide_font).place(x=10,y=185) ##O127
    cus7=tkinter.Entry(custom_window,width=5);cus7.place(x=80,y=187)
    cus7.insert(END,db_preset['B6'].value)
    tkinter.Label(custom_window,text="흐름어깨=          %",font=guide_font).place(x=10,y=215) ##O147
    cus8=tkinter.Entry(custom_window,width=5);cus8.place(x=80,y=217)
    cus8.insert(END,db_preset['B7'].value)
    tkinter.Label(custom_window,text="흐름벨트=          %",font=guide_font).place(x=10,y=245) ##O163
    cus9=tkinter.Entry(custom_window,width=5);cus9.place(x=80,y=247)
    cus9.insert(END,db_preset['B8'].value)
    tkinter.Label(custom_window,text="흐름신발=          %",font=guide_font).place(x=10,y=275) ##O179
    cus10=tkinter.Entry(custom_window,width=5);cus10.place(x=80,y=277)
    cus10.insert(END,db_preset['B9'].value)
    tkinter.Label(custom_window,text="흐름2셋=           %",font=guide_font).place(x=10,y=305) ##O295
    cus11=tkinter.Entry(custom_window,width=5);cus11.place(x=80,y=307)
    cus11.insert(END,db_preset['B10'].value)
    tkinter.Label(custom_window,text="흐름3셋=           %",font=guide_font).place(x=10,y=335) ##O296,O297
    cus12=tkinter.Entry(custom_window,width=5);cus12.place(x=80,y=337)
    cus12.insert(END,db_preset['B11'].value)

    tkinter.Label(custom_window,text="<버퍼설정>",font=mid_font,fg='blue').place(x=410,y=5)
    tkinter.Label(custom_window,text="노증폭/극마부/극찬작 기준 스탯에서\n자신의 스탯이 얼마나 가감되는지 기입\n(자기 캐릭 계수 보기 이용)",fg="Red").place(x=350,y=33)
    tkinter.Label(custom_window,text="1각스탯+          ",font=guide_font).place(x=320,y=80) ##
    c_stat=tkinter.Entry(custom_window,width=7);c_stat.place(x=390,y=82)
    c_stat.insert(END,db_preset['H1'].value)
    tkinter.Label(custom_window,text="축복스탯+          ",font=guide_font).place(x=470,y=80) ##
    b_stat=tkinter.Entry(custom_window,width=7);b_stat.place(x=540,y=82)
    b_stat.insert(END,db_preset['H6'].value)
    three=[0,1,2,3];two=[0,1,2]
    tkinter.Label(custom_window,text="축복칭호=",font=guide_font).place(x=320,y=110)
    b_style_lvl=tkinter.ttk.Combobox(custom_window,width=5,values=three); b_style_lvl.place(x=390,y=112) ##     
    b_style_lvl.set(db_preset['H2'].value)
    tkinter.Label(custom_window,text="1각칭호=",font=guide_font).place(x=470,y=110)
    c_style_lvl=tkinter.ttk.Combobox(custom_window,width=5,values=two); c_style_lvl.place(x=540,y=112) ##     
    c_style_lvl.set(db_preset['H3'].value)
    tkinter.Label(custom_window,text="축복플티=",font=guide_font).place(x=320,y=140)
    b_plt=tkinter.ttk.Combobox(custom_window,width=5,values=two); b_plt.place(x=390,y=142) ##
    b_plt.set(db_preset['H4'].value)
    tkinter.Label(custom_window,text="축복클쳐=",font=guide_font).place(x=470,y=140)
    b_cri=tkinter.ttk.Combobox(custom_window,width=5,values=[0,1]); b_cri.place(x=540,y=142) ##
    b_cri.set(db_preset['H5'].value)

    tkinter.Label(custom_window,text="<딜러속강>",font=mid_font).place(x=410,y=175)
    tkinter.Label(custom_window,text="마부총합=",font=guide_font).place(x=470,y=210)
    ele1=tkinter.Entry(custom_window,width=7); ele1.place(x=540,y=212) ##
    ele1.insert(END,db_preset['B14'].value)
    tkinter.Label(custom_window,text="오라속강=",font=guide_font).place(x=470,y=240)
    ele2=tkinter.Entry(custom_window,width=7); ele2.place(x=540,y=242) ##
    ele2.insert(END,db_preset['B15'].value)
    tkinter.Label(custom_window,text=" 젬 속강=",font=guide_font).place(x=470,y=270)
    ele3=tkinter.Entry(custom_window,width=7); ele3.place(x=540,y=272) ##
    ele3.insert(END,db_preset['B16'].value)
    tkinter.Label(custom_window,text="스킬속강= 자동",font=guide_font).place(x=320,y=210)
    ele4=tkinter.Entry(custom_window,width=7); ##ele4.place(x=390,y=212) ## 자속강 비활성화
    ele4.insert(END,db_preset['B17'].value)
    tkinter.Label(custom_window,text=" 몹 속저=",font=guide_font).place(x=320,y=240)
    ele5=tkinter.Entry(custom_window,width=7); ele5.place(x=390,y=242) ##
    ele5.insert(END,db_preset['B18'].value)
    tkinter.Label(custom_window,text="버퍼속깎=",font=guide_font).place(x=320,y=270)
    ele6=tkinter.Entry(custom_window,width=7); ele6.place(x=390,y=272) ##
    ele6.insert(END,db_preset['B19'].value)

    load_preset.close()
    save_command=lambda:save_custom(ele_type.get(),cool_con.get(),cus1.get(),cus2.get(),cus3.get(),cus4.get(),
                                    cus6.get(),cus7.get(),cus8.get(),cus9.get(),cus10.get(),cus11.get(),cus12.get(),
                                    c_stat.get(),b_stat.get(),b_style_lvl.get(),c_style_lvl.get(),b_plt.get(),b_cri.get(),
                                    ele1.get(),ele2.get(),ele3.get(),ele4.get(),ele5.get(),ele6.get())
    tkinter.Button(custom_window,text="저장하기",font=mid_font,command=save_command,bg="lightyellow").place(x=190,y=295)
    if auto==1:
        global auto_saved
        auto_saved=1
        save_custom(ele_type.get(),cool_con.get(),cus1.get(),cus2.get(),cus3.get(),cus4.get(),
                    cus6.get(),cus7.get(),cus8.get(),cus9.get(),cus10.get(),cus11.get(),cus12.get(),
                    c_stat.get(),b_stat.get(),b_style_lvl.get(),c_style_lvl.get(),b_plt.get(),b_cri.get(),
                    ele1.get(),ele2.get(),ele3.get(),ele4.get(),ele5.get(),ele6.get())
        print('자동저장')
        auto_saved=0
        auto=0
    
def save_custom(ele_type,cool_con,cus1,cus2,cus3,cus4,cus6,cus7,cus8,cus9,cus10,cus11,cus12,c_stat,b_stat,b_style_lvl,c_style_lvl,b_plt,b_cri,ele1,ele2,ele3,ele4,ele5,ele6):
    try:
        load_excel3=load_workbook("DATA.xlsx")
        load_preset1=load_workbook("preset.xlsx")
        db_custom1=load_preset1["custom"]
        db_save_one=load_excel3["one"]
        db_save_set=load_excel3["set"]
        
        db_custom1['B1']=ele_type
        if ele_type == '화':
            db_save_one['L181']=0;db_save_one['L165']=0;db_save_one['L149']=24;db_save_one['L129']=0
            db_save_one['L429']=0;db_save_one['L430']=0;db_save_one['L431']=20;db_save_one['L433']=0
        elif ele_type == '수':
            db_save_one['L181']=0;db_save_one['L165']=24;db_save_one['L149']=0;db_save_one['L129']=0
            db_save_one['L429']=20;db_save_one['L430']=0;db_save_one['L431']=0;db_save_one['L433']=0
        elif ele_type == '명':
            db_save_one['L181']=24;db_save_one['L165']=0;db_save_one['L149']=0;db_save_one['L129']=0
            db_save_one['L429']=0;db_save_one['L430']=20;db_save_one['L431']=0;db_save_one['L433']=0
        elif ele_type == '암':
            db_save_one['L181']=0;db_save_one['L165']=0;db_save_one['L149']=0;db_save_one['L129']=24
            db_save_one['L429']=0;db_save_one['L430']=0;db_save_one['L431']=0;db_save_one['L433']=20
        
        db_custom1['B3']=float(cus1);db_save_one['O164']=float(cus1)
        db_custom1['B4']=float(cus2);db_save_one['O180']=float(cus2)
        db_custom1['B5']=float(cus6);db_save_one['O100']=float(cus6);db_save_one['O101']=float(cus6)
        db_custom1['B6']=float(cus7);db_save_one['O127']=float(cus7)
        db_custom1['B7']=float(cus8);db_save_one['O147']=float(cus8)
        db_custom1['B8']=float(cus9);db_save_one['O163']=float(cus9)
        db_custom1['B9']=float(cus10);db_save_one['O179']=float(cus10)
        db_custom1['B10']=float(cus11);db_save_one['O295']=float(cus11)
        db_custom1['B11']=float(cus12);db_save_one['O296']=float(cus12);db_save_one['O297']=float(cus12)
        db_custom1['B12']=cus3
        db_custom1['B2']=cool_con
        if cus3=='전설↓':
            db_save_one['J86']=34;db_save_one['F120']=34;db_save_one['N140']=34;db_save_one['L156']=68;db_save_one['K172']=34;db_save_one['G276']=40;
        else:
            db_save_one['J86']=35;db_save_one['F120']=35;db_save_one['N140']=35;db_save_one['L156']=72;db_save_one['K172']=35;db_save_one['G276']=41;
        db_custom1['B13']=cus4
        db_save_one['N189']=int(cus4)+4;db_save_one['N190']=int(cus4)+4;db_save_one['K205']=int(cus4)+4;db_save_one['E214']=int(cus4)+4

        db_custom1['H1']=c_stat
        db_custom1['H6']=b_stat
        db_custom1['H2']=b_style_lvl
        db_custom1['H3']=c_style_lvl
        db_custom1['H4']=b_plt
        db_custom1['H5']=b_cri

        db_custom1['B14']=ele1
        db_custom1['B15']=ele2
        db_custom1['B16']=ele3
        db_custom1['B17']=ele4
        db_custom1['B18']=ele5
        db_custom1['B19']=ele6
        
        load_preset1.save("preset.xlsx")
        load_preset1.close()
        load_excel3.save("DATA.xlsx")
        load_excel3.close()
        custom_window.destroy()
        global auto_saved
        if auto_saved!=1:
            tkinter.messagebox.showinfo("알림","저장 완료")
    except PermissionError as error:
        tkinter.messagebox.showerror("에러","엑셀을 닫고 다시 시도해주세요.")


def load_checklist():
    ask_msg1=tkinter.messagebox.askquestion('확인',"저장된 내역을 불러오겠습니까?")
    for snum in range(0,10):
        if save_select.get() == save_name_list[snum]:
            ssnum1=snum
    if ask_msg1 == 'yes':
        load_preset3=load_workbook("preset.xlsx")
        db_load_check=load_preset3["one"]
        load_cell=db_load_check.cell
        k=1
        for i in range(1,296):
            if load_cell(i,2+ssnum1).value == 1:
                try:
                    select_item['tg{}'.format(load_cell(i,1).value)]=1
                except KeyError as error:
                    passss=1
            elif load_cell(i,2+ssnum1).value == 0:
                try:
                    select_item['tg{}'.format(load_cell(i,1).value)]=0
                except KeyError as error:
                    passss=1
        load_preset3.close()
        check_equipment()
        for i in range(101,136):
            check_set(i)

def save_checklist():
    ask_msg2=tkinter.messagebox.askquestion('확인',"저장하시겠습니까?")
    for snum in range(0,10):
        if save_select.get() == save_name_list[snum]:
            ssnum2=snum
    try:
        if ask_msg2 == 'yes':
            load_preset4=load_workbook("preset.xlsx")
            db_save_check=load_preset4["one"]
            save_cell=db_save_check.cell
            opt_save={}
            for i in range(1,296):
                opt_save[save_cell(i,1).value]=i

            for code in opt_save.keys():
                try:
                    if eval("select_item['tg{}']".format(code)) == 1:
                        save_cell(opt_save[code],2+ssnum2).value=1
                except KeyError as error:
                    passss1=1
                    
                try:
                    if eval("select_item['tg{}']".format(code)) == 0:
                        save_cell(opt_save[code],2+ssnum2).value=0
                except KeyError as error:
                    passss1=1
                
                passss=1
            load_preset4.save("preset.xlsx")
            load_preset4.close()
            tkinter.messagebox.showinfo("알림","저장 완료")
    except PermissionError as error:
        tkinter.messagebox.showerror("에러","엑셀을 닫고 다시 시도해주세요.")

def change_list_name():
    global change_window
    try:
        change_window.destroy()
    except:
        pass
    change_window=tkinter.Toplevel(self)
    change_window.geometry("190x320+750+200")
    tkinter.Label(change_window,text="1번슬롯").place(x=20,y=10)
    tkinter.Label(change_window,text="2번슬롯").place(x=20,y=35)
    tkinter.Label(change_window,text="3번슬롯").place(x=20,y=60)
    tkinter.Label(change_window,text="4번슬롯").place(x=20,y=85)
    tkinter.Label(change_window,text="5번슬롯").place(x=20,y=110)
    tkinter.Label(change_window,text="6번슬롯").place(x=20,y=135)
    tkinter.Label(change_window,text="7번슬롯").place(x=20,y=160)
    tkinter.Label(change_window,text="8번슬롯").place(x=20,y=185)
    tkinter.Label(change_window,text="9번슬롯").place(x=20,y=210)
    tkinter.Label(change_window,text="10번슬롯").place(x=20,y=235)
    entry1=tkinter.Entry(change_window,width=10);entry1.place(x=95,y=12);entry1.insert(END,save_name_list[0])
    entry2=tkinter.Entry(change_window,width=10);entry2.place(x=95,y=37);entry2.insert(END,save_name_list[1])
    entry3=tkinter.Entry(change_window,width=10);entry3.place(x=95,y=62);entry3.insert(END,save_name_list[2])
    entry4=tkinter.Entry(change_window,width=10);entry4.place(x=95,y=87);entry4.insert(END,save_name_list[3])
    entry5=tkinter.Entry(change_window,width=10);entry5.place(x=95,y=112);entry5.insert(END,save_name_list[4])
    entry6=tkinter.Entry(change_window,width=10);entry6.place(x=95,y=137);entry6.insert(END,save_name_list[5])
    entry7=tkinter.Entry(change_window,width=10);entry7.place(x=95,y=162);entry7.insert(END,save_name_list[6])
    entry8=tkinter.Entry(change_window,width=10);entry8.place(x=95,y=187);entry8.insert(END,save_name_list[7])
    entry9=tkinter.Entry(change_window,width=10);entry9.place(x=95,y=212);entry9.insert(END,save_name_list[8])
    entry10=tkinter.Entry(change_window,width=10);entry10.place(x=95,y=237);entry10.insert(END,save_name_list[9])

    tkinter.Button(change_window,text="저장",font=mid_font,command=lambda:change_savelist(entry1.get(),entry2.get(),entry3.get(),entry4.get(),entry5.get(),entry6.get(),entry7.get(),entry8.get(),entry9.get(),entry10.get())).place(x=60,y=270)

def change_savelist(in1,in2,in3,in4,in5,in6,in7,in8,in9,in10):
    in_list=[in1,in2,in3,in4,in5,in6,in7,in8,in9,in10]
    try:
        load_preset5=load_workbook("preset.xlsx", data_only=True)
        db_custom2=load_preset5["custom"]
        
        for i in range(1,11):
            db_custom2.cell(i,5).value=in_list[i-1]
        global save_name_list
        save_name_list=in_list
        load_preset5.save("preset.xlsx")
        load_preset5.close()
        save_select.set(save_name_list[0])
        save_select['values']=save_name_list
        change_window.destroy()
        tkinter.messagebox.showinfo("알림","저장 완료")
    except PermissionError as error:
        tkinter.messagebox.showerror("에러","엑셀을 닫고 다시 시도해주세요.")

def update_count():
    global count_num, count_all, show_number
    global showcon,all_list_list_num
    while True:
        showcon(text=str(count_num)+"유효/"+str(count_all)+"무효\n"+str(all_list_list_num)+"전체")
        time.sleep(0.1)

def update_count2():
    while True:
        global select_item
        a_num_all=0
        a_num=[0,0,0,0,0,0,0,0,0,0,0]
        for i in range(101,136):
            try:
                a_num[0]=a_num[0]+select_item['tg1{}0'.format(i)]+select_item['tg1{}1'.format(i)]
            except KeyError as error:
                p=0
            try:
                a_num[1]=a_num[1]+select_item['tg1{}0'.format(i+100)]
            except KeyError as error:
                p=0
            try:
                a_num[2]=a_num[2]+select_item['tg1{}0'.format(i+200)]
            except KeyError as error:
                p=0
            try:
                a_num[3]=a_num[3]+select_item['tg1{}0'.format(i+300)]
            except KeyError as error:
                p=0
            try:
                a_num[4]=a_num[4]+select_item['tg1{}0'.format(i+400)]
            except KeyError as error:
                p=0
            try:
                a_num[5]=a_num[5]+select_item['tg2{}0'.format(i)]+select_item['tg2{}1'.format(i)]
            except KeyError as error:
                p=0
            try:
                a_num[6]=a_num[6]+select_item['tg2{}0'.format(i+100)]
            except KeyError as error:
                p=0
            try:
                a_num[7]=a_num[7]+select_item['tg2{}0'.format(i+200)]
            except KeyError as error:
                p=0
            try:
                a_num[8]=a_num[8]+select_item['tg3{}0'.format(i)]
            except KeyError as error:
                p=0
            try:
                a_num[9]=a_num[9]+select_item['tg3{}0'.format(i+100)]
            except KeyError as error:
                p=0
            try:
                a_num[10]=a_num[10]+select_item['tg3{}0'.format(i+200)]+select_item['tg3{}1'.format(i+200)]
            except KeyError as error:
                p=0
                
        if a_num[0]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
        if a_num[1]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
        if a_num[2]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
        if a_num[3]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
        if a_num[4]==0:
            a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;

        if a_num[5]+a_num[6]+a_num[7]<2:
            a_num[5]=a_num[5]+1;a_num[6]=a_num[6]+1;a_num[7]=a_num[7]+1
        if a_num[8]+a_num[9]+a_num[10]<2:
            a_num[8]=a_num[8]+1;a_num[9]=a_num[9]+1;a_num[10]=a_num[10]+1
            
        if a_num[5]==0:
            a_num[5]=a_num[5]+1
        if a_num[6]==0:
            a_num[6]=a_num[6]+1
        if a_num[7]==0:
            a_num[7]=a_num[7]+1
            
        if a_num[8]==0:
            a_num[8]=a_num[8]+1
        if a_num[9]==0:
            a_num[9]=a_num[9]+1
        if a_num[10]==0:
            a_num[10]=a_num[10]+1
            
        a_num_all=a_num[0]*a_num[1]*a_num[2]*a_num[3]*a_num[4]*a_num[5]*a_num[6]*a_num[7]*a_num[8]*a_num[9]*a_num[10]
        showcon2(text="경우의 수(1초 갱신)= "+str(a_num_all))
        if a_num_all>10000000:
            show_count2['fg']="red"
        else:
            show_count2['fg']="white"
        time.sleep(1)

def update_thread():
    threading.Thread(target=update_count,daemon=True).start()
    threading.Thread(target=update_count2,daemon=True).start()

def timeline_select():
    global timeline_window
    try:
        timeline_window.destroy()
    except:
        pass
    timeline_window=tkinter.Toplevel(self)
    timeline_window.attributes("-topmost", True) 
    timeline_window.geometry("310x150+750+20")
    tkinter.Label(timeline_window,text="캐릭터명=\n(정확히)",font=guide_font).place(x=10,y=9)
    cha_name=tkinter.Entry(timeline_window,width=13)
    cha_name.place(x=80,y=12)
    tkinter.Label(timeline_window,text="서버명=",font=guide_font).place(x=10,y=59)
    sever_list=['카인','디레지에','바칼','힐더','안톤','카시야스','프레이','시로코']
    serv_name=tkinter.ttk.Combobox(timeline_window,values=sever_list,width=11)
    serv_name.place(x=80,y=62)
    serv_name.set('카인')
    load_timeline=tkinter.Button(timeline_window,command=lambda:show_timeline(cha_name.get(),serv_name.get()),text="불러오기",font=mid_font)
    load_timeline.place(x=200,y=25)
    tkinter.Label(timeline_window,text="타임라인에 있는 에픽만 불러옵니다(일부X)",fg="Red").place(x=10,y=100)
    tkinter.Label(timeline_window,text="서버 불안정때매 안되면 여러번 눌러보세요",fg="Red").place(x=10,y=120)

def show_timeline(name,server):
    
    server_dict={'안톤':'anton','바칼':'bakal','카인':'cain','카시야스':'casillas',
                '디레지에':'diregie','힐더':'hilder','프레이':'prey','시로코':'siroco'}
    try:
        sever_code=server_dict[server]
        cha_id_api=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters?characterName='+parse.quote(name)+'&apikey=' + apikey)
        cha_id_dic=loads(cha_id_api.read().decode("utf-8"))
        cha_id=cha_id_dic['rows'][0]['characterId']

    ##
        print(sever_code)
        print(cha_id)
        time.sleep(0.3)
        start_time='20200101T0000'
        time_now=time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
        now=time_now
        now_1='20200101T0000'
        now_2='20200101T0000'
        now_3='20200101T0000'
        now2='20200101T0000'
        now3='20200101T0000'
        now4='20200101T0000'
        if int(time_now[0:8]) >= 20200316:
            now='20200315T2359'
            now_1='20200316T0000'
            now2=time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
        if int(time_now[0:8]) >= 20200601:
            now2='20200531T2359'
            now_2='20200601T0000'
            now3=time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
        if int(time_now[0:8]) >= 20200816:
            now3='20200815T2359'
            now_3='20200816T0000'
            now4=time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
        time_code='504,505,506,507,508,510,511,512,513,514'
        timeline_list=[]
        for nows in [[now,start_time],[now2,now_1],[now3,now_2],[now4,now_3]]:
            if nows[0] != '20200101T0000':
                timeline=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters/'+cha_id+'/timeline?limit=100&code='+time_code+'&startDate='+nows[1]+'&endDate='+nows[0]+'&apikey='+apikey)
                timeline2=loads(timeline.read().decode("utf-8"))['timeline']
                show_next=timeline2['next']
                timeline_list=timeline_list+timeline2['rows']
                time.sleep(0.3)
                while show_next != None:
                    timeline_next=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters/'+cha_id+'/timeline?next='+show_next+'&apikey='+apikey)
                    timeline_next2=loads(timeline_next.read().decode("utf-8"))['timeline']
                    timeline_list=timeline_list+timeline_next2['rows']
                    time.sleep(0.3)
                    show_next=timeline_next2['next']
            
        all_item=[]
        for now in timeline_list:
            item=now['data']['itemId']
            all_item.append(item)
        xl=openpyxl.load_workbook("DATA.xlsx", data_only=True)
        sh=xl['one']
        
        reset()
            
        for i in range(76,257):
            api_cod=sh.cell(i,40).value
            if all_item.count(api_cod) != 0:
                select_item['tg{}'.format(str(sh.cell(i,1).value))]=1
        xl.close()
        check_equipment()
        for i in range(101,136):
            check_set(i)
        timeline_window.destroy()
        tkinter.messagebox.showinfo("주의","과거 메타몽했던 에픽도 전부 불러와집니다.\n알아서 빼주세요.\n\n초월한 에픽은 뜨지않습니다.\n알아서 넣으세요.")
    except urllib.error.HTTPError as error:
        tkinter.messagebox.showerror("에러","API 접근 실패(네트워크 오류)")

def reset():
    know_list2=['13390150','22390240','23390450','33390750','21390340','31390540','32390650']
    know_set_list=['22400150','22400250','22400350','22400450','22400550','21400640','31400750',
                   '31400850','31400950','31401050','31401150','32401240','32401340','32401440']
    know_jin_list=['11410100','11410110','11410120','11410130','11410140','11410150',
                   '21420100','21420110','21420120','21420130','21420140','21420150',
                   '33430100','33430110','33430120','33430130','33430140','33430150']
    for i in range(1101,3336):
        try:
            select_item['tg{}0'.format(i)]=0
        except KeyError as error:
            passss=1
        try:
            select_item['tg{}1'.format(i)]=0
        except KeyError as error:
            passss=1
    for i in know_list2+know_set_list+know_jin_list:
        select_item['tg{}'.format(i)]=0
    check_equipment()
    for i in range(101,136):
        check_set(i)

def guide_speed():
    tkinter.messagebox.showinfo("정확도 선택","빠름=단일 선택 부위를 전부 제거\n중간=단일은 포함하되, 신화에 우선권 부여\n느림=세트 수 우선권 완화, 신화 우선권 삭제\n\n메타몽모드=메타몽 경우의 수 추가\n중간=메타몽의 경우의 수만 계산\n느림=메타몽+기존 경우의 수 전부 계산")
                                

select_item={}
def click_equipment(code):
    if eval("select_item['tg{}']".format(code))==0:
        eval('select_{}'.format(code))['image']=image_list[str(code)]
        select_item['tg'+str('{}'.format(code))]=1
    elif eval("select_item['tg{}']".format(code))==1:
        eval('select_{}'.format(code))['image']=image_list2[str(code)]
        select_item['tg'+str('{}'.format(code))]=0
    if len(str(code))==5:
        check_set(int('1'+str(code)[2:4]))

def check_equipment():
    global select_item,select_13390150,select_22390240,select_23390450,select_33390750,select_21390340,select_31390540,select_32390650
    global select_22400150,select_22400250,select_22400350,select_22400450,select_22400550,select_21400640,select_31400750
    global select_31400850,select_31400950,select_31401050,select_31401150,select_32401240,select_32401340,select_32401440
    global select_11410100,select_11410110,select_11410120,select_11410130,select_11410140,select_11410150
    global select_21420100,select_21420110,select_21420120,select_21420130,select_21420140,select_21420150
    global select_33430100,select_33430110,select_33430120,select_33430130,select_33430140,select_33430150

    know_list2=['13390150','22390240','23390450','33390750','21390340','31390540','32390650']
    know_set_list=['22400150','22400250','22400350','22400450','22400550','21400640','31400750',
                   '31400850','31400950','31401050','31401150','32401240','32401340','32401440']
    know_jin_list=['11410100','11410110','11410120','11410130','11410140','11410150',
                   '21420100','21420110','21420120','21420130','21420140','21420150',
                   '33430100','33430110','33430120','33430130','33430140','33430150']
    for i in range(11010,33352):
        try:
            if eval("select_item['tg{}']".format(i))==0:
                eval('select_{}'.format(i))['image']=image_list2[str(i)]
            elif eval("select_item['tg{}']".format(i))==1:
                eval('select_{}'.format(i))['image']=image_list[str(i)]
        except:
            pass
    for i in know_list2+know_set_list+know_jin_list:
        try:
            if eval("select_item['tg{}']".format(i))==0:
                eval('select_{}'.format(i))['image']=image_list2[str(i)]
            elif eval("select_item['tg{}']".format(i))==1:
                eval('select_{}'.format(i))['image']=image_list[str(i)]
        except:
            pass

def click_set(code):
    code_add=code-100
    code_str=str(code)[1:3]
    set_checked=0
    if code >=116: ##악세/특장/스까면
        if 116<= code <=119:
            for i in range(21,24): ## 악세부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        elif 123>= code >=120:
            for i in range(31,34): ## 특장부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        elif 131>= code >=128:
            for i in [11,22,31]: ## 상목보부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        elif 127>= code >=124:
            for i in [12,21,32]: ## 하팔법부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        elif 135>= code >=132:
            for i in [15,23,33]: ## 신반귀부위에서
                try:
                    if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                        set_checked=set_checked+1 ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c=1
        if set_checked==3: ## 채택 숫자가 3이면
            for i in range(11,36): ##모든 부위에서
                try:
                    eval('select_'+str(i)+code_str+'0')['image']=image_list2[str(i)+code_str+'0'] ##이미지도 오프로 바꿈
                    select_item['tg'+str(i)+code_str+'0']=0 ##모든 체크를 0으로 만듬
                except KeyError as error:
                    c=1
            eval('set'+str(code))['image']=image_list_set2[str(code)] ##세트이미지도 오프로 바꿈
        else: ## 채택 숫자가 3미만이면
            for i in range(11,36): ##모든 부위에서
                try:
                    eval('select_'+str(i)+code_str+'0')['image']=image_list[str(i)+code_str+'0'] ##이미지도 온으로 바꿈
                    select_item['tg'+str(i)+code_str+'0']=1 ##모든 체크를 1으로 만듬
                except KeyError as error:
                    c=1
            eval('set'+str(code))['image']=image_list_set[str(code)] ##세트이미지도 온으로 바꿈

            
    else:
        for i in range(11,16): ## 방어구 부위에서
            try:
                if select_item['tg'+str(i)+code_str+'0']==1: ##채택된 숫자를 찾는다
                    set_checked=set_checked+1 ##그럼 변수에 +1을 더함
            except KeyError as error:
                c=1
        
        if set_checked==5: ## 채택 숫자가 5이면
            for i in range(11,16): ## 방어구 부위에서
                try:
                    eval('select_'+str(i)+code_str+'0')['image']=image_list2[str(i)+code_str+'0'] ##이미지도 오프로 바꿈
                    select_item['tg'+str(i)+code_str+'0']=0 ##모든 체크를 0으로 만듬
                except KeyError as error:
                    c=1
            eval('set'+str(code))['image']=image_list_set2[str(code)] ##세트이미지도 오프로 바꿈
            
        else: ## 채택 숫자가 5미만이면
            for i in range(11,16): ## 방어구 부위에서
                try:
                    eval('select_'+str(i)+code_str+'0')['image']=image_list[str(i)+code_str+'0'] ##이미지도 온으로 바꿈
                    select_item['tg'+str(i)+code_str+'0']=1 ##모든 체크를 1으로 만듬
                except KeyError as error:
                    c=1
            eval('set'+str(code))['image']=image_list_set[str(code)] ##세트이미지도 온으로 바꿈
                    
def check_set(code):
    code_str=str(code)[1:3]
    set_checked=0
    if code < 116:
        for i in [11,12,13,14,15]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 120:
        for i in [21,22,23]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 124:
        for i in [31,32,33]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 128:
        for i in [12,21,32]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 132:
        for i in [11,22,31]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
    elif code < 136:
        for i in [15,23,33]:
            if select_item['tg'+str(i)+code_str+'0']==1:
                set_checked=set_checked+1
                
    if code < 116:
        if set_checked==5:
            eval('set'+str(code))['image']=image_list_set[str(code)]
        else:
            eval('set'+str(code))['image']=image_list_set2[str(code)]
    else:
        if set_checked==3:
            eval('set'+str(code))['image']=image_list_set[str(code)]
        else:
            eval('set'+str(code))['image']=image_list_set2[str(code)]

def cha_select(jobs):
    global cha_window
    try:
        cha_window.destroy()
    except:
        pass
    cha_window=tkinter.Toplevel(self)
    cha_window.attributes("-topmost", True) 
    cha_window.geometry("350x200+750+20")
    tkinter.Label(cha_window,text="캐릭터명=\n(정확히)",font=guide_font).place(x=10,y=9)
    cha_name=tkinter.Entry(cha_window,width=15)
    cha_name.place(x=80,y=12)
    tkinter.Label(cha_window,text="서버명=",font=guide_font).place(x=10,y=49)
    sever_list=['카인','디레지에','바칼','힐더','안톤','카시야스','프레이','시로코']
    serv_name=tkinter.ttk.Combobox(cha_window,values=sever_list,width=13)
    serv_name.place(x=80,y=52)
    serv_name.set('카인')
    tkinter.Label(cha_window,text="직업=",font=guide_font).place(x=10,y=79)
    job_name=tkinter.ttk.Combobox(cha_window,values=jobs,width=13)
    job_name.place(x=80,y=82)
    job_name.set('가이아(2각)')

    
    load_timeline=tkinter.Button(cha_window,command=lambda:calc_my_cha(cha_name.get(),serv_name.get(),job_name.get()),text="불러오기",font=mid_font)
    load_timeline.place(x=240,y=25)
    tkinter.Label(cha_window,text="한캐릭씩 볼것!",font=mid_font,fg="Red").place(x=205,y=76)
    tkinter.Label(cha_window,text="오직 12부위 장비만 불러옵니다.(마부/강화,칭호/클쳐,압타 X)",fg="Red").place(x=7,y=110)
    tkinter.Label(cha_window,text="메인창과 호환을 위해 커스텀은 전부 메인창을 따라갑니다.",fg="Red").place(x=7,y=130)
    tkinter.Label(cha_window,text="100제 에픽이 아니면 전부 100제 레전으로 처리합니다.",fg="Red").place(x=7,y=150)
    tkinter.Label(cha_window,text="서버 불안정때매 안되면 여러번 눌러보세요",fg="Red").place(x=7,y=170)

def calc_my_cha(cha_name,serv_name,job_name):
    info_stat=0
    global show_cha_result
    try:
        show_cha_result.destroy()
    except NameError as error:
        pass
    server_dict={'안톤':'anton','바칼':'bakal','카인':'cain','카시야스':'casillas',
                '디레지에':'diregie','힐더':'hilder','프레이':'prey','시로코':'siroco'}
    try:
        sever_code=server_dict[serv_name]
        cha_id_api=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters?characterName='+parse.quote(cha_name)+'&apikey=' + apikey)
        cha_id_dic=loads(cha_id_api.read().decode("utf-8"))
        cha_id=cha_id_dic['rows'][0]['characterId']
        print(sever_code)
        print(cha_id)
        cha_image(cha_id,sever_code)
        time.sleep(0.3)

        cha_api=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters/'+cha_id+'/equip/equipment?apikey='+apikey)
        cha_api_dic=loads(cha_api.read().decode("utf-8"))
        cha_equ=cha_api_dic["equipment"]
        item_list=[]
        for i in range(0,20):
            try:
                if cha_equ[i]['slotName'] != '보조무기':
                    item_list.append(cha_equ[i]['itemName'])
                if cha_equ[i]['slotName'] == '무기':
                    wep_name=cha_equ[i]['itemName']
            except IndexError as error:
                pass

        ##싀칭박스
        if job_name[4:7] == '세인트' or job_name[4:7] == '세라핌' or job_name[4:7] == '헤카테':
            switch_api=urllib.request.urlopen('https://api.neople.co.kr/df/servers/'+sever_code+'/characters/'+cha_id+'/skill/buff/equip/equipment?apikey='+apikey)
            switch_api_dic=loads(switch_api.read().decode("utf-8"))
            swi_equ=switch_api_dic["skill"]['buff']['equipment']
            swi_list=[]
            for i in range(0,20):
                try:
                    if swi_equ[i]['slotName'] != '보조무기':
                        swi_list.append(swi_equ[i]['itemName'])
                    if swi_equ[i]['slotName'] == '무기':
                        swi_wep_name=swi_equ[i]['itemName']
                except IndexError as error:
                    pass
    except urllib.error.HTTPError as error:
        tkinter.messagebox.showerror("에러","API 접근 실패(네트워크 오류)")

    xl=openpyxl.load_workbook("DATA.xlsx", data_only=True)
    sh=xl['one']
    for_calc_list=[]
    for_calc_list2=[]
    for i in item_list:
        for j in range(1,257):
            if i==sh.cell(j,39).value:
                for_calc_list.append('{}'.format(str(sh.cell(j,1).value)))

    check_exist=[]
    wep_exist=0
    for i in for_calc_list:
        if len(i) != 6:
            check_exist.append(i[0:2])
        if len(i) == 6:
            wep_exist=1
    if wep_exist ==0:
        for_calc_list.append('111001')
        wep_name='흑천의 주인(검은 성전의 기억)'

    global default_legend,default_chawon,default_old
    for i in [11,12,13,14,15]:
        if check_exist.count(str(i))==0:
            if default_legend==1:
                for_calc_list.append(str(i)+'360')
            elif default_chawon==1:
                for_calc_list.append(str(i)+'440')
            elif default_old==1:
                for_calc_list.append(str(i)+'470')
    for i in [21,22,23]:
        if check_exist.count(str(i))==0:
            if default_legend==1:
                for_calc_list.append(str(i)+'370')
            elif default_chawon==1:
                for_calc_list.append(str(i)+'450')
            elif default_old==1:
                for_calc_list.append(str(i)+'480')
    for i in [31,32,33]:
        if check_exist.count(str(i))==0:
            if default_legend==1:
                for_calc_list.append(str(i)+'380')
            elif default_chawon==1:
                for_calc_list.append(str(i)+'460')
            elif default_old==1:
                for_calc_list.append(str(i)+'490')
            
    calc_now=[]
    for i in for_calc_list:
        if len(i)==6:
            wep_num=[i]
        else:
            calc_now.append(i)
    ##스위칭박스
    if job_name[4:7] == '세인트' or job_name[4:7] == '세라핌' or job_name[4:7] == '헤카테':
        check_exist2=[]
        wep_exist2=0
        for i in swi_list:
            for j in range(1,257):
                if i==sh.cell(j,39).value:
                    for_calc_list2.append('{}'.format(str(sh.cell(j,1).value)))
        for i in for_calc_list2:
            if len(i) != 6:
                check_exist2.append(i[0:2])
            if len(i) == 6:
                wep_exist2=1
                
        if wep_exist2 ==0:
            for_calc_list2.append('111001')
            wep_name2='흑천의 주인(검은 성전의 기억)'
            
        for i in [11,12,13,14,15]:
            if check_exist2.count(str(i))==0:
                if default_legend==1:
                    for_calc_list2.append(str(i)+'360')
                elif default_chawon==1:
                    for_calc_list2.append(str(i)+'440')
                elif default_old==1:
                    for_calc_list2.append(str(i)+'470')
        for i in [21,22,23]:
            if check_exist2.count(str(i))==0:
                if default_legend==1:
                    for_calc_list2.append(str(i)+'370')
                elif default_chawon==1:
                    for_calc_list2.append(str(i)+'450')
                elif default_old==1:
                    for_calc_list2.append(str(i)+'480')
        for i in [31,32,33]:
            if check_exist2.count(str(i))==0:
                if default_legend==1:
                    for_calc_list2.append(str(i)+'380')
                elif default_chawon==1:
                    for_calc_list2.append(str(i)+'460')
                elif default_old==1:
                    for_calc_list2.append(str(i)+'490')
                
        calc_now2=[]
        for i in for_calc_list2:
            if len(i)==6:
                wep_num2=[i]
            else:
                calc_now2.append(i)

    opt_one={}
    a=1
    for row in sh.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_one[sh.cell(a,1).value]=row_value_cut
        a=a+1
    c=1        
    db_buf=xl["buf"]
    opt_buf={}
    name_buf={}
    for row in db_buf.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = row_value[2:]
        opt_buf[db_buf.cell(c,1).value]=row_value_cut ## DB 불러오기 ##
        name_buf[db_buf.cell(c,1).value]=row_value
        c=c+1

    d=1        
    db_buflvl=xl["buflvl"]
    opt_buflvl={}
    for row in db_buflvl.rows:
        row_value=[]
        row_value_cut=[]
        for cell in row:
            row_value.append(cell.value)
            row_value_cut = [0] + row_value[1:]
        opt_buflvl[db_buflvl.cell(d,1).value]=row_value_cut
        d=d+1

    load_presetc=load_workbook("preset.xlsx", data_only=True)
    db_preset=load_presetc["custom"]
    ele_skill=int(opt_job_ele[job_name][1])
    ele_in=(int(db_preset["B14"].value)+int(db_preset["B15"].value)+int(db_preset["B16"].value)+
            int(ele_skill)-int(db_preset["B18"].value)+int(db_preset["B19"].value)+13)
    cool_eff=float(db_preset["B2"].value)/100

    betterang=int(sh["J86"].value)
    
    if job_name[-4:] == "(진각)":
        active_eff_one=15
        active_eff_set=18-3
    else:
        active_eff_one=21
        active_eff_set=27-3
    global time_select
    if time_select.get() == "60초(각성비중↓)":
        lvl_shift=6
    else:
        lvl_shift=0
    job_lv1=opt_job[job_name][11+lvl_shift]
    job_lv2=opt_job[job_name][12+lvl_shift]
    job_lv3=opt_job[job_name][13+lvl_shift]
    job_lv4=opt_job[job_name][14+lvl_shift]
    job_lv5=opt_job[job_name][15+lvl_shift]
    job_lv6=opt_job[job_name][16+lvl_shift]
    job_pas0=opt_job[job_name][0]
    job_pas1=opt_job[job_name][1]
    job_pas2=opt_job[job_name][2]
    job_pas3=opt_job[job_name][3]
    
    
    extra_dam=0
    extra_cri=0
    extra_bon=0
    extra_all=0
    extra_pas2=0
    global style_select, creature_select
    if style_select.get() == '증뎀칭호':
        extra_dam=extra_dam+10
    if style_select.get() == '추뎀칭호':
        extra_bon=extra_bon+10
    if creature_select.get() == '모공크리쳐':
        extra_all=extra_all+15
    if creature_select.get() == '크증크리쳐':
        extra_cri=extra_cri+18
        extra_pas2=extra_pas2+1
    
    set_list=["1"+str(calc_now[x][2:4]) for x in range(0,11)] 
    set_on=[];setapp=set_on.append
    setcount=set_list.count
    for i in range(101,136):
        if setcount(str(i))==2:
            setapp(str(i)+"1")
        if 4>=setcount(str(i))>=3:
            setapp(str(i)+"2")
        if setcount(str(i))==5:
            setapp(str(i)+"3")
    for i in range(136,150):
        if setcount(str(i))==2:
            setapp(str(i)+"0")
        if 4>=setcount(str(i))>=3:
            setapp(str(i)+"1")
        if setcount(str(i))==5:
            setapp(str(i)+"2")
    calc_wep=wep_num+calc_now
    for_calc=set_on+calc_wep

    if job_name[4:7] != '세인트' and job_name[4:7] != '세라핌' and job_name[4:7] != '헤카테':
                       
        skiper=0
        coolper=0
        damage=0
        base_array=np.array([0,0,extra_dam,extra_cri,extra_bon,0,extra_all,0,0,ele_in,0,1,0,0,0,0,0,0,extra_pas2,0,0,0,0,0,0,0,0,0])
        oneone=len(for_calc)
        oneonelist=[]
        for i in range(oneone):
            no_cut=opt_one.get(for_calc[i])               ## 11번 스증
            cut=np.array(no_cut[0:20]+no_cut[22:23]+no_cut[34:35]+no_cut[38:44])
            skiper=(skiper/100+1)*(cut[11]/100+1)*100-100
            coolper=(1-(100-coolper)/100*(100-cut[20])/100)*100
            oneonelist.append(cut)
        for i in range(oneone):
            base_array=base_array+oneonelist[i]
        # 코드 이름
        # 0추스탯 1추공 2증 3크 4추 5속추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 쿨감보정/21 2각캐특수액티브 /22~27 액티브레벨링

        if set_on.count('1201')==1 and calc_wep.count('32200')==1:
            base_array[3]=base_array[3]-5
        if calc_wep.count('33200')==1 and calc_wep.count('31200')==0:
            base_array[8]=base_array[8]-10
        if calc_wep.count('33230')==1 or calc_wep.count('33231')==1:
            if calc_wep.count('31230')==0:
                base_array[4]=base_array[4]-10
            if calc_wep.count('32230')==0:
                base_array[9]=base_array[9]-40
        if calc_wep.count('15340')==1 or calc_wep.count('23340')==1 or calc_wep.count('33340')==1 or calc_wep.count('33341')==1:
            if set_on.count('1341')==0 and set_on.count('1342') ==0:
                if calc_wep.count('15340')==1:
                    base_array[9]=base_array[9]-20
                elif calc_wep.count('23340')==1:
                    base_array[2]=base_array[2]-10
                elif calc_wep.count('33340')==1:
                    base_array[6]=base_array[6]-5
                else:
                    base_array[9]=base_array[9]-4
                    base_array[2]=base_array[2]-2
                    base_array[6]=base_array[6]-1
                    base_array[8]=base_array[8]-1.93
        if calc_wep.count('11111')==1:
            if set_on.count('1112')==1 or set_on.count('1113')==1:
                coolper=(1-(100-coolper)/100*(100-11)/100)*100
        if calc_wep.count('11301')==1:
            if calc_wep.count('22300')!=1:
                base_array[4]=base_array[4]-10
                base_array[7]=base_array[7]+10
            if calc_wep.count('31300')!=1:
                base_array[4]=base_array[4]-10
                base_array[7]=base_array[7]+10
        if calc_wep.count('11061')==1:
            if betterang ==34:
                if calc_wep.count('12060')==1:
                    base_array[3]=base_array[3]+1
                if calc_wep.count('13060')==1:
                    skiper=skiper/1.34*1.35
                if calc_wep.count('14060')==1:
                    base_array[9]=base_array[9]+4
                if calc_wep.count('15060')==1:
                    base_array[8]=base_array[8]+1
                if set_on.count('1063')==1:
                    base_array[4]=base_array[4]+1
        if set_on.count('1441') ==1:
            if calc_wep.count('11440')!=1: ##3셋 공3% 모공5% 감소
                base_array[7]=base_array[7]-3
                base_array[6]=base_array[6]-5
                        
        real_bon=base_array[4]+base_array[5]*(base_array[9]*0.0045+1.05)
        actlvl=((base_array[active_eff_one]+base_array[22]*job_lv1+base_array[23]*job_lv2+base_array[24]*job_lv3+
                base_array[25]*job_lv4+base_array[26]*job_lv5+base_array[27]*job_lv6)/100+1)
        paslvl=((100+base_array[16]*job_pas0)/100)*((100+base_array[17]*job_pas1)/100)*((100+base_array[18]*job_pas2)/100)*((100+base_array[19]*job_pas3)/100)
        damage=((base_array[2]/100+1)*(base_array[3]/100+1)*(real_bon/100+1)*(base_array[6]/100+1)*(base_array[7]/100+1)*
                (base_array[8]/100+1)*(base_array[9]*0.0045+1.05)*(base_array[10]/100+1)*(skiper/100+1)*(base_array[12]/100+1)*
                actlvl*paslvl*((54500+3.31*base_array[0])/54500)*((4800+base_array[1])/4800)/(1.05+0.0045*int(ele_skill)))

        damage_with_cool=damage*((100/(100-coolper)-1)*cool_eff+1)
        type_code='deal'

        final_list=[damage,damage_with_cool]
        show_my_cha(calc_now,final_list,type_code,job_name,wep_name,ele_skill,cha_name,info_stat)
    else:

        set_list2=["1"+str(calc_now2[x][2:4]) for x in range(0,11)] 
        set_on2=[];setapp2=set_on2.append
        setcount2=set_list2.count
        for i in range(101,135):
            if setcount2(str(i))==2:
                setapp2(str(i)+"1")
            if 4>=setcount2(str(i))>=3:
                setapp2(str(i)+"2")
            if setcount2(str(i))==5:
                setapp2(str(i)+"3")
        for i in range(136,150):
            if setcount2(str(i))==2:
                setapp2(str(i)+"0")
            if 4>=setcount2(str(i))>=3:
                setapp2(str(i)+"1")
            if setcount2(str(i))==5:
                setapp2(str(i)+"2")
        calc_wep2=wep_num2+calc_now2
        for_calc2=set_on2+calc_wep2
        
        base_b=10+int(db_preset['H2'].value)+int(db_preset['H4'].value)+int(db_preset['H5'].value)+1
        base_c=12+int(db_preset['H3'].value)+1
        base_pas0=0
        base_pas0_c=3
        base_pas0_b=0
        base_stat_s=4114
        base_stat_h=4180
        base_pas0_1=0
        lvlget=opt_buflvl.get
        #코드 이름
        #0 체정 1 지능
        #축복 2 스탯% 3 물공% 4 마공% 5 독공%
        #아포 6 고정 7 스탯%
        #8 축렙 9 포렙
        #10 아리아/보징증폭
        #11 전직패 12 보징 13 각패1 14 각패2 15 2각 16 각패3
        #17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨
        setget=opt_buf.get
        base_array=np.array([base_stat_h,base_stat_s,0,0,0,0,0,0,base_b,base_c,0,base_pas0,base_pas0_1,0,0,0,0,0,0,0,0])
        base_array2=np.array([base_stat_h,base_stat_s,0,0,0,0,0,0,base_b,base_c,0,base_pas0,base_pas0_1,0,0,0,0,0,0,0,0])
        calc_wep=wep_num+calc_now
        calc_wep2=wep_num2+calc_now2

        b_stat=10.24 ##탈리스만
        b_phy=0
        b_mag=0
        b_ind=0
        c_per=0
        b_stat2=10.24 ##탈리스만
        b_phy2=0
        b_mag2=0
        b_ind2=0
        c_per2=0
        oneone=len(for_calc)
        oneonelist=[]
        for i in range(oneone):
            no_cut=np.array(setget(for_calc[i]))             ## 2 3 4 5 7
            base_array=base_array+no_cut
            b_stat=(b_stat/100+1)*(no_cut[2]/100+1)*100-100
            b_phy=(b_phy/100+1)*(no_cut[3]/100+1)*100-100
            b_mag=(b_mag/100+1)*(no_cut[4]/100+1)*100-100
            b_ind=(b_ind/100+1)*(no_cut[5]/100+1)*100-100
            c_per=(c_per/100+1)*(no_cut[7]/100+1)*100-100
            oneonelist.append(no_cut)
            
        oneone2=len(for_calc2)
        oneonelist2=[]
        for i in range(oneone2):
            no_cut2=np.array(setget(for_calc2[i]))             ## 2 3 4 5 7
            base_array2=base_array2+no_cut2
            b_stat2=(b_stat2/100+1)*(no_cut2[2]/100+1)*100-100
            b_phy2=(b_phy2/100+1)*(no_cut2[3]/100+1)*100-100
            b_mag2=(b_mag2/100+1)*(no_cut2[4]/100+1)*100-100
            b_ind2=(b_ind2/100+1)*(no_cut2[5]/100+1)*100-100
            c_per2=(c_per2/100+1)*(no_cut2[7]/100+1)*100-100
            oneonelist2.append(no_cut)

        if set_on.count('1441') ==1:
            if calc_wep.count('11440')!=1: ##3셋 스탯160, 영축힘지8%, 물마독3% 감소
                base_array[0]=base_array[0]-160
                base_array[1]=base_array[1]-160
                b_stat=(b_stat/100+1)/1.08*100-100
                b_phy=(b_phy/100+1)/1.03*100-100
                b_mag=(b_mag/100+1)/1.03*100-100
                b_ind=(b_ind/100+1)/1.03*100-100
        if set_on2.count('1441') ==1:
            if calc_wep2.count('11440')!=1: ##3셋 스탯160, 영축힘지8%, 물마독3% 감소
                base_array2[0]=base_array2[0]-160
                base_array2[1]=base_array2[1]-160
                b_stat2=(b_stat2/100+1)/1.08*100-100
                b_phy2=(b_phy2/100+1)/1.03*100-100
                b_mag2=(b_mag2/100+1)/1.03*100-100
                b_ind2=(b_ind2/100+1)/1.03*100-100
        #0 체정 1 지능
        #축복 2 스탯% 3 물공% 4 마공% 5 독공%
        #아포 6 고정 7 스탯%
        #8 축렙 9 포렙
        #10 아리아/보징증폭
        #11 전직패 12 보징 13 각패1 14 각패2 15 2각 16 각패3
        #17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨
        if job_name[4:7] == "세인트":
            b_base_att=lvlget('hol_b_atta')[int(base_array2[8])]#
            stat_pas0lvl_b=lvlget('pas0')[int(base_array2[11])+base_pas0_b]+lvlget('hol_pas0_1')[int(base_array2[12])]#
            stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+base_pas0_c]+lvlget('hol_pas0_1')[int(base_array[12])]
            stat_pas1lvl=lvlget('hol_pas1')[int(base_array[13])]
            stat_pas1lvl2=lvlget('hol_pas1')[int(base_array2[13])]#
            stat_pas2lvl=lvlget('hol_act2')[int(base_array[15])]
            stat_pas2lvl2=lvlget('hol_act2')[int(base_array2[15])]#
            stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
            stat_pas3lvl2=lvlget('pas3')[int(base_array2[16])]#
            stat_b=base_array2[0]+stat_pas0lvl_b+stat_pas1lvl2+stat_pas2lvl2+stat_pas3lvl2+19*base_array2[10]+int(db_preset['H6'].value)
            stat_c=base_array[0]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+19*base_array[10]+int(db_preset['H1'].value)
            b_stat_calc=int(int(lvlget('hol_b_stat')[int(base_array2[8])]*(b_stat2/100+1))*(stat_b/630+1))
            b_phy_calc=int(int(b_base_att*(b_phy2/100+1))*(stat_b/630+1))
            b_mag_calc=int(int(b_base_att*(b_mag2/100+1))*(stat_b/630+1))
            b_ind_calc=int(int(b_base_att*(b_ind2/100+1))*(stat_b/630+1))
            b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
            c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
            pas1_calc=int(lvlget('hol_pas1_out')[int(base_array[13])]+213+base_array[17])
            pas1_out=str(int(lvlget('hol_pas1_out')[int(base_array[13])]+213+base_array[17]))+"  ("+str(int(20+base_array[13]))+"렙)"
            save1=str(b_stat_calc)+"/"+str(b_average)+"   ["+str(int(stat_b))+"("+str(int(base_array2[8]))+"렙)]"
            info_stat=int(stat_c-stat_pas2lvl-528-stat_pas1lvl-213-lvlget('hol_pas0_1')[int(base_array[12])])

        else:
            if job_name[4:7] == "세라핌":
                b_value=675
                aria=1.25+0.05*base_array2[10]
            if job_name[4:7] == "헤카테":
                b_value=665
                aria=(1.20+0.05*base_array2[10])*1.15
                            
            b_base_att=lvlget('se_b_atta')[int(base_array2[8])]#
            stat_pas0lvl_b=lvlget('pas0')[int(base_array2[11])+int(base_pas0_b)]#
            stat_pas0lvl_c=lvlget('pas0')[int(base_array[11])+int(base_pas0_c)]
            stat_pas1lvl=lvlget('se_pas1')[int(base_array[13])]+base_array[18]
            stat_pas1lvl2=lvlget('se_pas1')[int(base_array2[13])]+base_array2[18]#
            stat_pas2lvl=lvlget('se_pas2')[int(base_array[14])]
            stat_pas2lvl2=lvlget('se_pas2')[int(base_array2[14])]#
            stat_pas3lvl=lvlget('pas3')[int(base_array[16])]
            stat_pas3lvl2=lvlget('pas3')[int(base_array2[16])]#
            stat_b=base_array2[1]+stat_pas0lvl_b+stat_pas1lvl2+stat_pas2lvl2+stat_pas3lvl2+int(db_preset['H6'].value)
            stat_c=base_array[1]+stat_pas0lvl_c+stat_pas1lvl+stat_pas2lvl+stat_pas3lvl+int(db_preset['H1'].value)
            b_stat_calc=int(int(lvlget('se_b_stat')[int(base_array2[8])]*(b_stat2/100+1))*(stat_b/b_value+1)*aria)
            b_phy_calc=int(int(b_base_att*(b_phy2/100+1)*(stat_b/b_value+1))*aria)
            b_mag_calc=int(int(b_base_att*(b_mag2/100+1)*(stat_b/b_value+1))*aria)
            b_ind_calc=int(int(b_base_att*(b_ind2/100+1)*(stat_b/b_value+1))*aria)
            b_average=int((b_phy_calc+b_mag_calc+b_ind_calc)/3)
            c_calc=int(int((lvlget('c_stat')[int(base_array[9])]+base_array[6])*(c_per/100+1))*(stat_c/750+1))
            pas1_calc=int(stat_pas1lvl+442)
            pas1_out=str(int(stat_pas1lvl+442))+"  ("+str(int(20+base_array[13]))+"렙)"
            save1=str(b_stat_calc)+"("+str(int(b_stat_calc/aria))+")/"+str(b_average)+"("+str(int(b_average/aria))+")\n                  ["+str(int(stat_b))+"("+str(int(base_array2[8]))+"렙)]"
            info_stat=stat_c-pas1_calc
        ##1축 2포 3합
        save2=str(c_calc)+"    ["+str(int(stat_c))+"("+str(int(base_array[9]))+"렙)]"
        save3=pas1_out
        save4=((15000+pas1_calc+c_calc+b_stat_calc)/250+1)*(2650+b_average)
        type_code='buf'
        final_list=[save1,save2,save3,save4]
        show_my_cha(calc_now,final_list,type_code,job_name,wep_name,ele_skill,cha_name,info_stat,calc_now2)

    xl.close()
    load_presetc.close()
    
def cha_image(cha_code,serv_code):
    down_url='https://img-api.neople.co.kr/df/servers/'+serv_code+'/characters/'+cha_code+'?zoom=1?apikey='+apikey
    urllib.request.urlretrieve(down_url,'my_cha.png')

def show_my_cha(equipment,final_list,type_code,job_name,wep_name,ele_skill,cha_name,info_stat,*equipment2):
    global show_cha_result
    show_cha_result=Toplevel(self)
    show_cha_result.geometry("244x402+750+320")
    show_cha_result.resizable(False,False)
    global canvas,cha_bg,cha_img
    canvas = Canvas(show_cha_result, width=246, height=404, bd=0)
    canvas.place(x=-2,y=-2)
    cha_bg=tkinter.PhotoImage(file='ext_img/bg_cha_img.png')
    canvas.create_image(123,202,image=cha_bg)
    cha_img=tkinter.PhotoImage(file='my_cha.png')
    canvas.create_image(123,70,image=cha_img)
    global image_list,image_on
    image_on={}
    for i in [11,12,13,14,15,21,22,23,31,32,33]:
        for j in equipment:
            if j[0:2] == str(i):
                image_on[str(i)]=image_list[j]
    global img11,img12,img13,img14,img15,img21,img22,img23,img31,img32,img33
    img11=canvas.create_image(57,57,image=image_on['11'])
    img12=canvas.create_image(27,87,image=image_on['12'])
    img13=canvas.create_image(27,57,image=image_on['13'])
    img14=canvas.create_image(57,87,image=image_on['14'])
    img15=canvas.create_image(27,117,image=image_on['15'])
    img21=canvas.create_image(189,57,image=image_on['21'])
    img22=canvas.create_image(219,57,image=image_on['22'])
    img23=canvas.create_image(219,87,image=image_on['23'])
    img31=canvas.create_image(189,87,image=image_on['31'])
    img32=canvas.create_image(219,117,image=image_on['32'])
    img33=canvas.create_image(189,117,image=image_on['33'])

    canvas.create_text(123,197,text="직업명= "+job_name,font=guide_font,fill='white')
    canvas.create_text(123,217,text="무기= "+wep_name,font=guide_font,fill='white')
    canvas.create_text(58,157,text=cha_name,font=guide_font,fill='white')
    


    if type_code =='deal':
        canvas.create_text(123,237,text="스킬속강(역보정율)= "+str(ele_skill)+' (-'+str(int(100*(1-1.05/(1.05+int(ele_skill)*0.0045))))+"%)",font=guide_font,fill='white')
        global style_select, creature_select, time_select
        canvas.create_text(123,257,text="딜타임설정= "+time_select.get(),font=guide_font,fill='white')
        canvas.create_text(123,277,text=style_select.get()+" / "+creature_select.get(),font=guide_font,fill='white')
        canvas.create_text(123,307,text="순데미지= "+str(int(100*final_list[0]))+"%",font=mid_font,fill='white')
        canvas.create_text(123,342,text="쿨감반영= "+str(int(100*final_list[1]))+"%",font=mid_font,fill='white')

        canvas.create_text(123,377,text="칭호/크리쳐는 메인창에서 선택하세요.",font=guide_font,fill='red')

    if type_code =='buf':
        print(equipment2)
        global image_on2
        image_on2={}
        for i in [11,12,13,14,15,21,22,23,31,32,33]:
            for j in equipment2[0]:
                if j[0:2] == str(i):
                    image_on2[str(i)]=image_list[j]
        
        canvas.create_text(123,247,text="축복="+final_list[0],font=guide_font,fill='white')
        canvas.create_text(123-17,282,text="1각="+final_list[1],font=guide_font,fill='white')
        canvas.create_text(123-44,312,text="1각패="+final_list[2],font=guide_font,fill='white')
        canvas.create_text(123,340,text="총 버프력= "+str(int(final_list[3]/10)),font=mid_font,fill='white')
        canvas.create_text(123,368,text="예상 마을스탯="+str(int(info_stat)),font=guide_font,fill='white')
        canvas.create_text(123,387,text="스탯/스킬렙을 커스텀에서 맞추세요.",font=guide_font,fill='red')

        global toggle_cha, toggle_cha_but,toggle_swi_img,toggle_swi_img2
        toggle_cha=0
        toggle_swi_img=tkinter.PhotoImage(file='ext_img/info_swi.png')
        toggle_swi_img2=tkinter.PhotoImage(file='ext_img/info_swi2.png')
        toggle_cha_but=tkinter.Button(show_cha_result,borderwidth=0,activebackground=dark_main,bg=dark_main,image=toggle_swi_img,command=toggle_cha_swi)
        toggle_cha_but.image=toggle_swi_img
    
    canvas.image = cha_bg,cha_img,image_on['11'],image_on['12'],image_on['13'],image_on['14'],image_on['15'],image_on['21'],image_on['22'],image_on['23'],image_on['31'],image_on['32'],image_on['33']
    
    if type_code =='buf':
        toggle_cha_but.place(x=155,y=142)
        tkinter.messagebox.showinfo("주의할점!","버퍼는 '반드시' 통합커스텀에서 자신에 맞는 스탯과 스킬렙을 맞춰줘야합니다.(실제 적용스탯과 여기값의 차이로 직접기입)\n굳이 API로 불러오지 않는 이유는 이후 버퍼 계산시 자신에 맞는 기준값을 정하라는 의도입니다.\n본템과 스박이 다를 경우, 이중 증폭을 하지 않는 이상 추가 스탯은 축복스탯에만 넣어두세요\n\n솔플로 버프력 측정시 솔플마부가 적용되는점 주의!(마계회합 이용하면 편함)")

def toggle_cha_swi():
    global toggle_cha,cha_bg,cha_img, toggle_cha_but,toggle_swi_img,toggle_swi_img2
    global image_on,image_on2
    if toggle_cha == 0:
        image_ont=image_on2
        toggle_cha_but['image']=toggle_swi_img2
        toggle_cha=1
    elif toggle_cha == 1:
        image_ont=image_on
        toggle_cha_but['image']=toggle_swi_img
        toggle_cha=0
    global img11,img12,img13,img14,img15,img21,img22,img23,img31,img32,img33
    canvas.itemconfig(img11,image=image_ont['11'])
    canvas.itemconfig(img12,image=image_ont['12'])
    canvas.itemconfig(img13,image=image_ont['13'])
    canvas.itemconfig(img14,image=image_ont['14'])
    canvas.itemconfig(img15,image=image_ont['15'])
    canvas.itemconfig(img21,image=image_ont['21'])
    canvas.itemconfig(img22,image=image_ont['22'])
    canvas.itemconfig(img23,image=image_ont['23'])
    canvas.itemconfig(img31,image=image_ont['31'])
    canvas.itemconfig(img32,image=image_ont['32'])
    canvas.itemconfig(img33,image=image_ont['33'])
    canvas.image = cha_bg,cha_img,image_ont['11'],image_ont['12'],image_ont['13'],image_ont['14'],image_ont['15'],image_ont['21'],image_ont['22'],image_ont['23'],image_ont['31'],image_ont['32'],image_ont['33']


def stop_calc():
    global exit_calc
    exit_calc=1
    time.sleep(1)
    exit_calc=0


    
## 내부 구조 ##
know_list=['13390150','22390240','23390450','33390750','21390340','31390540','32390650']
know_set_list=['22400150','22400250','22400350','22400450','22400550','21400640','31400750',
               '31400850','31400950','31401050','31401150','32401240','32401340','32401440']
know_jin_list=['11410100','11410110','11410120','11410130','11410140','11410150',
               '21420100','21420110','21420120','21420130','21420140','21420150',
               '33430100','33430110','33430120','33430130','33430140','33430150']
image_list={}
image_list2={}
image_list_tag={}
image_list_set={}
image_list_set2={}
for i in know_list:
    image_list[i]=eval('PhotoImage(file="image/{}n.png")'.format(i))
    image_list2[i]=eval('PhotoImage(file="image/{}f.png")'.format(i))
    image_list_tag[i]=eval('PhotoImage(file="image/{}t.png")'.format(i))
for i in know_set_list:
    image_list[i]=eval('PhotoImage(file="image/{}n.png")'.format(i))
    image_list2[i]=eval('PhotoImage(file="image/{}f.png")'.format(i))
    image_list_tag[i]=eval('PhotoImage(file="image/{}t.png")'.format(i))
    for j in range(11,34):
        try:
            image_list['n'+str(j)+i[4:6]]=eval('PhotoImage(file="image/{}.png")'.format('n'+str(j)+i[4:6]))
            image_list_tag['n'+str(j)+i[4:6]]=eval('PhotoImage(file="image/{}.png")'.format('n'+str(j)+i[4:6]+'t'))
        except:
            pass
for i in know_jin_list:
    image_list[i]=eval('PhotoImage(file="image/{}n.png")'.format(i))
    image_list2[i]=eval('PhotoImage(file="image/{}f.png")'.format(i))
    image_list_tag[i]=eval('PhotoImage(file="image/{}t.png")'.format(i))
           
for i in range(1101,3350):
    try:
        image_list[str(i)+"0"]=eval('PhotoImage(file="image/{}0n.png")'.format(i))
        image_list_tag[str(i)+"0"]=eval('PhotoImage(file="image/{}0t.png")'.format(i))
    except TclError as error:
        passss=1
    try:
        image_list2[str(i)+"0"]=eval('PhotoImage(file="image/{}0f.png")'.format(i))
    except TclError as error:
        passss=1
    try:
        image_list[str(i)+"1"]=eval('PhotoImage(file="image/{}1n.gif")'.format(i))
        image_list2[str(i)+"1"]=eval('PhotoImage(file="image/{}1f.png")'.format(i))
        image_list_tag[str(i)+"1"]=eval('PhotoImage(file="image/{}1t.png")'.format(i))
    except TclError as error:
        passss=1
for i in range(1,36):
    image_list_set[str(100+i)]=eval('PhotoImage(file="set_name/{}.png")'.format(i+100))
    image_list_set2[str(100+i)]=eval('PhotoImage(file="set_name/{}f.png")'.format(i+100))
for i in range(1,18):
    image_list_set[str(200+i)]=eval('PhotoImage(file="set_name/{}.png")'.format(i+200))
image_list['99990']=PhotoImage(file="image/99990.png")
image_list2['99990']=PhotoImage(file="image/99990.png")
image_list_tag['99990']=PhotoImage(file="image/99990.png")



    
select_perfect=tkinter.ttk.Combobox(self,values=['단품제외(빠름)','단품포함(중간)','세트필터↓(느림)','메타몽모드(중간)','메타몽모드(느림)'],width=15)
select_perfect.place(x=145,y=11)
select_perfect.set('단품포함(중간)')
select_speed_img=PhotoImage(file="ext_img/select_speed.png")
tkinter.Button(self,command=guide_speed,image=select_speed_img,borderwidth=0,activebackground=dark_main,bg=dark_main).place(x=29,y=7)
reset_img=PhotoImage(file="ext_img/reset.png")
tkinter.Button(self,command=reset,image=reset_img,borderwidth=0,activebackground=dark_main,bg=dark_main).place(x=302,y=476)

wep_list=[]
for i in range(0,75):
    wep_list.append(name_one[str(i+111001)][1])
    
wep_image=PhotoImage(file="ext_img/wep.png")
wep_g=tkinter.Label(self,image=wep_image,borderwidth=0,activebackground=dark_main,bg=dark_main)
wep_g.place(x=29,y=55)
wep_select=tkinter.ttk.Combobox(self,width=30,values=wep_list)
wep_select.place(x=110,y=60)
wep_select.set('흑천의 주인')

time_select=tkinter.ttk.Combobox(self,width=13,values=['20초(각성비중↑)','60초(각성비중↓)'])
time_select.set('20초(각성비중↑)')
time_select.place(x=390-17,y=220+52)
jobup_select=tkinter.ttk.Combobox(self,width=13,values=jobs)
jobup_select.set('가이아(2각)')
jobup_select.place(x=390-17,y=190+52)
style_list=['증뎀칭호','추뎀칭호','기타(직접비교)']
style_select=tkinter.ttk.Combobox(self,width=13,values=style_list)
style_select.set('추뎀칭호')
style_select.place(x=390-17,y=250+52)
creature_list=['모공크리쳐','크증크리쳐','기타(직접비교)']
creature_select=tkinter.ttk.Combobox(self,width=13,values=creature_list)
creature_select.set('크증크리쳐')
creature_select.place(x=390-17,y=280+52)
req_cool=tkinter.ttk.Combobox(self,width=13,values=['X(순데미지)','O(쿨감보정)'])
req_cool.set('X(순데미지)')
req_cool.place(x=390-17,y=310+52)

calc_img=PhotoImage(file="ext_img/calc.png")
select_all=tkinter.Button(self,image=calc_img,borderwidth=0,activebackground=dark_main,command=calc_thread,bg=dark_main)
select_all.place(x=390-35,y=7)
stop_img=PhotoImage(file="ext_img/stop.png")
tkinter.Button(self,image=stop_img,borderwidth=0,activebackground=dark_main,command=stop_calc,bg=dark_main).place(x=390-35,y=62)

timeline_img=PhotoImage(file="ext_img/timeline.png")
select_custom=tkinter.Button(self,image=timeline_img,borderwidth=0,activebackground=dark_main,command=timeline_select,bg=dark_sub)
select_custom.place(x=345+165,y=340-100)
custom_img=PhotoImage(file="ext_img/custom.png")
select_custom2=tkinter.Button(self,image=custom_img,borderwidth=0,activebackground=dark_main,command=lambda:costum(0),bg=dark_sub)
select_custom2.place(x=435+165,y=340-100)

save_select=tkinter.ttk.Combobox(self,width=8,values=save_name_list)
save_select.place(x=345+165,y=410-100);save_select.set(save_name_list[0])
save_img=PhotoImage(file="ext_img/SAVE.png")
save=tkinter.Button(self,image=save_img,borderwidth=0,activebackground=dark_main,command=save_checklist,bg=dark_sub)
save.place(x=345+165,y=440-100)
load_img=PhotoImage(file="ext_img/LOAD.png")
load=tkinter.Button(self,image=load_img,borderwidth=0,activebackground=dark_main,command=load_checklist,bg=dark_sub)
load.place(x=435+165,y=440-100)
change_name_img=PhotoImage(file="ext_img/name_change.png")
change_list_but=tkinter.Button(self,image=change_name_img,borderwidth=0,activebackground=dark_main,command=change_list_name,bg=dark_sub)
change_list_but.place(x=435+165,y=405-100)
calc_me_img=PhotoImage(file="ext_img/calc_me.png")
calc_me=tkinter.Button(self,image=calc_me_img,borderwidth=0,activebackground=dark_main,command=lambda:cha_select(jobs),bg=dark_sub)
calc_me.place(x=300,y=400)


show_count=tkinter.Label(self,font=guide_font,fg="white",bg=dark_sub)
show_count.place(x=490,y=40)
showcon=show_count.configure
show_state=tkinter.Label(self,text="상태 표시 칸",font=guide_font,fg="white",bg=dark_sub)
show_state.place(x=490,y=20)
showsta=show_state.configure

show_count2=tkinter.Label(self,font=guide_font,fg="white",bg=dark_sub)
show_count2.place(x=430,y=480)
showcon2=show_count2.configure

set101=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['101'],command=lambda:click_set(101));set101.place(x=29,y=100)
set102=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['102'],command=lambda:click_set(102));set102.place(x=29,y=130)
set103=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['103'],command=lambda:click_set(103));set103.place(x=29,y=160)
set104=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['104'],command=lambda:click_set(104));set104.place(x=29,y=190)
set105=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['105'],command=lambda:click_set(105));set105.place(x=29,y=220)
set106=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['106'],command=lambda:click_set(106));set106.place(x=29,y=250)
set107=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['107'],command=lambda:click_set(107));set107.place(x=29,y=280)
set108=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['108'],command=lambda:click_set(108));set108.place(x=29,y=310)
set109=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['109'],command=lambda:click_set(109));set109.place(x=29,y=340)
set110=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['110'],command=lambda:click_set(110));set110.place(x=29,y=370)
set111=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['111'],command=lambda:click_set(111));set111.place(x=29,y=400)
set112=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['112'],command=lambda:click_set(112));set112.place(x=29,y=430)
set113=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['113'],command=lambda:click_set(113));set113.place(x=29,y=460)
set114=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['114'],command=lambda:click_set(114));set114.place(x=29,y=490)
set115=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['115'],command=lambda:click_set(115));set115.place(x=29,y=520) ##
set116=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['116'],command=lambda:click_set(116));set116.place(x=320-33,y=100)
set117=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['117'],command=lambda:click_set(117));set117.place(x=320-33,y=130)
set118=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['118'],command=lambda:click_set(118));set118.place(x=320-33,y=160)
set119=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['119'],command=lambda:click_set(119));set119.place(x=320-33,y=190) ##
set120=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['120'],command=lambda:click_set(120));set120.place(x=500-17,y=100)
set121=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['121'],command=lambda:click_set(121));set121.place(x=500-17,y=130)
set122=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['122'],command=lambda:click_set(122));set122.place(x=500-17,y=160)
set123=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['123'],command=lambda:click_set(123));set123.place(x=500-17,y=190) ##
set128=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['128'],command=lambda:click_set(128));set128.place(x=29,y=570)
set129=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['129'],command=lambda:click_set(129));set129.place(x=29,y=600)
set130=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['130'],command=lambda:click_set(130));set130.place(x=29,y=630)
set131=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['131'],command=lambda:click_set(131));set131.place(x=29,y=660) ##
set124=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['124'],command=lambda:click_set(124));set124.place(x=225,y=570)
set125=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['125'],command=lambda:click_set(125));set125.place(x=225,y=600)
set126=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['126'],command=lambda:click_set(126));set126.place(x=225,y=630)
set127=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['127'],command=lambda:click_set(127));set127.place(x=225,y=660) ##
set132=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['132'],command=lambda:click_set(132));set132.place(x=421,y=570)
set133=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['133'],command=lambda:click_set(133));set133.place(x=421,y=600)
set134=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['134'],command=lambda:click_set(134));set134.place(x=421,y=630)
set135=tkinter.Button(self,bg=dark_main,borderwidth=0,activebackground=dark_main,image=image_list_set2['135'],command=lambda:click_set(135));set135.place(x=421,y=660) ##




##지혜의 산물
know_set_list=['22400150','22400250','22400350','22400450','22400550','21400640','31400750',
               '31400850','31400950','31401050','31401150','32401240','32401340','32401440']
know_jin_list=['11410100','11410110','11410120','11410130','11410140','11410150',
               '21420100','21420110','21420120','21420130','21420140','21420150',
               '33430100','33430110','33430120','33430130','33430140','33430150']
select_item['tg13390150']=0;select_item['tg22400150']=0;select_item['tg31400850']=0
select_item['tg22390240']=0;select_item['tg22400250']=0;select_item['tg31400950']=0
select_item['tg23390450']=0;select_item['tg22400350']=0;select_item['tg31401050']=0
select_item['tg33390750']=0;select_item['tg22400450']=0;select_item['tg31401150']=0
select_item['tg21390340']=0;select_item['tg22400550']=0;select_item['tg32401240']=0
select_item['tg31390540']=0;select_item['tg21400640']=0;select_item['tg32401340']=0
select_item['tg32390650']=0;select_item['tg31400750']=0;select_item['tg32401440']=0

select_item['tg11410100']=0;select_item['tg11410110']=0;select_item['tg11410120']=0
select_item['tg11410130']=0;select_item['tg11410140']=0;select_item['tg11410150']=0
select_item['tg21420100']=0;select_item['tg21420110']=0;select_item['tg21420120']=0
select_item['tg21420130']=0;select_item['tg21420140']=0;select_item['tg21420150']=0
select_item['tg33430100']=0;select_item['tg33430110']=0;select_item['tg33430120']=0
select_item['tg33430130']=0;select_item['tg33430140']=0;select_item['tg33430150']=0
def know_epic():
    global select_item,select_13390150,select_22390240,select_23390450,select_33390750,select_21390340,select_31390540,select_32390650
    global select_22400150,select_22400250,select_22400350,select_22400450,select_22400550,select_21400640,select_31400750
    global select_31400850,select_31400950,select_31401050,select_31401150,select_32401240,select_32401340,select_32401440
    global select_11410100,select_11410110,select_11410120,select_11410130,select_11410140,select_11410150
    global select_21420100,select_21420110,select_21420120,select_21420130,select_21420140,select_21420150
    global select_33430100,select_33430110,select_33430120,select_33430130,select_33430140,select_33430150
    global know_window
    try:
        know_window.destroy()
    except:
        pass
    know_window=tkinter.Toplevel(self)
    know_window.attributes("-topmost", True) 
    know_window.geometry("545x405+750+20")
    know_window.resizable(False, False)
    know_window.configure(bg=dark_main)
    know_image_list={}
    for i in know_list+know_set_list+know_jin_list:
        if select_item['tg'+i]==0:
            know_image_list[i]=image_list2[i]
        else:
            know_image_list[i]=image_list[i]
    select_13390150=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['13390150'],command=lambda:click_equipment(13390150))
    select_13390150.place(x=303-290,y=20)
    select_22390240=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['22390240'],command=lambda:click_equipment(22390240))
    select_22390240.place(x=333-290,y=20)
    select_23390450=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['23390450'],command=lambda:click_equipment(23390450))
    select_23390450.place(x=363-290,y=20)
    select_33390750=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['33390750'],command=lambda:click_equipment(33390750))
    select_33390750.place(x=393-290,y=20)
    select_21390340=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['21390340'],command=lambda:click_equipment(21390340))
    select_21390340.place(x=424-290,y=20)
    select_31390540=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['31390540'],command=lambda:click_equipment(31390540))
    select_31390540.place(x=454-290,y=20)
    select_32390650=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['32390650'],command=lambda:click_equipment(32390650))
    select_32390650.place(x=484-290,y=20)

    tkinter.Label(know_window,bg=dark_main,image=image_list_set['201']).place(x=303-290,y=70)
    select_22400150=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['22400150'],command=lambda:click_equipment(22400150))
    select_22400150.place(x=303-290+63,y=70)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['202']).place(x=303-290,y=70+40)
    select_22400250=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['22400250'],command=lambda:click_equipment(22400250))
    select_22400250.place(x=303-290+63,y=70+40)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['203']).place(x=303-290,y=70+80)
    select_22400350=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['22400350'],command=lambda:click_equipment(22400350))
    select_22400350.place(x=303-290+63,y=70+80)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['204']).place(x=303-290,y=70+120)
    select_22400450=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['22400450'],command=lambda:click_equipment(22400450))
    select_22400450.place(x=303-290+63,y=70+120)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['205']).place(x=303-290,y=70+160)
    select_22400550=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['22400550'],command=lambda:click_equipment(22400550))
    select_22400550.place(x=303-290+63,y=70+160)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['206']).place(x=303-290,y=70+200)
    select_21400640=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['21400640'],command=lambda:click_equipment(21400640))
    select_21400640.place(x=303-290+63,y=70+200)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['207']).place(x=303-290,y=70+240)
    select_31400750=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['31400750'],command=lambda:click_equipment(31400750))
    select_31400750.place(x=303-290+63,y=70+240)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['208']).place(x=120,y=70)
    select_31400850=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['31400850'],command=lambda:click_equipment(31400850))
    select_31400850.place(x=120+63,y=70)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['209']).place(x=120,y=70+40)
    select_31400950=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['31400950'],command=lambda:click_equipment(31400950))
    select_31400950.place(x=120+63,y=70+40)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['210']).place(x=120,y=70+80)
    select_31401050=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['31401050'],command=lambda:click_equipment(31401050))
    select_31401050.place(x=120+63,y=70+80)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['211']).place(x=120,y=70+120)
    select_31401150=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['31401150'],command=lambda:click_equipment(31401150))
    select_31401150.place(x=120+63,y=70+120)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['212']).place(x=120,y=70+160)
    select_32401240=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['32401240'],command=lambda:click_equipment(32401240))
    select_32401240.place(x=120+63,y=70+160)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['213']).place(x=120,y=70+200)
    select_32401340=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['32401340'],command=lambda:click_equipment(32401340))
    select_32401340.place(x=120+63,y=70+200)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['214']).place(x=120,y=70+240)
    select_32401440=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['32401440'],command=lambda:click_equipment(32401440))
    select_32401440.place(x=120+63,y=70+240)

    tkinter.Label(know_window,bg=dark_main,image=image_list_set['215']).place(x=250,y=69+200)
    select_11410100=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['11410100'],command=lambda:click_equipment(11410100))
    select_11410100.place(x=280+45,y=70+200)
    select_11410110=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['11410110'],command=lambda:click_equipment(11410110))
    select_11410110.place(x=280+75,y=70+200)
    select_11410120=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['11410120'],command=lambda:click_equipment(11410120))
    select_11410120.place(x=280+105,y=70+200)
    select_11410130=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['11410130'],command=lambda:click_equipment(11410130))
    select_11410130.place(x=280+135,y=70+200)
    select_11410140=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['11410140'],command=lambda:click_equipment(11410140))
    select_11410140.place(x=280+165,y=70+200)
    select_11410150=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['11410150'],command=lambda:click_equipment(11410150))
    select_11410150.place(x=280+195,y=70+200)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['216']).place(x=250,y=69+240)
    select_21420100=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['21420100'],command=lambda:click_equipment(21420100))
    select_21420100.place(x=280+45,y=70+240)
    select_21420110=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['21420110'],command=lambda:click_equipment(21420110))
    select_21420110.place(x=280+75,y=70+240)
    select_21420120=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['21420120'],command=lambda:click_equipment(21420120))
    select_21420120.place(x=280+105,y=70+240)
    select_21420130=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['21420130'],command=lambda:click_equipment(21420130))
    select_21420130.place(x=280+135,y=70+240)
    select_21420140=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['21420140'],command=lambda:click_equipment(21420140))
    select_21420140.place(x=280+165,y=70+240)
    select_21420150=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['21420150'],command=lambda:click_equipment(21420150))
    select_21420150.place(x=280+195,y=70+240)
    tkinter.Label(know_window,bg=dark_main,image=image_list_set['217']).place(x=250,y=69+280)
    select_33430100=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['33430100'],command=lambda:click_equipment(33430100))
    select_33430100.place(x=280+45,y=70+280)
    select_33430110=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['33430110'],command=lambda:click_equipment(33430110))
    select_33430110.place(x=280+75,y=70+280)
    select_33430120=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['33430120'],command=lambda:click_equipment(33430120))
    select_33430120.place(x=280+105,y=70+280)
    select_33430130=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['33430130'],command=lambda:click_equipment(33430130))
    select_33430130.place(x=280+135,y=70+280)
    select_33430140=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['33430140'],command=lambda:click_equipment(33430140))
    select_33430140.place(x=280+165,y=70+280)
    select_33430150=tkinter.Button(know_window,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=know_image_list['33430150'],command=lambda:click_equipment(33430150))
    select_33430150.place(x=280+195,y=70+280)

    tkinter.Label(know_window,bg=dark_main,font=mid_font,fg='white',text="<주의할 점>").place(x=315,y=20)
    tkinter.Label(know_window,bg=dark_main,fg='white',text=("세트 산물은 증/크증이 겹치는 경우가 있습니다.\n칭호/크리쳐 선택에 유의하세요.\n(중복안되게 계산식 처리 해놨음)\n\n"
                                                            +"불마/엘드 셋의 '마딜 전용'옵션은\n따로 구분되어 계산되지 않습니다.\n알아서 빼주세요.\n\n"
                                                            +"스탯 옵션은 버프+가호 받은 기준입니다.\n가호 미적용 스탯이 많은 산물 특성상,\n수련방 솔플 효율과 굉장히 다를수 있습니다.")).place(x=250,y=70)

know_image=PhotoImage(file="set_name/know_name.png")
tkinter.Button(self,bg=dark_main,image=know_image,command=know_epic).place(x=302,y=520)

##디폴트 변경
default_legend=1
default_chawon=0
default_old=0
default_img1n=PhotoImage(file="ext_img/default1n.png")
default_img1f=PhotoImage(file="ext_img/default1f.png")
default_img2n=PhotoImage(file="ext_img/default2n.png")
default_img2f=PhotoImage(file="ext_img/default2f.png")
default_img3n=PhotoImage(file="ext_img/default3n.png")
default_img3f=PhotoImage(file="ext_img/default3f.png")
def change_default(value):
    global default_legend,default_chawon,default_old
    if value==0:
        default_legend=1
        default_chawon=0
        default_old=0
        select_default_lengend['image']=default_img1n
        select_default_chawon['image']=default_img2f
        select_default_old['image']=default_img3f
    elif value==1:
        default_legend=0
        default_chawon=1
        default_old=0
        select_default_lengend['image']=default_img1f
        select_default_chawon['image']=default_img2n
        select_default_old['image']=default_img3f
    elif value==2:
        default_legend=0
        default_chawon=0
        default_old=1
        select_default_lengend['image']=default_img1f
        select_default_chawon['image']=default_img2f
        select_default_old['image']=default_img3n
select_default_lengend=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=default_img1n,command=lambda:change_default(0))
select_default_lengend.place(x=492+15,y=521)
select_default_chawon=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=default_img2f,command=lambda:change_default(1))
select_default_chawon.place(x=522+15,y=521)
select_default_old=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=default_img3f,command=lambda:change_default(2))
select_default_old.place(x=552+15,y=521)

default_tag_img=PhotoImage(file="ext_img/default_tag.png")
tkinter.Label(self,bg=dark_main,image=default_tag_img).place(x=431,y=520)


##상의
select_item['tg11010']=0;select_11010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11010'],command=lambda:click_equipment(11010))
select_11010.place(x=100,y=100)
select_item['tg11011']=0;select_11011=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11011'],command=lambda:click_equipment(11011))
select_11011.place(x=130,y=100)
select_item['tg11020']=0;select_11020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11020'],command=lambda:click_equipment(11020))
select_11020.place(x=100,y=130)
select_item['tg11021']=0;select_11021=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11021'],command=lambda:click_equipment(11021))
select_11021.place(x=130,y=130)
select_item['tg11030']=0;select_11030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11030'],command=lambda:click_equipment(11030))
select_11030.place(x=100,y=160)
select_item['tg11031']=0;select_11031=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11031'],command=lambda:click_equipment(11031))
select_11031.place(x=130,y=160)
select_item['tg11040']=0;select_11040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11040'],command=lambda:click_equipment(11040))
select_11040.place(x=100,y=190)
select_item['tg11041']=0;select_11041=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11041'],command=lambda:click_equipment(11041))
select_11041.place(x=130,y=190)
select_item['tg11050']=0;select_11050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11050'],command=lambda:click_equipment(11050))
select_11050.place(x=100,y=220)
select_item['tg11051']=0;select_11051=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11051'],command=lambda:click_equipment(11051))
select_11051.place(x=130,y=220)
select_item['tg11060']=0;select_11060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11060'],command=lambda:click_equipment(11060))
select_11060.place(x=100,y=250)
select_item['tg11061']=0;select_11061=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11061'],command=lambda:click_equipment(11061))
select_11061.place(x=130,y=250)
select_item['tg11070']=0;select_11070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11070'],command=lambda:click_equipment(11070))
select_11070.place(x=100,y=280)
select_item['tg11071']=0;select_11071=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11071'],command=lambda:click_equipment(11071))
select_11071.place(x=130,y=280)
select_item['tg11080']=0;select_11080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11080'],command=lambda:click_equipment(11080))
select_11080.place(x=100,y=310)
select_item['tg11081']=0;select_11081=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11081'],command=lambda:click_equipment(11081))
select_11081.place(x=130,y=310)
select_item['tg11090']=0;select_11090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11090'],command=lambda:click_equipment(11090))
select_11090.place(x=100,y=340)
select_item['tg11091']=0;select_11091=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11091'],command=lambda:click_equipment(11091))
select_11091.place(x=130,y=340)
select_item['tg11100']=0;select_11100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11100'],command=lambda:click_equipment(11100))
select_11100.place(x=100,y=370)
select_item['tg11101']=0;select_11101=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11101'],command=lambda:click_equipment(11101))
select_11101.place(x=130,y=370)
select_item['tg11110']=0;select_11110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11110'],command=lambda:click_equipment(11110))
select_11110.place(x=100,y=400)
select_item['tg11111']=0;select_11111=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11111'],command=lambda:click_equipment(11111))
select_11111.place(x=130,y=400)
select_item['tg11120']=0;select_11120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11120'],command=lambda:click_equipment(11120))
select_11120.place(x=100,y=430)
select_item['tg11121']=0;select_11121=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11121'],command=lambda:click_equipment(11121))
select_11121.place(x=130,y=430)
select_item['tg11130']=0;select_11130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11130'],command=lambda:click_equipment(11130))
select_11130.place(x=100,y=460)
select_item['tg11131']=0;select_11131=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11131'],command=lambda:click_equipment(11131))
select_11131.place(x=130,y=460)
select_item['tg11140']=0;select_11140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11140'],command=lambda:click_equipment(11140))
select_11140.place(x=100,y=490)
select_item['tg11141']=0;select_11141=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11141'],command=lambda:click_equipment(11141))
select_11141.place(x=130,y=490)
select_item['tg11150']=0;select_11150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11150'],command=lambda:click_equipment(11150))
select_11150.place(x=100,y=520)
select_item['tg11151']=0;select_11151=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11151'],command=lambda:click_equipment(11151))
select_11151.place(x=130,y=520)

select_item['tg11280']=0;select_11280=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11280'],command=lambda:click_equipment(11280))
select_11280.place(x=100,y=570)
select_item['tg11281']=0;select_11281=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11281'],command=lambda:click_equipment(11281))
select_11281.place(x=130,y=570)
select_item['tg11290']=0;select_11290=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11290'],command=lambda:click_equipment(11290))
select_11290.place(x=100,y=600)
select_item['tg11291']=0;select_11291=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11291'],command=lambda:click_equipment(11291))
select_11291.place(x=130,y=600)
select_item['tg11300']=0;select_11300=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11300'],command=lambda:click_equipment(11300))
select_11300.place(x=100,y=630)
select_item['tg11301']=0;select_11301=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11301'],command=lambda:click_equipment(11301))
select_11301.place(x=130,y=630)
select_item['tg11310']=0;select_11310=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11310'],command=lambda:click_equipment(11310))
select_11310.place(x=100,y=660)
select_item['tg11311']=0;select_11311=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['11311'],command=lambda:click_equipment(11311))
select_11311.place(x=130,y=660)
##하의
select_item['tg12010']=0;select_12010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12010'],command=lambda:click_equipment(12010))
select_12010.place(x=161,y=100)
select_item['tg12020']=0;select_12020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12020'],command=lambda:click_equipment(12020))
select_12020.place(x=161,y=130)
select_item['tg12030']=0;select_12030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12030'],command=lambda:click_equipment(12030))
select_12030.place(x=161,y=160)
select_item['tg12040']=0;select_12040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12040'],command=lambda:click_equipment(12040))
select_12040.place(x=161,y=190)
select_item['tg12050']=0;select_12050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12050'],command=lambda:click_equipment(12050))
select_12050.place(x=161,y=220)
select_item['tg12060']=0;select_12060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12060'],command=lambda:click_equipment(12060))
select_12060.place(x=161,y=250)
select_item['tg12070']=0;select_12070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12070'],command=lambda:click_equipment(12070))
select_12070.place(x=161,y=280)
select_item['tg12080']=0;select_12080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12080'],command=lambda:click_equipment(12080))
select_12080.place(x=161,y=310)
select_item['tg12090']=0;select_12090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12090'],command=lambda:click_equipment(12090))
select_12090.place(x=161,y=340)
select_item['tg12100']=0;select_12100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12100'],command=lambda:click_equipment(12100))
select_12100.place(x=161,y=370)
select_item['tg12110']=0;select_12110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12110'],command=lambda:click_equipment(12110))
select_12110.place(x=161,y=400)
select_item['tg12120']=0;select_12120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12120'],command=lambda:click_equipment(12120))
select_12120.place(x=161,y=430)
select_item['tg12130']=0;select_12130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12130'],command=lambda:click_equipment(12130))
select_12130.place(x=161,y=460)
select_item['tg12140']=0;select_12140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12140'],command=lambda:click_equipment(12140))
select_12140.place(x=161,y=490)
select_item['tg12150']=0;select_12150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12150'],command=lambda:click_equipment(12150))
select_12150.place(x=161,y=520)
select_item['tg12240']=0;select_12240=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12240'],command=lambda:click_equipment(12240))
select_12240.place(x=296,y=570)
select_item['tg12250']=0;select_12250=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12250'],command=lambda:click_equipment(12250))
select_12250.place(x=296,y=600)
select_item['tg12260']=0;select_12260=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12260'],command=lambda:click_equipment(12260))
select_12260.place(x=296,y=630)
select_item['tg12270']=0;select_12270=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['12270'],command=lambda:click_equipment(12270))
select_12270.place(x=296,y=660)
##어깨
select_item['tg13010']=0;select_13010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13010'],command=lambda:click_equipment(13010))
select_13010.place(x=192,y=100)
select_item['tg13020']=0;select_13020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13020'],command=lambda:click_equipment(13020))
select_13020.place(x=192,y=130)
select_item['tg13030']=0;select_13030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13030'],command=lambda:click_equipment(13030))
select_13030.place(x=192,y=160)
select_item['tg13040']=0;select_13040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13040'],command=lambda:click_equipment(13040))
select_13040.place(x=192,y=190)
select_item['tg13050']=0;select_13050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13050'],command=lambda:click_equipment(13050))
select_13050.place(x=192,y=220)
select_item['tg13060']=0;select_13060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13060'],command=lambda:click_equipment(13060))
select_13060.place(x=192,y=250)
select_item['tg13070']=0;select_13070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13070'],command=lambda:click_equipment(13070))
select_13070.place(x=192,y=280)
select_item['tg13080']=0;select_13080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13080'],command=lambda:click_equipment(13080))
select_13080.place(x=192,y=310)
select_item['tg13090']=0;select_13090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13090'],command=lambda:click_equipment(13090))
select_13090.place(x=192,y=340)
select_item['tg13100']=0;select_13100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13100'],command=lambda:click_equipment(13100))
select_13100.place(x=192,y=370)
select_item['tg13110']=0;select_13110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13110'],command=lambda:click_equipment(13110))
select_13110.place(x=192,y=400)
select_item['tg13120']=0;select_13120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13120'],command=lambda:click_equipment(13120))
select_13120.place(x=192,y=430)
select_item['tg13130']=0;select_13130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13130'],command=lambda:click_equipment(13130))
select_13130.place(x=192,y=460)
select_item['tg13140']=0;select_13140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13140'],command=lambda:click_equipment(13140))
select_13140.place(x=192,y=490)
select_item['tg13150']=0;select_13150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['13150'],command=lambda:click_equipment(13150))
select_13150.place(x=192,y=520)
##벨트
select_item['tg14010']=0;select_14010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14010'],command=lambda:click_equipment(14010))
select_14010.place(x=223,y=100)
select_item['tg14020']=0;select_14020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14020'],command=lambda:click_equipment(14020))
select_14020.place(x=223,y=130)
select_item['tg14030']=0;select_14030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14030'],command=lambda:click_equipment(14030))
select_14030.place(x=223,y=160)
select_item['tg14040']=0;select_14040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14040'],command=lambda:click_equipment(14040))
select_14040.place(x=223,y=190)
select_item['tg14050']=0;select_14050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14050'],command=lambda:click_equipment(14050))
select_14050.place(x=223,y=220)
select_item['tg14060']=0;select_14060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14060'],command=lambda:click_equipment(14060))
select_14060.place(x=223,y=250)
select_item['tg14070']=0;select_14070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14070'],command=lambda:click_equipment(14070))
select_14070.place(x=223,y=280)
select_item['tg14080']=0;select_14080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14080'],command=lambda:click_equipment(14080))
select_14080.place(x=223,y=310)
select_item['tg14090']=0;select_14090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14090'],command=lambda:click_equipment(14090))
select_14090.place(x=223,y=340)
select_item['tg14100']=0;select_14100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14100'],command=lambda:click_equipment(14100))
select_14100.place(x=223,y=370)
select_item['tg14110']=0;select_14110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14110'],command=lambda:click_equipment(14110))
select_14110.place(x=223,y=400)
select_item['tg14120']=0;select_14120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14120'],command=lambda:click_equipment(14120))
select_14120.place(x=223,y=430)
select_item['tg14130']=0;select_14130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14130'],command=lambda:click_equipment(14130))
select_14130.place(x=223,y=460)
select_item['tg14140']=0;select_14140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14140'],command=lambda:click_equipment(14140))
select_14140.place(x=223,y=490)
select_item['tg14150']=0;select_14150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['14150'],command=lambda:click_equipment(14150))
select_14150.place(x=223,y=520)
##신발
select_item['tg15010']=0;select_15010=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15010'],command=lambda:click_equipment(15010))
select_15010.place(x=254,y=100)
select_item['tg15020']=0;select_15020=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15020'],command=lambda:click_equipment(15020))
select_15020.place(x=254,y=130)
select_item['tg15030']=0;select_15030=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15030'],command=lambda:click_equipment(15030))
select_15030.place(x=254,y=160)
select_item['tg15040']=0;select_15040=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15040'],command=lambda:click_equipment(15040))
select_15040.place(x=254,y=190)
select_item['tg15050']=0;select_15050=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15050'],command=lambda:click_equipment(15050))
select_15050.place(x=254,y=220)
select_item['tg15060']=0;select_15060=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15060'],command=lambda:click_equipment(15060))
select_15060.place(x=254,y=250)
select_item['tg15070']=0;select_15070=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15070'],command=lambda:click_equipment(15070))
select_15070.place(x=254,y=280)
select_item['tg15080']=0;select_15080=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15080'],command=lambda:click_equipment(15080))
select_15080.place(x=254,y=310)
select_item['tg15090']=0;select_15090=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15090'],command=lambda:click_equipment(15090))
select_15090.place(x=254,y=340)
select_item['tg15100']=0;select_15100=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15100'],command=lambda:click_equipment(15100))
select_15100.place(x=254,y=370)
select_item['tg15110']=0;select_15110=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15110'],command=lambda:click_equipment(15110))
select_15110.place(x=254,y=400)
select_item['tg15120']=0;select_15120=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15120'],command=lambda:click_equipment(15120))
select_15120.place(x=254,y=430)
select_item['tg15130']=0;select_15130=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15130'],command=lambda:click_equipment(15130))
select_15130.place(x=254,y=460)
select_item['tg15140']=0;select_15140=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15140'],command=lambda:click_equipment(15140))
select_15140.place(x=254,y=490)
select_item['tg15150']=0;select_15150=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15150'],command=lambda:click_equipment(15150))
select_15150.place(x=254,y=520)
select_item['tg15320']=0;select_15320=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15320'],command=lambda:click_equipment(15320))
select_15320.place(x=492,y=570)
select_item['tg15330']=0;select_15330=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15330'],command=lambda:click_equipment(15330))
select_15330.place(x=492,y=600)
select_item['tg15340']=0;select_15340=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15340'],command=lambda:click_equipment(15340))
select_15340.place(x=492,y=630)
select_item['tg15350']=0;select_15350=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['15350'],command=lambda:click_equipment(15350))
select_15350.place(x=492,y=660)
##팔찌
select_item['tg21160']=0;select_21160=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21160'],command=lambda:click_equipment(21160))
select_21160.place(x=370-12,y=100)
select_item['tg21161']=0;select_21161=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21161'],command=lambda:click_equipment(21161))
select_21161.place(x=370-12+30,y=100)
select_item['tg21170']=0;select_21170=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21170'],command=lambda:click_equipment(21170))
select_21170.place(x=370-12,y=130)
select_item['tg21171']=0;select_21171=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21171'],command=lambda:click_equipment(21171))
select_21171.place(x=370-12+30,y=130)
select_item['tg21180']=0;select_21180=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21180'],command=lambda:click_equipment(21180))
select_21180.place(x=370-12,y=160)
select_item['tg21181']=0;select_21181=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21181'],command=lambda:click_equipment(21181))
select_21181.place(x=370-12+30,y=160)
select_item['tg21190']=0;select_21190=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21190'],command=lambda:click_equipment(21190))
select_21190.place(x=370-12,y=190)
select_item['tg21191']=0;select_21191=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21191'],command=lambda:click_equipment(21191))
select_21191.place(x=370-12+30,y=190)
select_item['tg21240']=0;select_21240=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21240'],command=lambda:click_equipment(21240))
select_21240.place(x=327,y=570)
select_item['tg21241']=0;select_21241=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21241'],command=lambda:click_equipment(21241))
select_21241.place(x=357,y=570)
select_item['tg21250']=0;select_21250=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21250'],command=lambda:click_equipment(21250))
select_21250.place(x=327,y=600)
select_item['tg21251']=0;select_21251=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21251'],command=lambda:click_equipment(21251))
select_21251.place(x=357,y=600)
select_item['tg21260']=0;select_21260=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21260'],command=lambda:click_equipment(21260))
select_21260.place(x=327,y=630)
select_item['tg21261']=0;select_21261=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21261'],command=lambda:click_equipment(21261))
select_21261.place(x=357,y=630)
select_item['tg21270']=0;select_21270=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21270'],command=lambda:click_equipment(21270))
select_21270.place(x=327,y=660)
select_item['tg21271']=0;select_21271=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['21271'],command=lambda:click_equipment(21271))
select_21271.place(x=357,y=660)
##목걸이
select_item['tg22160']=0;select_22160=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22160'],command=lambda:click_equipment(22160))
select_22160.place(x=419,y=100)
select_item['tg22170']=0;select_22170=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22170'],command=lambda:click_equipment(22170))
select_22170.place(x=419,y=130)
select_item['tg22180']=0;select_22180=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22180'],command=lambda:click_equipment(22180))
select_22180.place(x=419,y=160)
select_item['tg22190']=0;select_22190=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22190'],command=lambda:click_equipment(22190))
select_22190.place(x=419,y=190)
select_item['tg22280']=0;select_22280=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22280'],command=lambda:click_equipment(22280))
select_22280.place(x=161,y=570)
select_item['tg22290']=0;select_22290=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22290'],command=lambda:click_equipment(22290))
select_22290.place(x=161,y=600)
select_item['tg22300']=0;select_22300=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22300'],command=lambda:click_equipment(22300))
select_22300.place(x=161,y=630)
select_item['tg22310']=0;select_22310=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['22310'],command=lambda:click_equipment(22310))
select_22310.place(x=161,y=660)
##반지
select_item['tg23160']=0;select_23160=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23160'],command=lambda:click_equipment(23160))
select_23160.place(x=450,y=100)
select_item['tg23170']=0;select_23170=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23170'],command=lambda:click_equipment(23170))
select_23170.place(x=450,y=130)
select_item['tg23180']=0;select_23180=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23180'],command=lambda:click_equipment(23180))
select_23180.place(x=450,y=160)
select_item['tg23190']=0;select_23190=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23190'],command=lambda:click_equipment(23190))
select_23190.place(x=450,y=190)
select_item['tg23320']=0;select_23320=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23320'],command=lambda:click_equipment(23320))
select_23320.place(x=523,y=570)
select_item['tg23330']=0;select_23330=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23330'],command=lambda:click_equipment(23330))
select_23330.place(x=523,y=600)
select_item['tg23340']=0;select_23340=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23340'],command=lambda:click_equipment(23340))
select_23340.place(x=523,y=630)
select_item['tg23350']=0;select_23350=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['23350'],command=lambda:click_equipment(23350))
select_23350.place(x=523,y=660)
##보조장비
select_item['tg31200']=0;select_31200=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31200'],command=lambda:click_equipment(31200))
select_31200.place(x=554,y=100)
select_item['tg31210']=0;select_31210=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31210'],command=lambda:click_equipment(31210))
select_31210.place(x=554,y=130)
select_item['tg31220']=0;select_31220=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31220'],command=lambda:click_equipment(31220))
select_31220.place(x=554,y=160)
select_item['tg31230']=0;select_31230=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31230'],command=lambda:click_equipment(31230))
select_31230.place(x=554,y=190)
select_item['tg31280']=0;select_31280=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31280'],command=lambda:click_equipment(31280))
select_31280.place(x=192,y=570)
select_item['tg31290']=0;select_31290=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31290'],command=lambda:click_equipment(31290))
select_31290.place(x=192,y=600)
select_item['tg31300']=0;select_31300=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31300'],command=lambda:click_equipment(31300))
select_31300.place(x=192,y=630)
select_item['tg31310']=0;select_31310=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['31310'],command=lambda:click_equipment(31310))
select_31310.place(x=192,y=660)
##마법석
select_item['tg32200']=0;select_32200=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32200'],command=lambda:click_equipment(32200))
select_32200.place(x=585,y=100)
select_item['tg32210']=0;select_32210=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32210'],command=lambda:click_equipment(32210))
select_32210.place(x=585,y=130)
select_item['tg32220']=0;select_32220=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32220'],command=lambda:click_equipment(32220))
select_32220.place(x=585,y=160)
select_item['tg32230']=0;select_32230=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32230'],command=lambda:click_equipment(32230))
select_32230.place(x=585,y=190)
select_item['tg32240']=0;select_32240=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32240'],command=lambda:click_equipment(32240))
select_32240.place(x=388,y=570)
select_item['tg32250']=0;select_32250=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32250'],command=lambda:click_equipment(32250))
select_32250.place(x=388,y=600)
select_item['tg32260']=0;select_32260=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32260'],command=lambda:click_equipment(32260))
select_32260.place(x=388,y=630)
select_item['tg32270']=0;select_32270=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['32270'],command=lambda:click_equipment(32270))
select_32270.place(x=388,y=660)
##귀걸이
select_item['tg33200']=0;select_33200=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33200'],command=lambda:click_equipment(33200))
select_33200.place(x=616,y=100)
select_item['tg33201']=0;select_33201=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33201'],command=lambda:click_equipment(33201))
select_33201.place(x=646,y=100)
select_item['tg33210']=0;select_33210=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33210'],command=lambda:click_equipment(33210))
select_33210.place(x=616,y=130)
select_item['tg33211']=0;select_33211=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33211'],command=lambda:click_equipment(33211))
select_33211.place(x=646,y=130)
select_item['tg33220']=0;select_33220=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33220'],command=lambda:click_equipment(33220))
select_33220.place(x=616,y=160)
select_item['tg33221']=0;select_33221=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33221'],command=lambda:click_equipment(33221))
select_33221.place(x=646,y=160)
select_item['tg33230']=0;select_33230=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33230'],command=lambda:click_equipment(33230))
select_33230.place(x=616,y=190)
select_item['tg33231']=0;select_33231=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33231'],command=lambda:click_equipment(33231))
select_33231.place(x=646,y=190)
select_item['tg33320']=0;select_33320=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33320'],command=lambda:click_equipment(33320))
select_33320.place(x=554,y=570)
select_item['tg33321']=0;select_33321=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33321'],command=lambda:click_equipment(33321))
select_33321.place(x=584,y=570)
select_item['tg33330']=0;select_33330=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33330'],command=lambda:click_equipment(33330))
select_33330.place(x=554,y=600)
select_item['tg33331']=0;select_33331=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33331'],command=lambda:click_equipment(33331))
select_33331.place(x=584,y=600)
select_item['tg33340']=0;select_33340=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33340'],command=lambda:click_equipment(33340))
select_33340.place(x=554,y=630)
select_item['tg33341']=0;select_33341=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33341'],command=lambda:click_equipment(33341))
select_33341.place(x=584,y=630)
select_item['tg33350']=0;select_33350=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33350'],command=lambda:click_equipment(33350))
select_33350.place(x=554,y=660)
select_item['tg33351']=0;select_33351=tkinter.Button(self,relief='flat',borderwidth=0,activebackground=dark_main,bg=dark_main,image=image_list2['33351'],command=lambda:click_equipment(33351))
select_33351.place(x=584,y=660)


def donate():
    webbrowser.open('https://twip.kr/dawnclass16')
donate_image=PhotoImage(file='ext_img/donate.png')
donate_bt=tkinter.Button(self,image=donate_image, command=donate,borderwidth=0,bg=dark_main,activebackground=dark_main)
donate_bt.place(x=625,y=550-28)
def dunfaoff():
    webbrowser.open('https://dunfaoff.com/')
dunfaoff_image=PhotoImage(file='ext_img/dunfaoff.png')
dunfaoff_url=tkinter.Button(self,image=dunfaoff_image, command=dunfaoff,borderwidth=0,bg=dark_main,activebackground=dark_main)
dunfaoff_url.place(x=535,y=410)

def blog():
    webbrowser.open('https://blog.naver.com/dawnclass16/221837654941')
blog_image=PhotoImage(file='ext_img/blog.png')
blog_url=tkinter.Button(self,image=blog_image, command=blog,borderwidth=0,bg=dark_main,activebackground=dark_main)
blog_url.place(x=615,y=410)
    
def hamjung():
    tkinter.messagebox.showinfo("제작자 크레딧","총제작자=Dawnclass(새벽반)\n이미지/그래픽=경철부동산\n직업/버퍼DB=대략볼록할철\n서버제공=던파오프\n기타조언=히든 도비 4,5,6호\n\n오류 제보는 블로그 덧글이나 던조 쪽지로")
maker_image=PhotoImage(file='ext_img/maker.png')
maker=tkinter.Button(self,image=maker_image, command=hamjung,borderwidth=0,bg=dark_main,activebackground=dark_main)
version=tkinter.Label(self,text='V '+str(now_version)+'\n'+ver_time,font=guide_font,fg="white",bg=dark_main)

maker.place(x=625,y=590)
version.place(x=630,y=650)

try:
    now_version_num=int(now_version[0]+now_version[2]+now_version[4])
    html = urllib.request.urlopen("https://drive.google.com/open?id=1p8ZdzW_NzGKHHOtfPTuZSr1YgSEVtYCj")
    bsObject = BeautifulSoup(html, "html.parser")
    for meta in bsObject.head.find_all('meta'):
        if meta.get('content').count('zip')==1:
            net_version=str(meta.get('content'))[-9:-4]
            print('최신 업데이트 버전='+net_version)
    net_version_num=int(net_version[0]+net_version[2]+net_version[4])
    if now_version_num < net_version_num:
        ask_update=tkinter.messagebox.askquestion('업데이트',"최신버전이 존재합니다. 이동하시겠습니까?")
        if ask_update == 'yes':
            webbrowser.open('https://drive.google.com/open?id=1p8ZdzW_NzGKHHOtfPTuZSr1YgSEVtYCj')
    else:
        pass
except:
    print("업데이트 체크 실패(네트워크 오류)")

if auto_custom==1:
    costum(1)

if __name__ == "__main__":
    update_thread()
self.mainloop()



self.quit()

